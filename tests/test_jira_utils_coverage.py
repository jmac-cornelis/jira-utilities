import argparse
import csv
import json
import sys
from pathlib import Path
from types import SimpleNamespace
from typing import Any, cast
from unittest.mock import MagicMock

import openpyxl
import pytest

import jira_utils


class DummyResponse:
    def __init__(self, status_code=200, payload=None, text='', headers=None):
        self.status_code = status_code
        self._payload = payload if payload is not None else {}
        self.text = text
        self.headers = headers or {}

    def json(self):
        return self._payload


def _silence_cli(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(jira_utils, 'output', lambda *_args, **_kwargs: None)
    monkeypatch.setattr(jira_utils, 'show_jql', lambda _jql: None)
    monkeypatch.setattr(jira_utils.time, 'sleep', lambda _seconds: None)


def _mock_auth(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        jira_utils,
        'get_jira_credentials',
        lambda: ('engineer@cornelisnetworks.com', 'token-123'),
    )


def test_user_resolver_caches_results(mock_jira):
    resolver = jira_utils.UserResolver(jira=mock_jira)

    first = resolver.resolve('John Doe', project_key='STL')
    second = resolver.resolve('John Doe', project_key='STL')

    assert first == 'acct-jdoe'
    assert second == 'acct-jdoe'
    assert mock_jira.search_assignable_users_for_issues.call_count == 1


def test_user_resolver_handles_connection_error(monkeypatch: pytest.MonkeyPatch):
    resolver = jira_utils.UserResolver(jira=None)
    monkeypatch.setattr(jira_utils, 'get_connection', lambda: (_ for _ in ()).throw(RuntimeError('down')))

    result = resolver.resolve('Jane Doe', project_key='STL')

    assert result is None
    report = resolver.get_resolution_report()
    assert report[0]['status'] == 'error'


def test_user_resolver_resolve_plan_and_report(mock_jira):
    resolver = jira_utils.UserResolver(jira=mock_jira)
    plan = {
        'project_key': 'STL',
        'epics': [
            {
                'assignee': 'John Doe',
                'stories': [{'assignee': 'Jane Smith'}, {'assignee': 'Unknown Person'}],
            }
        ],
    }

    resolved = resolver.resolve_plan(plan)

    assert resolved['epics'][0]['assignee'] == 'acct-jdoe'
    assert resolved['epics'][0]['stories'][0]['assignee'] == 'acct-jane'
    assert resolved['epics'][0]['stories'][1]['assignee'] is None
    lines = resolver.format_resolution_report()
    assert any('Assignee Resolution:' in line for line in lines)


def test_get_project_fields_collects_create_edit_and_transition_data(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)

    mock_jira.createmeta.return_value = {
        'projects': [
            {
                'issuetypes': [
                    {
                        'name': 'Bug',
                        'fields': {
                            'summary': {
                                'name': 'Summary',
                                'required': True,
                                'schema': {'type': 'string'},
                            },
                            'description': {
                                'name': 'Description',
                                'required': False,
                                'schema': {'type': 'string'},
                            },
                        },
                    }
                ]
            }
        ]
    }
    mock_jira.search_issues.return_value = [SimpleNamespace(key='STL-123')]
    mock_jira.editmeta.return_value = {
        'fields': {
            'summary': {'name': 'Summary', 'required': True, 'schema': {'type': 'string'}},
            'labels': {'name': 'Labels', 'required': False, 'schema': {'type': 'array'}},
        }
    }
    mock_jira.transitions.return_value = [
        {
            'id': '31',
            'name': 'Done',
            'to': {'name': 'Closed'},
            'fields': {
                'resolution': {
                    'name': 'Resolution',
                    'required': True,
                    'schema': {'type': 'string'},
                }
            },
        }
    ]

    data = jira_utils.get_project_fields(mock_jira, 'STL', ['bug'])

    assert data['project'] == 'STL'
    issue_type = data['issue_types'][0]
    assert issue_type['name'] == 'Bug'
    assert any(field['key'] == 'summary' for field in issue_type['create_fields'])
    assert any(field['key'] == 'labels' for field in issue_type['edit_fields'])
    assert issue_type['transitions'][0]['name'] == 'Done'


def test_get_project_components_with_activity_filter_and_dump(
    mock_jira,
    issue_factory,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    mock_jira.project_components.return_value = [
        SimpleNamespace(
            name='Fabric',
            id='2001',
            lead=SimpleNamespace(displayName='Lead A', emailAddress='lead.a@cornelisnetworks.com'),
            description='Fabric features',
        ),
        SimpleNamespace(
            name='Driver',
            id='2002',
            lead=None,
            description='Driver features',
        ),
    ]

    active_issue = issue_factory(key='STL-777', components=['Fabric'])

    def _fake_post(_url: str, auth=None, headers=None, json=None):
        return DummyResponse(status_code=200, payload={'issues': [active_issue], 'nextPageToken': None})

    monkeypatch.setattr(jira_utils.requests, 'post', _fake_post)

    out_prefix = tmp_path / 'components_dump'
    jira_utils.get_project_components(
        mock_jira,
        'STL',
        date_filter='week',
        dump_file=str(out_prefix),
        dump_format='json',
    )

    dumped = json.loads((tmp_path / 'components_dump.json').read_text(encoding='utf-8'))
    assert len(dumped) == 1
    assert dumped[0]['name'] == 'Fabric'
    assert dumped[0]['ticket_count'] == 1


def test_get_children_hierarchy_recurses_with_pagination_and_dump(
    mock_jira,
    issue_factory,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    root = issue_factory(key='STL-1', summary='Root')
    child_a = issue_factory(key='STL-2', summary='Child A')
    child_b = issue_factory(key='STL-3', summary='Child B')

    issues_by_key = {'STL-1': root, 'STL-2': child_a, 'STL-3': child_b}
    mock_jira.issue.side_effect = lambda key: SimpleNamespace(raw=issues_by_key[key])

    responses = [
        DummyResponse(status_code=429, headers={'Retry-After': '0'}),
        DummyResponse(status_code=200, payload={'issues': [child_a], 'nextPageToken': 'next-1'}),
        DummyResponse(status_code=200, payload={'issues': [child_b], 'nextPageToken': None}),
        DummyResponse(status_code=200, payload={'issues': [], 'nextPageToken': None}),
        DummyResponse(status_code=200, payload={'issues': [], 'nextPageToken': None}),
    ]

    def _fake_post(_url: str, auth=None, headers=None, json=None):
        return responses.pop(0)

    monkeypatch.setattr(jira_utils.requests, 'post', _fake_post)

    out_prefix = tmp_path / 'children'
    jira_utils.get_children_hierarchy(
        mock_jira,
        project_key='STL',
        root_key='STL-1',
        limit=5,
        dump_file=str(out_prefix),
        dump_format='csv',
        table_format='indented',
    )

    content = (tmp_path / 'children.csv').read_text(encoding='utf-8')
    assert 'Depth 0' in content
    assert 'Depth 1' in content
    assert 'STL-1' in content


def test_get_related_issues_recursive_dump_includes_relation_metadata(
    mock_jira,
    issue_factory,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    root = issue_factory(
        key='STL-10',
        summary='Root',
        issuelinks=[
            {
                'type': {'name': 'blocks', 'outward': 'blocks', 'inward': 'is blocked by'},
                'outwardIssue': {'key': 'STL-20'},
            }
        ],
    )
    linked = issue_factory(key='STL-20', summary='Linked', issuelinks=[])
    child = issue_factory(key='STL-30', summary='Child', issuelinks=[])

    issues_by_key = {'STL-10': root, 'STL-20': linked, 'STL-30': child}
    mock_jira.issue.side_effect = lambda key, fields=None: SimpleNamespace(raw=issues_by_key[key])

    monkeypatch.setattr(
        jira_utils.requests,
        'post',
        lambda _url, auth=None, headers=None, json=None: DummyResponse(
            status_code=200,
            payload={'issues': [{'key': 'STL-30'}], 'nextPageToken': None},
        ),
    )

    out_prefix = tmp_path / 'related'
    jira_utils.get_related_issues(
        mock_jira,
        project_key='STL',
        root_key='STL-10',
        hierarchy=1,
        limit=10,
        dump_file=str(out_prefix),
        dump_format='json',
        table_format='flat',
    )

    dumped = json.loads((tmp_path / 'related.json').read_text(encoding='utf-8'))
    by_key = {row['key']: row for row in dumped}
    assert by_key['STL-20']['link_via'] == 'blocks'
    assert by_key['STL-20']['relation'] == 'link'
    assert by_key['STL-30']['relation'] == 'child'


def test_get_related_issues_direct_mode(mock_jira, issue_factory, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    root = issue_factory(
        key='STL-50',
        summary='Root',
        issuelinks=[
            {
                'type': {'name': 'relates', 'outward': 'relates to', 'inward': 'relates to'},
                'outwardIssue': {'key': 'STL-51'},
            }
        ],
    )
    linked = issue_factory(key='STL-51', summary='Linked', issuelinks=[])

    issues_by_key = {'STL-50': root, 'STL-51': linked}
    mock_jira.issue.side_effect = lambda key, fields=None: SimpleNamespace(raw=issues_by_key[key])

    monkeypatch.setattr(
        jira_utils.requests,
        'post',
        lambda _url, auth=None, headers=None, json=None: DummyResponse(
            status_code=200,
            payload={'issues': [], 'nextPageToken': None},
        ),
    )

    jira_utils.get_related_issues(mock_jira, project_key='STL', root_key='STL-50', hierarchy=None)


def test_get_tickets_handles_rate_limit_pagination_and_limit(mock_jira, issue_factory, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    i1 = issue_factory(key='STL-100')
    i2 = issue_factory(key='STL-101')
    i3 = issue_factory(key='STL-102')
    i4 = issue_factory(key='STL-103')

    responses = [
        DummyResponse(status_code=429, headers={'Retry-After': '0'}),
        DummyResponse(status_code=200, payload={'issues': [i1, i2], 'nextPageToken': 'page-2'}),
        DummyResponse(status_code=200, payload={'issues': [i3, i4], 'nextPageToken': None}),
    ]
    monkeypatch.setattr(jira_utils.requests, 'post', lambda *_args, **_kwargs: responses.pop(0))

    issues = jira_utils.get_tickets(mock_jira, 'STL', limit=3)

    assert [issue['key'] for issue in issues] == ['STL-100', 'STL-101', 'STL-102']


def test_get_release_tickets_delegates_for_wildcard(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)

    delegated = MagicMock(return_value=['delegated'])
    monkeypatch.setattr(jira_utils, 'get_releases_tickets', delegated)

    result = jira_utils.get_release_tickets(mock_jira, 'STL', '12.*', limit=25)

    assert result == ['delegated']
    delegated.assert_called_once()


def test_get_releases_tickets_handles_no_matching_releases(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)

    mock_jira.project_versions.return_value = [SimpleNamespace(name='11.0.0'), SimpleNamespace(name='10.9.0')]

    result = jira_utils.get_releases_tickets(mock_jira, 'STL', '12.*')

    assert result is None


def test_get_no_release_tickets_and_dump_json(
    mock_jira,
    issue_factory,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    issue = issue_factory(key='STL-900', fix_versions=[])
    monkeypatch.setattr(
        jira_utils.requests,
        'post',
        lambda *_args, **_kwargs: DummyResponse(
            status_code=200,
            payload={'issues': [issue], 'nextPageToken': None},
        ),
    )

    out_prefix = tmp_path / 'no_release'
    issues = jira_utils.get_no_release_tickets(
        mock_jira,
        'STL',
        issue_types=['Bug'],
        statuses=['Open'],
        limit=10,
        dump_file=str(out_prefix),
        dump_format='json',
    )

    assert len(issues) == 1
    assert (tmp_path / 'no_release.json').exists()


def test_get_ticket_totals_builds_status_include_exclude_jql(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    monkeypatch.setattr(
        jira_utils.requests,
        'post',
        lambda *_args, **_kwargs: DummyResponse(status_code=200, payload={'count': 42}),
    )

    result = jira_utils.get_ticket_totals(
        mock_jira,
        'STL',
        issue_types=['Bug'],
        statuses=['Open', '^Closed'],
        date_filter='week',
    )

    assert result['count'] == 42
    assert 'status IN ("Open")' in result['jql']
    assert 'status NOT IN ("Closed")' in result['jql']


def test_dump_tickets_to_file_latest_comments_json(issue_factory, tmp_path: Path):
    issue = issue_factory(
        key='STL-444',
        comments=[
            {
                'id': '1',
                'author': {'displayName': 'A'},
                'created': '2026-01-01T10:00:00.000+0000',
                'updated': '2026-01-01T10:00:00.000+0000',
                'body': {
                    'type': 'doc',
                    'content': [
                        {'type': 'paragraph', 'content': [{'type': 'text', 'text': 'older'}]}
                    ],
                },
            },
            {
                'id': '2',
                'author': {'displayName': 'B'},
                'created': '2026-01-02T08:30:00.000+0000',
                'updated': '2026-01-02T08:30:00.000+0000',
                'body': {
                    'type': 'doc',
                    'content': [
                        {'type': 'paragraph', 'content': [{'type': 'text', 'text': 'newer'}]}
                    ],
                },
            },
        ],
    )

    path = jira_utils.dump_tickets_to_file(
        [issue],
        str(tmp_path / 'tickets_comments'),
        'json',
        include_comments='latest',
    )

    rows = json.loads(Path(path).read_text(encoding='utf-8'))
    assert len(rows[0]['comments']) == 1
    assert rows[0]['comments'][0]['body'] == 'newer'


def test_dump_tickets_to_file_excel_and_write_excel(issue_factory, tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)

    issue = issue_factory(key='STL-500', summary='Excel ticket')
    output_path = jira_utils.dump_tickets_to_file([issue], str(tmp_path / 'tickets_excel'), 'excel')

    wb = openpyxl.load_workbook(output_path)
    ws = wb['Tickets']
    assert ws.cell(row=2, column=1).value == 'STL-500'
    assert ws.cell(row=2, column=1).hyperlink is not None
    assert ws.freeze_panes == 'A2'
    wb.close()


def test_write_excel_indented_layout_uses_depth_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)

    rows = [
        {'key': 'STL-1', 'summary': 'root', 'status': 'Open', 'priority': 'P1-Critical', 'depth': 0},
        {'key': 'STL-2', 'summary': 'child', 'status': 'Open', 'priority': 'P1-Critical', 'depth': 1},
    ]
    out_file = tmp_path / 'indented.xlsx'

    jira_utils._write_excel(rows, str(out_file), table_format='indented')

    wb = openpyxl.load_workbook(out_file)
    ws = wb['Tickets']
    headers = [ws.cell(row=1, column=1).value, ws.cell(row=1, column=2).value]
    assert headers == ['Depth 0', 'Depth 1']
    assert ws.cell(row=2, column=1).value == 'STL-1'
    assert ws.cell(row=3, column=2).value == 'STL-2'
    wb.close()


def test_load_tickets_from_csv_supports_missing_extension(tmp_path: Path):
    csv_path = tmp_path / 'input.csv'
    csv_path.write_text('key,summary\nSTL-1,Sample\n', encoding='utf-8')

    tickets = jira_utils.load_tickets_from_csv(str(tmp_path / 'input'))

    assert tickets == [{'key': 'STL-1', 'summary': 'Sample'}]


def test_load_tickets_from_csv_requires_key_column(tmp_path: Path):
    csv_path = tmp_path / 'bad.csv'
    csv_path.write_text('summary\nSample\n', encoding='utf-8')

    with pytest.raises(ValueError):
        jira_utils.load_tickets_from_csv(str(csv_path))


def test_create_ticket_dry_run_resolves_assignee(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)

    resolver = MagicMock()
    resolver.resolve.return_value = 'acct-123'
    monkeypatch.setattr(jira_utils, 'get_user_resolver', lambda: resolver)

    jira_utils.create_ticket(
        mock_jira,
        project_key='STL',
        summary='Create me',
        issue_type='Bug',
        description='description text',
        assignee='John Doe',
        components=['Fabric'],
        fix_versions=['12.1.0'],
        labels=['triage'],
        parent_key='STL-1',
        product_family=['CN5000'],
        dry_run=True,
    )

    resolver.resolve.assert_called_once_with('John Doe', project_key='STL')
    mock_jira.create_issue.assert_not_called()


def test_create_ticket_retries_with_adjusted_fields(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)

    resolver = MagicMock()
    resolver.resolve.return_value = None
    monkeypatch.setattr(jira_utils, 'get_user_resolver', lambda: resolver)

    mock_jira.create_issue.side_effect = [
        Exception('Description is required customfield_28434 customfield_28382'),
        SimpleNamespace(key='STL-222'),
    ]

    jira_utils.create_ticket(
        mock_jira,
        project_key='STL',
        summary='Retry me',
        issue_type='Task',
        product_family=['CN5000'],
        dry_run=False,
    )

    assert mock_jira.create_issue.call_count == 2
    second_fields = mock_jira.create_issue.call_args_list[1].kwargs['fields']
    assert 'description' in second_fields
    assert 'customfield_28382' in second_fields


def test_bulk_update_tickets_executes_all_operations(
    mock_jira,
    fake_issue_resource_factory,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)

    input_file = tmp_path / 'bulk_update.csv'
    input_file.write_text('key\nSTL-123\n', encoding='utf-8')

    issue_resource = fake_issue_resource_factory(key='STL-123')
    mock_jira.issue.side_effect = None
    mock_jira.issue.return_value = issue_resource
    mock_jira.transitions.return_value = [{'id': '41', 'name': 'Closed'}]

    jira_utils.bulk_update_tickets(
        mock_jira,
        str(input_file),
        set_release='12.1.1',
        remove_release=True,
        transition='Closed',
        assign='unassigned',
        dry_run=False,
    )

    assert {'fixVersions': [{'name': '12.1.1'}]} in issue_resource.updated_fields
    assert {'fixVersions': []} in issue_resource.updated_fields
    mock_jira.transition_issue.assert_called_once()
    mock_jira.assign_issue.assert_called_once_with(issue_resource, None)


def test_bulk_delete_tickets_force_execute_with_retry_and_errors(
    mock_jira,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    input_file = tmp_path / 'bulk_delete.csv'
    input_file.write_text('key\nSTL-1\nSTL-2\n', encoding='utf-8')

    calls = []
    responses = [
        DummyResponse(status_code=429, headers={'Retry-After': '0'}),
        DummyResponse(status_code=204),
        DummyResponse(status_code=400, text='cannot delete'),
    ]

    def _fake_delete(url: str, auth=None, headers=None, params=None):
        calls.append((url, params))
        return responses.pop(0)

    monkeypatch.setattr(jira_utils.requests, 'delete', _fake_delete)

    jira_utils.bulk_delete_tickets(
        mock_jira,
        str(input_file),
        delete_subtasks=True,
        dry_run=False,
        force=True,
    )

    assert calls[0][1] == {'deleteSubtasks': 'true'}
    assert len(calls) == 3


def test_bulk_delete_tickets_aborts_when_confirmation_fails(
    mock_jira,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    input_file = tmp_path / 'bulk_delete_abort.csv'
    input_file.write_text('key\nSTL-1\n', encoding='utf-8')

    monkeypatch.setattr('builtins.input', lambda _prompt='': 'NOPE')
    delete_mock = MagicMock()
    monkeypatch.setattr(jira_utils.requests, 'delete', delete_mock)

    jira_utils.bulk_delete_tickets(mock_jira, str(input_file), dry_run=False, force=False)

    delete_mock.assert_not_called()


def test_list_filters_paginates_and_filters_owner_me(monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    responses = [
        DummyResponse(status_code=429, headers={'Retry-After': '0'}),
        DummyResponse(
            status_code=200,
            payload={
                'values': [
                    {
                        'id': '1',
                        'name': 'Mine',
                        'jql': 'project = STL',
                        'favourite': True,
                        'owner': {'displayName': 'Me', 'emailAddress': 'engineer@cornelisnetworks.com'},
                    }
                ],
                'total': 2,
            },
        ),
        DummyResponse(
            status_code=200,
            payload={
                'values': [
                    {
                        'id': '2',
                        'name': 'Other',
                        'jql': 'project = CN',
                        'favourite': False,
                        'owner': {'displayName': 'Other', 'emailAddress': 'other@cornelisnetworks.com'},
                    }
                ],
                'total': 2,
            },
        ),
    ]

    monkeypatch.setattr(jira_utils.requests, 'get', lambda *_args, **_kwargs: responses.pop(0))

    rows = jira_utils.list_filters(jira=MagicMock(), owner='me', favourite_only=False)

    assert len(rows) == 1
    assert rows[0]['id'] == '1'


def test_list_filters_favourite_endpoint(monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    monkeypatch.setattr(
        jira_utils.requests,
        'get',
        lambda _url, auth=None, headers=None, params=None: DummyResponse(
            status_code=200,
            payload=[
                {
                    'id': '9',
                    'name': 'Fav',
                    'jql': 'project = STL',
                    'favourite': True,
                    'owner': {'displayName': 'Analyst', 'emailAddress': 'analyst@cornelisnetworks.com'},
                }
            ],
        ),
    )

    rows = jira_utils.list_filters(jira=MagicMock(), owner='analyst', favourite_only=True)

    assert len(rows) == 1
    assert rows[0]['name'] == 'Fav'


def test_get_filter_returns_details(monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    payload = {
        'id': '123',
        'name': 'Open Bugs',
        'favourite': True,
        'description': 'Useful filter',
        'jql': 'project = STL',
        'viewUrl': 'https://example/view',
        'searchUrl': 'https://example/search',
        'owner': {'displayName': 'Owner'},
        'sharePermissions': [
            {'type': 'global'},
            {'type': 'project', 'project': {'name': 'Storage', 'key': 'STL'}},
            {'type': 'group', 'group': {'name': 'jira-users'}},
            {'type': 'authenticated'},
            {'type': 'custom'},
        ],
    }
    monkeypatch.setattr(
        jira_utils.requests,
        'get',
        lambda *_args, **_kwargs: DummyResponse(status_code=200, payload=payload),
    )

    result = jira_utils.get_filter(jira=MagicMock(), filter_id='123')

    assert result['id'] == '123'
    assert result['name'] == 'Open Bugs'


def test_run_filter_fetches_jql_and_executes_query(
    issue_factory,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    issue = issue_factory(key='STL-808')

    def _fake_get(url: str, auth=None, headers=None, params=None):
        assert url.endswith('/rest/api/3/filter/123')
        return DummyResponse(status_code=200, payload={'name': 'By ID', 'jql': 'project = "STL"'})

    monkeypatch.setattr(jira_utils.requests, 'get', _fake_get)
    monkeypatch.setattr(
        jira_utils.requests,
        'post',
        lambda *_args, **_kwargs: DummyResponse(
            status_code=200,
            payload={'issues': [issue], 'nextPageToken': None},
        ),
    )

    out_prefix = tmp_path / 'run_filter'
    issues = jira_utils.run_filter(
        jira=MagicMock(),
        filter_id='123',
        limit=10,
        dump_file=str(out_prefix),
        dump_format='csv',
    )

    assert len(issues) == 1
    assert (tmp_path / 'run_filter.csv').exists()


def test_list_dashboards_paginates(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    responses = [
        DummyResponse(status_code=200, payload={'values': [{'id': '1', 'name': 'A'}], 'total': 2}),
        DummyResponse(status_code=200, payload={'values': [{'id': '2', 'name': 'B'}], 'total': 2}),
    ]

    def _fake_get(url: str, auth=None, headers=None, params=None):
        assert url.endswith('/rest/api/3/dashboard/search')
        return responses.pop(0)

    monkeypatch.setattr(jira_utils.requests, 'get', _fake_get)

    jira_utils.list_dashboards(mock_jira, owner='me', shared=True)

    assert len(responses) == 0


def test_dashboard_crud_operations(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    def _fake_get(url: str, auth=None, headers=None, params=None):
        if url.endswith('/rest/api/3/dashboard/101'):
            return DummyResponse(
                status_code=200,
                payload={
                    'id': '101',
                    'name': 'Ops Dashboard',
                    'description': 'desc',
                    'isFavourite': True,
                    'view': 'https://example/dashboard/101',
                    'owner': {'displayName': 'Owner'},
                    'sharePermissions': [{'type': 'global'}],
                },
            )
        raise AssertionError(f'Unexpected GET URL: {url}')

    def _fake_post(url: str, auth=None, headers=None, json=None):
        payload = dict(json or {})
        if url.endswith('/rest/api/3/dashboard'):
            return DummyResponse(status_code=201, payload={'id': '201', 'name': payload.get('name', ''), 'view': 'v'})
        if url.endswith('/rest/api/3/dashboard/101/copy'):
            return DummyResponse(status_code=201, payload={'id': '301', 'name': payload.get('name', ''), 'view': 'copy'})
        raise AssertionError(f'Unexpected POST URL: {url}')

    def _fake_put(url: str, auth=None, headers=None, json=None):
        assert url.endswith('/rest/api/3/dashboard/101')
        payload = dict(json or {})
        return DummyResponse(status_code=200, payload={'id': '101', **payload})

    def _fake_delete(url: str, auth=None, headers=None):
        assert url.endswith('/rest/api/3/dashboard/101')
        return DummyResponse(status_code=204)

    monkeypatch.setattr(jira_utils.requests, 'get', _fake_get)
    monkeypatch.setattr(jira_utils.requests, 'post', _fake_post)
    monkeypatch.setattr(jira_utils.requests, 'put', _fake_put)
    monkeypatch.setattr(jira_utils.requests, 'delete', _fake_delete)

    jira_utils.get_dashboard(mock_jira, '101')
    jira_utils.create_dashboard(mock_jira, name='New Board', description='new', share_permissions='[{"type":"global"}]')
    jira_utils.update_dashboard(mock_jira, dashboard_id='101', name='Updated', description='changed')
    jira_utils.delete_dashboard(mock_jira, dashboard_id='101', force=True)
    jira_utils.copy_dashboard(mock_jira, dashboard_id='101', name='Copy Name', description='copy')


def test_create_dashboard_rejects_invalid_share_permissions_json(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    with pytest.raises(jira_utils.JiraDashboardError):
        jira_utils.create_dashboard(mock_jira, name='Broken', share_permissions='not-json')


def test_gadget_operations(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    def _fake_get(url: str, auth=None, headers=None, params=None):
        if url.endswith('/rest/api/3/dashboard/55'):
            return DummyResponse(status_code=200, payload={'name': 'Dashboard 55'})
        if url.endswith('/rest/api/3/dashboard/55/gadget'):
            return DummyResponse(
                status_code=200,
                payload={
                    'gadgets': [
                        {
                            'id': '501',
                            'moduleKey': 'mod',
                            'title': 'G1',
                            'position': {'row': 0, 'column': 1},
                            'color': 'blue',
                        }
                    ]
                },
            )
        raise AssertionError(f'Unexpected GET URL: {url}')

    def _fake_post(url: str, auth=None, headers=None, json=None):
        assert url.endswith('/rest/api/3/dashboard/55/gadget')
        payload = dict(json or {})
        return DummyResponse(
            status_code=201,
            payload={
                'id': '999',
                'moduleKey': payload.get('moduleKey', ''),
                'title': 'New Gadget',
                'position': payload.get('position', {'row': 0, 'column': 0}),
                'color': payload.get('color', 'blue'),
            },
        )

    def _fake_put(url: str, auth=None, headers=None, json=None):
        assert url.endswith('/rest/api/3/dashboard/55/gadget/999')
        return DummyResponse(status_code=200, payload={})

    def _fake_delete(url: str, auth=None, headers=None):
        assert url.endswith('/rest/api/3/dashboard/55/gadget/999')
        return DummyResponse(status_code=204)

    monkeypatch.setattr(jira_utils.requests, 'get', _fake_get)
    monkeypatch.setattr(jira_utils.requests, 'post', _fake_post)
    monkeypatch.setattr(jira_utils.requests, 'put', _fake_put)
    monkeypatch.setattr(jira_utils.requests, 'delete', _fake_delete)

    jira_utils.list_gadgets(mock_jira, dashboard_id='55')
    jira_utils.add_gadget(
        mock_jira,
        dashboard_id='55',
        module_key='mod-key',
        position='1,2',
        color='green',
        properties='{"a": 1}',
    )
    jira_utils.update_gadget(mock_jira, dashboard_id='55', gadget_id='999', position='0,0', color='red')
    jira_utils.remove_gadget(mock_jira, dashboard_id='55', gadget_id='999')


def test_add_gadget_validation_errors(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    with pytest.raises(jira_utils.JiraDashboardError):
        jira_utils.add_gadget(mock_jira, dashboard_id='1', module_key='mod', position='bad-position')

    with pytest.raises(jira_utils.JiraDashboardError):
        jira_utils.add_gadget(mock_jira, dashboard_id='1', module_key='mod', color='orange')


def test_update_gadget_requires_changes(mock_jira, monkeypatch: pytest.MonkeyPatch):
    _silence_cli(monkeypatch)
    _mock_auth(monkeypatch)

    with pytest.raises(jira_utils.JiraDashboardError):
        jira_utils.update_gadget(mock_jira, dashboard_id='1', gadget_id='2')


def test_handle_args_create_ticket_json_and_comments(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    ticket_json = {
        'project_key': 'STL',
        'summary': 'Ticket from JSON',
        'issue_type': 'Task',
        'description': 'Desc from file',
        'components': ['Fabric'],
        'fix_versions': ['12.1.0'],
        'labels': ['triage'],
    }
    json_path = tmp_path / 'ticket.json'
    json_path.write_text(json.dumps(ticket_json), encoding='utf-8')

    monkeypatch.setattr(
        sys,
        'argv',
        [
            'jira_utils.py',
            '--project',
            'STL',
            '--get-tickets',
            '--get-comments',
            'latest',
            '--create-ticket',
            str(json_path),
            '--no-formatting',
        ],
    )

    args = jira_utils.handle_args()

    assert args.project == 'STL'
    assert args.summary == 'Ticket from JSON'
    assert args.issue_type == 'Task'
    assert args.dump_format == 'json'
    assert args.dump_file == 'tickets_with_comments'
    assert jira_utils._include_comments == 'latest'
    assert jira_utils._no_formatting is True


@pytest.mark.parametrize(
    'argv',
    [
        ['jira_utils.py', '--owner', 'me', '--list'],
        ['jira_utils.py', '--favourite', '--list'],
        ['jira_utils.py', '--bulk-update', '--input-file', 'x.csv'],
        ['jira_utils.py', '--position', '0,1', '--list'],
    ],
)
def test_handle_args_invalid_combinations(monkeypatch: pytest.MonkeyPatch, argv):
    monkeypatch.setattr(sys, 'argv', argv)

    with pytest.raises(SystemExit):
        jira_utils.handle_args()


def test_handle_args_env_file_must_exist(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(sys, 'argv', ['jira_utils.py', '--list', '--env', 'missing.env'])

    with pytest.raises(SystemExit):
        jira_utils.handle_args()


def _build_main_args(**overrides):
    defaults = {
        'list_projects': False,
        'get_workflow': False,
        'get_issue_types': False,
        'get_fields_specified': False,
        'get_fields': None,
        'issue_types': None,
        'project': 'STL',
        'get_versions': False,
        'get_components': False,
        'date': None,
        'dump_file': None,
        'dump_format': 'csv',
        'get_children': None,
        'get_related': None,
        'hierarchy': None,
        'limit': None,
        'table_format': 'flat',
        'releases': None,
        'get_tickets': False,
        'release_tickets_specified': False,
        'release_tickets': None,
        'status': None,
        'no_release': False,
        'total': False,
        'jql_specified': False,
        'jql': None,
        'create_ticket': None,
        'summary': None,
        'issue_type': None,
        'ticket_description': None,
        'assignee_id': None,
        'components': None,
        'fix_versions': None,
        'labels': None,
        'parent': None,
        'dry_run': True,
        'bulk_update': False,
        'input_file': None,
        'set_release': None,
        'remove_release': False,
        'transition': None,
        'assign': None,
        'max_updates': None,
        'bulk_delete': False,
        'delete_subtasks': False,
        'max_deletes': None,
        'force': False,
        'dashboards': False,
        'owner': None,
        'shared': False,
        'dashboard': None,
        'add_gadget': None,
        'remove_gadget': None,
        'update_gadget': None,
        'position': None,
        'color': None,
        'gadget_properties': None,
        'create_dashboard': None,
        'description': None,
        'share_permissions': None,
        'update_dashboard': None,
        'name': None,
        'delete_dashboard': None,
        'copy_dashboard': None,
        'gadgets': None,
        'list_filters': False,
        'favourite_only': False,
        'get_filter': None,
        'run_filter': None,
    }
    defaults.update(overrides)
    return argparse.Namespace(**defaults)


def test_main_routes_many_actions(monkeypatch: pytest.MonkeyPatch):
    args = _build_main_args(
        list_projects=True,
        get_workflow=True,
        get_issue_types=True,
        get_fields_specified=True,
        get_versions=True,
        get_components=True,
        get_children='STL-10',
        get_related='STL-10',
        release_tickets_specified=True,
        release_tickets='12.1.0',
        no_release=True,
        total=True,
        get_tickets=True,
        jql_specified=True,
        jql='project = STL',
        create_ticket='',
        summary='S',
        issue_type='Task',
        bulk_update=True,
        input_file='bulk.csv',
        set_release='12.1.1',
        bulk_delete=True,
        dashboards=True,
        dashboard='101',
        add_gadget='module',
        create_dashboard='db',
        update_dashboard='101',
        delete_dashboard='101',
        copy_dashboard='101',
        gadgets='101',
        list_filters=True,
        get_filter='10',
        run_filter='11',
    )

    jira = MagicMock(name='jira')

    monkeypatch.setattr(jira_utils, 'handle_args', lambda: args)
    monkeypatch.setattr(jira_utils, 'connect_to_jira', lambda: jira)
    monkeypatch.setattr(jira_utils, 'display_jql', MagicMock())

    for fn_name in [
        'list_projects',
        'get_project_workflows',
        'get_project_issue_types',
        'get_project_fields',
        'get_project_versions',
        'get_project_components',
        'get_children_hierarchy',
        'get_related_issues',
        'get_release_tickets',
        'get_no_release_tickets',
        'get_ticket_totals',
        'get_tickets',
        'run_jql_query',
        'create_ticket',
        'bulk_update_tickets',
        'bulk_delete_tickets',
        'list_dashboards',
        'add_gadget',
        'create_dashboard',
        'update_dashboard',
        'delete_dashboard',
        'copy_dashboard',
        'list_gadgets',
        'list_filters',
        'get_filter',
        'run_filter',
    ]:
        monkeypatch.setattr(jira_utils, fn_name, MagicMock())

    jira_utils.main()

    cast(Any, jira_utils.list_projects).assert_called_once_with(jira)
    cast(Any, jira_utils.get_project_workflows).assert_called_once_with(jira, 'STL')
    cast(Any, jira_utils.get_tickets).assert_called_once()
    cast(Any, jira_utils.add_gadget).assert_called_once()
    cast(Any, jira_utils.run_filter).assert_called_once()
    cast(Any, jira_utils.display_jql).assert_called_once()


def test_main_routes_release_pattern_branch(monkeypatch: pytest.MonkeyPatch):
    args = _build_main_args(project='STL', releases='12.*', get_tickets=True)

    monkeypatch.setattr(jira_utils, 'handle_args', lambda: args)
    monkeypatch.setattr(jira_utils, 'connect_to_jira', lambda: MagicMock(name='jira'))
    monkeypatch.setattr(jira_utils, 'get_releases_tickets', MagicMock())
    monkeypatch.setattr(jira_utils, 'display_jql', MagicMock())

    jira_utils.main()

    cast(Any, jira_utils.get_releases_tickets).assert_called_once()


def test_main_exits_on_credentials_error(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(jira_utils, 'handle_args', lambda: _build_main_args())
    monkeypatch.setattr(jira_utils, 'output', lambda *_args, **_kwargs: None)

    def _raise_credentials_error():
        raise jira_utils.JiraCredentialsError('bad credentials')

    monkeypatch.setattr(jira_utils, 'connect_to_jira', _raise_credentials_error)

    with pytest.raises(SystemExit):
        jira_utils.main()
