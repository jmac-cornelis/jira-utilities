import csv
import sys
from pathlib import Path
from typing import Any, cast

import openpyxl

import excel_utils


def test_convert_from_csv_creates_excel_structure(tmp_path):
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(
        "key,status,priority,summary\n"
        "STL-1,Open,P1-Critical,First row\n"
        "STL-2,Closed,P0-Stopper,Second row\n",
        encoding="utf-8",
    )

    out_path = excel_utils.convert_from_csv(str(csv_path), str(tmp_path / "converted.xlsx"))

    wb = openpyxl.load_workbook(out_path)
    ws = wb[wb.sheetnames[0]]

    assert out_path.endswith("converted.xlsx")
    assert ws.cell(row=1, column=1).value == "key"
    assert ws.cell(row=2, column=1).value == "STL-1"
    assert ws.cell(row=3, column=4).value == "Second row"
    hyperlink = ws.cell(row=2, column=1).hyperlink
    assert hyperlink is not None
    target = hyperlink.target or ""
    assert target.endswith("/browse/STL-1")
    assert ws.freeze_panes == "A2"
    wb.close()


def test_create_dashboard_sheet_adds_countif_formulas():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Tickets"

    headers = ["key", "status", "priority"]
    ws.append(headers)
    ws.append(["STL-1", "Open", "P1-Critical"])
    ws.append(["STL-2", "Open", "P0-Stopper"])
    ws.append(["STL-3", "Closed", "P1-Critical"])

    excel_utils._create_dashboard_sheet(wb, ws, headers, dashboard_columns=["status", "priority"])

    dash = wb["Dashboard"]

    assert dash.cell(row=1, column=1).value == "Dashboard"
    assert dash.cell(row=2, column=1).value == "status"
    assert dash.cell(row=3, column=2).value == "Count"

    formulas = [dash.cell(row=4, column=2).value, dash.cell(row=5, column=2).value]
    assert any(str(formula).startswith("=COUNTIF(") for formula in formulas)

    status_total_formula = dash.cell(row=6, column=2).value
    assert status_total_formula == "=SUM(B4:B5)"


def test_validate_and_repair_csv_repairs_unquoted_comma(tmp_path):
    csv_path = tmp_path / "bad.csv"
    csv_path.write_text(
        "key,summary,status\n"
        "STL-1,Hello, world,Open\n"
        "STL-2,Good row,Closed\n",
        encoding="utf-8",
    )

    repaired, stats = excel_utils._validate_and_repair_csv(str(csv_path))

    assert repaired is True
    assert stats["repaired_rows"] == 1

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))

    assert rows[1] == ["STL-1", "Hello, world", "Open"]
    assert len(rows[1]) == 3


def test_concat_merge_sheet_merges_rows_and_union_columns(temp_excel_file, tmp_path):
    file_a = temp_excel_file(
        "a.xlsx",
        ["key", "status"],
        [["STL-1", "Open"]],
    )
    file_b = temp_excel_file(
        "b.xlsx",
        ["key", "priority"],
        [["STL-2", "P1-Critical"]],
    )

    output_file = tmp_path / "merged.xlsx"
    excel_utils.concat_merge_sheet([str(file_a), str(file_b)], str(output_file))

    wb = openpyxl.load_workbook(output_file)
    ws = wb["Merged"]

    headers = [ws.cell(row=1, column=idx).value for idx in range(1, 4)]
    assert headers == ["key", "status", "priority"]
    assert ws.cell(row=2, column=1).value == "STL-1"
    assert ws.cell(row=3, column=1).value == "STL-2"
    assert ws.cell(row=2, column=3).value in (None, "")
    assert ws.cell(row=3, column=2).value in (None, "")
    wb.close()


def test_concat_add_sheet_creates_one_sheet_per_input(temp_excel_file, tmp_path):
    file_a = temp_excel_file("alpha.xlsx", ["key", "status"], [["STL-1", "Open"]])
    file_b = temp_excel_file("beta.xlsx", ["key", "status"], [["STL-2", "Closed"]])

    output_file = tmp_path / "added.xlsx"
    excel_utils.concat_add_sheet([str(file_a), str(file_b)], str(output_file))

    wb = openpyxl.load_workbook(output_file)

    assert "alpha" in wb.sheetnames
    assert "beta" in wb.sheetnames
    assert wb["alpha"].cell(row=2, column=1).value == "STL-1"
    assert wb["beta"].cell(row=2, column=1).value == "STL-2"

    wb.close()


def test_build_excel_map_consolidated_function(monkeypatch, temp_excel_file, tmp_path):
    output_file = tmp_path / "map.xlsx"

    root_a = {"key": "STL-100", "fields": {"summary": "Root A"}}
    child_a1 = {"key": "STL-101", "fields": {"summary": "Child A1"}}
    root_b = {"key": "STL-200", "fields": {"summary": "Root B"}}
    child_b1 = {"key": "STL-201", "fields": {"summary": "Child B1"}}
    grandchild = {"key": "STL-202", "fields": {"summary": "Grandchild B1"}}

    related_by_root = {
        "STL-100": [
            {"issue": root_a, "depth": 0, "via": "", "relation": "root", "from_key": ""},
            {"issue": child_a1, "depth": 1, "via": "parent", "relation": "child", "from_key": "STL-100"},
        ],
        "STL-200": [
            {"issue": root_b, "depth": 0, "via": "", "relation": "root", "from_key": ""},
            {"issue": child_b1, "depth": 1, "via": "parent", "relation": "child", "from_key": "STL-200"},
        ],
    }

    children_by_key = {
        "STL-101": [
            {"issue": child_a1, "depth": 0},
        ],
        "STL-201": [
            {"issue": child_b1, "depth": 0},
            {"issue": grandchild, "depth": 1},
        ],
    }

    class FakeJiraUtils:
        def __init__(self):
            self.calls = {
                "related": [],
                "children": [],
                "dump": [],
                "validate": [],
            }

        def get_connection(self):
            return object()

        def validate_project(self, jira, project_key):
            self.calls["validate"].append((jira, project_key))

        def _get_related_data(self, jira, root_key, hierarchy=None, limit=None):
            self.calls["related"].append((root_key, hierarchy, limit))
            return related_by_root[root_key]

        def _get_children_data(self, jira, ticket_key, limit=None):
            self.calls["children"].append((ticket_key, limit))
            return children_by_key[ticket_key]

        def dump_tickets_to_file(self, issues, dump_file, dump_format, extra_fields=None, table_format="flat"):
            self.calls["dump"].append((dump_file, dump_format, table_format, len(issues)))
            rows = [[issue["key"], issue.get("fields", {}).get("summary", "")] for issue in issues]
            temp_excel_file(Path(dump_file).name, ["key", "summary"], rows)
            source = tmp_path / Path(dump_file).name
            if source != Path(dump_file):
                Path(dump_file).write_bytes(source.read_bytes())

    fake_jira_utils = FakeJiraUtils()

    prior_jira_module = sys.modules.get("jira_utils")
    monkeypatch.setitem(sys.modules, "jira_utils", cast(Any, fake_jira_utils))

    messages = []
    result = excel_utils.build_excel_map(
        ticket_keys=["stl-100", "STL-200"],
        hierarchy_depth=3,
        limit=25,
        output_file=str(output_file),
        project_key="STL",
        output_callback=messages.append,
    )

    assert output_file.exists()
    assert result["output_file"] == str(output_file)
    assert result["sheet_count"] == 3
    assert result["depth1_tickets"] == 2
    assert result["related_count"] == 4
    assert result["root_tickets"] == ["STL-100", "STL-200"]
    assert result["hierarchy_depth"] == 3

    assert fake_jira_utils.calls["related"] == [
        ("STL-100", 1, 25),
        ("STL-200", 1, 25),
    ]
    assert fake_jira_utils.calls["children"] == [
        ("STL-101", None),
        ("STL-201", None),
    ]
    assert len(fake_jira_utils.calls["dump"]) == 3
    assert fake_jira_utils.calls["validate"][0][1] == "STL"

    wb = openpyxl.load_workbook(output_file)
    assert wb.sheetnames[0] == "Tickets"
    assert "STL-101" in wb.sheetnames
    assert "STL-201" in wb.sheetnames
    wb.close()

    assert any(str(msg).startswith("Step 1/4") for msg in messages)
    assert any(str(msg).startswith("Output: ") for msg in messages)

    if prior_jira_module is not None:
        monkeypatch.setitem(sys.modules, "jira_utils", prior_jira_module)
