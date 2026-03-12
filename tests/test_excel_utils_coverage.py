import csv
import json
import sys
from argparse import Namespace
from pathlib import Path
from typing import Any, cast

import openpyxl
import pytest

import excel_utils


def create_test_excel(path: Path, headers: list[str], data: list[list[Any]], sheet_name: str = 'Data') -> Path:
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = sheet_name
    ws.append(headers)
    for row in data:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


def create_test_csv(path: Path, headers: list[str], data: list[list[Any]]) -> Path:
    with path.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(data)
    return path


def make_main_args(**overrides: Any) -> Namespace:
    base = {
        'concat': None,
        'method': 'merge-sheet',
        'output_file': None,
        'convert_to_csv': None,
        'convert_from_csv': None,
        'diff': None,
        'to_plan_json': None,
        'jira_url': excel_utils.DEFAULT_JIRA_BASE_URL,
        'dashboard_columns': None,
        'project_key': '',
        'product_family': '',
        'feature_name': '',
    }
    base.update(overrides)
    return Namespace(**base)


def count_conditional_rules(ws: Any) -> int:
    return sum(len(cf.rules) for cf in ws.conditional_formatting)


def test_error_classes_capture_message_text():
    err = excel_utils.Error('base error')
    excel_err = excel_utils.ExcelFileError('excel error')

    assert err.message == 'base error'
    assert str(err) == 'base error'
    assert excel_err.message == 'excel error'
    assert isinstance(excel_err, excel_utils.Error)


def test_apply_cell_format_handles_empty_and_hyperlink_defaults():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)

    plain_cell = ws.cell(row=1, column=1, value='plain')
    excel_utils._apply_cell_format(plain_cell, None)
    assert plain_cell.hyperlink is None

    link_cell = ws.cell(row=1, column=2, value='STL-1')
    excel_utils._apply_cell_format(
        link_cell,
        {
            'hyperlink': 'https://example.invalid/browse/STL-1',
            'font': None,
            'fill': None,
            'alignment': None,
            'border': None,
            'number_format': None,
        },
    )

    assert link_cell.hyperlink is not None
    assert (link_cell.hyperlink.target or '').endswith('/browse/STL-1')
    assert link_cell.font.underline == 'single'

    wb.close()


def test_load_excel_file_raises_for_missing_and_invalid_input(tmp_path):
    missing = tmp_path / 'missing.xlsx'

    with pytest.raises(excel_utils.ExcelFileError, match='File not found'):
        excel_utils._load_excel_file(str(missing))

    invalid = tmp_path / 'invalid.xlsx'
    invalid.write_text('not an xlsx workbook', encoding='utf-8')

    with pytest.raises(excel_utils.ExcelFileError, match='Failed to load Excel file'):
        excel_utils._load_excel_file(str(invalid))


def test_concat_add_sheet_handles_duplicate_truncated_sheet_names(tmp_path):
    common_prefix = 'a' * 31
    file_a = create_test_excel(tmp_path / f'{common_prefix}111.xlsx', ['key'], [['STL-1']])
    file_b = create_test_excel(tmp_path / f'{common_prefix}222.xlsx', ['key'], [['STL-2']])
    file_c = create_test_excel(tmp_path / f'{common_prefix}333.xlsx', ['key'], [['STL-3']])

    output_file = tmp_path / 'sheets.xlsx'
    excel_utils.concat_add_sheet([str(file_a), str(file_b), str(file_c)], str(output_file))

    wb = openpyxl.load_workbook(output_file)

    assert common_prefix in wb.sheetnames
    assert f'{common_prefix[:28]}_2' in wb.sheetnames
    assert f'{common_prefix[:28]}_3' in wb.sheetnames

    assert wb[common_prefix].cell(row=2, column=1).value == 'STL-1'
    assert wb[f'{common_prefix[:28]}_2'].cell(row=2, column=1).value == 'STL-2'
    assert wb[f'{common_prefix[:28]}_3'].cell(row=2, column=1).value == 'STL-3'

    wb.close()


def test_convert_to_csv_supports_default_and_extensionless_output(tmp_path):
    input_xlsx = create_test_excel(
        tmp_path / 'tickets.xlsx',
        ['key', 'summary'],
        [['STL-1', 'First'], ['STL-2', 'Second']],
    )

    excel_utils.convert_to_csv(str(input_xlsx))

    default_csv = tmp_path / 'tickets.csv'
    assert default_csv.exists()
    with default_csv.open('r', newline='', encoding='utf-8') as handle:
        default_rows = list(csv.reader(handle))
    assert default_rows[0] == ['key', 'summary']
    assert default_rows[1] == ['STL-1', 'First']

    excel_utils.convert_to_csv(str(input_xlsx), str(tmp_path / 'custom_output'))

    custom_csv = tmp_path / 'custom_output.csv'
    assert custom_csv.exists()
    with custom_csv.open('r', newline='', encoding='utf-8') as handle:
        custom_rows = list(csv.reader(handle))
    assert custom_rows[2] == ['STL-2', 'Second']


def test_diff_files_detects_added_removed_changed_and_same(tmp_path):
    file_a = create_test_excel(
        tmp_path / 'a.xlsx',
        ['key', 'status', 'summary'],
        [
            ['STL-1', 'Open', 'Old summary'],
            ['STL-2', 'Closed', 'Removed row'],
            ['STL-4', 'Open', 'Unchanged row'],
        ],
    )
    file_b = create_test_excel(
        tmp_path / 'b.xlsx',
        ['key', 'status', 'summary'],
        [
            ['STL-1', 'Open', 'New summary'],
            ['STL-3', 'Verify', 'Added row'],
            ['STL-4', 'Open', 'Unchanged row'],
        ],
    )

    excel_utils.diff_files([str(file_a), str(file_b)], str(tmp_path / 'diff_report'))

    out_path = tmp_path / 'diff_report.xlsx'
    wb = openpyxl.load_workbook(out_path)
    summary_ws = wb['Summary']
    diff_ws = wb['Diff']

    assert summary_ws.cell(row=2, column=4).value == 1
    assert summary_ws.cell(row=2, column=5).value == 1
    assert summary_ws.cell(row=2, column=6).value == 1
    assert summary_ws.cell(row=2, column=7).value == 1

    statuses = [str(diff_ws.cell(row=r, column=1).value) for r in range(2, diff_ws.max_row + 1)]
    assert sorted(statuses) == ['ADDED', 'CHANGED', 'REMOVED']

    detail_cells = [
        str(diff_ws.cell(row=r, column=diff_ws.max_column).value)
        for r in range(2, diff_ws.max_row + 1)
    ]
    assert any('summary:' in details for details in detail_cells)

    wb.close()


def test_diff_files_identical_inputs_emit_header_only_diff_sheet(tmp_path, monkeypatch):
    file_a = create_test_excel(
        tmp_path / 'base_a.xlsx',
        ['key', 'status'],
        [['STL-10', 'Open'], ['STL-11', 'Closed']],
    )
    file_b = create_test_excel(
        tmp_path / 'base_b.xlsx',
        ['key', 'status'],
        [['STL-10', 'Open'], ['STL-11', 'Closed']],
    )

    monkeypatch.chdir(tmp_path)
    excel_utils.diff_files([str(file_a), str(file_b)])

    report = tmp_path / 'diff_output.xlsx'
    wb = openpyxl.load_workbook(report)

    assert wb['Diff'].max_row == 1
    assert wb['Summary'].cell(row=2, column=7).value == 2

    wb.close()


def test_validate_and_repair_csv_reports_clean_and_padded_rows(tmp_path):
    clean_csv = create_test_csv(
        tmp_path / 'clean.csv',
        ['key', 'summary', 'status'],
        [['STL-1', 'Good row', 'Open'], ['STL-2', 'Also good', 'Closed']],
    )

    repaired_clean, clean_stats = excel_utils._validate_and_repair_csv(str(clean_csv))

    assert repaired_clean is False
    assert clean_stats['ok_rows'] == 2
    assert clean_stats['padded_rows'] == 0

    short_csv = tmp_path / 'short.csv'
    short_csv.write_text('key,summary,status\nSTL-3,Needs padding\n', encoding='utf-8')

    repaired_short, short_stats = excel_utils._validate_and_repair_csv(str(short_csv))

    assert repaired_short is True
    assert short_stats['padded_rows'] == 1

    with short_csv.open('r', newline='', encoding='utf-8') as handle:
        rows = list(csv.reader(handle))
    assert rows[1] == ['STL-3', 'Needs padding', '']


def test_validate_and_repair_csv_handles_many_extra_columns_fallback(tmp_path):
    csv_path = tmp_path / 'extra_cols.csv'
    csv_path.write_text(
        'key,summary,status\n'
        'STL-1,one,two,three,four,five,six,seven\n',
        encoding='utf-8',
    )

    repaired, stats = excel_utils._validate_and_repair_csv(str(csv_path))

    assert repaired is True
    assert stats['repaired_rows'] == 1

    with csv_path.open('r', newline='', encoding='utf-8') as handle:
        rows = list(csv.reader(handle))

    assert rows[1][0] == 'STL-1'
    assert rows[1][1] == 'one'
    assert rows[1][2] == 'two,three,four,five,six,seven'


def test_create_dashboard_sheet_skips_when_dashboard_columns_not_provided():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = 'Tickets'
    ws.append(['key', 'status'])
    ws.append(['STL-1', 'Open'])

    excel_utils._create_dashboard_sheet(wb, ws, ['key', 'status'], dashboard_columns=None)

    assert 'Dashboard' not in wb.sheetnames
    wb.close()


def test_create_dashboard_sheet_skips_when_no_requested_columns_match():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = 'Tickets'
    ws.append(['key', 'status'])
    ws.append(['STL-1', 'Open'])

    excel_utils._create_dashboard_sheet(wb, ws, ['key', 'status'], dashboard_columns=['priority'])

    assert 'Dashboard' not in wb.sheetnames
    wb.close()


def test_create_dashboard_sheet_builds_countblank_and_escaped_countif_formulas():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = 'Data Sheet'
    ws.append(['key', 'status'])
    ws.append(['STL-1', 'Open'])
    ws.append(['STL-2', '   '])
    ws.append(['STL-3', 'Escaped "quote"'])

    excel_utils._create_dashboard_sheet(
        wb,
        ws,
        ['key', 'status'],
        dashboard_columns=['status', 'not-real-column'],
    )

    dash = wb['Dashboard']
    formulas = [dash.cell(row=r, column=2).value for r in range(4, 7)]

    assert any(str(formula).startswith('=COUNTBLANK(') for formula in formulas)
    assert any('""quote""' in str(formula) for formula in formulas)
    assert any("'Data Sheet'!B$2:B$4" in str(formula) for formula in formulas)
    assert dash.cell(row=7, column=2).value == '=SUM(B4:B6)'

    wb.close()


def test_create_dashboard_sheet_with_header_only_data_creates_total_row():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = 'Tickets'
    ws.append(['key', 'status'])

    excel_utils._create_dashboard_sheet(wb, ws, ['key', 'status'], dashboard_columns=['status'])

    dash = wb['Dashboard']
    assert dash.cell(row=4, column=1).value == 'Total'
    assert dash.cell(row=4, column=2).value == '=SUM(B4:B3)'

    wb.close()


def test_status_and_priority_conditional_formatting_add_expected_rule_count():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.append(['key', 'status', 'priority'])
    ws.append(['STL-1', 'Open', 'P1-Critical'])
    ws.append(['STL-2', 'Closed', 'P0-Stopper'])

    excel_utils._apply_status_conditional_formatting(ws, ['key', 'status', 'priority'])
    excel_utils._apply_priority_conditional_formatting(ws, ['key', 'status', 'priority'])

    assert count_conditional_rules(ws) == (
        len(excel_utils.STATUS_FILL_COLORS) + len(excel_utils.PRIORITY_FILL_COLORS)
    )

    ranges = {str(cf.sqref) for cf in ws.conditional_formatting}
    assert 'B2:B3' in ranges
    assert 'C2:C3' in ranges

    wb.close()


def test_conditional_formatting_skips_missing_columns_and_header_only_sheets():
    wb = openpyxl.Workbook()
    ws_missing = cast(Any, wb.active)
    ws_missing.title = 'MissingColumns'
    ws_missing.append(['key'])

    excel_utils._apply_status_conditional_formatting(ws_missing, ['key'])
    excel_utils._apply_priority_conditional_formatting(ws_missing, ['key'])
    assert count_conditional_rules(ws_missing) == 0

    ws_header_only = wb.create_sheet('HeaderOnly')
    ws_header_only.append(['status', 'priority'])

    excel_utils._apply_status_conditional_formatting(ws_header_only, ['status', 'priority'])
    excel_utils._apply_priority_conditional_formatting(ws_header_only, ['status', 'priority'])
    assert count_conditional_rules(ws_header_only) == 0

    wb.close()


def test_auto_fit_columns_applies_minimum_and_maximum_width_bounds():
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.append(['short', 'long'])
    ws.append(['x', 'y' * 200])

    excel_utils._auto_fit_columns(ws)

    assert int(ws.column_dimensions['A'].width) == 10
    assert int(ws.column_dimensions['B'].width) == 50

    wb.close()


def test_convert_from_csv_missing_file_raises_excel_file_error(tmp_path):
    missing = tmp_path / 'missing.csv'

    with pytest.raises(excel_utils.ExcelFileError, match='File not found'):
        excel_utils.convert_from_csv(str(missing))


def test_convert_from_csv_invalid_encoding_raises_excel_file_error(tmp_path):
    bad_csv = tmp_path / 'bad_encoding.csv'
    bad_csv.write_bytes(b'\xff\xfe\xfd')

    with pytest.raises(excel_utils.ExcelFileError, match='Failed to read CSV file'):
        excel_utils.convert_from_csv(str(bad_csv))


def test_convert_from_csv_empty_file_raises_no_header_error(tmp_path):
    empty_csv = tmp_path / 'empty.csv'
    empty_csv.write_text('', encoding='utf-8')

    with pytest.raises(excel_utils.ExcelFileError, match='has no header row'):
        excel_utils.convert_from_csv(str(empty_csv))


def test_convert_from_csv_creates_dashboard_and_appends_xlsx_extension(tmp_path):
    csv_path = create_test_csv(
        tmp_path / 'dash.csv',
        ['key', 'status'],
        [['STL-1', 'Open'], ['STL-2', '   ']],
    )

    output_path = excel_utils.convert_from_csv(
        str(csv_path),
        str(tmp_path / 'dashboard_output'),
        jira_base_url='https://example.atlassian.net/',
        dashboard_columns=['status'],
    )

    assert output_path.endswith('.xlsx')
    assert Path(output_path).exists()

    wb = openpyxl.load_workbook(output_path)
    data_ws = wb[wb.sheetnames[0]]
    dash_ws = wb['Dashboard']

    key_link = data_ws.cell(row=2, column=1).hyperlink
    assert key_link is not None
    assert key_link.target == 'https://example.atlassian.net/browse/STL-1'

    formula_cells = [str(dash_ws.cell(row=r, column=2).value) for r in range(4, dash_ws.max_row + 1)]
    assert any(formula.startswith('=COUNTBLANK(') for formula in formula_cells)

    wb.close()


def test_convert_to_plan_json_from_csv_writes_expected_plan_payload(tmp_path):
    csv_path = create_test_csv(
        tmp_path / 'plan.csv',
        ['key', 'project', 'issue_type', 'summary', 'component', 'labels', 'depth', 'product_family'],
        [
            ['', 'STL', 'Epic', 'Fabric Enablement', 'Fabric', 'roadmap', '0', 'CN5000'],
            ['', 'STL', 'Story', 'Implement lane-state checks', 'Driver', 'execution', '1', 'CN5000'],
        ],
    )

    output_path = excel_utils.convert_to_plan_json(
        str(csv_path),
        project_key='STL',
        product_family='CN6000',
        feature_name='Lane State Feature',
    )

    assert output_path.endswith('.json')
    payload = json.loads(Path(output_path).read_text(encoding='utf-8'))

    assert payload['project_key'] == 'STL'
    assert payload['feature_name'] == 'Lane State Feature'
    assert payload['product_family'] == ['CN6000']
    assert payload['total_epics'] == 1
    assert payload['total_stories'] == 1
    assert payload['total_tickets'] == 2


def test_convert_to_plan_json_missing_input_raises_excel_file_error(tmp_path):
    missing = tmp_path / 'does_not_exist.csv'

    with pytest.raises(excel_utils.ExcelFileError, match='Input file not found'):
        excel_utils.convert_to_plan_json(str(missing))


def test_handle_args_convert_from_csv_parses_dashboard_and_jira_flags(monkeypatch, tmp_path):
    csv_path = create_test_csv(tmp_path / 'input.csv', ['key', 'status'], [['STL-1', 'Open']])

    monkeypatch.setattr(
        sys,
        'argv',
        [
            'excel_utils.py',
            '--quiet',
            '--convert-from-csv',
            str(csv_path),
            '--d-columns',
            'status',
            'priority',
            '--jira-url',
            'none',
        ],
    )

    args = excel_utils.handle_args()

    assert args.convert_from_csv == str(csv_path)
    assert args.dashboard_columns == ['status', 'priority']
    assert args.jira_url == 'none'
    assert args.quiet is True


def test_handle_args_convert_to_csv_parses_verbose_flag(monkeypatch, tmp_path):
    xlsx_path = create_test_excel(tmp_path / 'data.xlsx', ['key'], [['STL-2']])

    monkeypatch.setattr(
        sys,
        'argv',
        ['excel_utils.py', '--verbose', '--convert-to-csv', str(xlsx_path)],
    )

    args = excel_utils.handle_args()

    assert args.convert_to_csv == str(xlsx_path)
    assert args.verbose is True


def test_handle_args_diff_parses_input_list(monkeypatch, tmp_path):
    file_a = create_test_excel(tmp_path / 'a.xlsx', ['key'], [['STL-1']])
    file_b = create_test_excel(tmp_path / 'b.xlsx', ['key'], [['STL-2']])

    monkeypatch.setattr(
        sys,
        'argv',
        ['excel_utils.py', '--diff', str(file_a), str(file_b), '--output', str(tmp_path / 'changes.xlsx')],
    )

    args = excel_utils.handle_args()

    assert args.diff == [str(file_a), str(file_b)]
    assert args.output_file == str(tmp_path / 'changes.xlsx')


def test_handle_args_concat_parses_method_and_appends_output_extension(monkeypatch, tmp_path):
    file_a = create_test_excel(tmp_path / 'x.xlsx', ['key'], [['STL-1']])
    file_b = create_test_excel(tmp_path / 'y.xlsx', ['key'], [['STL-2']])

    monkeypatch.setattr(
        sys,
        'argv',
        [
            'excel_utils.py',
            '--concat',
            str(file_a),
            str(file_b),
            '--method',
            'add-sheet',
            '--output',
            str(tmp_path / 'combined'),
        ],
    )

    args = excel_utils.handle_args()

    assert args.concat == [str(file_a), str(file_b)]
    assert args.method == 'add-sheet'
    assert args.output_file == f'{tmp_path / "combined"}.xlsx'


def test_handle_args_to_plan_json_parses_plan_options(monkeypatch, tmp_path):
    csv_path = create_test_csv(
        tmp_path / 'plan.csv',
        ['key', 'project', 'issue_type', 'summary'],
        [['', 'STL', 'Epic', 'Feature']],
    )

    monkeypatch.setattr(
        sys,
        'argv',
        [
            'excel_utils.py',
            '--to-plan-json',
            str(csv_path),
            '--project',
            'STL',
            '--product-family',
            'CN5000',
            '--feature-name',
            'Feature Plan',
            '--output',
            str(tmp_path / 'plan_out.json'),
        ],
    )

    args = excel_utils.handle_args()

    assert args.to_plan_json == str(csv_path)
    assert args.project_key == 'STL'
    assert args.product_family == 'CN5000'
    assert args.feature_name == 'Feature Plan'
    assert args.output_file == str(tmp_path / 'plan_out.json')


def test_main_dispatches_concat_modes(monkeypatch):
    merge_called: dict[str, Any] = {}
    add_called: dict[str, Any] = {}

    def fake_merge(files: list[str], out_file: str | None):
        merge_called['args'] = (files, out_file)

    def fake_add(files: list[str], out_file: str | None):
        add_called['args'] = (files, out_file)

    monkeypatch.setattr(excel_utils, 'concat_merge_sheet', fake_merge)
    monkeypatch.setattr(excel_utils, 'concat_add_sheet', fake_add)

    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(concat=['a.xlsx', 'b.xlsx'], method='merge-sheet', output_file='merge.xlsx'),
    )
    excel_utils.main()

    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(concat=['a.xlsx', 'b.xlsx'], method='add-sheet', output_file='add.xlsx'),
    )
    excel_utils.main()

    assert merge_called['args'] == (['a.xlsx', 'b.xlsx'], 'merge.xlsx')
    assert add_called['args'] == (['a.xlsx', 'b.xlsx'], 'add.xlsx')


def test_main_dispatches_convert_to_csv_and_diff(monkeypatch):
    convert_calls: dict[str, Any] = {}
    diff_calls: dict[str, Any] = {}

    def fake_convert(input_file: str, output_file: str | None):
        convert_calls['args'] = (input_file, output_file)

    def fake_diff(files: list[str], output_file: str | None):
        diff_calls['args'] = (files, output_file)

    monkeypatch.setattr(excel_utils, 'convert_to_csv', fake_convert)
    monkeypatch.setattr(excel_utils, 'diff_files', fake_diff)

    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(convert_to_csv='tickets.xlsx', output_file='tickets.csv'),
    )
    excel_utils.main()

    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(diff=['a.xlsx', 'b.xlsx'], output_file='changes.xlsx'),
    )
    excel_utils.main()

    assert convert_calls['args'] == ('tickets.xlsx', 'tickets.csv')
    assert diff_calls['args'] == (['a.xlsx', 'b.xlsx'], 'changes.xlsx')


def test_main_convert_from_csv_normalizes_none_jira_url(monkeypatch):
    captured: dict[str, Any] = {}

    def fake_convert_from_csv(
        input_file: str,
        output_file: str | None,
        jira_base_url: str | None = excel_utils.DEFAULT_JIRA_BASE_URL,
        dashboard_columns: list[str] | None = None,
    ):
        captured['input_file'] = input_file
        captured['output_file'] = output_file
        captured['jira_base_url'] = jira_base_url
        captured['dashboard_columns'] = dashboard_columns

    monkeypatch.setattr(excel_utils, 'convert_from_csv', fake_convert_from_csv)
    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(
            convert_from_csv='input.csv',
            output_file='output.xlsx',
            jira_url='none',
            dashboard_columns=['status'],
        ),
    )

    excel_utils.main()

    assert captured['input_file'] == 'input.csv'
    assert captured['output_file'] == 'output.xlsx'
    assert captured['jira_base_url'] is None
    assert captured['dashboard_columns'] == ['status']


def test_main_dispatches_to_plan_json(monkeypatch):
    captured: dict[str, Any] = {}

    def fake_to_plan_json(
        input_file: str,
        output_file: str | None = None,
        project_key: str = '',
        product_family: str = '',
        feature_name: str = '',
    ):
        captured['input_file'] = input_file
        captured['output_file'] = output_file
        captured['project_key'] = project_key
        captured['product_family'] = product_family
        captured['feature_name'] = feature_name

    monkeypatch.setattr(excel_utils, 'convert_to_plan_json', fake_to_plan_json)
    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(
            to_plan_json='plan.csv',
            output_file='plan.json',
            project_key='STL',
            product_family='CN5000',
            feature_name='Feature Name',
        ),
    )

    excel_utils.main()

    assert captured == {
        'input_file': 'plan.csv',
        'output_file': 'plan.json',
        'project_key': 'STL',
        'product_family': 'CN5000',
        'feature_name': 'Feature Name',
    }


def test_main_exits_with_code_1_on_excel_file_error(monkeypatch):
    def raise_excel_error(_input_file: str, _output_file: str | None):
        raise excel_utils.ExcelFileError('bad excel input')

    monkeypatch.setattr(excel_utils, 'convert_to_csv', raise_excel_error)
    monkeypatch.setattr(excel_utils, 'output', lambda _msg='': None)
    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(convert_to_csv='bad.xlsx', output_file='out.csv'),
    )

    with pytest.raises(SystemExit) as excinfo:
        excel_utils.main()

    assert excinfo.value.code == 1


def test_main_exits_with_code_1_on_unexpected_error(monkeypatch):
    def raise_runtime_error(_input_file: str, _output_file: str | None):
        raise RuntimeError('unexpected boom')

    monkeypatch.setattr(excel_utils, 'convert_to_csv', raise_runtime_error)
    monkeypatch.setattr(excel_utils, 'output', lambda _msg='': None)
    monkeypatch.setattr(
        excel_utils,
        'handle_args',
        lambda: make_main_args(convert_to_csv='bad.xlsx', output_file='out.csv'),
    )

    with pytest.raises(SystemExit) as excinfo:
        excel_utils.main()

    assert excinfo.value.code == 1
