##########################################################################################
#
# Tests for core/reporting.py
#
# Covers:
#   - _next_day() helper
#   - tickets_created_on() — JQL construction + issue_to_dict mapping
#   - bugs_missing_field() — date-scoped and all-open variants
#   - status_changes_by_actor() — REST API pagination, automation classification
#   - daily_report() — composite orchestration
#   - export_daily_report() — Excel and CSV export
#
##########################################################################################

import csv
import json
import os
import tempfile
from datetime import date, timedelta
from types import SimpleNamespace
from typing import Any, Dict, List, Optional
from unittest.mock import MagicMock, patch

import pytest


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_issue_dict(
    key: str = 'STL-1',
    summary: str = 'Test ticket',
    issue_type: str = 'Bug',
    status: str = 'Open',
    priority: str = 'P1-Critical',
    assignee: str = 'Jane Dev',
    created: str = '2026-03-12',
) -> dict:
    """Return a minimal dict matching the shape of core.tickets.issue_to_dict output."""
    return {
        'key': key,
        'summary': summary,
        'issue_type': issue_type,
        'status': status,
        'priority': priority,
        'assignee': assignee,
        'created': created,
    }


def _make_raw_issue(
    key: str = 'STL-1',
    summary: str = 'Test ticket',
    issue_type: str = 'Bug',
    status: str = 'Open',
) -> dict:
    """Return a raw Jira REST API issue dict (used by paginated_jql_search)."""
    return {
        'key': key,
        'fields': {
            'summary': summary,
            'issuetype': {'name': issue_type},
            'status': {'name': status},
            'priority': {'name': 'P1-Critical'},
            'assignee': {'displayName': 'Jane Dev'},
            'reporter': {'displayName': 'John Reporter'},
            'created': '2026-03-12T10:00:00.000+0000',
            'updated': '2026-03-12T12:00:00.000+0000',
            'project': {'key': 'STL'},
            'fixVersions': [],
            'versions': [],
            'components': [],
            'labels': [],
            'issuelinks': [],
        },
    }


def _make_changelog_response(
    issues: list[dict],
    next_page_token: Optional[str] = None,
) -> dict:
    """Build a mock REST API /search/jql response with changelog data."""
    return {
        'issues': issues,
        'nextPageToken': next_page_token,
    }


def _make_changelog_issue(
    key: str,
    histories: list[dict],
) -> dict:
    """Build a single issue dict with changelog histories."""
    return {
        'key': key,
        'changelog': {'histories': histories},
    }


def _make_history(
    created: str,
    author_name: str,
    author_email: str,
    from_status: str,
    to_status: str,
) -> dict:
    """Build a single changelog history entry with a status transition."""
    return {
        'created': created,
        'author': {
            'displayName': author_name,
            'emailAddress': author_email,
        },
        'items': [
            {
                'field': 'status',
                'fromString': from_status,
                'toString': to_status,
            }
        ],
    }


# ---------------------------------------------------------------------------
# _next_day
# ---------------------------------------------------------------------------

class TestNextDay:
    def test_basic_next_day(self):
        from core.reporting import _next_day
        assert _next_day('2026-03-12') == '2026-03-13'

    def test_month_boundary(self):
        from core.reporting import _next_day
        assert _next_day('2026-03-31') == '2026-04-01'

    def test_year_boundary(self):
        from core.reporting import _next_day
        assert _next_day('2026-12-31') == '2027-01-01'

    def test_leap_year(self):
        from core.reporting import _next_day
        # 2028 is a leap year
        assert _next_day('2028-02-28') == '2028-02-29'
        assert _next_day('2028-02-29') == '2028-03-01'


# ---------------------------------------------------------------------------
# tickets_created_on
# ---------------------------------------------------------------------------

class TestTicketsCreatedOn:
    @patch('core.reporting.paginated_jql_search')
    @patch('core.reporting.issue_to_dict')
    def test_returns_issue_dicts(self, mock_to_dict, mock_search):
        """tickets_created_on should call paginated_jql_search and map results."""
        from core.reporting import tickets_created_on

        raw1 = _make_raw_issue(key='STL-10', summary='Ticket A')
        raw2 = _make_raw_issue(key='STL-11', summary='Ticket B')
        mock_search.return_value = [raw1, raw2]
        mock_to_dict.side_effect = lambda i: {'key': i['key'], 'summary': i['fields']['summary']}

        jira = MagicMock()
        result = tickets_created_on(jira, 'STL', '2026-03-12')

        assert len(result) == 2
        assert result[0]['key'] == 'STL-10'
        assert result[1]['key'] == 'STL-11'

    @patch('core.reporting.paginated_jql_search')
    @patch('core.reporting.issue_to_dict')
    def test_jql_uses_explicit_date_bounds(self, mock_to_dict, mock_search):
        """JQL should use >= target_date AND < next_day (no '+' character)."""
        from core.reporting import tickets_created_on

        mock_search.return_value = []
        mock_to_dict.return_value = {}

        jira = MagicMock()
        tickets_created_on(jira, 'STL', '2026-03-12')

        # Inspect the JQL passed to paginated_jql_search
        call_args = mock_search.call_args
        jql = call_args[0][1]  # second positional arg
        assert '2026-03-12' in jql
        assert '2026-03-13' in jql
        # Must NOT contain '+' which breaks Jira enhanced search
        assert '+' not in jql

    @patch('core.reporting.paginated_jql_search')
    @patch('core.reporting.issue_to_dict')
    def test_empty_result(self, mock_to_dict, mock_search):
        """Should return empty list when no tickets found."""
        from core.reporting import tickets_created_on

        mock_search.return_value = []
        jira = MagicMock()
        result = tickets_created_on(jira, 'STL', '2026-03-12')
        assert result == []


# ---------------------------------------------------------------------------
# bugs_missing_field
# ---------------------------------------------------------------------------

class TestBugsMissingField:
    @patch('core.reporting.paginated_jql_search')
    @patch('core.reporting.issue_to_dict')
    def test_with_target_date(self, mock_to_dict, mock_search):
        """When target_date is given, flagged bugs are date-scoped."""
        from core.reporting import bugs_missing_field

        raw = _make_raw_issue(key='STL-20', issue_type='Bug')
        # First call = flagged (date-scoped), second call = total open
        mock_search.side_effect = [[raw], [raw, _make_raw_issue(key='STL-21')]]
        mock_to_dict.side_effect = lambda i: {'key': i['key']}

        jira = MagicMock()
        result = bugs_missing_field(jira, 'STL', field='affectedVersion',
                                    target_date='2026-03-12')

        assert result['field'] == 'affectedVersion'
        assert len(result['flagged']) == 1
        assert result['flagged'][0]['key'] == 'STL-20'
        assert result['total_open_count'] == 2

    @patch('core.reporting.paginated_jql_search')
    @patch('core.reporting.issue_to_dict')
    def test_without_target_date(self, mock_to_dict, mock_search):
        """When target_date is None, flagged = all open bugs missing field."""
        from core.reporting import bugs_missing_field

        mock_search.side_effect = [[], []]
        jira = MagicMock()
        result = bugs_missing_field(jira, 'STL', field='fixVersion')

        assert result['field'] == 'fixVersion'
        assert result['flagged'] == []
        assert result['total_open_count'] == 0

    @patch('core.reporting.paginated_jql_search')
    @patch('core.reporting.issue_to_dict')
    def test_jql_contains_field_is_empty(self, mock_to_dict, mock_search):
        """JQL should include 'field is EMPTY' clause."""
        from core.reporting import bugs_missing_field

        mock_search.side_effect = [[], []]
        mock_to_dict.return_value = {}

        jira = MagicMock()
        bugs_missing_field(jira, 'STL', field='component', target_date='2026-03-12')

        # Both calls should have 'component is EMPTY' in the JQL
        for call in mock_search.call_args_list:
            jql = call[0][1]
            assert 'component is EMPTY' in jql

    @patch('core.reporting.paginated_jql_search')
    @patch('core.reporting.issue_to_dict')
    def test_total_excludes_closed_statuses(self, mock_to_dict, mock_search):
        """The total open count query should exclude Closed/Done/Resolved."""
        from core.reporting import bugs_missing_field

        mock_search.side_effect = [[], []]
        mock_to_dict.return_value = {}

        jira = MagicMock()
        bugs_missing_field(jira, 'STL', target_date='2026-03-12')

        # Second call is the total-open query
        total_jql = mock_search.call_args_list[1][0][1]
        assert 'NOT IN' in total_jql
        assert 'Closed' in total_jql


# ---------------------------------------------------------------------------
# status_changes_by_actor
# ---------------------------------------------------------------------------

class TestStatusChangesByActor:
    @patch('core.reporting.requests.get')
    @patch('core.reporting.os.getenv', return_value='https://test.atlassian.net')
    def test_classifies_automation_vs_human(self, mock_getenv, mock_get):
        """Transitions by bot accounts should be classified as automation."""
        from core.reporting import status_changes_by_actor

        histories = [
            _make_history('2026-03-12T10:00:00.000+0000', 'SCM Bot', 'scm-bot@cornelis.com',
                          'Open', 'In Progress'),
            _make_history('2026-03-12T11:00:00.000+0000', 'Jane Dev', 'jane@cornelis.com',
                          'In Progress', 'Verify'),
        ]
        issue = _make_changelog_issue('STL-30', histories)
        response_data = _make_changelog_response([issue])

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = response_data
        mock_resp.raise_for_status = MagicMock()
        mock_get.return_value = mock_resp

        result = status_changes_by_actor(
            'STL', '2026-03-12',
            email='test@test.com', api_token='token123',
        )

        assert len(result['automation']) == 1
        assert result['automation'][0]['author'] == 'SCM Bot'
        assert len(result['human']) == 1
        assert result['human'][0]['author'] == 'Jane Dev'
        assert result['total'] == 2

    @patch('core.reporting.requests.get')
    @patch('core.reporting.os.getenv', return_value='https://test.atlassian.net')
    def test_filters_by_target_date(self, mock_getenv, mock_get):
        """Only transitions created on target_date should be included."""
        from core.reporting import status_changes_by_actor

        histories = [
            # On target date — should be included
            _make_history('2026-03-12T10:00:00.000+0000', 'Jane', 'jane@test.com',
                          'Open', 'In Progress'),
            # Different date — should be excluded
            _make_history('2026-03-11T23:59:59.000+0000', 'Jane', 'jane@test.com',
                          'New', 'Open'),
        ]
        issue = _make_changelog_issue('STL-40', histories)
        response_data = _make_changelog_response([issue])

        mock_resp = MagicMock()
        mock_resp.json.return_value = response_data
        mock_resp.raise_for_status = MagicMock()
        mock_get.return_value = mock_resp

        result = status_changes_by_actor(
            'STL', '2026-03-12',
            email='test@test.com', api_token='token123',
        )

        assert result['total'] == 1
        assert result['human'][0]['from'] == 'Open'

    @patch('core.reporting.requests.get')
    @patch('core.reporting.os.getenv', return_value='https://test.atlassian.net')
    def test_pagination(self, mock_getenv, mock_get):
        """Should follow nextPageToken for multi-page results."""
        from core.reporting import status_changes_by_actor

        # Page 1: has nextPageToken
        issue1 = _make_changelog_issue('STL-50', [
            _make_history('2026-03-12T10:00:00.000+0000', 'Jane', 'jane@test.com',
                          'Open', 'In Progress'),
        ])
        page1 = _make_changelog_response([issue1], next_page_token='page2')

        # Page 2: no nextPageToken (last page)
        issue2 = _make_changelog_issue('STL-51', [
            _make_history('2026-03-12T11:00:00.000+0000', 'Bob', 'bob@test.com',
                          'In Progress', 'Verify'),
        ])
        page2 = _make_changelog_response([issue2])

        mock_resp1 = MagicMock()
        mock_resp1.json.return_value = page1
        mock_resp1.raise_for_status = MagicMock()

        mock_resp2 = MagicMock()
        mock_resp2.json.return_value = page2
        mock_resp2.raise_for_status = MagicMock()

        mock_get.side_effect = [mock_resp1, mock_resp2]

        result = status_changes_by_actor(
            'STL', '2026-03-12',
            email='test@test.com', api_token='token123',
        )

        assert result['total'] == 2
        assert mock_get.call_count == 2

    @patch('core.reporting.requests.get')
    @patch('core.reporting.os.getenv', return_value='https://test.atlassian.net')
    def test_custom_automation_keywords(self, mock_getenv, mock_get):
        """Custom automation_keywords should override defaults."""
        from core.reporting import status_changes_by_actor

        histories = [
            _make_history('2026-03-12T10:00:00.000+0000', 'CI Runner', 'ci@internal.com',
                          'Open', 'In Progress'),
        ]
        issue = _make_changelog_issue('STL-60', histories)
        response_data = _make_changelog_response([issue])

        mock_resp = MagicMock()
        mock_resp.json.return_value = response_data
        mock_resp.raise_for_status = MagicMock()
        mock_get.return_value = mock_resp

        # Default keywords won't match 'ci@internal.com'
        result_default = status_changes_by_actor(
            'STL', '2026-03-12',
            email='test@test.com', api_token='token123',
        )
        assert len(result_default['automation']) == 0
        assert len(result_default['human']) == 1

        # Custom keywords should match
        mock_get.return_value = mock_resp  # reset
        result_custom = status_changes_by_actor(
            'STL', '2026-03-12',
            email='test@test.com', api_token='token123',
            automation_keywords=['ci@'],
        )
        assert len(result_custom['automation']) == 1
        assert len(result_custom['human']) == 0

    @patch('core.reporting.requests.get')
    def test_falls_back_to_get_jira_credentials(self, mock_get):
        """When email/api_token not provided, should call get_jira_credentials()."""
        from core.reporting import status_changes_by_actor

        response_data = _make_changelog_response([])
        mock_resp = MagicMock()
        mock_resp.json.return_value = response_data
        mock_resp.raise_for_status = MagicMock()
        mock_get.return_value = mock_resp

        with patch('core.reporting.os.getenv', return_value='https://test.atlassian.net'):
            with patch('jira_utils.get_jira_credentials',
                       return_value=('auto@test.com', 'auto-token')):
                result = status_changes_by_actor('STL', '2026-03-12')

        assert result['total'] == 0
        # Verify the auth tuple used the fallback credentials
        call_kwargs = mock_get.call_args
        assert call_kwargs[1]['auth'] == ('auto@test.com', 'auto-token')

    @patch('core.reporting.requests.get')
    @patch('core.reporting.os.getenv', return_value='https://test.atlassian.net')
    def test_empty_changelog(self, mock_getenv, mock_get):
        """Issues with no changelog histories should produce zero transitions."""
        from core.reporting import status_changes_by_actor

        issue = {'key': 'STL-70', 'changelog': {'histories': []}}
        response_data = _make_changelog_response([issue])

        mock_resp = MagicMock()
        mock_resp.json.return_value = response_data
        mock_resp.raise_for_status = MagicMock()
        mock_get.return_value = mock_resp

        result = status_changes_by_actor(
            'STL', '2026-03-12',
            email='test@test.com', api_token='token123',
        )
        assert result['total'] == 0
        assert result['automation'] == []
        assert result['human'] == []


# ---------------------------------------------------------------------------
# daily_report (composite)
# ---------------------------------------------------------------------------

class TestDailyReport:
    @patch('core.reporting.status_changes_by_actor')
    @patch('core.reporting.bugs_missing_field')
    @patch('core.reporting.tickets_created_on')
    def test_calls_all_three_queries(self, mock_created, mock_bugs, mock_changes):
        """daily_report should call all three sub-functions and combine results."""
        from core.reporting import daily_report

        mock_created.return_value = [_make_issue_dict(key='STL-100')]
        mock_bugs.return_value = {
            'field': 'affectedVersion',
            'flagged': [_make_issue_dict(key='STL-101')],
            'total_open_count': 5,
        }
        mock_changes.return_value = {
            'automation': [{'key': 'STL-102', 'from': 'Open', 'to': 'In Progress',
                            'author': 'Bot', 'email': 'bot@test.com', 'time': '2026-03-12T10:00'}],
            'human': [],
            'total': 1,
        }

        jira = MagicMock()
        result = daily_report(jira, 'STL', '2026-03-12')

        assert result['date'] == '2026-03-12'
        assert result['project'] == 'STL'
        assert len(result['created_tickets']) == 1
        assert len(result['bugs_missing_field']['flagged']) == 1
        assert result['status_changes']['total'] == 1

        # Verify each sub-function was called with correct args
        mock_created.assert_called_once_with(jira, 'STL', '2026-03-12')
        mock_bugs.assert_called_once_with(jira, 'STL', field='affectedVersion',
                                          target_date='2026-03-12')
        mock_changes.assert_called_once_with('STL', '2026-03-12',
                                             automation_keywords=None)

    @patch('core.reporting.status_changes_by_actor')
    @patch('core.reporting.bugs_missing_field')
    @patch('core.reporting.tickets_created_on')
    def test_defaults_to_today(self, mock_created, mock_bugs, mock_changes):
        """When target_date is None, should default to today's date."""
        from core.reporting import daily_report

        mock_created.return_value = []
        mock_bugs.return_value = {'field': 'affectedVersion', 'flagged': [],
                                  'total_open_count': 0}
        mock_changes.return_value = {'automation': [], 'human': [], 'total': 0}

        jira = MagicMock()
        result = daily_report(jira, 'STL')

        today = date.today().isoformat()
        assert result['date'] == today

    @patch('core.reporting.status_changes_by_actor')
    @patch('core.reporting.bugs_missing_field')
    @patch('core.reporting.tickets_created_on')
    def test_custom_missing_field(self, mock_created, mock_bugs, mock_changes):
        """missing_field parameter should be forwarded to bugs_missing_field."""
        from core.reporting import daily_report

        mock_created.return_value = []
        mock_bugs.return_value = {'field': 'fixVersion', 'flagged': [],
                                  'total_open_count': 0}
        mock_changes.return_value = {'automation': [], 'human': [], 'total': 0}

        jira = MagicMock()
        daily_report(jira, 'STL', '2026-03-12', missing_field='fixVersion')

        mock_bugs.assert_called_once_with(jira, 'STL', field='fixVersion',
                                          target_date='2026-03-12')


# ---------------------------------------------------------------------------
# export_daily_report — Excel
# ---------------------------------------------------------------------------

class TestExportExcel:
    def _make_report(self) -> dict:
        return {
            'date': '2026-03-12',
            'project': 'STL',
            'created_tickets': [
                _make_issue_dict(key='STL-200', summary='Created ticket A'),
                _make_issue_dict(key='STL-201', summary='Created ticket B'),
            ],
            'bugs_missing_field': {
                'field': 'affectedVersion',
                'flagged': [
                    _make_issue_dict(key='STL-210', summary='Bug without AV'),
                ],
                'total_open_count': 3,
            },
            'status_changes': {
                'automation': [
                    {'key': 'STL-220', 'from': 'Open', 'to': 'In Progress',
                     'author': 'Bot', 'email': 'bot@test.com',
                     'time': '2026-03-12T10:00:00.000+0000'},
                ],
                'human': [
                    {'key': 'STL-221', 'from': 'In Progress', 'to': 'Verify',
                     'author': 'Jane', 'email': 'jane@test.com',
                     'time': '2026-03-12T11:00:00.000+0000'},
                ],
                'total': 2,
            },
        }

    def test_creates_xlsx_file(self):
        """export_daily_report with fmt='excel' should create an .xlsx file."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report')
            result = export_daily_report(report, path, fmt='excel')

            assert result.endswith('.xlsx')
            assert os.path.exists(result)
            assert os.path.getsize(result) > 0

    def test_xlsx_has_expected_sheets(self):
        """Excel workbook should have Summary, Created Tickets, Bugs, Changes sheets."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report.xlsx')
            result = export_daily_report(report, path, fmt='excel')

            from openpyxl import load_workbook
            wb = load_workbook(result)
            sheet_names = wb.sheetnames

            assert 'Summary' in sheet_names
            assert 'Created Tickets' in sheet_names
            assert 'Bugs Missing Field' in sheet_names
            assert 'Status Changes' in sheet_names

    def test_xlsx_created_tickets_row_count(self):
        """Created Tickets sheet should have header + 2 data rows."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report.xlsx')
            export_daily_report(report, path, fmt='excel')

            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb['Created Tickets']
            # header row + 2 data rows = 3
            assert ws.max_row == 3

    def test_xlsx_adds_extension_if_missing(self):
        """If output_path doesn't end with .xlsx, it should be appended."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'no_extension')
            result = export_daily_report(report, path, fmt='excel')
            assert result.endswith('.xlsx')


# ---------------------------------------------------------------------------
# export_daily_report — CSV
# ---------------------------------------------------------------------------

class TestExportCsv:
    def _make_report(self) -> dict:
        return {
            'date': '2026-03-12',
            'project': 'STL',
            'created_tickets': [
                _make_issue_dict(key='STL-300', summary='CSV ticket A'),
            ],
            'bugs_missing_field': {
                'field': 'affectedVersion',
                'flagged': [
                    _make_issue_dict(key='STL-310', summary='CSV bug'),
                ],
                'total_open_count': 1,
            },
            'status_changes': {
                'automation': [],
                'human': [
                    {'key': 'STL-320', 'from': 'Open', 'to': 'Verify',
                     'author': 'Jane', 'email': 'jane@test.com',
                     'time': '2026-03-12T10:00:00.000+0000'},
                ],
                'total': 1,
            },
        }

    def test_creates_three_csv_files(self):
        """CSV export should create _created.csv, _bugs.csv, _changes.csv."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report')
            base = export_daily_report(report, path, fmt='csv')

            assert os.path.exists(f'{base}_created.csv')
            assert os.path.exists(f'{base}_bugs.csv')
            assert os.path.exists(f'{base}_changes.csv')

    def test_created_csv_content(self):
        """_created.csv should have header + 1 data row with correct key."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report')
            base = export_daily_report(report, path, fmt='csv')

            with open(f'{base}_created.csv', 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)

            assert len(rows) == 1
            assert rows[0]['key'] == 'STL-300'

    def test_bugs_csv_includes_missing_field_column(self):
        """_bugs.csv should include a 'missing_field' column."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report')
            base = export_daily_report(report, path, fmt='csv')

            with open(f'{base}_bugs.csv', 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)

            assert len(rows) == 1
            assert rows[0]['missing_field'] == 'affectedVersion'

    def test_changes_csv_has_is_automation_column(self):
        """_changes.csv should have an is_automation column."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report')
            base = export_daily_report(report, path, fmt='csv')

            with open(f'{base}_changes.csv', 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)
                rows = list(reader)

            assert 'is_automation' in headers
            assert len(rows) == 1
            # Human change → is_automation should be 'False'
            assert rows[0][-1] == 'False'

    def test_strips_csv_extension_from_base(self):
        """If output_path ends with .csv, it should be stripped for the base."""
        from core.reporting import export_daily_report

        report = self._make_report()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, 'test_report.csv')
            base = export_daily_report(report, path, fmt='csv')

            # base should NOT end with .csv
            assert not base.endswith('.csv')
            assert os.path.exists(f'{base}_created.csv')


# ---------------------------------------------------------------------------
# export_daily_report — invalid format
# ---------------------------------------------------------------------------

class TestExportInvalidFormat:
    def test_raises_on_unsupported_format(self):
        """export_daily_report should raise ValueError for unknown formats."""
        from core.reporting import export_daily_report

        with pytest.raises(ValueError, match='Unsupported export format'):
            export_daily_report({}, 'output', fmt='pdf')
