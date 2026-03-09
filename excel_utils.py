##########################################################################################
#
# Script name: excel_utils.py
#
# Description: Excel utilities for concatenating and manipulating .xlsx workbooks.
#              Designed for standalone CLI use and integration with the agent pipeline.
#
#              Also provides --to-plan-json to convert a flat or indented CSV/Excel
#              file into the feature-plan JSON format used by the Jira ticket
#              creation pipeline (see tools/plan_export_tools.py).
#
# Author: John Macdonald
#
# Usage:
#   python excel_utils.py --concat fileA.xlsx fileB.xlsx --method merge-sheet --output merged.xlsx
#   python excel_utils.py --concat fileA.xlsx fileB.xlsx --method add-sheet --output combined.xlsx
#   python excel_utils.py --to-plan-json plan.csv -o plan.json
#   python excel_utils.py --to-plan-json plan.xlsx --project XYZ -o plan.json
#
##########################################################################################

import argparse
import csv
import logging
import sys
import os
from collections import OrderedDict
from datetime import date

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from copy import copy
except ImportError:
    print('ERROR: openpyxl is required. Install with: pip install openpyxl')
    sys.exit(1)

# ****************************************************************************************
# Global data and configuration
# ****************************************************************************************

# Logging config
log = logging.getLogger(os.path.basename(sys.argv[0]))
log.setLevel(logging.DEBUG)

# File handler for logging
fh = logging.FileHandler('excel_utils.log', mode='w')
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter(
    '%(asctime)-15s [%(funcName)25s:%(lineno)-5s] %(levelname)-8s %(message)s')
fh.setFormatter(formatter)
log.addHandler(fh)

log.debug('Global data and configuration for this script...')

# Output control - set by handle_args()
_quiet_mode = False
_no_formatting = False

# Public API surface — used by `from excel_utils import *` and tooling introspection.
__all__ = [
    # Color maps
    'STATUS_FILL_COLORS', 'PRIORITY_FILL_COLORS',
    # Formatting
    '_apply_status_conditional_formatting',
    '_apply_priority_conditional_formatting',
    '_apply_header_style', '_auto_fit_columns',
    # Concatenation
    'concat_merge_sheet', 'concat_add_sheet',
    # Conversion
    'convert_to_csv', 'convert_from_csv',
    # Plan JSON conversion (CSV/Excel → feature-plan JSON)
    'convert_to_plan_json',
    # Diff
    'diff_files',
    # Dashboard
    '_create_dashboard_sheet',
    # Validation
    '_validate_and_repair_csv',
    # Exceptions
    'Error', 'ExcelFileError',
    # Display
    'output',
]


def output(message=''):
    '''
    Print user-facing output, respecting quiet mode.
    Always logs to file regardless of quiet mode.

    Input:
        message: String to output (default empty for blank line).

    Output:
        None; prints to stdout unless in quiet mode.

    Side Effects:
        Always logs message to log file at INFO level.
    '''
    # Log to file only (bypass stdout handler by writing directly to file handler)
    if message:
        record = logging.LogRecord(
            name=log.name,
            level=logging.INFO,
            pathname=__file__,
            lineno=0,
            msg=f'OUTPUT: {message}',
            args=(),
            exc_info=None,
            func='output'
        )
        fh.emit(record)

    # Print to stdout unless quiet mode
    if not _quiet_mode:
        print(message)


# ****************************************************************************************
# Status conditional formatting colors
# ****************************************************************************************

# Status-to-fill color mapping for Excel conditional formatting.
# These are embedded as dynamic Excel rules so they update if the user edits
# status values in the spreadsheet.
STATUS_FILL_COLORS = {
    'Open':        'CCE5FF',   # Light blue
    'In Progress': 'CCFFCC',   # Light green
    'Verify':      'E5CCFF',   # Light purple
    'Ready':       'FFFFCC',   # Light yellow
    'Closed':      'FFFFFF',   # White
}

# Priority-to-fill/font color mapping for Excel conditional formatting.
# Each entry maps a priority value to (fill_hex, font_hex).
PRIORITY_FILL_COLORS = {
    'P0-Stopper':  ('FF0000', 'FFFFFF'),   # Red fill, white text
    'P1-Critical': ('FFFF00', '000000'),   # Yellow fill, black text
}


# ****************************************************************************************
# Custom exceptions
# ****************************************************************************************

class Error(Exception):
    '''Base exception for excel_utils errors.'''
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)


class ExcelFileError(Error):
    '''Raised when an Excel file cannot be read or is invalid.'''
    def __init__(self, message):
        super().__init__(message)


# ****************************************************************************************
# Core functions
# ****************************************************************************************

def _read_sheet_data(ws):
    '''
    Read all data from a worksheet into a list of row dicts.

    Uses the first row as column headers. Returns (headers, rows, cell_formats)
    where headers is a list of strings, rows is a list of dicts keyed by header
    (values only), and cell_formats is a parallel list of dicts keyed by header
    containing per-cell formatting metadata (hyperlink, font, fill, alignment,
    border, number_format).

    Input:
        ws: openpyxl Worksheet object.

    Output:
        Tuple of (headers: list[str], rows: list[dict], cell_formats: list[dict]).
        Each entry in cell_formats is a dict mapping header -> dict with keys:
            hyperlink, font, fill, alignment, border, number_format.
        Formatting values are copies safe to apply to new cells.
    '''
    log.debug(f'Reading sheet: {ws.title} ({ws.max_row} rows, {ws.max_column} cols)')

    headers = []
    rows = []
    cell_formats = []

    for row_idx, row in enumerate(ws.iter_rows(), 1):
        if row_idx == 1:
            # First row is the header — extract display values
            headers = [str(cell.value) if cell.value is not None else f'Column_{i}'
                       for i, cell in enumerate(row, 1)]
            continue

        # Skip completely empty rows
        if all(cell.value is None for cell in row):
            continue

        row_dict = {}
        fmt_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                col_name = headers[col_idx]
                row_dict[col_name] = cell.value

                # Preserve formatting metadata for this cell
                fmt_dict[col_name] = {
                    'hyperlink': cell.hyperlink.target if cell.hyperlink else None,
                    'font': copy(cell.font) if cell.font else None,
                    'fill': copy(cell.fill) if cell.fill else None,
                    'alignment': copy(cell.alignment) if cell.alignment else None,
                    'border': copy(cell.border) if cell.border else None,
                    'number_format': cell.number_format,
                }

        rows.append(row_dict)
        cell_formats.append(fmt_dict)

    log.debug(f'Read {len(rows)} data rows with {len(headers)} columns from sheet "{ws.title}"')
    return headers, rows, cell_formats


def _apply_cell_format(dest_cell, fmt):
    '''
    Apply saved formatting metadata to a destination cell.

    Input:
        dest_cell: openpyxl Cell object to format.
        fmt: Dict with keys: hyperlink, font, fill, alignment, border, number_format.
             Any key may be None, in which case that property is skipped.

    Side Effects:
        Modifies the destination cell's formatting properties in place.
    '''
    if not fmt:
        return

    if fmt.get('hyperlink'):
        dest_cell.hyperlink = fmt['hyperlink']
        # If no explicit font was saved, apply a default link style
        if fmt.get('font'):
            dest_cell.font = copy(fmt['font'])
        else:
            dest_cell.font = Font(color='0563C1', underline='single')
    elif fmt.get('font'):
        dest_cell.font = copy(fmt['font'])

    if fmt.get('fill'):
        dest_cell.fill = copy(fmt['fill'])

    if fmt.get('alignment'):
        dest_cell.alignment = copy(fmt['alignment'])

    if fmt.get('border'):
        dest_cell.border = copy(fmt['border'])

    if fmt.get('number_format'):
        dest_cell.number_format = fmt['number_format']


def _load_excel_file(file_path):
    '''
    Load an Excel file and return the workbook.

    Input:
        file_path: Path to the .xlsx file.

    Output:
        openpyxl Workbook object.

    Raises:
        ExcelFileError: If the file cannot be read.
    '''
    log.debug(f'Loading Excel file: {file_path}')

    if not os.path.exists(file_path):
        raise ExcelFileError(f'File not found: {file_path}')

    try:
        wb = load_workbook(file_path, data_only=True)
        log.debug(f'Loaded workbook with sheets: {wb.sheetnames}')
        return wb
    except Exception as e:
        raise ExcelFileError(f'Failed to load Excel file "{file_path}": {e}')


def _apply_header_style(ws, num_cols):
    '''
    Apply standard header styling to the first row of a worksheet.

    Input:
        ws: openpyxl Worksheet object.
        num_cols: Number of columns to style.

    Side Effects:
        Modifies the first row of the worksheet with header formatting.
    '''
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _apply_status_conditional_formatting(ws, fieldnames):
    '''
    Add Excel conditional formatting rules to the status column.

    Each rule highlights the status cell with a fill color based on its value.
    Rules are dynamic — they are evaluated by Excel/LibreOffice when the file
    is opened, so they update if the user edits status values.

    Input:
        ws: openpyxl Worksheet object (already populated with data).
        fieldnames: List of column header names (to locate the status column).

    Side Effects:
        Adds conditional formatting rules to the worksheet.
    '''
    from openpyxl.formatting.rule import CellIsRule

    # Find the status column index (1-based)
    status_col_idx = None
    for idx, name in enumerate(fieldnames, 1):
        if name.lower() == 'status':
            status_col_idx = idx
            break

    if status_col_idx is None:
        log.debug('No "status" column found — skipping conditional formatting')
        return

    col_letter = get_column_letter(status_col_idx)
    # Apply rules from row 2 (skip header) to the last data row
    last_row = ws.max_row
    if last_row < 2:
        return

    cell_range = f'{col_letter}2:{col_letter}{last_row}'
    log.debug(f'Applying status conditional formatting to range {cell_range}')

    for status_value, hex_color in STATUS_FILL_COLORS.items():
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        rule = CellIsRule(
            operator='equal',
            formula=[f'"{status_value}"'],
            fill=fill,
        )
        ws.conditional_formatting.add(cell_range, rule)

    log.debug(f'Added {len(STATUS_FILL_COLORS)} conditional formatting rules for status column')


def _apply_priority_conditional_formatting(ws, fieldnames):
    '''
    Add Excel conditional formatting rules to the priority column.

    Highlights priority cells based on their value:
      - P0-Stopper:  red fill, white text
      - P1-Critical: yellow fill, black text

    Input:
        ws: openpyxl Worksheet object (already populated with data).
        fieldnames: List of column header names (to locate the priority column).

    Side Effects:
        Adds conditional formatting rules to the worksheet.
    '''
    from openpyxl.formatting.rule import CellIsRule

    # Find the priority column index (1-based)
    priority_col_idx = None
    for idx, name in enumerate(fieldnames, 1):
        if name.lower() == 'priority':
            priority_col_idx = idx
            break

    if priority_col_idx is None:
        log.debug('No "priority" column found — skipping priority conditional formatting')
        return

    col_letter = get_column_letter(priority_col_idx)
    last_row = ws.max_row
    if last_row < 2:
        return

    cell_range = f'{col_letter}2:{col_letter}{last_row}'
    log.debug(f'Applying priority conditional formatting to range {cell_range}')

    for priority_value, (fill_hex, font_hex) in PRIORITY_FILL_COLORS.items():
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
        font = Font(color=font_hex)
        rule = CellIsRule(
            operator='equal',
            formula=[f'"{priority_value}"'],
            fill=fill,
            font=font,
        )
        ws.conditional_formatting.add(cell_range, rule)

    log.debug(f'Added {len(PRIORITY_FILL_COLORS)} conditional formatting rules for priority column')


def _auto_fit_columns(ws):
    '''
    Auto-fit column widths based on content (approximate).

    Input:
        ws: openpyxl Worksheet object.

    Side Effects:
        Adjusts column widths in the worksheet.
    '''
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        # Sample up to 100 rows for width estimation
        for row_idx in range(1, min(ws.max_row + 1, 102)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        # Cap at 50 characters, minimum 10
        adjusted_width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def concat_merge_sheet(input_files, output_file):
    '''
    Concatenate multiple Excel files into a single worksheet.

    All rows from each file are appended into one sheet. Columns are unioned
    across all files: if fileA has columnA and fileB has columnB, the output
    sheet has both columnA and columnB with blanks where the source file did
    not have that column.

    Input:
        input_files: List of paths to .xlsx files.
        output_file: Path for the output .xlsx file.

    Output:
        None; writes the merged workbook to output_file.

    Raises:
        ExcelFileError: If any input file cannot be read.
    '''
    log.debug(f'Entering concat_merge_sheet(input_files={input_files}, output_file={output_file})')

    # Phase 1: Read all files and collect the union of all column headers
    # We preserve column order: columns from the first file come first, then
    # new columns from subsequent files are appended in the order encountered.
    all_file_data = []  # list of (filename, headers, rows, cell_formats)
    ordered_columns = []  # union of all headers, preserving first-seen order
    seen_columns = set()

    for file_path in input_files:
        wb = _load_excel_file(file_path)
        # Read the first (active) sheet — preserving cell formatting
        ws = wb.active
        headers, rows, cell_formats = _read_sheet_data(ws)
        all_file_data.append((os.path.basename(file_path), headers, rows, cell_formats))
        wb.close()

        # Add new columns to the ordered list
        for h in headers:
            if h not in seen_columns:
                ordered_columns.append(h)
                seen_columns.add(h)

    log.debug(f'Union of columns ({len(ordered_columns)}): {ordered_columns}')

    # Phase 2: Build the output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Merged'

    # Write header row
    for col_idx, col_name in enumerate(ordered_columns, 1):
        out_ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows from all files, preserving per-cell formatting
    current_row = 2
    total_rows = 0

    for filename, headers, rows, cell_formats in all_file_data:
        log.debug(f'Writing {len(rows)} rows from "{filename}"')
        for row_idx, row_data in enumerate(rows):
            # Get the parallel formatting dict for this row (if available)
            fmt_row = cell_formats[row_idx] if row_idx < len(cell_formats) else {}

            for col_idx, col_name in enumerate(ordered_columns, 1):
                value = row_data.get(col_name, None)
                cell = out_ws.cell(row=current_row, column=col_idx, value=value)

                # Apply original cell formatting (hyperlinks, font, fill, etc.)
                cell_fmt = fmt_row.get(col_name)
                if cell_fmt:
                    _apply_cell_format(cell, cell_fmt)

            current_row += 1
            total_rows += 1

    # Apply styling (unless --no-formatting)
    if not _no_formatting:
        _apply_header_style(out_ws, len(ordered_columns))
        _auto_fit_columns(out_ws)

        # Freeze header row and add auto-filter
        out_ws.freeze_panes = 'A2'
        out_ws.auto_filter.ref = out_ws.dimensions

        # Apply conditional formatting (dynamic Excel rules)
        _apply_status_conditional_formatting(out_ws, ordered_columns)
        _apply_priority_conditional_formatting(out_ws, ordered_columns)

    # Save
    out_wb.save(output_file)
    log.info(f'Wrote {total_rows} rows ({len(ordered_columns)} columns) to: {output_file}')
    log.debug(f'Input files: {len(input_files)}')
    for f in input_files:
        log.debug(f'  - {f}')
    log.debug(f'Output file: {output_file}')
    log.debug(f'Total rows: {total_rows}')
    log.debug(f'Total columns: {len(ordered_columns)}')

    output('')
    output('=' * 80)
    output('Excel Merge Complete (merge-sheet)')
    output('=' * 80)
    output(f'Input files:    {len(input_files)}')
    for f in input_files:
        output(f'  - {f}')
    output(f'Output file:    {output_file}')
    output(f'Total rows:     {total_rows}')
    output(f'Total columns:  {len(ordered_columns)}')
    output('=' * 80)
    output('')


def concat_add_sheet(input_files, output_file):
    '''
    Concatenate multiple Excel files by adding each as a separate worksheet.

    Each input file becomes a new sheet in the output workbook. The sheet name
    is derived from the input filename (without extension), truncated to 31
    characters (Excel sheet name limit). Duplicate sheet names get a numeric
    suffix.

    Input:
        input_files: List of paths to .xlsx files.
        output_file: Path for the output .xlsx file.

    Output:
        None; writes the combined workbook to output_file.

    Raises:
        ExcelFileError: If any input file cannot be read.
    '''
    log.debug(f'Entering concat_add_sheet(input_files={input_files}, output_file={output_file})')

    out_wb = Workbook()
    # Remove the default empty sheet (we'll add named sheets)
    default_ws = out_wb.active
    out_wb.remove(default_ws)

    used_names = set()

    for file_path in input_files:
        wb = _load_excel_file(file_path)
        ws = wb.active
        headers, rows, cell_formats = _read_sheet_data(ws)
        wb.close()

        # Derive sheet name from filename (without extension), max 31 chars
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        sheet_name = base_name[:31]

        # Handle duplicate sheet names by appending a numeric suffix
        if sheet_name in used_names:
            suffix = 2
            while f'{sheet_name[:28]}_{suffix}' in used_names:
                suffix += 1
            sheet_name = f'{sheet_name[:28]}_{suffix}'
        used_names.add(sheet_name)

        log.debug(f'Adding sheet "{sheet_name}" with {len(rows)} rows from "{file_path}"')

        out_ws = out_wb.create_sheet(title=sheet_name)

        # Write header row
        for col_idx, col_name in enumerate(headers, 1):
            out_ws.cell(row=1, column=col_idx, value=col_name)

        # Write data rows, preserving per-cell formatting
        for row_idx, row_data in enumerate(rows, 2):
            # Get the parallel formatting dict for this row (if available)
            data_idx = row_idx - 2  # rows list is 0-based
            fmt_row = cell_formats[data_idx] if data_idx < len(cell_formats) else {}

            for col_idx, col_name in enumerate(headers, 1):
                value = row_data.get(col_name, None)
                cell = out_ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply original cell formatting (hyperlinks, font, fill, etc.)
                cell_fmt = fmt_row.get(col_name)
                if cell_fmt:
                    _apply_cell_format(cell, cell_fmt)

        # Apply styling (unless --no-formatting)
        if not _no_formatting:
            _apply_header_style(out_ws, len(headers))
            _auto_fit_columns(out_ws)

            # Freeze header row and add auto-filter
            out_ws.freeze_panes = 'A2'
            if out_ws.dimensions:
                out_ws.auto_filter.ref = out_ws.dimensions

            # Apply conditional formatting (dynamic Excel rules)
            _apply_status_conditional_formatting(out_ws, headers)
            _apply_priority_conditional_formatting(out_ws, headers)

    # Save
    out_wb.save(output_file)
    log.info(f'Wrote {len(input_files)} sheets to: {output_file}')
    log.debug(f'Input files: {len(input_files)}')
    for f in input_files:
        log.debug(f'  - {f}')
    log.debug(f'Output file: {output_file}')
    log.debug(f'Sheets: {", ".join(used_names)}')

    output('')
    output('=' * 80)
    output('Excel Merge Complete (add-sheet)')
    output('=' * 80)
    output(f'Input files:  {len(input_files)}')
    for f in input_files:
        output(f'  - {f}')
    output(f'Output file:  {output_file}')
    output(f'Sheets:       {", ".join(used_names)}')
    output('=' * 80)
    output('')


# ****************************************************************************************
# CSV validation and repair
# ****************************************************************************************

def _validate_and_repair_csv(input_file):
    '''
    Validate a CSV file for column-count consistency and attempt to repair
    misaligned rows in-place.

    LLM-generated CSV files frequently contain rows where a field (typically
    summary, assignee, or fix_version) has an unquoted comma, producing one
    or more extra delimiter fields.  This shifts every subsequent column to
    the right and breaks the Excel output.

    The repair strategy for rows with TOO MANY fields:
      1. Identify "anchor" columns whose values are highly recognisable
         (e.g. Jira keys like STL-NNNNN, issue types like "Bug", status
         values like "In Progress", priority values like "P0-Stopper").
      2. Walk the row fields and find the best alignment of anchors to
         headers.  Where extra fields appear between two anchors, merge
         them back into a single quoted value (they were one cell that
         got split by a bare comma).
      3. If heuristic alignment fails, fall back to merging the LAST
         extra fields (rightmost), which is usually the summary or
         fix_version — the most common offenders.

    For rows with TOO FEW fields: pad with empty strings on the right.

    Input:
        input_file: Path to the .csv file (modified in-place when repairs
                    are needed).

    Output:
        Tuple of (repaired: bool, stats: dict).
        stats keys: total_rows, ok_rows, repaired_rows, padded_rows,
                    unfixable_rows.

    Side Effects:
        Overwrites input_file with the repaired CSV when any rows are
        changed.  The original is logged but not preserved (the caller
        should keep the LLM raw output in llm_output.md).
    '''
    log.debug(f'Entering _validate_and_repair_csv(input_file={input_file})')

    # ------------------------------------------------------------------
    # Phase 1: Read raw lines and parse with csv.reader (not DictReader)
    #          so we get positional field lists.
    # ------------------------------------------------------------------
    with open(input_file, 'r', encoding='utf-8', newline='') as f:
        raw_rows = list(csv.reader(f))

    if len(raw_rows) < 2:
        log.debug('CSV has fewer than 2 rows — nothing to validate')
        return False, {'total_rows': max(len(raw_rows) - 1, 0),
                       'ok_rows': max(len(raw_rows) - 1, 0),
                       'repaired_rows': 0, 'padded_rows': 0,
                       'unfixable_rows': 0}

    header = raw_rows[0]
    expected = len(header)
    log.debug(f'CSV header has {expected} columns: {header}')

    # ------------------------------------------------------------------
    # Build anchor-detection helpers.
    # For each header position, define a recogniser function that returns
    # True when a cell value "looks right" for that column.
    # ------------------------------------------------------------------
    # Jira key pattern: PROJECT-DIGITS (e.g. STL-76636)
    import re as _re
    _jira_key_re = _re.compile(r'^[A-Z]{2,10}-\d+$')

    # Known categorical values per column name (lowered).
    _known_values = {
        'issue_type': {'bug', 'story', 'task', 'epic', 'sub-task', 'subtask',
                       'improvement', 'new feature', 'change request'},
        'status':     {'open', 'in progress', 'closed', 'verify', 'ready',
                       'to do', 'done', 'resolved', 'reopened', 'in review'},
        'priority':   {'p0-stopper', 'p1-critical', 'p2-major', 'p3-minor',
                       'p4-trivial', 'blocker', 'critical', 'major', 'minor',
                       'trivial'},
        'project':    {'stl', 'stlsb', 'cn', 'opx'},
        'product':    {'nic', 'switch'},
        'module':     {'driver', 'bts', 'fw', 'opx', 'gpu'},
    }

    def _score_alignment(fields, header_list):
        '''Return a score (higher = better) for how well *fields* align to *header_list*.'''
        score = 0
        for i, hdr in enumerate(header_list):
            if i >= len(fields):
                break
            val = (fields[i] or '').strip()
            hdr_low = hdr.strip().lower()

            # Jira key column
            if hdr_low == 'key' and _jira_key_re.match(val):
                score += 10
            # Known categorical columns
            elif hdr_low in _known_values and val.lower() in _known_values[hdr_low]:
                score += 5
            # Date-like column (updated)
            elif hdr_low == 'updated' and _re.match(r'^\d{4}-\d{2}-\d{2}', val):
                score += 5
            # Non-empty value in a column that usually has data
            elif val and hdr_low in ('customer', 'summary', 'assignee', 'fix_version'):
                score += 1
        return score

    # ------------------------------------------------------------------
    # Phase 2: Check each data row and repair if needed.
    # ------------------------------------------------------------------
    stats = {'total_rows': len(raw_rows) - 1, 'ok_rows': 0,
             'repaired_rows': 0, 'padded_rows': 0, 'unfixable_rows': 0}
    any_changed = False

    for row_idx in range(1, len(raw_rows)):
        fields = raw_rows[row_idx]
        n = len(fields)

        if n == expected:
            stats['ok_rows'] += 1
            continue

        if n < expected:
            # Too few fields — pad with empty strings on the right.
            log.warning(f'CSV row {row_idx + 1}: {n} fields (expected {expected}) — '
                        f'padding {expected - n} empty field(s)')
            fields.extend([''] * (expected - n))
            raw_rows[row_idx] = fields
            stats['padded_rows'] += 1
            any_changed = True
            continue

        # Too many fields — attempt heuristic merge.
        extra = n - expected
        log.warning(f'CSV row {row_idx + 1}: {n} fields (expected {expected}), '
                    f'{extra} extra — attempting merge repair')

        # Strategy: try every possible way to merge `extra` adjacent field
        # pairs and pick the merge that produces the best anchor score.
        # For efficiency, limit to merging at most 3 extra fields (covers
        # the vast majority of LLM CSV errors).
        best_score = -1
        best_fields = None

        if extra <= 4:
            # Generate candidate merges.  Each candidate is defined by a
            # set of merge-start indices: at each such index i, fields[i]
            # and fields[i+1] are joined with a comma.  We need exactly
            # `extra` merges.
            from itertools import combinations
            merge_candidates = list(range(n - 1))  # possible merge points
            for merge_points in combinations(merge_candidates, extra):
                # Build merged field list
                candidate = []
                skip_until = -1
                i = 0
                while i < n:
                    if i in merge_points:
                        # Merge this field with the next one(s) in a
                        # contiguous run of merge points starting at i.
                        merged = fields[i]
                        while i in merge_points:
                            i += 1
                            if i < n:
                                merged += ',' + fields[i]
                        candidate.append(merged)
                    else:
                        candidate.append(fields[i])
                    i += 1

                if len(candidate) != expected:
                    continue  # shouldn't happen, but guard

                score = _score_alignment(candidate, header)
                if score > best_score:
                    best_score = score
                    best_fields = candidate
        else:
            # Too many extras for combinatorial search — fall back to
            # merging the rightmost extra fields into the last-but-one
            # column (usually summary or fix_version).
            # Find the last anchor column from the right to anchor the
            # merge boundary.
            best_fields = fields[:expected - 1]
            # Join all remaining fields into the last column
            best_fields.append(','.join(fields[expected - 1:]))

        if best_fields and len(best_fields) == expected:
            raw_rows[row_idx] = best_fields
            stats['repaired_rows'] += 1
            any_changed = True
            log.info(f'CSV row {row_idx + 1}: repaired by merging {extra} extra field(s) '
                     f'(alignment score={best_score})')
        else:
            # Could not repair — leave as-is (DictReader will handle overflow)
            stats['unfixable_rows'] += 1
            log.warning(f'CSV row {row_idx + 1}: could not repair {extra} extra field(s)')

    # ------------------------------------------------------------------
    # Phase 3: Write back if anything changed.
    # ------------------------------------------------------------------
    if any_changed:
        log.info(f'Rewriting repaired CSV: {input_file} '
                 f'(repaired={stats["repaired_rows"]}, padded={stats["padded_rows"]})')
        with open(input_file, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(raw_rows)
    else:
        log.debug('All CSV rows have correct column count — no repairs needed')

    return any_changed, stats


# ****************************************************************************************
# Convert functions
# ****************************************************************************************

def convert_to_csv(input_file, output_file=None):
    '''
    Convert an Excel (.xlsx) file to a comma-delimited CSV file.

    Reads the first (active) sheet and writes all rows to CSV. The header row
    is preserved as the first CSV row.

    Input:
        input_file: Path to the .xlsx file.
        output_file: Optional path for the output .csv file. If None, the
                     output filename is derived from the input filename by
                     replacing the extension with .csv.

    Output:
        None; writes the CSV file to disk.

    Raises:
        ExcelFileError: If the input file cannot be read.
    '''
    log.debug(f'Entering convert_to_csv(input_file={input_file}, output_file={output_file})')

    wb = _load_excel_file(input_file)
    ws = wb.active
    headers, rows, _ = _read_sheet_data(ws)
    wb.close()

    # Derive output filename if not provided
    if not output_file:
        base = os.path.splitext(input_file)[0]
        output_file = f'{base}.csv'

    # Ensure .csv extension
    if not output_file.endswith('.csv'):
        output_file = f'{output_file}.csv'

    log.debug(f'Writing {len(rows)} rows to CSV: {output_file}')

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    log.info(f'Converted "{input_file}" to CSV: {output_file} ({len(rows)} rows)')
    log.debug(f'Input file: {input_file}')
    log.debug(f'Output file: {output_file}')
    log.debug(f'Rows: {len(rows)}')
    log.debug(f'Columns: {len(headers)}')

    output('')
    output('=' * 80)
    output('Excel → CSV Conversion Complete')
    output('=' * 80)
    output(f'Input file:   {input_file}')
    output(f'Output file:  {output_file}')
    output(f'Rows:         {len(rows)}')
    output(f'Columns:      {len(headers)}')
    output('=' * 80)
    output('')


def _create_dashboard_sheet(wb, ws_data, headers, dashboard_columns=None):
    '''
    Create a "Dashboard" sheet with dynamic, formula-driven summary tables that
    count occurrences of key categorical columns from the data sheet.

    Each summary table uses COUNTIF formulas referencing the data sheet so that
    the dashboard updates automatically when data values change.  Tables are
    laid out side-by-side horizontally with a gap column between them.

    Input:
        wb: openpyxl Workbook to add the Dashboard sheet to.
        ws_data: The data worksheet to read values from.
        headers: List of column header names from the data sheet.
        dashboard_columns: Optional list of column names to summarize.  Names
                           are matched case-insensitively against the CSV
                           headers.  When None or empty, no dashboard is
                           created.

    Output:
        None; adds a "Dashboard" sheet to the workbook.

    Side Effects:
        Creates and populates a new worksheet named "Dashboard" in the workbook.
    '''
    log.debug('Entering _create_dashboard_sheet()')

    # If no dashboard columns requested, skip entirely.
    if not dashboard_columns:
        log.debug('Dashboard: no dashboard_columns specified — skipping dashboard sheet')
        return

    # Build a case-insensitive header lookup: lowered_name -> (col_index_1based, original_name)
    header_lookup = {}
    for idx, h in enumerate(headers, 1):
        header_lookup[h.strip().lower()] = (idx, h)

    # The data sheet name (quoted for use in formulas with special characters).
    data_sheet_name = ws_data.title
    # Excel sheet references with spaces or special chars need single-quoting.
    quoted_sheet = f"'{data_sheet_name}'" if (' ' in data_sheet_name or
                                               "'" in data_sheet_name) else data_sheet_name

    # Collect distinct values and column references for each requested category.
    # category_data entries: (display_title, col_letter, sorted_distinct_values)
    category_data = []
    for col_name in dashboard_columns:
        match_key = col_name.strip().lower()
        entry = header_lookup.get(match_key)
        if entry is None:
            log.debug(f'Dashboard: column "{col_name}" not found in headers — skipping')
            continue

        col_idx, original_name = entry
        col_letter = get_column_letter(col_idx)

        # Collect distinct values from the data rows (row 2 onwards).
        counts = {}
        for row_idx in range(2, ws_data.max_row + 1):
            val = ws_data.cell(row=row_idx, column=col_idx).value
            val_str = str(val).strip() if val else '(blank)'
            if not val_str:
                val_str = '(blank)'
            counts[val_str] = counts.get(val_str, 0) + 1

        # Sort by count descending, then alphabetically for deterministic layout.
        sorted_values = sorted(counts.items(), key=lambda x: (-x[1], x[0]))
        # Use the original header name (preserving case) as the display title.
        category_data.append((original_name, col_letter, sorted_values))
        log.debug(f'Dashboard: "{original_name}" ({col_letter}) has '
                  f'{len(sorted_values)} distinct values')

    if not category_data:
        log.debug('Dashboard: no matching columns found — skipping dashboard sheet')
        return

    # Create the Dashboard sheet
    ws_dash = wb.create_sheet(title='Dashboard')

    # Determine the data range extent (last data row on the data sheet).
    last_data_row = ws_data.max_row

    # Styling
    title_font = Font(bold=True, size=12, color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # Layout: tables side-by-side starting at row 2, with a gap column between.
    # Row 1 is reserved for a dashboard title.
    ws_dash.cell(row=1, column=1, value='Dashboard').font = Font(
        bold=True, size=14, color='1F4E79')

    current_col = 1  # Starting column for the first table

    for display_title, data_col_letter, sorted_values in category_data:
        start_col = current_col

        # Table title row (row 2)
        cell = ws_dash.cell(row=2, column=start_col, value=display_title)
        cell.font = title_font

        # Table header row (row 3): category name | Count
        hdr_val = ws_dash.cell(row=3, column=start_col, value=display_title)
        hdr_val.font = header_font
        hdr_val.fill = header_fill
        hdr_val.alignment = header_alignment
        hdr_val.border = thin_border

        hdr_count = ws_dash.cell(row=3, column=start_col + 1, value='Count')
        hdr_count.font = header_font
        hdr_count.fill = header_fill
        hdr_count.alignment = header_alignment
        hdr_count.border = thin_border

        # Data range on the data sheet for COUNTIF (e.g. 'Sheet1'!E$2:E$500).
        # Using absolute row references so the formula is stable.
        data_range = (f'{quoted_sheet}!{data_col_letter}$2:'
                      f'{data_col_letter}${last_data_row}')

        # Data rows (row 4 onwards) — each count cell is a COUNTIF formula.
        for i, (val, _static_count) in enumerate(sorted_values):
            row_num = 4 + i

            # Value cell: the distinct category value
            val_cell = ws_dash.cell(row=row_num, column=start_col, value=val)
            val_cell.border = thin_border

            # Count cell: dynamic COUNTIF formula referencing the data sheet.
            # For "(blank)" entries we use COUNTBLANK instead.
            if val == '(blank)':
                formula = f'=COUNTBLANK({data_range})'
            else:
                # Escape double-quotes inside the value for the Excel formula.
                escaped_val = val.replace('"', '""')
                formula = f'=COUNTIF({data_range},"{escaped_val}")'

            count_cell = ws_dash.cell(row=row_num, column=start_col + 1)
            count_cell.value = formula
            count_cell.border = thin_border
            count_cell.alignment = Alignment(horizontal='center')

        # Total row: SUM of the count column above.
        total_row = 4 + len(sorted_values)
        count_col_letter = get_column_letter(start_col + 1)
        sum_range = f'{count_col_letter}4:{count_col_letter}{total_row - 1}'

        total_label = ws_dash.cell(row=total_row, column=start_col, value='Total')
        total_label.font = total_font
        total_label.fill = total_fill
        total_label.border = thin_border

        total_value = ws_dash.cell(row=total_row, column=start_col + 1)
        total_value.value = f'=SUM({sum_range})'
        total_value.font = total_font
        total_value.fill = total_fill
        total_value.border = thin_border
        total_value.alignment = Alignment(horizontal='center')

        # Auto-fit the two columns for this table
        max_val_len = max((len(str(v)) for v, _ in sorted_values), default=len(display_title))
        max_val_len = max(max_val_len, len(display_title))
        ws_dash.column_dimensions[get_column_letter(start_col)].width = min(max_val_len + 4, 35)
        ws_dash.column_dimensions[get_column_letter(start_col + 1)].width = 10

        # Move to next table position (2 data columns + 1 gap column)
        current_col = start_col + 3

    log.info(f'Created Dashboard sheet with {len(category_data)} formula-driven summary tables')


# Default Jira URL — used when callers don't supply an explicit jira_base_url.
DEFAULT_JIRA_BASE_URL = 'https://cornelisnetworks.atlassian.net'


def convert_from_csv(input_file, output_file=None, jira_base_url=DEFAULT_JIRA_BASE_URL,
                     dashboard_columns=None):
    '''
    Convert a comma-delimited CSV file to an Excel (.xlsx) file.

    Reads the CSV, creates a styled workbook with header formatting,
    auto-fit columns, frozen header row, auto-filter, and status
    conditional formatting (if a status column is present).

    When jira_base_url is provided, any column whose header contains "key"
    (case-insensitive) will have its cells rendered as clickable Jira
    hyperlinks (e.g. https://cornelisnetworks.atlassian.net/browse/STL-76582).

    When dashboard_columns is provided, a "Dashboard" sheet is added with
    formula-driven summary tables (COUNTIF) for each listed column.  Column
    names are matched case-insensitively against the CSV headers.

    Input:
        input_file: Path to the .csv file.
        output_file: Optional path for the output .xlsx file. If None, the
                     output filename is derived from the input filename by
                     replacing the extension with .xlsx.
        jira_base_url: Jira instance URL (e.g.
                       "https://cornelisnetworks.atlassian.net"). Defaults to
                       the Cornelis Atlassian URL.  Ticket-key cells become
                       clickable hyperlinks.  Pass None to disable links.
        dashboard_columns: Optional list of column name strings to create
                           summary tables for on a Dashboard sheet.

    Output:
        str: The resolved output file path.

    Raises:
        ExcelFileError: If the input file cannot be read.
    '''
    log.debug(f'Entering convert_from_csv(input_file={input_file}, output_file={output_file}, '
              f'jira_base_url={jira_base_url}, dashboard_columns={dashboard_columns})')

    if not os.path.exists(input_file):
        raise ExcelFileError(f'File not found: {input_file}')

    # ---- Validate and repair CSV column alignment before parsing ----
    # LLM-generated CSV files frequently have rows with unquoted commas in
    # fields like summary, assignee, or fix_version.  This causes extra
    # delimiters that shift columns to the right.  _validate_and_repair_csv
    # detects and merges those split fields back together so that every row
    # has exactly the same number of columns as the header.
    try:
        repaired, repair_stats = _validate_and_repair_csv(input_file)
        if repaired:
            log.info(f'CSV repair applied: {repair_stats}')
            output(f'  CSV repair: fixed {repair_stats["repaired_rows"]} row(s) with '
                   f'extra columns, padded {repair_stats["padded_rows"]} row(s) with '
                   f'missing columns'
                   + (f', {repair_stats["unfixable_rows"]} row(s) could not be repaired'
                      if repair_stats['unfixable_rows'] else ''))
    except Exception as repair_err:
        # Repair is best-effort — log the error but continue with the
        # original CSV so the pipeline is not blocked.
        log.warning(f'CSV repair failed (continuing with original): {repair_err}')

    # Read CSV
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = list(reader)
    except Exception as e:
        raise ExcelFileError(f'Failed to read CSV file "{input_file}": {e}')

    if not headers:
        raise ExcelFileError(f'CSV file "{input_file}" has no header row')

    # Derive output filename if not provided
    if not output_file:
        base = os.path.splitext(input_file)[0]
        output_file = f'{base}.xlsx'

    # Ensure .xlsx extension
    if not output_file.endswith('.xlsx'):
        output_file = f'{output_file}.xlsx'

    log.debug(f'Writing {len(rows)} rows to Excel: {output_file}')

    # Normalise the Jira base URL (strip trailing slash) when provided so we
    # can build browse links like  <base>/browse/STL-76582.
    jira_url = jira_base_url.rstrip('/') if jira_base_url else None

    # Identify which column indices (0-based) hold Jira ticket keys.
    # Convention: any header whose lowered name is exactly "key" qualifies.
    key_col_indices = {
        idx for idx, h in enumerate(headers) if h.strip().lower() == 'key'
    }
    if jira_url and key_col_indices:
        log.debug(f'Jira hyperlink columns (0-based): {sorted(key_col_indices)}')

    # Font style for hyperlink cells (blue, underlined)
    link_font = Font(color='0563C1', underline='single')

    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(input_file))[0][:31]

    # Write header row
    for col_idx, col_name in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows — ticket-key cells become clickable Jira links when
    # jira_base_url was supplied and the column header is "key".
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(headers, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Render ticket key as a clickable hyperlink
            if jira_url and (col_idx - 1) in key_col_indices and value:
                cell.hyperlink = f'{jira_url}/browse/{value}'
                cell.font = link_font

    # Apply styling (unless --no-formatting)
    if not _no_formatting:
        _apply_header_style(ws, len(headers))
        _auto_fit_columns(ws)

        # Freeze header row and add auto-filter
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # Apply conditional formatting
        _apply_status_conditional_formatting(ws, headers)
        _apply_priority_conditional_formatting(ws, headers)

    # Create a Dashboard summary sheet with formula-driven pivot tables for
    # the requested columns.  The dashboard uses COUNTIF formulas that
    # reference the data sheet so values update automatically.
    if not _no_formatting and dashboard_columns:
        _create_dashboard_sheet(wb, ws, headers, dashboard_columns=dashboard_columns)

    wb.save(output_file)
    log.info(f'Converted "{input_file}" to Excel: {output_file} ({len(rows)} rows)')
    log.debug(f'Input file: {input_file}')
    log.debug(f'Output file: {output_file}')
    log.debug(f'Rows: {len(rows)}')
    log.debug(f'Columns: {len(headers)}')

    output('')
    output('=' * 80)
    output('CSV → Excel Conversion Complete')
    output('=' * 80)
    output(f'Input file:   {input_file}')
    output(f'Output file:  {output_file}')
    output(f'Rows:         {len(rows)}')
    output(f'Columns:      {len(headers)}')
    output('=' * 80)
    output('')

    # Return the resolved output path so callers (e.g., workflow automation) can
    # chain subsequent steps without re-deriving the filename.
    return output_file


# ****************************************************************************************
# Plan JSON conversion (CSV/Excel → feature-plan JSON)
# ****************************************************************************************

def convert_to_plan_json(input_file, output_file=None, project_key='',
                         product_family='', feature_name=''):
    '''
    Convert a flat or indented CSV/Excel file into the feature-plan JSON format
    used by the Jira ticket creation pipeline.

    This delegates to the core conversion logic in tools/plan_export_tools.py
    (plan_file_to_json / write_plan_json) but provides a standalone entry point
    that does not require the agent framework.

    The table format (flat vs indented) is auto-detected from the column headers:
      - "Indented" format has columns named ``Depth 0``, ``Depth 1``, … instead
        of a ``key`` column.
      - "Flat" format has a ``key`` column and optionally a ``depth`` column.

    Input:
        input_file:     Path to a .csv or .xlsx file containing a feature plan.
        output_file:    Path for the output JSON file.  If None, derived from
                        input_file by replacing the extension with .json.
        project_key:    Override the Jira project key (auto-detected from rows
                        if empty).
        product_family: Override the product family (auto-detected if empty).
        feature_name:   Feature name for the plan (defaults to first epic summary).

    Output:
        The resolved output file path.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If the file extension is unsupported or the file is empty.
        ImportError: If openpyxl is needed but not installed (for .xlsx files).
    '''
    log.debug(f'Entering convert_to_plan_json(input_file={input_file})')

    if not os.path.exists(input_file):
        raise ExcelFileError(f'Input file not found: {input_file}')

    # Derive output path if not provided
    if not output_file:
        base, _ = os.path.splitext(input_file)
        output_file = f'{base}.json'

    # Import the plan_export_tools conversion functions.
    # These live in tools/ but are pure-Python with no agent framework dependency.
    try:
        from tools.plan_export_tools import plan_file_to_json, write_plan_json
    except ImportError:
        # Fallback: if tools package is not on the path, try a direct import.
        # This handles the case where excel_utils.py is run from the repo root.
        import importlib.util
        _spec = importlib.util.spec_from_file_location(
            'plan_export_tools',
            os.path.join(os.path.dirname(__file__), 'tools', 'plan_export_tools.py'),
        )
        if _spec is None or _spec.loader is None:
            raise ImportError(
                'Cannot import plan_export_tools. Ensure tools/plan_export_tools.py '
                'is available on the Python path.'
            )
        _mod = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        plan_file_to_json = _mod.plan_file_to_json
        write_plan_json = _mod.write_plan_json

    # Convert the CSV/Excel file to a plan dict
    plan = plan_file_to_json(
        input_file,
        project_key=project_key,
        product_family=product_family,
        feature_name=feature_name,
    )

    # Write the plan dict to JSON
    write_plan_json(plan, output_file)

    # Summary output
    n_epics = plan.get('total_epics', 0)
    n_stories = plan.get('total_stories', 0)
    n_tickets = plan.get('total_tickets', 0)

    output('')
    output('=' * 80)
    output(f'Plan JSON Conversion')
    output('=' * 80)
    output(f'Input file:   {input_file}')
    output(f'Output file:  {output_file}')
    output(f'Project:      {plan.get("project_key", "?")}')
    output(f'Feature:      {plan.get("feature_name", "?")}')
    output(f'Product:      {plan.get("product_family", "?")}')
    output(f'Epics:        {n_epics}')
    output(f'Stories:      {n_stories}')
    output(f'Total tickets:{n_tickets}')
    output('=' * 80)
    output('')

    return output_file


# ****************************************************************************************
# Diff function
# ****************************************************************************************

def diff_files(input_files, output_file=None):
    '''
    Diff two or more Excel (.xlsx) files and produce a summary report.

    Compares files by building a union of all column headers, then comparing
    rows across files. The diff output is an Excel workbook with:
      - A "Summary" sheet showing file-level statistics
      - A "Diff" sheet showing row-by-row differences with a source column

    Rows are matched by a key column. The key column is determined by priority:
      1. "key" (Jira ticket key)
      2. First column in the union header set

    For each row, the diff marks it as:
      - ADDED:   row exists in file B but not file A
      - REMOVED: row exists in file A but not file B
      - CHANGED: row exists in both but values differ (changed cells noted)
      - SAME:    row exists in both with identical values

    When more than 2 files are provided, each consecutive pair is diffed
    (file1 vs file2, file2 vs file3, etc.) and results are combined.

    Input:
        input_files: List of paths to .xlsx files (minimum 2).
        output_file: Optional path for the output .xlsx diff report.
                     Defaults to "diff_output.xlsx".

    Output:
        None; writes the diff report to disk.

    Raises:
        ExcelFileError: If any input file cannot be read.
    '''
    log.debug(f'Entering diff_files(input_files={input_files}, output_file={output_file})')

    if not output_file:
        output_file = 'diff_output.xlsx'
    if not output_file.endswith('.xlsx'):
        output_file = f'{output_file}.xlsx'

    # Phase 1: Read all files
    all_file_data = []  # list of (filename, headers, rows_as_dicts)
    all_headers = OrderedDict()  # preserves first-seen order

    for file_path in input_files:
        wb = _load_excel_file(file_path)
        ws = wb.active
        headers, rows, _ = _read_sheet_data(ws)
        wb.close()
        fname = os.path.basename(file_path)
        all_file_data.append((fname, headers, rows))
        for h in headers:
            all_headers[h] = True

    union_headers = list(all_headers.keys())

    # Determine the key column for matching rows
    key_col = None
    for candidate in ['key', 'Key', 'KEY']:
        if candidate in union_headers:
            key_col = candidate
            break
    if key_col is None:
        # Fall back to the first column
        key_col = union_headers[0] if union_headers else None

    log.debug(f'Using key column: {key_col}')
    log.debug(f'Union headers ({len(union_headers)}): {union_headers}')

    # Phase 2: Perform pairwise diffs
    diff_rows = []  # list of dicts with _diff_status, _diff_source, _diff_details + data cols
    summary_rows = []  # per-pair summary stats

    for i in range(len(all_file_data) - 1):
        fname_a, headers_a, rows_a = all_file_data[i]
        fname_b, headers_b, rows_b = all_file_data[i + 1]

        # Index rows by key
        index_a = {}
        for row in rows_a:
            k = str(row.get(key_col, '')) if key_col else ''
            if k:
                index_a[k] = row

        index_b = {}
        for row in rows_b:
            k = str(row.get(key_col, '')) if key_col else ''
            if k:
                index_b[k] = row

        pair_label = f'{fname_a} → {fname_b}'
        counts = {'added': 0, 'removed': 0, 'changed': 0, 'same': 0}

        # Check rows in A
        for k, row_a in index_a.items():
            if k in index_b:
                row_b = index_b[k]
                # Compare values across union headers
                changes = []
                for h in union_headers:
                    val_a = str(row_a.get(h, '') or '')
                    val_b = str(row_b.get(h, '') or '')
                    if val_a != val_b:
                        changes.append(f'{h}: "{val_a}" → "{val_b}"')

                if changes:
                    diff_row = {h: row_b.get(h, '') for h in union_headers}
                    diff_row['_diff_status'] = 'CHANGED'
                    diff_row['_diff_source'] = pair_label
                    diff_row['_diff_details'] = '; '.join(changes)
                    diff_rows.append(diff_row)
                    counts['changed'] += 1
                else:
                    counts['same'] += 1
            else:
                # Row in A but not in B → REMOVED
                diff_row = {h: row_a.get(h, '') for h in union_headers}
                diff_row['_diff_status'] = 'REMOVED'
                diff_row['_diff_source'] = pair_label
                diff_row['_diff_details'] = f'Not in {fname_b}'
                diff_rows.append(diff_row)
                counts['removed'] += 1

        # Check rows in B that are not in A → ADDED
        for k, row_b in index_b.items():
            if k not in index_a:
                diff_row = {h: row_b.get(h, '') for h in union_headers}
                diff_row['_diff_status'] = 'ADDED'
                diff_row['_diff_source'] = pair_label
                diff_row['_diff_details'] = f'Not in {fname_a}'
                diff_rows.append(diff_row)
                counts['added'] += 1

        summary_rows.append({
            'Comparison': pair_label,
            'File A Rows': len(rows_a),
            'File B Rows': len(rows_b),
            'Added': counts['added'],
            'Removed': counts['removed'],
            'Changed': counts['changed'],
            'Same': counts['same'],
        })

        log.debug(f'Pair "{pair_label}": added={counts["added"]}, removed={counts["removed"]}, '
                   f'changed={counts["changed"]}, same={counts["same"]}')

    # Phase 3: Write the diff report workbook
    out_wb = Workbook()

    # --- Summary sheet ---
    ws_summary = out_wb.active
    ws_summary.title = 'Summary'
    summary_headers = ['Comparison', 'File A Rows', 'File B Rows', 'Added', 'Removed', 'Changed', 'Same']
    for col_idx, h in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=col_idx, value=h)
    for row_idx, srow in enumerate(summary_rows, 2):
        for col_idx, h in enumerate(summary_headers, 1):
            ws_summary.cell(row=row_idx, column=col_idx, value=srow.get(h, ''))
    if not _no_formatting:
        _apply_header_style(ws_summary, len(summary_headers))
        _auto_fit_columns(ws_summary)
        ws_summary.freeze_panes = 'A2'

    # --- Diff sheet ---
    ws_diff = out_wb.create_sheet(title='Diff')
    # Diff columns: status, source, then union data columns, then details
    diff_headers = ['Status', 'Source'] + union_headers + ['Details']
    for col_idx, h in enumerate(diff_headers, 1):
        ws_diff.cell(row=1, column=col_idx, value=h)

    # Status fill colors for diff rows
    status_fills = {
        'ADDED':   PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'),  # green
        'REMOVED': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),  # red
        'CHANGED': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # yellow
    }

    for row_idx, drow in enumerate(diff_rows, 2):
        status = drow.get('_diff_status', '')
        source = drow.get('_diff_source', '')
        details = drow.get('_diff_details', '')

        ws_diff.cell(row=row_idx, column=1, value=status)
        ws_diff.cell(row=row_idx, column=2, value=source)

        for col_offset, h in enumerate(union_headers):
            ws_diff.cell(row=row_idx, column=3 + col_offset, value=drow.get(h, ''))

        ws_diff.cell(row=row_idx, column=len(diff_headers), value=details)

        # Color the status cell (unless --no-formatting)
        if not _no_formatting:
            fill = status_fills.get(status)
            if fill:
                ws_diff.cell(row=row_idx, column=1).fill = fill

    if not _no_formatting:
        _apply_header_style(ws_diff, len(diff_headers))
        _auto_fit_columns(ws_diff)
        ws_diff.freeze_panes = 'A2'
        ws_diff.auto_filter.ref = ws_diff.dimensions

    out_wb.save(output_file)
    log.info(f'Wrote diff report to: {output_file} ({len(diff_rows)} diff rows)')
    log.debug(f'Files compared: {len(input_files)}')
    for f in input_files:
        log.debug(f'  - {f}')
    log.debug(f'Key column: {key_col}')
    log.debug(f'Diff rows: {len(diff_rows)}')
    for srow in summary_rows:
        log.debug(f'  {srow["Comparison"]}: +{srow["Added"]} -{srow["Removed"]} ~{srow["Changed"]} ={srow["Same"]}')
    log.debug(f'Output file: {output_file}')

    output('')
    output('=' * 80)
    output('Excel Diff Report')
    output('=' * 80)
    output(f'Files compared: {len(input_files)}')
    for f in input_files:
        output(f'  - {f}')
    output(f'Key column:     {key_col}')
    output(f'Diff rows:      {len(diff_rows)}')
    for srow in summary_rows:
        output(f'  {srow["Comparison"]}: +{srow["Added"]} -{srow["Removed"]} ~{srow["Changed"]} ={srow["Same"]}')
    output(f'Output file:    {output_file}')
    output('=' * 80)
    output('')


# ****************************************************************************************
# Handle the arguments
# ****************************************************************************************

def handle_args():
    '''
    Parse CLI arguments and configure console logging handlers.

    Input:
        None directly; reads flags from sys.argv.

    Output:
        argparse.Namespace containing parsed arguments.

    Side Effects:
        Attaches a stream handler to the module logger with formatting and
        level derived from the parsed arguments.
    '''
    log.debug('Entering handle_args()')

    parser = argparse.ArgumentParser(
        description='Excel utilities for concatenating, converting, and diffing .xlsx workbooks.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s --concat fileA.xlsx fileB.xlsx --method merge-sheet --output merged.xlsx
      Merge all rows from fileA and fileB into a single sheet in merged.xlsx.
      Columns are unioned: missing columns get blank cells.

  %(prog)s --concat fileA.xlsx fileB.xlsx --method add-sheet --output combined.xlsx
      Add fileA as one sheet and fileB as another sheet in combined.xlsx.
      Sheet names are derived from filenames.

  %(prog)s --concat *.xlsx --method merge-sheet --output all_data.xlsx
      Merge all .xlsx files in the current directory into one sheet.

  %(prog)s --convert-to-csv data.xlsx
      Convert data.xlsx to data.csv (comma-delimited).

  %(prog)s --convert-to-csv data.xlsx --output custom_name.csv
      Convert data.xlsx to custom_name.csv.

  %(prog)s --convert-from-csv data.csv
      Convert data.csv to data.xlsx with styling and conditional formatting.

  %(prog)s --diff fileA.xlsx fileB.xlsx --output changes.xlsx
      Diff two Excel files and produce a report showing added/removed/changed rows.

  %(prog)s --diff v1.xlsx v2.xlsx v3.xlsx
      Pairwise diff across three files (v1→v2, v2→v3).
        ''')

    parser.add_argument(
        '-v',
        '--verbose',
        action='store_true',
        help='Enable verbose output to stdout.')

    parser.add_argument(
        '-q',
        '--quiet',
        action='store_true',
        help='Minimal stdout.')

    # --- Action arguments (mutually exclusive) ---
    parser.add_argument(
        '--concat',
        nargs='+',
        metavar='FILE',
        help='List of Excel (.xlsx) files to concatenate.')

    parser.add_argument(
        '--convert-to-csv',
        type=str,
        metavar='FILE',
        dest='convert_to_csv',
        help='Convert an Excel (.xlsx) file to a comma-delimited CSV file.')

    parser.add_argument(
        '--convert-from-csv',
        type=str,
        metavar='FILE',
        dest='convert_from_csv',
        help='Convert a comma-delimited CSV file to an Excel (.xlsx) file.')

    parser.add_argument(
        '--diff',
        nargs='+',
        metavar='FILE',
        help='Diff two or more Excel (.xlsx) files and produce a diff report.')

    parser.add_argument(
        '--to-plan-json',
        type=str,
        metavar='FILE',
        dest='to_plan_json',
        help='Convert a flat or indented CSV/Excel file into the feature-plan '
             'JSON format used by the Jira ticket creation pipeline.')

    # --- Plan JSON option arguments ---
    parser.add_argument(
        '--project',
        type=str,
        default='',
        dest='project_key',
        help='Override the Jira project key (used with --to-plan-json). '
             'Auto-detected from rows if not provided.')

    parser.add_argument(
        '--product-family',
        type=str,
        default='',
        dest='product_family',
        help='Override the product family (used with --to-plan-json). '
             'Auto-detected from rows if not provided.')

    parser.add_argument(
        '--feature-name',
        type=str,
        default='',
        dest='feature_name',
        help='Feature name for the plan (used with --to-plan-json). '
             'Defaults to the first epic summary.')

    # --- Option arguments ---
    parser.add_argument(
        '--method',
        type=str,
        choices=['merge-sheet', 'add-sheet'],
        default='merge-sheet',
        metavar='METHOD',
        help='Concatenation method: "merge-sheet" (all rows into one sheet) or '
             '"add-sheet" (each file becomes a separate sheet). Default: merge-sheet.')

    parser.add_argument(
        '--output',
        '-o',
        type=str,
        metavar='FILE',
        dest='output_file',
        help='Output filename (default depends on action).')

    parser.add_argument(
        '--no-formatting',
        action='store_true',
        dest='no_formatting',
        help='Disable all Excel formatting (header styling, conditional formatting, '
             'auto-fit columns). Produces a plain data-only workbook.')

    parser.add_argument(
        '--jira-url',
        type=str,
        metavar='URL',
        dest='jira_url',
        default=DEFAULT_JIRA_BASE_URL,
        help='Jira instance URL for clickable "key" column hyperlinks in '
             '--convert-from-csv. Defaults to %(default)s. '
             'Pass "none" to disable links.')

    parser.add_argument(
        '--d-columns',
        nargs='+',
        metavar='COL',
        dest='dashboard_columns',
        default=None,
        help='Column names for the Dashboard summary sheet (used with '
             '--convert-from-csv). Each named column gets a COUNTIF-based '
             'pivot table. Names are case-insensitive. '
             'Example: --d-columns Phase Customer Product Module Priority')

    args = parser.parse_args()

    # Configure stdout logging based on arguments (always add handler, level varies)
    ch = logging.StreamHandler(sys.stdout)
    if args.verbose:
        ch.setLevel(logging.DEBUG)
    elif args.quiet:
        ch.setLevel(logging.ERROR)
    else:
        ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    log.addHandler(ch)

    # Set quiet mode for output function
    global _quiet_mode
    _quiet_mode = args.quiet

    # Set no-formatting mode
    global _no_formatting
    _no_formatting = args.no_formatting

    log.debug('Checking script requirements...')

    # Determine which action was requested
    actions = []
    if args.concat:
        actions.append('concat')
    if args.convert_to_csv:
        actions.append('convert_to_csv')
    if args.convert_from_csv:
        actions.append('convert_from_csv')
    if args.diff:
        actions.append('diff')
    if args.to_plan_json:
        actions.append('to_plan_json')

    if len(actions) == 0:
        parser.print_help()
        sys.exit(1)

    if len(actions) > 1:
        parser.error('Only one action may be specified at a time '
                      '(--concat, --convert-to-csv, --convert-from-csv, --diff, --to-plan-json)')

    # --- Validate per-action constraints ---
    if args.concat:
        # Validate that at least 2 files are provided for concat
        if len(args.concat) < 2:
            parser.error('--concat requires at least 2 input files')
        # Validate that all input files exist
        for file_path in args.concat:
            if not os.path.exists(file_path):
                parser.error(f'Input file not found: {file_path}')
        # Default output filename for concat
        if not args.output_file:
            args.output_file = 'concat_output.xlsx'
        # Ensure output has .xlsx extension
        if not args.output_file.endswith('.xlsx'):
            args.output_file = f'{args.output_file}.xlsx'

    if args.convert_to_csv:
        if not os.path.exists(args.convert_to_csv):
            parser.error(f'Input file not found: {args.convert_to_csv}')

    if args.convert_from_csv:
        if not os.path.exists(args.convert_from_csv):
            parser.error(f'Input file not found: {args.convert_from_csv}')

    if args.diff:
        if len(args.diff) < 2:
            parser.error('--diff requires at least 2 input files')
        for file_path in args.diff:
            if not os.path.exists(file_path):
                parser.error(f'Input file not found: {file_path}')

    if args.to_plan_json:
        if not os.path.exists(args.to_plan_json):
            parser.error(f'Input file not found: {args.to_plan_json}')
        ext = os.path.splitext(args.to_plan_json)[1].lower()
        if ext not in ('.csv', '.xlsx', '.xls'):
            parser.error(f'--to-plan-json requires a .csv or .xlsx file, got: {ext}')

    log.info('++++++++++++++++++++++++++++++++++++++++++++++')
    log.info(f'+  {os.path.basename(sys.argv[0])}')
    log.info(f'+  Python Version: {sys.version.split()[0]}')
    log.info(f'+  Today is: {date.today()}')

    # Log action-specific input/output summary in the startup box
    if args.concat:
        log.info(f'+  Action: concat ({args.method})')
        log.info(f'+  Input files: {len(args.concat)}')
        for f in args.concat:
            log.info(f'+    - {f}')
        log.info(f'+  Output file: {args.output_file}')
    elif args.convert_to_csv:
        log.info(f'+  Action: convert-to-csv')
        log.info(f'+  Input file: {args.convert_to_csv}')
        if args.output_file:
            log.info(f'+  Output file: {args.output_file}')
    elif args.convert_from_csv:
        log.info(f'+  Action: convert-from-csv')
        log.info(f'+  Input file: {args.convert_from_csv}')
        if args.output_file:
            log.info(f'+  Output file: {args.output_file}')
    elif args.diff:
        log.info(f'+  Action: diff')
        log.info(f'+  Input files: {len(args.diff)}')
        for f in args.diff:
            log.info(f'+    - {f}')
        if args.output_file:
            log.info(f'+  Output file: {args.output_file}')
    elif args.to_plan_json:
        log.info(f'+  Action: to-plan-json')
        log.info(f'+  Input file: {args.to_plan_json}')
        if args.output_file:
            log.info(f'+  Output file: {args.output_file}')
        if args.project_key:
            log.info(f'+  Project key: {args.project_key}')
        if args.product_family:
            log.info(f'+  Product family: {args.product_family}')
        if args.feature_name:
            log.info(f'+  Feature name: {args.feature_name}')

    log.info('++++++++++++++++++++++++++++++++++++++++++++++')

    return args


# ****************************************************************************************
# Main
# ****************************************************************************************

def main():
    '''
    Entrypoint that wires together dependencies and launches the CLI.

    Sequence:
        1. Parse command line arguments
        2. Execute requested action(s)

    Output:
        Exit code 0 on success, 1 on failure.
    '''
    args = handle_args()
    log.debug('Entering main()')

    try:
        if args.concat:
            if args.method == 'merge-sheet':
                concat_merge_sheet(args.concat, args.output_file)
            elif args.method == 'add-sheet':
                concat_add_sheet(args.concat, args.output_file)

        elif args.convert_to_csv:
            convert_to_csv(args.convert_to_csv, args.output_file)

        elif args.convert_from_csv:
            # Allow --jira-url none to disable links entirely.
            jira_url = getattr(args, 'jira_url', DEFAULT_JIRA_BASE_URL)
            if isinstance(jira_url, str) and jira_url.lower() == 'none':
                jira_url = None
            convert_from_csv(args.convert_from_csv, args.output_file,
                             jira_base_url=jira_url,
                             dashboard_columns=getattr(args, 'dashboard_columns', None))

        elif args.diff:
            diff_files(args.diff, args.output_file)

        elif args.to_plan_json:
            convert_to_plan_json(
                args.to_plan_json,
                output_file=args.output_file or None,
                project_key=getattr(args, 'project_key', ''),
                product_family=getattr(args, 'product_family', ''),
                feature_name=getattr(args, 'feature_name', ''),
            )

    except ExcelFileError as e:
        log.error(e.message)
        output('')
        output(f'ERROR: {e.message}')
        output('')
        sys.exit(1)
    except Exception as e:
        log.error(f'Unexpected error: {e}')
        output(f'ERROR: {e}')
        sys.exit(1)

    log.info('Operation complete.')


if __name__ == '__main__':
    main()
