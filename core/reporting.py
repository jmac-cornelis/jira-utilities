##########################################################################################
#
# Module: core/reporting.py
#
# Description: Pure-logic daily reporting functions.
#              No print statements, no CLI concerns — returns structured dicts
#              that any consumer (CLI, agent tool, MCP endpoint) can format.
#
# Functions:
#   tickets_created_on()       — all tickets created on a given date
#   bugs_missing_field()       — bugs missing a required field (e.g. affectedVersion)
#   status_changes_by_actor()  — status transitions split by automation vs human
#   daily_report()             — composite: runs all three queries
#   export_daily_report()      — writes report to Excel (multi-sheet) or CSV
#
# Author: Cornelis Networks
#
##########################################################################################

from __future__ import annotations

import csv
import logging
import os
from datetime import date, datetime, timedelta
from typing import Any, Optional

import requests

from core.queries import build_tickets_jql, paginated_jql_search
from core.tickets import issue_to_dict

log = logging.getLogger(__name__)

# Default keywords used to identify automation/bot accounts in changelog authors.
DEFAULT_AUTOMATION_KEYWORDS: list[str] = ['scm-bot', 'automation', 'bot@', 'scm@']


# ---------------------------------------------------------------------------
# Helper: compute next-day string to avoid JQL "+" character
# ---------------------------------------------------------------------------

def _next_day(target_date: str) -> str:
    """Return the ISO date string for the day after *target_date* (YYYY-MM-DD).

    Jira Cloud's enhanced search API URL-encodes the ``+`` in ``"date" + 1d``
    as a space, causing a JQL parse error.  Using an explicit upper bound
    avoids this entirely.
    """
    d = datetime.strptime(target_date, '%Y-%m-%d').date()
    return (d + timedelta(days=1)).isoformat()


# ---------------------------------------------------------------------------
# 1. Tickets created on a date
# ---------------------------------------------------------------------------

def tickets_created_on(
    jira: Any,
    project: str,
    target_date: str,
) -> list[dict[str, Any]]:
    """Return issue dicts for all tickets created on *target_date*.

    Args:
        jira: Active ``jira.JIRA`` connection.
        project: Jira project key (e.g. ``"STL"``).
        target_date: Date string ``YYYY-MM-DD``.

    Returns:
        List of dicts produced by :func:`core.tickets.issue_to_dict`.
    """
    nd = _next_day(target_date)
    jql = build_tickets_jql(
        project,
        date_filter=f'AND created >= "{target_date}" AND created < "{nd}"',
    )
    log.info('tickets_created_on JQL: %s', jql)

    issues = paginated_jql_search(jira, jql)
    return [issue_to_dict(i) for i in issues]


# ---------------------------------------------------------------------------
# 2. Bugs missing a required field
# ---------------------------------------------------------------------------

def bugs_missing_field(
    jira: Any,
    project: str,
    field: str = 'affectedVersion',
    target_date: Optional[str] = None,
) -> dict[str, Any]:
    """Find bugs missing a required field.

    Args:
        jira: Active ``jira.JIRA`` connection.
        project: Jira project key.
        field: JQL field name to check (``affectedVersion``, ``fixVersion``,
               ``component``, ``assignee``, etc.).
        target_date: If given, ``flagged`` contains only bugs created on that
                     date.  Otherwise ``flagged`` contains all open bugs.

    Returns:
        ``{"flagged": [...], "total_open_count": int, "field": str}``
    """
    # --- Flagged bugs (date-scoped or all open) ---
    if target_date:
        nd = _next_day(target_date)
        date_filter = f'AND created >= "{target_date}" AND created < "{nd}"'
    else:
        date_filter = None

    jql_flagged = build_tickets_jql(
        project,
        issue_types=['Bug'],
        date_filter=date_filter,
        jql_extra=f'{field} is EMPTY',
    )
    log.info('bugs_missing_field (flagged) JQL: %s', jql_flagged)
    flagged_issues = paginated_jql_search(jira, jql_flagged)
    flagged = [issue_to_dict(i) for i in flagged_issues]

    # --- Total open count (always all open bugs, regardless of date) ---
    jql_all = build_tickets_jql(
        project,
        issue_types=['Bug'],
        statuses={'exclude': ['Closed', 'Done', 'Resolved']},
        jql_extra=f'{field} is EMPTY',
    )
    log.info('bugs_missing_field (total) JQL: %s', jql_all)
    all_issues = paginated_jql_search(jira, jql_all)

    return {
        'field': field,
        'flagged': flagged,
        'total_open_count': len(all_issues),
    }


# ---------------------------------------------------------------------------
# 3. Status changes by actor (automation vs human)
# ---------------------------------------------------------------------------

def status_changes_by_actor(
    project: str,
    target_date: str,
    *,
    jira_url: Optional[str] = None,
    email: Optional[str] = None,
    api_token: Optional[str] = None,
    automation_keywords: Optional[list[str]] = None,
) -> dict[str, Any]:
    """Query Jira REST API for status changes on *target_date*, split by actor.

    Uses the v3 enhanced search endpoint (``/rest/api/3/search/jql``) with
    ``expand=changelog`` because the jira-python ``enhanced_search_issues``
    method does not support changelog expansion.

    Args:
        project: Jira project key.
        target_date: Date string ``YYYY-MM-DD``.
        jira_url: Base Jira URL.  Falls back to ``JIRA_URL`` env var.
        email: Jira account email.  Falls back to ``get_jira_credentials()``.
        api_token: Jira API token.  Falls back to ``get_jira_credentials()``.
        automation_keywords: Substrings to match in author email/name to
                             classify a transition as automation-triggered.

    Returns:
        ``{"automation": [...], "human": [...], "total": int}``
        Each transition dict: ``{key, from, to, author, email, time}``.
    """
    # Resolve credentials lazily so callers don't need to pass them.
    if email is None or api_token is None:
        from jira_utils import get_jira_credentials
        _email, _token = get_jira_credentials()
        email = email or _email
        api_token = api_token or _token

    if jira_url is None:
        jira_url = os.getenv('JIRA_URL', '').rstrip('/')

    keywords = automation_keywords or DEFAULT_AUTOMATION_KEYWORDS

    nd = _next_day(target_date)
    jql = (
        f'project = "{project}" '
        f'AND updated >= "{target_date}" AND updated < "{nd}"'
    )
    log.info('status_changes_by_actor JQL: %s', jql)

    url = f'{jira_url}/rest/api/3/search/jql'
    all_transitions: list[dict[str, Any]] = []
    next_page_token: Optional[str] = None

    while True:
        params: dict[str, Any] = {
            'jql': jql,
            'maxResults': 50,
            'expand': 'changelog',
        }
        if next_page_token:
            params['nextPageToken'] = next_page_token

        resp = requests.get(url, params=params, auth=(email, api_token))
        resp.raise_for_status()
        data = resp.json()

        for issue in data.get('issues', []):
            key = issue['key']
            changelog = issue.get('changelog', {})
            for history in changelog.get('histories', []):
                created = history.get('created', '')
                # Only include transitions that occurred on target_date.
                if not created.startswith(target_date):
                    continue

                author = history.get('author', {})
                display_name = author.get('displayName', 'unknown')
                email_addr = author.get('emailAddress', '')

                for item in history.get('items', []):
                    if item.get('field') == 'status':
                        all_transitions.append({
                            'key': key,
                            'from': item.get('fromString', ''),
                            'to': item.get('toString', ''),
                            'author': display_name,
                            'email': email_addr,
                            'time': created,
                        })

        next_page_token = data.get('nextPageToken')
        if not next_page_token:
            break

    # Classify transitions as automation or human.
    auto: list[dict[str, Any]] = []
    human: list[dict[str, Any]] = []
    for t in all_transitions:
        is_auto = any(
            kw in t['email'].lower() or kw in t['author'].lower()
            for kw in keywords
        )
        (auto if is_auto else human).append(t)

    return {
        'automation': sorted(auto, key=lambda x: x['time'], reverse=True),
        'human': sorted(human, key=lambda x: x['time'], reverse=True),
        'total': len(all_transitions),
    }


# ---------------------------------------------------------------------------
# 4. Composite daily report
# ---------------------------------------------------------------------------

def daily_report(
    jira: Any,
    project: str,
    target_date: Optional[str] = None,
    *,
    missing_field: str = 'affectedVersion',
    automation_keywords: Optional[list[str]] = None,
) -> dict[str, Any]:
    """Run all three daily queries and return a combined report dict.

    Args:
        jira: Active ``jira.JIRA`` connection.
        project: Jira project key.
        target_date: Date string ``YYYY-MM-DD``.  Defaults to today.
        missing_field: Field to check for bugs (default ``affectedVersion``).
        automation_keywords: Override automation detection keywords.

    Returns:
        Combined dict with keys: ``date``, ``project``, ``created_tickets``,
        ``bugs_missing_field``, ``status_changes``.
    """
    if target_date is None:
        target_date = date.today().isoformat()

    created = tickets_created_on(jira, project, target_date)
    bugs = bugs_missing_field(jira, project, field=missing_field,
                              target_date=target_date)
    changes = status_changes_by_actor(
        project, target_date,
        automation_keywords=automation_keywords,
    )

    return {
        'date': target_date,
        'project': project,
        'created_tickets': created,
        'bugs_missing_field': bugs,
        'status_changes': changes,
    }


# ---------------------------------------------------------------------------
# 5. Export helper
# ---------------------------------------------------------------------------

def export_daily_report(
    report: dict[str, Any],
    output_path: str,
    fmt: str = 'excel',
) -> str:
    """Write a daily report dict to Excel (multi-sheet) or CSV.

    Args:
        report: Dict returned by :func:`daily_report`.
        output_path: Destination file path (extension added if missing).
        fmt: ``"excel"`` or ``"csv"``.

    Returns:
        The resolved output path (or base path for CSV).
    """
    if fmt == 'excel':
        return _export_excel(report, output_path)
    elif fmt == 'csv':
        return _export_csv(report, output_path)
    else:
        raise ValueError(f'Unsupported export format: {fmt!r}')


def _export_excel(report: dict[str, Any], output_path: str) -> str:
    """Write a multi-sheet Excel workbook from the report dict."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError(
            'openpyxl is required for Excel export. '
            'Install with: pip install openpyxl'
        )

    if not output_path.endswith('.xlsx'):
        output_path += '.xlsx'

    jira_url = os.getenv('JIRA_URL', 'https://cornelisnetworks.atlassian.net').rstrip('/')

    wb = Workbook()

    # --- Shared styling ---
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    link_font = Font(color='0563C1', underline='single')

    def _style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value else ''
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

    def _write_key_link(ws, row_idx, col_idx, key):
        """Write a ticket key as a clickable hyperlink."""
        cell = ws.cell(row=row_idx, column=col_idx, value=key)
        cell.hyperlink = f'{jira_url}/browse/{key}'
        cell.font = link_font

    # ---- Sheet 1: Created Tickets ----
    ws1 = wb.active
    ws1.title = 'Created Tickets'
    headers1 = ['key', 'issue_type', 'status', 'priority', 'assignee', 'summary', 'created']
    ws1.append(headers1)
    for ticket in report.get('created_tickets', []):
        ws1.append([ticket.get(h, '') for h in headers1])
    # Add hyperlinks to key column
    for row_idx in range(2, ws1.max_row + 1):
        key_val = ws1.cell(row=row_idx, column=1).value
        if key_val:
            _write_key_link(ws1, row_idx, 1, key_val)
    _style_header(ws1)
    _auto_width(ws1)
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = ws1.dimensions

    # ---- Sheet 2: Bugs Missing Field ----
    ws2 = wb.create_sheet('Bugs Missing Field')
    bugs_data = report.get('bugs_missing_field', {})
    field_name = bugs_data.get('field', 'affectedVersion')
    headers2 = ['key', 'status', 'priority', 'assignee', 'summary', 'created', 'missing_field']
    ws2.append(headers2)
    for ticket in bugs_data.get('flagged', []):
        row = [ticket.get(h, '') for h in headers2[:-1]]
        row.append(field_name)
        ws2.append(row)
    for row_idx in range(2, ws2.max_row + 1):
        key_val = ws2.cell(row=row_idx, column=1).value
        if key_val:
            _write_key_link(ws2, row_idx, 1, key_val)
    _style_header(ws2)
    _auto_width(ws2)
    ws2.freeze_panes = 'A2'
    if ws2.max_row > 1:
        ws2.auto_filter.ref = ws2.dimensions

    # ---- Sheet 3: Status Changes ----
    ws3 = wb.create_sheet('Status Changes')
    headers3 = ['key', 'from_status', 'to_status', 'author', 'email', 'time', 'is_automation']
    ws3.append(headers3)
    for t in report.get('status_changes', {}).get('automation', []):
        ws3.append([t['key'], t['from'], t['to'], t['author'], t['email'], t['time'], True])
    for t in report.get('status_changes', {}).get('human', []):
        ws3.append([t['key'], t['from'], t['to'], t['author'], t['email'], t['time'], False])
    for row_idx in range(2, ws3.max_row + 1):
        key_val = ws3.cell(row=row_idx, column=1).value
        if key_val:
            _write_key_link(ws3, row_idx, 1, key_val)
    _style_header(ws3)
    _auto_width(ws3)
    ws3.freeze_panes = 'A2'
    if ws3.max_row > 1:
        ws3.auto_filter.ref = ws3.dimensions

    # ---- Summary sheet (first position) ----
    ws_summary = wb.create_sheet('Summary', 0)
    ws_summary.append(['Daily Jira Report'])
    ws_summary.append(['Project', report.get('project', '')])
    ws_summary.append(['Date', report.get('date', '')])
    ws_summary.append([])
    ws_summary.append(['Metric', 'Count'])
    ws_summary.append(['Tickets created', len(report.get('created_tickets', []))])
    ws_summary.append([
        f'Bugs missing {field_name}',
        len(bugs_data.get('flagged', [])),
    ])
    ws_summary.append([
        f'Total open bugs missing {field_name}',
        bugs_data.get('total_open_count', 0),
    ])
    changes = report.get('status_changes', {})
    ws_summary.append(['Automation status changes', len(changes.get('automation', []))])
    ws_summary.append(['Human status changes', len(changes.get('human', []))])
    ws_summary.append(['Total status changes', changes.get('total', 0)])
    # Style the summary title
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15

    wb.save(output_path)
    log.info('Wrote daily report (excel) to: %s', output_path)
    return output_path


def _export_csv(report: dict[str, Any], output_path: str) -> str:
    """Write 3 CSV files from the report dict."""
    # Strip extension if present to use as base name
    base = output_path
    for ext in ('.csv', '.xlsx'):
        if base.endswith(ext):
            base = base[:-len(ext)]
            break

    # --- Created tickets ---
    created_path = f'{base}_created.csv'
    headers1 = ['key', 'issue_type', 'status', 'priority', 'assignee', 'summary', 'created']
    with open(created_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers1, extrasaction='ignore')
        writer.writeheader()
        for ticket in report.get('created_tickets', []):
            writer.writerow(ticket)
    log.info('Wrote %d rows to %s', len(report.get('created_tickets', [])), created_path)

    # --- Bugs missing field ---
    bugs_path = f'{base}_bugs.csv'
    bugs_data = report.get('bugs_missing_field', {})
    headers2 = ['key', 'status', 'priority', 'assignee', 'summary', 'created', 'missing_field']
    with open(bugs_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers2, extrasaction='ignore')
        writer.writeheader()
        for ticket in bugs_data.get('flagged', []):
            row = dict(ticket)
            row['missing_field'] = bugs_data.get('field', 'affectedVersion')
            writer.writerow(row)
    log.info('Wrote %d rows to %s', len(bugs_data.get('flagged', [])), bugs_path)

    # --- Status changes ---
    changes_path = f'{base}_changes.csv'
    headers3 = ['key', 'from_status', 'to_status', 'author', 'email', 'time', 'is_automation']
    with open(changes_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers3)
        changes = report.get('status_changes', {})
        for t in changes.get('automation', []):
            writer.writerow([t['key'], t['from'], t['to'], t['author'], t['email'], t['time'], True])
        for t in changes.get('human', []):
            writer.writerow([t['key'], t['from'], t['to'], t['author'], t['email'], t['time'], False])
    total_changes = len(changes.get('automation', [])) + len(changes.get('human', []))
    log.info('Wrote %d rows to %s', total_changes, changes_path)

    return base
