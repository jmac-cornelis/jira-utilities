##########################################################################################
#
# Module: tools/plan_export_tools.py
#
# Description: Converts between feature-plan JSON (as produced by FeaturePlanBuilderAgent)
#              and the standard CSV / Excel format used by jira_utils.dump_tickets_to_file().
#
#              Supports BOTH directions:
#                JSON → CSV/Excel  (plan_json_to_rows, write_plan_csv, write_plan_excel)
#                CSV/Excel → JSON  (read_plan_rows, plan_rows_to_json)
#
#              Works as:
#                1. An @tool()-decorated function usable inside the agentic pipeline.
#                2. A standalone CLI:  python -m tools.plan_export_tools plan.json [-o out.csv]
#                                      python -m tools.plan_export_tools plan.csv --to-json [-o out.json]
#
# Author: Cornelis Networks
#
##########################################################################################

import csv
import json
import logging
import os
import re
from typing import Any, Dict, List, Optional, Tuple

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Graceful import of the @tool decorator (mirrors pattern in excel_tools.py)
# ---------------------------------------------------------------------------
try:
    from tools.base import tool, ToolResult, BaseTool
except ImportError:
    log.warning('tools.base not available; plan_export_tools will not register @tool decorators')

    def tool(**kwargs):  # type: ignore[misc]
        '''No-op decorator fallback when tools.base is unavailable.'''
        def _identity(fn):
            return fn
        return _identity

    class ToolResult:  # type: ignore[no-redef]
        '''Minimal stub so the module can still be imported standalone.'''
        def __init__(self, **kwargs):
            for k, v in kwargs.items():
                setattr(self, k, v)

        @classmethod
        def success(cls, data=None, message=''):
            return cls(status='success', data=data or {}, message=message, error=None)

        @classmethod
        def error(cls, error='', data=None):
            return cls(status='error', data=data or {}, message='', error=error)

    class BaseTool:  # type: ignore[no-redef]
        pass


# ---------------------------------------------------------------------------
# Constants — base CSV columns matching jira_utils.dump_tickets_to_file()
# ---------------------------------------------------------------------------
BASE_FIELDS: List[str] = [
    'key', 'project', 'issue_type', 'status', 'priority', 'summary',
    'assignee', 'reporter', 'created', 'updated', 'resolved',
    'fix_version', 'affects_version', 'component', 'customer',
]

# Extra columns specific to feature-plan exports (appended after base fields)
PLAN_EXTRA_FIELDS: List[str] = [
    'depth',
    'product_family',
    'complexity',
    'confidence',
    'labels',
    'acceptance_criteria',
    'dependencies',
    'parent_epic',
]


# ============================================================================
# Core conversion logic (pure function — no I/O)
# ============================================================================

def plan_json_to_rows(
    plan: Dict[str, Any],
    *,
    include_description: bool = False,
) -> List[Dict[str, str]]:
    '''Convert a feature-plan dict into a flat list of row dicts.

    Each row uses the same column vocabulary as jira_utils.dump_tickets_to_file()
    so the output can be consumed by bulk_update_tickets, convert_from_csv (Excel),
    or any other downstream tool that expects that schema.

    Args:
        plan: The feature-plan dict (as stored in feature_plan.json).
        include_description: If True, add a ``description`` extra column with the
            full ticket description text.  Defaults to False because descriptions
            can be very long and make the CSV unwieldy.

    Returns:
        List of row dicts keyed by column name.
    '''
    rows: List[Dict[str, str]] = []
    project_key = plan.get('project_key', '')
    # product_family is plan-level (e.g. "CN5000") — applies to every ticket
    product_family = plan.get('product_family', '')

    for epic_idx, epic in enumerate(plan.get('epics', [])):
        # ----- Epic row -----
        epic_row: Dict[str, str] = {
            'key':              epic.get('key') or '',
            'project':          project_key,
            'issue_type':       'Epic',
            'status':           '',          # not yet created
            'priority':         '',
            'summary':          epic.get('summary', ''),
            'assignee':         '',
            'reporter':         '',
            'created':          '',
            'updated':          '',
            'resolved':         '',
            'fix_version':      '',
            'affects_version':  '',
            'component':        '; '.join(epic.get('components') or []),
            'customer':         '',
            # Plan-specific extras
            'depth':            '0',
            'product_family':   product_family,
            'complexity':       '',
            'confidence':       '',
            'labels':           '; '.join(epic.get('labels') or []),
            'acceptance_criteria': '',
            'dependencies':     '',
            'parent_epic':      '',
        }
        if include_description:
            epic_row['description'] = epic.get('description', '')

        rows.append(epic_row)

        # ----- Story rows under this Epic -----
        for story in epic.get('stories', []):
            story_row: Dict[str, str] = {
                'key':              story.get('key') or '',
                'project':          project_key,
                'issue_type':       'Story',
                'status':           '',
                'priority':         '',
                'summary':          story.get('summary', ''),
                'assignee':         story.get('assignee') or '',
                'reporter':         '',
                'created':          '',
                'updated':          '',
                'resolved':         '',
                'fix_version':      '',
                'affects_version':  '',
                'component':        '; '.join(story.get('components') or []),
                'customer':         '',
                # Plan-specific extras
                'depth':            '1',
                'product_family':   product_family,
                'complexity':       story.get('complexity', ''),
                'confidence':       story.get('confidence', ''),
                'labels':           '; '.join(story.get('labels') or []),
                'acceptance_criteria': '; '.join(story.get('acceptance_criteria') or []),
                'dependencies':     '; '.join(story.get('dependencies') or []),
                'parent_epic':      epic.get('summary', ''),
            }
            if include_description:
                story_row['description'] = story.get('description', '')

            rows.append(story_row)

    return rows


def _resolve_output_path(input_path: str, output_path: Optional[str], fmt: str) -> str:
    '''Derive the output file path from the input path when not explicitly given.

    If *output_path* is provided it is treated as a **basename without extension**;
    the correct extension (.csv or .xlsx) is appended automatically.  When
    *output_path* is ``None`` or empty the basename is derived from *input_path*.
    '''
    ext = '.csv' if fmt == 'csv' else '.xlsx'
    if output_path:
        # Strip any extension the caller may have included, then apply the right one
        base, _ = os.path.splitext(output_path)
        return f'{base}{ext}'
    base, _ = os.path.splitext(input_path)
    return f'{base}{ext}'


# ============================================================================
# CSV / Excel writers
# ============================================================================

def write_plan_csv(
    rows: List[Dict[str, str]],
    output_path: str,
    *,
    table_format: str = 'indented',
) -> str:
    '''Write plan rows to a CSV file using the jira_utils column convention.

    Args:
        rows: List of row dicts (from plan_json_to_rows).
        output_path: Destination file path.
        table_format: ``'indented'`` (default) replaces the key column with
            per-depth columns (Depth 0, Depth 1, …) matching jira_utils
            indented format; ``'flat'`` keeps depth as a regular column.

    Returns:
        The resolved output path.
    '''
    if not rows:
        # Write header-only CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(BASE_FIELDS)
        log.info(f'Wrote empty CSV (headers only) to: {output_path}')
        return output_path

    # Determine which extra columns are actually populated
    all_keys = set(BASE_FIELDS)
    for r in rows:
        all_keys.update(r.keys())
    extra_columns = sorted(k for k in all_keys if k not in BASE_FIELDS)

    # ------------------------------------------------------------------
    # Indented table format: replace key + depth with Depth N columns
    # ------------------------------------------------------------------
    if table_format == 'indented' and any(r.get('depth') for r in rows):
        max_depth = 0
        for r in rows:
            try:
                max_depth = max(max_depth, int(r.get('depth', 0)))
            except (ValueError, TypeError):
                pass

        depth_columns = [f'Depth {i}' for i in range(max_depth + 1)]
        content_fields = [f for f in BASE_FIELDS if f != 'key']
        # Remove 'depth' from extras since it is represented by depth columns
        indented_extras = [c for c in extra_columns if c != 'depth']
        fieldnames = depth_columns + content_fields + indented_extras

        indented_rows: List[Dict[str, str]] = []
        for r in rows:
            new_row: Dict[str, str] = {}
            d = 0
            try:
                d = int(r.get('depth', 0))
            except (ValueError, TypeError):
                d = 0
            for col in depth_columns:
                new_row[col] = ''
            new_row[f'Depth {d}'] = r.get('key', '') or r.get('summary', '')

            for f in content_fields:
                new_row[f] = r.get(f, '')
            for col in indented_extras:
                new_row[col] = r.get(col, '')
            indented_rows.append(new_row)

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(indented_rows)

        log.info(f'Wrote {len(indented_rows)} rows (indented, max depth {max_depth}) to: {output_path}')
        return output_path

    # ------------------------------------------------------------------
    # Flat table format (default)
    # ------------------------------------------------------------------
    fieldnames = BASE_FIELDS + extra_columns
    # Ensure every row has all fields
    for r in rows:
        for col in fieldnames:
            r.setdefault(col, '')

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    log.info(f'Wrote {len(rows)} rows (flat) to: {output_path}')
    return output_path


def write_plan_excel(
    rows: List[Dict[str, str]],
    output_path: str,
    *,
    table_format: str = 'indented',
) -> str:
    '''Write plan rows to an Excel file, delegating to jira_utils._write_excel().

    Falls back to CSV if openpyxl / jira_utils is unavailable.

    Returns:
        The resolved output path.
    '''
    try:
        from jira_utils import _write_excel
        _write_excel(rows, output_path, extra_fields=None, table_format=table_format)
        log.info(f'Wrote {len(rows)} rows (Excel, {table_format}) to: {output_path}')
        return output_path
    except ImportError:
        log.warning('jira_utils._write_excel not available; falling back to CSV')
        csv_path = output_path.replace('.xlsx', '.csv')
        return write_plan_csv(rows, csv_path, table_format=table_format)


# ============================================================================
# Reverse direction: CSV / Excel → feature-plan JSON
# ============================================================================

def _detect_table_format(headers: List[str]) -> str:
    '''Detect whether a CSV/Excel uses "flat" or "indented" table format.

    "Indented" format has columns named ``Depth 0``, ``Depth 1``, … instead of
    a ``key`` column.  "Flat" format has a ``key`` column and optionally a
    ``depth`` column.

    Args:
        headers: List of column header strings from the first row.

    Returns:
        ``'indented'`` or ``'flat'``.
    '''
    # Indented format is identified by the presence of "Depth N" columns
    depth_col_pattern = re.compile(r'^Depth\s+\d+$', re.IGNORECASE)
    has_depth_cols = any(depth_col_pattern.match(h) for h in headers)
    return 'indented' if has_depth_cols else 'flat'


def _read_csv_rows(input_path: str) -> Tuple[List[Dict[str, str]], str]:
    '''Read a CSV file and return (rows, detected_table_format).

    Each row is a dict keyed by column header.  The table format is
    auto-detected from the header row.

    Args:
        input_path: Path to the CSV file.

    Returns:
        Tuple of (list of row dicts, 'flat' or 'indented').

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the CSV is empty or has no header row.
    '''
    if not os.path.exists(input_path):
        raise FileNotFoundError(f'CSV file not found: {input_path}')

    with open(input_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f'CSV file is empty or has no header row: {input_path}')
        table_format = _detect_table_format(list(reader.fieldnames))
        rows = list(reader)

    return rows, table_format


def _read_excel_rows(input_path: str) -> Tuple[List[Dict[str, str]], str]:
    '''Read an Excel (.xlsx) file and return (rows, detected_table_format).

    Reads the first sheet (or a sheet named "Tickets" / "Plan" if present).
    Each row is a dict keyed by column header.

    Args:
        input_path: Path to the .xlsx file.

    Returns:
        Tuple of (list of row dicts, 'flat' or 'indented').

    Raises:
        FileNotFoundError: If the file does not exist.
        ImportError: If openpyxl is not installed.
        ValueError: If the workbook is empty.
    '''
    if not os.path.exists(input_path):
        raise FileNotFoundError(f'Excel file not found: {input_path}')

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            'openpyxl is required to read Excel files. '
            'Install with: pip install openpyxl'
        )

    wb = load_workbook(input_path, read_only=True, data_only=True)

    # Prefer a sheet named "Tickets" or "Plan"; fall back to the first sheet
    ws = None
    for preferred in ('Tickets', 'Plan'):
        if preferred in wb.sheetnames:
            ws = wb[preferred]
            break
    if ws is None:
        ws = wb.active

    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw or len(rows_raw) < 1:
        raise ValueError(f'Excel workbook has no data: {input_path}')

    # First row is the header
    headers = [str(h).strip() if h is not None else '' for h in rows_raw[0]]
    table_format = _detect_table_format(headers)

    rows: List[Dict[str, str]] = []
    for raw_row in rows_raw[1:]:
        row_dict: Dict[str, str] = {}
        for col_idx, value in enumerate(raw_row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = str(value).strip() if value is not None else ''
        # Skip completely empty rows
        if any(v for v in row_dict.values()):
            rows.append(row_dict)

    return rows, table_format


def read_plan_rows(input_path: str) -> Tuple[List[Dict[str, str]], str]:
    '''Read plan rows from a CSV or Excel file, auto-detecting the format.

    Dispatches to _read_csv_rows or _read_excel_rows based on file extension.

    Args:
        input_path: Path to a .csv or .xlsx file.

    Returns:
        Tuple of (list of row dicts, 'flat' or 'indented').

    Raises:
        ValueError: If the file extension is not .csv or .xlsx.
    '''
    ext = os.path.splitext(input_path)[1].lower()
    if ext == '.csv':
        return _read_csv_rows(input_path)
    elif ext in ('.xlsx', '.xls'):
        return _read_excel_rows(input_path)
    else:
        raise ValueError(
            f'Unsupported file extension "{ext}". Expected .csv or .xlsx.'
        )


def _normalize_flat_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    '''Ensure flat-format rows have a ``depth`` field.

    If the rows already have a ``depth`` column, they are returned as-is.
    Otherwise, depth is inferred from ``issue_type``:
      - Epic → depth 0
      - Story / Task / Sub-task → depth 1
      - Anything else → depth 1

    Args:
        rows: List of flat-format row dicts.

    Returns:
        The same list with ``depth`` populated on every row.
    '''
    # Check if depth is already present and populated
    has_depth = any(r.get('depth', '').strip() for r in rows)
    if has_depth:
        return rows

    # Infer depth from issue_type
    for r in rows:
        issue_type = (r.get('issue_type') or '').strip().lower()
        if issue_type in ('epic', 'initiative'):
            r['depth'] = '0'
        else:
            r['depth'] = '1'

    return rows


def _normalize_indented_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    '''Convert indented-format rows into flat-format rows with key + depth.

    Indented format has ``Depth 0``, ``Depth 1``, … columns where the ticket
    key (or summary for new tickets) appears in the column matching its depth.
    This function collapses those columns back into ``key`` and ``depth`` fields.

    Args:
        rows: List of indented-format row dicts.

    Returns:
        List of flat-format row dicts with ``key`` and ``depth`` fields.
    '''
    # Identify all Depth N columns and their numeric indices
    depth_col_pattern = re.compile(r'^Depth\s+(\d+)$', re.IGNORECASE)
    depth_columns: Dict[int, str] = {}  # {depth_int: column_name}
    if rows:
        for col_name in rows[0].keys():
            m = depth_col_pattern.match(col_name)
            if m:
                depth_columns[int(m.group(1))] = col_name

    if not depth_columns:
        # No depth columns found — treat as flat
        return rows

    flat_rows: List[Dict[str, str]] = []
    for r in rows:
        flat_row: Dict[str, str] = {}

        # Find which depth column is populated for this row
        found_depth = 0
        found_key = ''
        for depth_int in sorted(depth_columns.keys()):
            col_name = depth_columns[depth_int]
            val = (r.get(col_name) or '').strip()
            if val:
                found_depth = depth_int
                found_key = val
                break  # Only one depth column should be populated per row

        flat_row['key'] = found_key
        flat_row['depth'] = str(found_depth)

        # Copy all non-depth columns
        for col_name, val in r.items():
            if not depth_col_pattern.match(col_name):
                flat_row[col_name] = val

        flat_rows.append(flat_row)

    return flat_rows


def plan_rows_to_json(
    rows: List[Dict[str, str]],
    table_format: str = 'auto',
    *,
    project_key: str = '',
    product_family: str = '',
    feature_name: str = '',
) -> Dict[str, Any]:
    '''Convert flat row dicts (from CSV/Excel) into a feature-plan JSON dict.

    This is the reverse of plan_json_to_rows().  It reconstructs the hierarchical
    epic→story structure from the flat/indented table representation.

    The function auto-detects whether the rows use flat or indented format
    (unless *table_format* is explicitly set).

    Depth semantics:
      - depth 0 → Epic
      - depth 1 → Story (child of the most recent depth-0 Epic)
      - depth 2+ → Story (child of the most recent depth-(N-1) item)

    Args:
        rows: List of row dicts (from read_plan_rows or similar).
        table_format: ``'flat'``, ``'indented'``, or ``'auto'`` (default).
        project_key: Override the project key (uses row data if empty).
        product_family: Override the product family (uses row data if empty).
        feature_name: Feature name for the plan (defaults to first epic summary).

    Returns:
        A dict matching the JiraPlan schema:
        ``{project_key, feature_name, product_family, epics: [{summary, description,
        components, labels, stories: [{summary, description, ...}]}], ...}``
    '''
    if not rows:
        return {
            'project_key': project_key,
            'feature_name': feature_name,
            'product_family': product_family,
            'epics': [],
            'total_epics': 0,
            'total_stories': 0,
            'total_tickets': 0,
        }

    # --- Auto-detect and normalize table format ---
    if table_format == 'auto':
        headers = list(rows[0].keys())
        table_format = _detect_table_format(headers)

    if table_format == 'indented':
        rows = _normalize_indented_rows(rows)
    else:
        rows = _normalize_flat_rows(rows)

    # --- Extract project_key and product_family from rows if not provided ---
    if not project_key:
        for r in rows:
            pk = (r.get('project') or '').strip()
            if pk:
                project_key = pk
                break

    if not product_family:
        for r in rows:
            pf = (r.get('product_family') or '').strip()
            if pf:
                product_family = pf
                break

    # Normalise product_family into a list — the field may arrive as a
    # semicolon-delimited string (e.g. "CN5000; CN6000") from CSV/Excel.
    # Jira's Product Family custom field expects each value as a separate
    # option, so we split here and store as a list.
    if isinstance(product_family, str) and product_family:
        product_family = [v.strip() for v in product_family.split(';') if v.strip()]

    # --- Group rows into epics and stories by depth ---
    epics: List[Dict[str, Any]] = []
    current_epic: Optional[Dict[str, Any]] = None

    for r in rows:
        depth = 0
        try:
            depth = int(r.get('depth', 0))
        except (ValueError, TypeError):
            depth = 0

        issue_type = (r.get('issue_type') or '').strip()

        # Parse semicolon-separated list fields back into lists
        components = [c.strip() for c in (r.get('component') or '').split(';') if c.strip()]
        labels = [l.strip() for l in (r.get('labels') or '').split(';') if l.strip()]
        acceptance_criteria = [
            a.strip() for a in (r.get('acceptance_criteria') or '').split(';') if a.strip()
        ]
        dependencies = [d.strip() for d in (r.get('dependencies') or '').split(';') if d.strip()]

        # Determine the summary — for indented format the key field may contain
        # the summary (for new tickets without a Jira key) or the Jira key
        key_val = (r.get('key') or '').strip()
        summary = (r.get('summary') or '').strip()

        # If summary is empty but key looks like a summary (not a Jira key pattern),
        # use key as summary
        jira_key_pattern = re.compile(r'^[A-Z][A-Z0-9]+-\d+$')
        actual_key = ''
        if key_val:
            if jira_key_pattern.match(key_val):
                actual_key = key_val
            elif not summary:
                summary = key_val

        if depth == 0 or issue_type.lower() in ('epic', 'initiative'):
            # Start a new epic
            current_epic = {
                'key': actual_key,
                'summary': summary,
                'description': r.get('description', ''),
                'components': components,
                'labels': labels,
                'stories': [],
            }
            epics.append(current_epic)
        else:
            # Story (or deeper item) — attach to the current epic
            story = {
                'key': actual_key,
                'summary': summary,
                'description': r.get('description', ''),
                'components': components,
                'labels': labels,
                'assignee': (r.get('assignee') or '').strip() or None,
                'complexity': (r.get('complexity') or '').strip(),
                'confidence': (r.get('confidence') or '').strip(),
                'acceptance_criteria': acceptance_criteria,
                'dependencies': dependencies,
                'parent_epic_summary': (r.get('parent_epic') or '').strip(),
            }

            if current_epic is not None:
                current_epic['stories'].append(story)
            else:
                # No epic yet — create a synthetic one
                current_epic = {
                    'key': '',
                    'summary': 'Ungrouped Stories',
                    'description': '',
                    'components': [],
                    'labels': [],
                    'stories': [story],
                }
                epics.append(current_epic)

    # --- Derive feature_name if not provided ---
    if not feature_name and epics:
        feature_name = epics[0].get('summary', '')

    # --- Compute totals ---
    total_epics = len(epics)
    total_stories = sum(len(e.get('stories', [])) for e in epics)

    plan: Dict[str, Any] = {
        'project_key': project_key,
        'feature_name': feature_name,
        'product_family': product_family,
        'epics': epics,
        'total_epics': total_epics,
        'total_stories': total_stories,
        'total_tickets': total_epics + total_stories,
    }

    return plan


def plan_file_to_json(
    input_path: str,
    *,
    project_key: str = '',
    product_family: str = '',
    feature_name: str = '',
) -> Dict[str, Any]:
    '''Read a CSV or Excel plan file and convert it to a feature-plan JSON dict.

    This is the high-level convenience function that combines read_plan_rows()
    and plan_rows_to_json().

    Args:
        input_path: Path to a .csv or .xlsx file containing a feature plan.
        project_key: Override the project key (auto-detected from rows if empty).
        product_family: Override the product family (auto-detected if empty).
        feature_name: Feature name for the plan (defaults to first epic summary).

    Returns:
        A dict matching the JiraPlan schema, ready for _run_execute_plan().
    '''
    rows, table_format = read_plan_rows(input_path)
    log.info(
        f'Read {len(rows)} rows from {input_path} '
        f'(detected table_format={table_format})'
    )
    return plan_rows_to_json(
        rows,
        table_format=table_format,
        project_key=project_key,
        product_family=product_family,
        feature_name=feature_name,
    )


def write_plan_json(plan: Dict[str, Any], output_path: str) -> str:
    '''Write a feature-plan dict to a JSON file.

    Args:
        plan: The feature-plan dict (as returned by plan_rows_to_json).
        output_path: Destination file path.

    Returns:
        The output path written to.
    '''
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(plan, f, indent=2, ensure_ascii=False)
    log.info(f'Wrote plan JSON ({plan.get("total_tickets", 0)} tickets) to: {output_path}')
    return output_path


# ============================================================================
# @tool-decorated entry point for the agentic pipeline
# ============================================================================

@tool(
    name='plan_to_csv',
    description=(
        'Convert a feature-plan JSON file (as produced by the Feature Planning pipeline) '
        'into a CSV file matching the standard Jira CSV format used by dump_tickets_to_file(). '
        'Supports flat and indented table formats. Returns the output file path.'
    ),
    parameters={
        'input_path': 'Path to the feature-plan JSON file (e.g. feature_plan.json)',
        'output_path': 'Output file basename without extension. Extension is added automatically. Defaults to <input_basename>',
        'table_format': "Table layout: 'indented' (default) or 'flat'",
        'include_description': 'Include full description column (default: false)',
        'output_format': "Output format: 'csv' (default) or 'excel'",
    },
)
def plan_to_csv(
    input_path: str,
    output_path: str = '',
    table_format: str = 'indented',
    include_description: bool = False,
    output_format: str = 'csv',
) -> ToolResult:
    '''Convert a feature-plan JSON to the standard Jira CSV/Excel format.

    The CSV is always written.  When *output_format* is ``'excel'`` an
    additional ``.xlsx`` file is produced alongside the CSV.

    Args:
        input_path: Path to the feature-plan JSON file.
        output_path: Output basename without extension (auto-derived if empty).
        table_format: 'indented' (default) or 'flat'.
        include_description: Whether to include the description column.
        output_format: 'csv' or 'excel'.

    Returns:
        ToolResult with the output file path(s) and row count.
    '''
    # --- Validate input ---
    if not os.path.exists(input_path):
        return ToolResult.error(error=f'Input file not found: {input_path}')

    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            plan = json.load(f)
    except json.JSONDecodeError as e:
        return ToolResult.error(error=f'Invalid JSON in {input_path}: {e}')

    # --- Convert ---
    rows = plan_json_to_rows(plan, include_description=include_description)
    if not rows:
        return ToolResult.error(error='Plan contains no epics or stories')

    # --- Resolve output paths ---
    fmt = output_format.lower().strip()
    csv_path = _resolve_output_path(input_path, output_path or None, 'csv')

    # --- Always write CSV ---
    try:
        write_plan_csv(rows, csv_path, table_format=table_format)
    except Exception as e:
        return ToolResult.error(error=f'Failed to write CSV: {e}')

    written_paths = [csv_path]

    # --- Optionally write Excel alongside the CSV ---
    if fmt == 'excel':
        xlsx_path = _resolve_output_path(input_path, output_path or None, 'excel')
        try:
            write_plan_excel(rows, xlsx_path, table_format=table_format)
            written_paths.append(xlsx_path)
        except Exception as e:
            log.warning(f'Excel write failed (CSV still written): {e}')

    return ToolResult.success(
        data={
            'output_path': csv_path,
            'output_paths': written_paths,
            'total_rows': len(rows),
            'epics': sum(1 for r in rows if r.get('issue_type') == 'Epic'),
            'stories': sum(1 for r in rows if r.get('issue_type') == 'Story'),
            'format': fmt,
            'table_format': table_format,
        },
        message=f'Exported {len(rows)} rows to {", ".join(written_paths)}',
    )


@tool(
    name='plan_json_to_dict_rows',
    description=(
        'Convert a feature-plan JSON (dict or file path) into a list of flat row dicts '
        'matching the standard Jira CSV schema. Useful for in-memory pipeline processing '
        'without writing to disk.'
    ),
    parameters={
        'plan_or_path': 'Either a dict (the plan) or a string file path to the JSON',
        'include_description': 'Include full description text (default: false)',
    },
)
def plan_json_to_dict_rows(
    plan_or_path,
    include_description: bool = False,
) -> ToolResult:
    '''Return plan rows as a list of dicts (no file I/O).

    Args:
        plan_or_path: A dict (the plan itself) or a string path to a JSON file.
        include_description: Whether to include the description column.

    Returns:
        ToolResult whose data['rows'] is the list of row dicts.
    '''
    # Accept either a dict or a file path
    if isinstance(plan_or_path, str):
        if not os.path.exists(plan_or_path):
            return ToolResult.error(error=f'File not found: {plan_or_path}')
        try:
            with open(plan_or_path, 'r', encoding='utf-8') as f:
                plan = json.load(f)
        except json.JSONDecodeError as e:
            return ToolResult.error(error=f'Invalid JSON: {e}')
    elif isinstance(plan_or_path, dict):
        plan = plan_or_path
    else:
        return ToolResult.error(error=f'Expected dict or file path, got {type(plan_or_path).__name__}')

    rows = plan_json_to_rows(plan, include_description=include_description)
    return ToolResult.success(
        data={
            'rows': rows,
            'total_rows': len(rows),
            'epics': sum(1 for r in rows if r.get('issue_type') == 'Epic'),
            'stories': sum(1 for r in rows if r.get('issue_type') == 'Story'),
        },
        message=f'Converted plan to {len(rows)} rows',
    )


# ---------------------------------------------------------------------------
# @tool: CSV/Excel → JSON (reverse direction)
# ---------------------------------------------------------------------------

@tool(
    name='plan_file_to_plan_json',
    description=(
        'Convert a feature-plan CSV or Excel file (flat or indented table format) '
        'into the standard feature-plan JSON dict used by the execution pipeline. '
        'Auto-detects flat vs indented format from the column headers. '
        'Optionally writes the JSON to disk.'
    ),
    parameters={
        'input_path': 'Path to a .csv or .xlsx file containing a feature plan',
        'output_path': 'Optional path to write the resulting JSON file. If empty, no file is written.',
        'project_key': 'Override the project key (auto-detected from rows if empty)',
        'product_family': 'Override the product family (auto-detected from rows if empty)',
        'feature_name': 'Feature name for the plan (defaults to first epic summary)',
    },
)
def plan_file_to_plan_json(
    input_path: str,
    output_path: str = '',
    project_key: str = '',
    product_family: str = '',
    feature_name: str = '',
) -> ToolResult:
    '''Convert a CSV/Excel plan file into a feature-plan JSON dict.

    Supports both flat and indented table formats.  The table format is
    auto-detected from the column headers.

    Args:
        input_path: Path to a .csv or .xlsx file.
        output_path: Optional path to write the JSON output.
        project_key: Override the project key.
        product_family: Override the product family.
        feature_name: Feature name for the plan.

    Returns:
        ToolResult with the plan dict and optional output path.
    '''
    if not os.path.exists(input_path):
        return ToolResult.error(error=f'Input file not found: {input_path}')

    try:
        plan = plan_file_to_json(
            input_path,
            project_key=project_key,
            product_family=product_family,
            feature_name=feature_name,
        )
    except (ValueError, ImportError, FileNotFoundError) as e:
        return ToolResult.error(error=str(e))
    except Exception as e:
        return ToolResult.error(error=f'Failed to convert {input_path}: {e}')

    # Optionally write the JSON to disk
    written_path = ''
    if output_path:
        try:
            written_path = write_plan_json(plan, output_path)
        except Exception as e:
            return ToolResult.error(
                error=f'Converted successfully but failed to write JSON: {e}'
            )

    return ToolResult.success(
        data={
            'plan': plan,
            'output_path': written_path,
            'total_epics': plan.get('total_epics', 0),
            'total_stories': plan.get('total_stories', 0),
            'total_tickets': plan.get('total_tickets', 0),
            'project_key': plan.get('project_key', ''),
            'feature_name': plan.get('feature_name', ''),
        },
        message=(
            f'Converted {input_path} → {plan.get("total_tickets", 0)} tickets '
            f'({plan.get("total_epics", 0)} epics, {plan.get("total_stories", 0)} stories)'
            + (f' → {written_path}' if written_path else '')
        ),
    )


# ============================================================================
# BaseTool class for agent registration
# ============================================================================

class PlanExportTools(BaseTool):
    '''Collection of plan-export tools for agent use.'''

    @tool(description='Convert a feature-plan JSON to Jira-compatible CSV')
    def plan_to_csv(self, input_path: str, output_path: str = '',
                    table_format: str = 'indented', include_description: bool = False,
                    output_format: str = 'csv') -> ToolResult:
        return plan_to_csv(input_path, output_path, table_format,
                           include_description, output_format)

    @tool(description='Convert a feature-plan JSON to flat row dicts (in-memory)')
    def plan_json_to_dict_rows(self, plan_or_path, include_description: bool = False) -> ToolResult:
        return plan_json_to_dict_rows(plan_or_path, include_description)

    @tool(description='Convert a CSV/Excel plan file to feature-plan JSON')
    def plan_file_to_plan_json(self, input_path: str, output_path: str = '',
                               project_key: str = '', product_family: str = '',
                               feature_name: str = '') -> ToolResult:
        return plan_file_to_plan_json(input_path, output_path, project_key,
                                      product_family, feature_name)


# ============================================================================
# Standalone CLI entry point
# ============================================================================

def _cli_main() -> None:
    '''CLI entry point for bidirectional plan conversion.

    JSON → CSV/Excel:
        python -m tools.plan_export_tools plan.json [-o out.csv]
        python -m tools.plan_export_tools plan.json -f excel

    CSV/Excel → JSON:
        python -m tools.plan_export_tools plan.csv --to-json [-o out.json]
        python -m tools.plan_export_tools plan.xlsx --to-json
    '''
    import argparse

    parser = argparse.ArgumentParser(
        prog='plan_export_tools',
        description=(
            'Convert between feature-plan JSON and Jira-compatible CSV/Excel.\n\n'
            'Default direction: JSON → CSV/Excel.\n'
            'Use --to-json to reverse: CSV/Excel → JSON.'
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s plan.json                        # JSON → indented CSV
  %(prog)s plan.json -f excel               # JSON → CSV + Excel
  %(prog)s plan.json -t flat                # JSON → flat CSV
  %(prog)s plan.csv --to-json               # CSV → JSON (auto-detect format)
  %(prog)s plan.xlsx --to-json -o out.json  # Excel → JSON with explicit output
  %(prog)s plan.csv --to-json --project XYZ # CSV → JSON, override project key
        ''',
    )
    parser.add_argument(
        'input',
        help='Path to the input file (JSON, CSV, or Excel)',
    )
    parser.add_argument(
        '-o', '--outfile',
        default='',
        help='Output file path (default: derived from input basename)',
    )
    parser.add_argument(
        '-f', '--format',
        choices=['csv', 'excel'],
        default='csv',
        help='Output format for JSON→CSV/Excel (default: csv)',
    )
    parser.add_argument(
        '-t', '--table-format',
        choices=['flat', 'indented'],
        default='indented',
        help="Table layout: 'indented' (default) or 'flat'",
    )
    parser.add_argument(
        '--include-description',
        action='store_true',
        default=False,
        help='Include the full description column (JSON→CSV only)',
    )
    parser.add_argument(
        '--to-json',
        action='store_true',
        default=False,
        dest='to_json',
        help='Reverse direction: convert CSV/Excel → feature-plan JSON',
    )
    parser.add_argument(
        '--project',
        default='',
        dest='project_key',
        help='Override the project key (used with --to-json)',
    )
    parser.add_argument(
        '--product-family',
        default='',
        dest='product_family',
        help='Override the product family (used with --to-json)',
    )
    parser.add_argument(
        '--feature-name',
        default='',
        dest='feature_name',
        help='Feature name for the plan (used with --to-json)',
    )
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        default=False,
        help='Enable verbose logging',
    )

    args = parser.parse_args()

    # Configure logging
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(levelname)-8s %(message)s',
    )

    # ---------------------------------------------------------------
    # Reverse direction: CSV/Excel → JSON
    # ---------------------------------------------------------------
    if args.to_json:
        # Derive output path if not provided
        out = args.outfile
        if not out:
            base, _ = os.path.splitext(args.input)
            out = f'{base}.json'

        result = plan_file_to_plan_json(
            input_path=args.input,
            output_path=out,
            project_key=args.project_key,
            product_family=args.product_family,
            feature_name=args.feature_name,
        )

        if hasattr(result, 'error') and result.error:
            print(f'ERROR: {result.error}')
            raise SystemExit(1)

        data = result.data if hasattr(result, 'data') else {}
        print(f'Output:       {data.get("output_path", out)}')
        print(f'Project:      {data.get("project_key", "?")}')
        print(f'Feature:      {data.get("feature_name", "?")}')
        print(f'Epics:        {data.get("total_epics", 0)}')
        print(f'Stories:      {data.get("total_stories", 0)}')
        print(f'Total tickets:{data.get("total_tickets", 0)}')
        return

    # ---------------------------------------------------------------
    # Default direction: JSON → CSV/Excel
    # ---------------------------------------------------------------
    result = plan_to_csv(
        input_path=args.input,
        output_path=args.outfile,
        table_format=args.table_format,
        include_description=args.include_description,
        output_format=args.format,
    )

    # Print result (works with both real ToolResult and stub)
    if hasattr(result, 'error') and result.error:
        print(f'ERROR: {result.error}')
        raise SystemExit(1)

    data = result.data if hasattr(result, 'data') else {}
    for p in data.get('output_paths', [data.get('output_path', '?')]):
        print(f'Output:  {p}')
    print(f'Rows:    {data.get("total_rows", 0)}')
    print(f'Epics:   {data.get("epics", 0)}')
    print(f'Stories: {data.get("stories", 0)}')
    print(f'Format:  {data.get("format", "?")} ({data.get("table_format", "?")})')


if __name__ == '__main__':
    _cli_main()
