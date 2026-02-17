##########################################################################################
#
# Module: tools/excel_tools.py
#
# Description: Agent tool wrappers for Excel map building and Excel utility operations.
#              Wraps the build-excel-map pipeline from pm_agent.py and excel_utils.py
#              functions as agent-callable tools.
#
# Author: Cornelis Networks
#
##########################################################################################

import logging
import os
import sys

log = logging.getLogger(os.path.basename(sys.argv[0]))

from typing import Optional, List, Dict, Any

try:
    from tools.base import tool, ToolResult, BaseTool
except ImportError:
    log.warning('tools.base not available; excel_tools will not register @tool decorators')
    # Provide a no-op decorator so the module can still be imported
    def tool(**kwargs):
        def decorator(func):
            return func
        return decorator
    class ToolResult:
        pass
    class BaseTool:
        pass

# Lazy imports to avoid circular dependencies and heavy startup cost
JIRA_UTILS_AVAILABLE = False
EXCEL_UTILS_AVAILABLE = False

try:
    import jira_utils
    JIRA_UTILS_AVAILABLE = True
except ImportError:
    log.debug('jira_utils not available for excel_tools')

try:
    import excel_utils
    EXCEL_UTILS_AVAILABLE = True
except ImportError:
    log.debug('excel_utils not available for excel_tools')


# ****************************************************************************************
# Agent tools
# ****************************************************************************************

@tool(
    description='Build a multi-sheet Excel workbook mapping one or more root tickets and '
                'all their related issues child hierarchies. Sheet 1 ("Tickets") is the '
                'merged get-related overview. Sheets 2..N are per-ticket get-children results.',
    parameters={
        'ticket_keys': 'List of root Jira ticket keys (e.g., ["STL-74071", "STL-76297"])',
        'hierarchy_depth': 'Depth for related issue traversal (default: 1)',
        'limit': 'Optional max tickets per step',
        'output_file': 'Output filename (default: {ticket_key}.xlsx)',
    }
)
def build_excel_map(
    ticket_keys: List[str],
    hierarchy_depth: int = 1,
    limit: int = None,
    output_file: str = None,
) -> ToolResult:
    '''
    Build a multi-sheet Excel workbook mapping one or more tickets and their
    related issues' child hierarchies.

    Orchestrates:
      1. For each root ticket, _get_related_data() and merge results (dedup by key)
      2. Writes Tickets sheet (indented format) from merged data
      3. _get_children_data() for each depth=1 ticket
      4. Assembles all sheets into one workbook

    Input:
        ticket_keys: List of root Jira ticket keys.
        hierarchy_depth: Depth for related issue traversal (default 1).
        limit: Optional max tickets per step.
        output_file: Output filename (default: {first_key}.xlsx or combined).

    Output:
        ToolResult with the output file path and summary.
    '''
    if not JIRA_UTILS_AVAILABLE:
        return ToolResult.error('jira_utils is not available')

    # Accept a single string for convenience
    if isinstance(ticket_keys, str):
        ticket_keys = [ticket_keys]

    import tempfile
    import shutil
    from copy import copy

    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        return ToolResult.error('openpyxl is required. Install with: pip install openpyxl')

    keys = [k.upper() for k in ticket_keys]
    if output_file:
        out_file = output_file
    elif len(keys) == 1:
        out_file = f'{keys[0]}.xlsx'
    else:
        out_file = f'{"_".join(keys)}.xlsx'
    if not out_file.endswith('.xlsx'):
        out_file = f'{out_file}.xlsx'

    temp_dir = tempfile.mkdtemp(prefix='excel_map_')
    temp_files = []

    try:
        # Step 1: Connect to Jira
        jira = jira_utils.connect_to_jira()

        # Step 2: Get related issues for each root ticket, merge with dedup
        merged_data = []
        seen_keys = set()

        for root_key in keys:
            related_data = jira_utils._get_related_data(jira, root_key, hierarchy=hierarchy_depth, limit=limit)
            for item in related_data:
                issue_key = item['issue'].get('key', '')
                if issue_key and issue_key not in seen_keys:
                    seen_keys.add(issue_key)
                    merged_data.append(item)

        # Write Tickets sheet to temp file
        map_temp = os.path.join(temp_dir, '_map_temp.xlsx')
        temp_files.append(map_temp)

        map_extras = {
            item['issue'].get('key', ''): {
                'depth': item.get('depth'),
                'via': item.get('via'),
                'relation': item.get('relation'),
                'from_key': item.get('from_key'),
            }
            for item in merged_data
        }
        jira_utils.dump_tickets_to_file(
            [item['issue'] for item in merged_data],
            map_temp, 'excel', map_extras, table_format='indented'
        )

        # Step 3: Get children for each depth=1 ticket
        depth1_keys = [item['issue'].get('key', '') for item in merged_data if item['depth'] == 1]

        children_temps = []
        for ticket in depth1_keys:
            try:
                children_data = jira_utils._get_children_data(jira, ticket, limit=None)
                child_temp = os.path.join(temp_dir, f'temp_{ticket}.xlsx')
                temp_files.append(child_temp)

                child_extras = {
                    item['issue'].get('key', ''): {'depth': item.get('depth')}
                    for item in children_data
                }
                jira_utils.dump_tickets_to_file(
                    [item['issue'] for item in children_data],
                    child_temp, 'excel', child_extras, table_format='indented'
                )
                children_temps.append((ticket, child_temp, len(children_data)))
            except Exception as e:
                log.warning(f'Failed to get children for {ticket}: {e}')

        # Step 4: Assemble final workbook
        final_wb = Workbook()
        final_wb.remove(final_wb.active)

        total_rows = 0
        sheet_count = 0

        def _copy_sheet(src_wb_path, dest_wb, sheet_name):
            nonlocal total_rows, sheet_count
            src_wb = load_workbook(src_wb_path)
            src_ws = src_wb.active
            safe_name = sheet_name[:31]
            dest_ws = dest_wb.create_sheet(title=safe_name)

            row_count = 0
            for row in src_ws.iter_rows():
                row_count += 1
                for cell in row:
                    dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.font:
                        dest_cell.font = copy(cell.font)
                    if cell.fill:
                        dest_cell.fill = copy(cell.fill)
                    if cell.alignment:
                        dest_cell.alignment = copy(cell.alignment)
                    if cell.border:
                        dest_cell.border = copy(cell.border)
                    if cell.number_format:
                        dest_cell.number_format = cell.number_format
                    if cell.hyperlink:
                        dest_cell.hyperlink = cell.hyperlink

            for col_letter, dim in src_ws.column_dimensions.items():
                dest_ws.column_dimensions[col_letter].width = dim.width

            for merged_range in src_ws.merged_cells.ranges:
                dest_ws.merge_cells(str(merged_range))

            for cf_rule in src_ws.conditional_formatting:
                for rule in cf_rule.rules:
                    dest_ws.conditional_formatting.add(str(cf_rule), rule)

            if src_ws.freeze_panes:
                dest_ws.freeze_panes = src_ws.freeze_panes

            src_wb.close()
            data_rows = max(0, row_count - 1)
            total_rows += data_rows
            sheet_count += 1
            return data_rows

        # Sheet 1: Tickets (merged overview)
        _copy_sheet(map_temp, final_wb, 'Tickets')

        # Sheets 2..N: Per-ticket children
        for ticket, child_temp, _ in children_temps:
            _copy_sheet(child_temp, final_wb, ticket)

        final_wb.save(out_file)
        final_wb.close()

        return ToolResult.success({
            'output_file': out_file,
            'sheet_count': sheet_count,
            'total_rows': total_rows,
            'depth1_tickets': len(depth1_keys),
            'related_count': len(merged_data),
            'root_tickets': keys,
        })

    except Exception as e:
        log.error(f'build_excel_map failed: {e}', exc_info=True)
        return ToolResult.error(str(e))

    finally:
        # Always clean up temp files for agent tool usage
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass


@tool(
    description='Concatenate multiple Excel files into one using merge-sheet or add-sheet method.',
    parameters={
        'input_files': 'List of Excel file paths to concatenate',
        'output_file': 'Output Excel file path',
        'method': 'Concatenation method: merge-sheet or add-sheet',
    }
)
def concat_excel(
    input_files: List[str],
    output_file: str,
    method: str = 'merge-sheet',
) -> ToolResult:
    '''
    Concatenate multiple Excel files into one workbook.

    Input:
        input_files: List of Excel file paths.
        output_file: Output file path.
        method: 'merge-sheet' (combine all into one sheet) or 'add-sheet' (each file as a sheet).

    Output:
        ToolResult with the output file path.
    '''
    if not EXCEL_UTILS_AVAILABLE:
        return ToolResult.error('excel_utils is not available')

    try:
        if method == 'merge-sheet':
            excel_utils.concat_merge_sheet(input_files, output_file)
        elif method == 'add-sheet':
            excel_utils.concat_add_sheet(input_files, output_file)
        else:
            return ToolResult.error(f'Unknown method: {method}. Use merge-sheet or add-sheet.')

        return ToolResult.success({
            'output_file': output_file,
            'input_count': len(input_files),
            'method': method,
        })
    except Exception as e:
        log.error(f'concat_excel failed: {e}', exc_info=True)
        return ToolResult.error(str(e))


@tool(
    description='Convert an Excel file to CSV format.',
    parameters={
        'input_file': 'Excel file path to convert',
        'output_file': 'Optional output CSV file path',
    }
)
def excel_to_csv(
    input_file: str,
    output_file: str = None,
) -> ToolResult:
    '''
    Convert an Excel file to CSV.

    Input:
        input_file: Path to the .xlsx file.
        output_file: Optional output path (default: replaces .xlsx with .csv).

    Output:
        ToolResult with the output file path.
    '''
    if not EXCEL_UTILS_AVAILABLE:
        return ToolResult.error('excel_utils is not available')

    try:
        result_path = excel_utils.convert_to_csv(input_file, output_file)
        return ToolResult.success({'output_file': result_path})
    except Exception as e:
        log.error(f'excel_to_csv failed: {e}', exc_info=True)
        return ToolResult.error(str(e))


@tool(
    description='Convert a CSV file to Excel format with styling.',
    parameters={
        'input_file': 'CSV file path to convert',
        'output_file': 'Optional output Excel file path',
    }
)
def csv_to_excel(
    input_file: str,
    output_file: str = None,
) -> ToolResult:
    '''
    Convert a CSV file to Excel with header styling and auto-fit columns.

    Input:
        input_file: Path to the .csv file.
        output_file: Optional output path (default: replaces .csv with .xlsx).

    Output:
        ToolResult with the output file path.
    '''
    if not EXCEL_UTILS_AVAILABLE:
        return ToolResult.error('excel_utils is not available')

    try:
        result_path = excel_utils.convert_from_csv(input_file, output_file)
        return ToolResult.success({'output_file': result_path})
    except Exception as e:
        log.error(f'csv_to_excel failed: {e}', exc_info=True)
        return ToolResult.error(str(e))


@tool(
    description='Diff two Excel files and produce a comparison report.',
    parameters={
        'input_files': 'List of two Excel file paths to compare',
        'output_file': 'Optional output file path for the diff report',
    }
)
def diff_excel(
    input_files: List[str],
    output_file: str = None,
) -> ToolResult:
    '''
    Compare two Excel files and produce a diff report.

    Input:
        input_files: List of exactly two Excel file paths.
        output_file: Optional output path for the diff report.

    Output:
        ToolResult with the diff summary.
    '''
    if not EXCEL_UTILS_AVAILABLE:
        return ToolResult.error('excel_utils is not available')

    if len(input_files) != 2:
        return ToolResult.error('diff_excel requires exactly 2 input files')

    try:
        result_path = excel_utils.diff_files(input_files, output_file)
        return ToolResult.success({'output_file': result_path})
    except Exception as e:
        log.error(f'diff_excel failed: {e}', exc_info=True)
        return ToolResult.error(str(e))


# ****************************************************************************************
# Tool class for registration with agent framework
# ****************************************************************************************

class ExcelTools(BaseTool):
    '''Collection of Excel-related agent tools.'''

    @tool(description='Build a multi-sheet Excel map from one or more root tickets')
    def build_excel_map(
        self, ticket_keys: List[str], hierarchy_depth: int = 1,
        limit: int = None, output_file: str = None
    ) -> ToolResult:
        return build_excel_map(ticket_keys, hierarchy_depth, limit, output_file)

    @tool(description='Concatenate Excel files')
    def concat_excel(
        self, input_files: List[str], output_file: str, method: str = 'merge-sheet'
    ) -> ToolResult:
        return concat_excel(input_files, output_file, method)

    @tool(description='Convert Excel to CSV')
    def excel_to_csv(self, input_file: str, output_file: str = None) -> ToolResult:
        return excel_to_csv(input_file, output_file)

    @tool(description='Convert CSV to Excel')
    def csv_to_excel(self, input_file: str, output_file: str = None) -> ToolResult:
        return csv_to_excel(input_file, output_file)

    @tool(description='Diff two Excel files')
    def diff_excel(self, input_files: List[str], output_file: str = None) -> ToolResult:
        return diff_excel(input_files, output_file)
