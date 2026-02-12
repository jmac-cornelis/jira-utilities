##########################################################################################
#
# Script name: jira_utils.py
#
# Description: Jira utilities for interacting with Cornelis Networks Jira instance.
#
# Author: John Macdonald
#
# Credentials:
#   This script uses Jira API tokens for authentication. To set up:
#   1. Generate an API token at: https://id.atlassian.com/manage-profile/security/api-tokens
#   2. Set environment variables:
#      export JIRA_EMAIL="your.email@cornelisnetworks.com"
#      export JIRA_API_TOKEN="your_api_token_here"
#   
#   NEVER commit credentials to version control.
#
##########################################################################################

import argparse
import logging
import sys
import os
import time
import json
import csv
from datetime import date, datetime, timedelta
import re
import requests

from dotenv import load_dotenv

# Load environment variables from the default .env if present.
#
# CLI users can override this at runtime with --env; see handle_args().
# We use override=False here so real process environment variables (e.g. set by CI)
# still take precedence.
load_dotenv(override=False)

try:
    from jira import JIRA
except ImportError:
    print('Error: jira package not installed. Run: pip install jira')
    sys.exit(1)

# ****************************************************************************************
# Global data and configuration
# ****************************************************************************************
# Set global variables here and log.debug them below

DEFAULT_JIRA_URL = 'https://cornelisnetworks.atlassian.net'

# Jira configuration (allow override via env / .env)
JIRA_URL = os.getenv('JIRA_URL', DEFAULT_JIRA_URL)

# Logging config
log = logging.getLogger(os.path.basename(sys.argv[0]))
log.setLevel(logging.DEBUG)

# File handler for logging
fh = logging.FileHandler('jira_utils.log', mode='w')
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter(
    '%(asctime)-15s [%(funcName)25s:%(lineno)-5s] %(levelname)-8s %(message)s')
fh.setFormatter(formatter)
log.addHandler(fh)  # Add file handler to logger                

log.debug(f'Global data and configuration for this script...')
log.debug(f'JIRA_URL: {JIRA_URL}')

# Output control - set by handle_args()
_quiet_mode = False
_show_jql = False

def output(message=''):
    '''
    Print user-facing output, respecting quiet mode.
    Always logs to file regardless of quiet mode.
    
    For tables and user-facing output:
    - stdout: Clean output without logger prefix (via print)
    - log file: Full logger format with timestamps (written directly to file handler)
    
    Input:
        message: String to output (default empty for blank line).
    
    Output:
        None; prints to stdout unless in quiet mode.
    
    Side Effects:
        Always logs message to log file at INFO level.
    '''
    # Log to file only (bypass stdout handler by writing directly to file handler)
    if message:
        record = logging.LogRecord(
            name=log.name,
            level=logging.INFO,
            pathname=__file__,
            lineno=0,
            msg=f'OUTPUT: {message}',
            args=(),
            exc_info=None,
            func='output'
        )
        fh.emit(record)
    
    # Only print to stdout if not in quiet mode (clean output)
    if not _quiet_mode:
        print(message)


def print_ticket_table_header():
    '''Print the header row for ticket tables.'''
    output('-' * 170)
    output(f'{"Key":<15} {"Type":<12} {"Status":<12} {"Priority":<10} {"Created":<12} {"Updated":<12} {"Fix Version":<15} {"Assignee":<18} {"Summary":<30}')
    output('-' * 170)


def print_ticket_row(issue):
    '''
    Print a single ticket row in the standard table format.
    
    Input:
        issue: Issue dict from Jira API.
    '''
    key = issue.get('key', 'N/A')
    fields = issue.get('fields', {})
    
    issue_type = fields.get('issuetype', {}).get('name', 'N/A') if fields.get('issuetype') else 'N/A'
    status = fields.get('status', {}).get('name', 'N/A') if fields.get('status') else 'N/A'
    priority = fields.get('priority', {}).get('name', 'N/A') if fields.get('priority') else 'N/A'
    
    created = fields.get('created', 'N/A')
    if created and created != 'N/A':
        try:
            created_dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
            created = created_dt.strftime('%Y-%m-%d')
        except:
            created = created[:10] if len(created) >= 10 else created
    
    updated = fields.get('updated', 'N/A')
    if updated and updated != 'N/A':
        try:
            updated_dt = datetime.fromisoformat(updated.replace('Z', '+00:00'))
            updated = updated_dt.strftime('%Y-%m-%d')
        except:
            updated = updated[:10] if len(updated) >= 10 else updated
    
    fix_versions = fields.get('fixVersions', [])
    fix_version = ', '.join([v.get('name', '') for v in fix_versions]) if fix_versions else 'N/A'
    
    assignee = fields.get('assignee', {})
    assignee_name = assignee.get('displayName', 'Unassigned') if assignee else 'Unassigned'
    summary = fields.get('summary', 'N/A') or 'N/A'
    
    # Truncate for display
    if len(issue_type) > 10:
        issue_type = issue_type[:10] + '..'
    if len(status) > 10:
        status = status[:10] + '..'
    if len(priority) > 8:
        priority = priority[:8] + '..'
    if len(fix_version) > 13:
        fix_version = fix_version[:13] + '..'
    if len(assignee_name) > 16:
        assignee_name = assignee_name[:16] + '..'
    if len(summary) > 28:
        summary = summary[:28] + '..'
    
    output(f'{key:<15} {issue_type:<12} {status:<12} {priority:<10} {created:<12} {updated:<12} {fix_version:<15} {assignee_name:<18} {summary:<30}')


def print_ticket_table_footer(count):
    '''Print the footer row for ticket tables.'''
    output('=' * 170)
    output(f'Total: {count} tickets')
    output('')


def print_dashboard_table_header():
    '''Print the header row for dashboard tables.'''
    output('-' * 130)
    output(f'{"ID":<12} {"Name":<35} {"Owner":<25} {"Shared":<10} {"Favourite":<10} {"Description":<35}')
    output('-' * 130)


def print_dashboard_row(dashboard):
    '''
    Print a single dashboard row in the standard table format.
    
    Input:
        dashboard: Dashboard dict from Jira API.
    '''
    dash_id = str(dashboard.get('id', 'N/A'))
    name = dashboard.get('name', 'N/A') or 'N/A'
    
    # Extract owner info
    owner = dashboard.get('owner', {})
    owner_name = owner.get('displayName', 'N/A') if owner else 'N/A'
    
    # Check if shared (has sharePermissions with entries)
    share_permissions = dashboard.get('sharePermissions', [])
    is_shared = 'Yes' if share_permissions else 'No'
    
    # Check if favourite
    is_favourite = 'Yes' if dashboard.get('isFavourite', False) else 'No'
    
    description = dashboard.get('description', '') or ''
    
    # Truncate for display
    if len(name) > 33:
        name = name[:33] + '..'
    if len(owner_name) > 23:
        owner_name = owner_name[:23] + '..'
    if len(description) > 33:
        description = description[:33] + '..'
    
    output(f'{dash_id:<12} {name:<35} {owner_name:<25} {is_shared:<10} {is_favourite:<10} {description:<35}')


def print_dashboard_table_footer(count):
    '''Print the footer row for dashboard tables.'''
    output('=' * 130)
    output(f'Total: {count} dashboards')
    output('')


def print_gadget_table_header():
    '''Print the header row for gadget tables.'''
    output('-' * 120)
    output(f'{"ID":<12} {"Module Key":<40} {"Title":<30} {"Position":<12} {"Color":<15}')
    output('-' * 120)


def print_gadget_row(gadget):
    '''
    Print a single gadget row in the standard table format.
    
    Input:
        gadget: Gadget dict from Jira API.
    '''
    gadget_id = str(gadget.get('id', 'N/A'))
    module_key = gadget.get('moduleKey', 'N/A') or 'N/A'
    title = gadget.get('title', 'N/A') or 'N/A'
    
    # Extract position
    position = gadget.get('position', {})
    row = position.get('row', 0)
    col = position.get('column', 0)
    position_str = f'{row},{col}'
    
    color = gadget.get('color', 'N/A') or 'N/A'
    
    # Truncate for display
    if len(module_key) > 38:
        module_key = module_key[:38] + '..'
    if len(title) > 28:
        title = title[:28] + '..'
    
    output(f'{gadget_id:<12} {module_key:<40} {title:<30} {position_str:<12} {color:<15}')


def print_gadget_table_footer(count):
    '''Print the footer row for gadget tables.'''
    output('=' * 120)
    output(f'Total: {count} gadgets')
    output('')


# Store the last JQL query for display at end of operation
_last_jql = None

def show_jql(jql):
    '''
    Store the JQL query for display at end of operation if --show-jql flag is set.
    
    Input:
        jql: The JQL query string to store.
    
    Output:
        None; stores JQL for later display.
    '''
    global _last_jql
    if _show_jql:
        _last_jql = jql
        log.debug(f'Stored JQL for end-of-operation display: {jql}')


def display_jql():
    '''
    Display the stored JQL query at end of operation and write to jql.txt.
    Called from main() before "Operation complete" message.
    
    Output:
        None; prints JQL to stdout and writes to jql.txt if _show_jql is True.
    '''
    global _last_jql
    if _show_jql and _last_jql:
        log.info(f'JQL: {_last_jql}')
        output('')
        output('=' * 80)
        output('Equivalent JQL Query:')
        output('=' * 80)
        output(_last_jql)
        output('=' * 80)
        output('')
        
        # Write to jql.txt
        try:
            with open('jql.txt', 'w') as f:
                f.write(_last_jql)
            log.info('Wrote JQL to jql.txt')
        except Exception as e:
            log.warning(f'Failed to write jql.txt: {e}')


# ****************************************************************************************
# Exceptions
# ****************************************************************************************

class Error(Exception):
    '''
    Base class for exceptions in this module.
    '''
    pass

class JiraConnectionError(Error):
    '''
    Exception raised when Jira connection fails.
    '''
    def __init__(self, message):
        self.message = f'Jira connection failed: {message}'
        super().__init__(self.message)

class JiraCredentialsError(Error):
    '''
    Exception raised when Jira credentials are missing or invalid.
    '''
    def __init__(self, message):
        self.message = f'Jira credentials error: {message}'
        super().__init__(self.message)

class JiraProjectError(Error):
    '''
    Exception raised when Jira project operations fail.
    '''
    def __init__(self, message):
        self.message = f'Jira project error: {message}'
        super().__init__(self.message)


class JiraDashboardError(Error):
    '''
    Exception raised when Jira dashboard operations fail.
    '''
    def __init__(self, message):
        self.message = f'Jira dashboard error: {message}'
        super().__init__(self.message)


# ****************************************************************************************
# Functions
# ****************************************************************************************

def get_jira_credentials():
    '''
    Retrieve Jira credentials from environment variables.

    Input:
        None directly; reads from environment variables JIRA_EMAIL and JIRA_API_TOKEN.

    Output:
        Tuple of (email, api_token) strings.

    Raises:
        JiraCredentialsError: If required environment variables are not set.
    '''
    log.debug('Entering get_jira_credentials()')
    email = os.environ.get('JIRA_EMAIL')
    api_token = os.environ.get('JIRA_API_TOKEN')
    
    if not email:
        raise JiraCredentialsError('JIRA_EMAIL environment variable not set')
    if not api_token:
        raise JiraCredentialsError('JIRA_API_TOKEN environment variable not set')
    
    log.debug(f'Retrieved credentials for: {email}')
    return email, api_token


def connect_to_jira():
    '''
    Establish connection to Jira instance using API token authentication.

    Input:
        None directly; uses credentials from environment variables.

    Output:
        JIRA object connected to the Cornelis Networks Jira instance.

    Raises:
        JiraConnectionError: If connection to Jira fails.
        JiraCredentialsError: If credentials are missing.
    '''
    log.debug('Entering connect_to_jira()')
    email, api_token = get_jira_credentials()
    
    log.info(f'Connecting to Jira at {JIRA_URL}...')
    try:
        jira = JIRA(
            server=JIRA_URL,
            basic_auth=(email, api_token),
            options={'rest_api_version': '3'}
        )
        log.info('Successfully connected to Jira')
        return jira
    except Exception as e:
        raise JiraConnectionError(str(e))


def list_projects(jira):
    '''
    List all available Jira projects.

    Input:
        jira: JIRA object with active connection.

    Output:
        None; prints project list to stdout.

    Side Effects:
        Logs project information and prints formatted project list.
    '''
    log.debug('Entering list_projects()')
    
    try:
        projects = jira.projects()
        log.debug(f'Found {len(projects)} projects')
        
        output('')
        output('=' * 80)
        output(f'{"Key":<15} {"Name":<40} {"Lead":<20}')
        output('=' * 80)
        
        for project in sorted(projects, key=lambda p: p.key):
            lead = getattr(project, 'lead', None)
            lead_name = lead.displayName if lead else 'N/A'
            output(f'{project.key:<15} {project.name:<40} {lead_name:<20}')
        
        output('=' * 80)
        output(f'Total projects: {len(projects)}')
        output('')
        
    except Exception as e:
        log.error(f'Failed to list projects: {e}')
        raise


def validate_project(jira, project_key):
    '''
    Validate that a project exists and return the project object.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').

    Output:
        Project object if found.

    Raises:
        JiraProjectError: If project does not exist or cannot be accessed.
    '''
    log.debug(f'Entering validate_project(project_key={project_key})')
    try:
        project = jira.project(project_key)
        log.debug(f'Project validated: {project.name}')
        return project
    except Exception as e:
        raise JiraProjectError(f'Project "{project_key}" not found or not accessible: {e}')


def get_project_workflows(jira, project_key):
    '''
    Get and display the status workflow for a project.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').

    Output:
        None; prints workflow statuses to stdout.

    Side Effects:
        Logs workflow information and prints formatted status list.
    '''
    log.debug(f'Entering get_project_workflows(project_key={project_key})')
    
    project = validate_project(jira, project_key)
    
    try:
        # Get all statuses available in the project
        statuses = jira.statuses()
        log.debug(f'Found {len(statuses)} total statuses')
        
        # Get issue types for the project to find relevant workflows
        issue_types = jira.project(project_key).issueTypes
        
        output('')
        output('=' * 80)
        output(f'Workflow Statuses for Project: {project_key} ({project.name})')
        output('=' * 80)
        
        # Group statuses by category
        status_categories = {}
        for status in statuses:
            category = status.statusCategory.name
            if category not in status_categories:
                status_categories[category] = []
            status_categories[category].append(status)
        
        for category in sorted(status_categories.keys()):
            output(f'\n{category}:')
            output('-' * 40)
            for status in sorted(status_categories[category], key=lambda s: s.name):
                output(f'  {status.name:<30} (ID: {status.id})')
        
        output('')
        output('=' * 80)
        output(f'Total statuses: {len(statuses)}')
        output('')
        
    except JiraProjectError:
        raise
    except Exception as e:
        log.error(f'Failed to get workflows: {e}')
        raise


def get_project_issue_types(jira, project_key):
    '''
    Get and display the issue types (ticket types) for a project.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').

    Output:
        None; prints issue types to stdout.

    Side Effects:
        Logs issue type information and prints formatted list.
    '''
    log.debug(f'Entering get_project_issue_types(project_key={project_key})')
    
    project = validate_project(jira, project_key)
    
    try:
        issue_types = project.issueTypes
        log.debug(f'Found {len(issue_types)} issue types')
        
        output('')
        output('=' * 80)
        output(f'Issue Types for Project: {project_key} ({project.name})')
        output('=' * 80)
        output(f'{"Name":<25} {"ID":<12} {"Subtask":<10} {"Description":<30}')
        output('-' * 80)
        
        for issue_type in sorted(issue_types, key=lambda t: t.name):
            is_subtask = getattr(issue_type, 'subtask', False)
            description = getattr(issue_type, 'description', 'N/A') or 'N/A'
            # Truncate description if too long
            if len(description) > 27:
                description = description[:27] + '...'
            output(f'{issue_type.name:<25} {issue_type.id:<12} {str(is_subtask):<10} {description:<30}')
        
        output('=' * 80)
        output(f'Total issue types: {len(issue_types)}')
        output('')
        
    except JiraProjectError:
        raise
    except Exception as e:
        log.error(f'Failed to get issue types: {e}')
        raise


def get_project_fields(jira, project_key, issue_type_names=None):
    '''
    Get and display create, edit, and transition fields for issue types in a project.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        issue_type_names: List of issue type names to show, or None/empty for all.

    Output:
        None; prints fields per issue type to stdout.

    Side Effects:
        Logs field information and prints formatted field list.
    '''
    log.debug(f'Entering get_project_fields(project_key={project_key}, issue_type_names={issue_type_names})')
    
    project = validate_project(jira, project_key)
    
    try:
        # Get all issue types for the project
        all_issue_types = {it.name: it for it in project.issueTypes}
        log.debug(f'Found {len(all_issue_types)} issue types in project')
        
        # Determine which issue types to process (case-insensitive)
        if issue_type_names:
            # Normalize issue types using case-insensitive matching
            types_to_process = normalize_issue_types(jira, project_key, issue_type_names)
        else:
            types_to_process = sorted(all_issue_types.keys())
        
        log.debug(f'Processing issue types: {types_to_process}')
        
        # Get create metadata for all issue types
        create_meta = jira.createmeta(
            projectKeys=project_key,
            expand='projects.issuetypes.fields'
        )
        
        # Build lookup for create fields by issue type name
        create_fields_by_type = {}
        if create_meta['projects']:
            project_meta = create_meta['projects'][0]
            for it in project_meta['issuetypes']:
                create_fields_by_type[it['name']] = it.get('fields', {})
        
        # For edit and transition fields, we need to find an existing issue of each type
        # or use the editmeta endpoint with a sample issue
        for issue_type_name in types_to_process:
            output('')
            output('=' * 100)
            output(f'Issue Type: {issue_type_name}')
            output(f'Project: {project_key} ({project.name})')
            output('=' * 100)
            
            # === CREATE SCREEN FIELDS ===
            output(f'\n  CREATE SCREEN FIELDS:')
            output(f'  ' + '-' * 90)
            output(f'    {"Field Name":<30} {"Field Key":<25} {"Required":<10} {"Type":<15}')
            output(f'    {"-" * 30} {"-" * 25} {"-" * 10} {"-" * 15}')
            
            create_fields = create_fields_by_type.get(issue_type_name, {})
            if create_fields:
                sorted_fields = sorted(
                    create_fields.items(),
                    key=lambda x: (not x[1].get('required', False), x[1].get('name', x[0]))
                )
                for field_key, field_info in sorted_fields:
                    _print_field_row(field_key, field_info)
            else:
                output(f'    No create fields found.')
            
            # === EDIT SCREEN FIELDS ===
            output(f'\n  EDIT SCREEN FIELDS:')
            output(f'  ' + '-' * 90)
            
            # Find an existing issue of this type to get edit metadata
            sample_issue = _find_sample_issue(jira, project_key, issue_type_name)
            if sample_issue:
                log.debug(f'Found sample issue {sample_issue.key} for edit metadata')
                edit_meta = jira.editmeta(sample_issue.key)
                edit_fields = edit_meta.get('fields', {})
                
                output(f'    {"Field Name":<30} {"Field Key":<25} {"Required":<10} {"Type":<15}')
                output(f'    {"-" * 30} {"-" * 25} {"-" * 10} {"-" * 15}')
                
                sorted_fields = sorted(
                    edit_fields.items(),
                    key=lambda x: (not x[1].get('required', False), x[1].get('name', x[0]))
                )
                for field_key, field_info in sorted_fields:
                    _print_field_row(field_key, field_info)
            else:
                output(f'    No existing {issue_type_name} issues found to retrieve edit fields.')
                output(f'    Create an issue of this type to see edit screen fields.')
            
            # === TRANSITION FIELDS ===
            output(f'\n  TRANSITIONS (Status Changes):')
            output(f'  ' + '-' * 90)
            
            if sample_issue:
                transitions = jira.transitions(sample_issue.key, expand='transitions.fields')
                if transitions:
                    for transition in transitions:
                        trans_name = transition['name']
                        trans_id = transition['id']
                        trans_to = transition.get('to', {}).get('name', 'Unknown')
                        output(f'\n    Transition: "{trans_name}" (ID: {trans_id}) -> {trans_to}')
                        
                        trans_fields = transition.get('fields', {})
                        if trans_fields:
                            output(f'      {"Field Name":<28} {"Field Key":<23} {"Required":<10} {"Type":<15}')
                            output(f'      {"-" * 28} {"-" * 23} {"-" * 10} {"-" * 15}')
                            sorted_fields = sorted(
                                trans_fields.items(),
                                key=lambda x: (not x[1].get('required', False), x[1].get('name', x[0]))
                            )
                            for field_key, field_info in sorted_fields:
                                _print_field_row(field_key, field_info, indent=6)
                        else:
                            output(f'      (No additional fields required)')
                else:
                    output(f'    No transitions available from current status.')
            else:
                output(f'    No existing {issue_type_name} issues found to retrieve transitions.')
            
            output('')
        
        output('=' * 100)
        output('')
        
    except JiraProjectError:
        raise
    except Exception as e:
        log.error(f'Failed to get fields: {e}')
        raise


def normalize_issue_types(jira, project_key, issue_type_names):
    '''
    Normalize issue type names to match Jira's case (case-insensitive matching).

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project.
        issue_type_names: List of issue type names (case-insensitive).

    Output:
        List of issue type names with correct case as defined in Jira.

    Raises:
        JiraProjectError: If an issue type is not found.
    '''
    log.debug(f'Entering normalize_issue_types(project_key={project_key}, issue_type_names={issue_type_names})')
    if not issue_type_names:
        log.debug('No issue type names provided, returning None')
        return None
    
    project = jira.project(project_key)
    # Build case-insensitive lookup
    type_lookup = {it.name.lower(): it.name for it in project.issueTypes}
    log.debug(f'Available issue types: {list(type_lookup.values())}')
    
    normalized = []
    for name in issue_type_names:
        name_lower = name.lower()
        if name_lower in type_lookup:
            normalized.append(type_lookup[name_lower])
            log.debug(f'Normalized issue type "{name}" -> "{type_lookup[name_lower]}"')
        else:
            available = ', '.join(sorted(type_lookup.values()))
            log.error(f'Issue type "{name}" not found')
            raise JiraProjectError(f'Issue type "{name}" not found. Available: {available}')
    
    log.debug(f'Normalized issue types: {normalized}')
    return normalized


def normalize_statuses(jira, status_names):
    '''
    Normalize status names to match Jira's case (case-insensitive matching).
    Supports negation with ^ prefix (e.g., ^Closed to exclude Closed status).

    Input:
        jira: JIRA object with active connection.
        status_names: List of status names (case-insensitive). Use ^ prefix to exclude.

    Output:
        Dict with 'include' and 'exclude' lists of normalized status names.
        For backward compatibility, if no exclusions, returns just the list.

    Raises:
        JiraProjectError: If a status is not found.
    '''
    log.debug(f'Entering normalize_statuses(status_names={status_names})')
    if not status_names:
        log.debug('No status names provided, returning None')
        return None
    
    # Get all statuses from Jira
    statuses = jira.statuses()
    log.debug(f'Retrieved {len(statuses)} statuses from Jira')
    # Build case-insensitive lookup
    status_lookup = {s.name.lower(): s.name for s in statuses}
    
    # Separate includes and excludes (^ prefix for exclusions)
    includes = []
    excludes = []
    
    for name in status_names:
        # Check for negation prefix
        if name.startswith('^'):
            actual_name = name[1:]  # Remove the ^ prefix
            is_exclude = True
        else:
            actual_name = name
            is_exclude = False
        
        name_lower = actual_name.lower()
        if name_lower in status_lookup:
            normalized_name = status_lookup[name_lower]
            if is_exclude:
                excludes.append(normalized_name)
                log.debug(f'Normalized excluded status "{name}" -> "^{normalized_name}"')
            else:
                includes.append(normalized_name)
                log.debug(f'Normalized status "{name}" -> "{normalized_name}"')
        else:
            available = ', '.join(sorted(status_lookup.values()))
            log.error(f'Status "{actual_name}" not found')
            raise JiraProjectError(f'Status "{actual_name}" not found. Available: {available}')
    
    # If there are exclusions, return a dict with both lists
    if excludes:
        result = {'include': includes, 'exclude': excludes}
        log.debug(f'Normalized statuses with exclusions: {result}')
        return result
    
    # For backward compatibility, return just the list if no exclusions
    log.debug(f'Normalized statuses: {includes}')
    return includes


def _build_status_jql(normalized_statuses):
    '''
    Build JQL clause(s) for status filtering, handling both includes and excludes.
    
    Input:
        normalized_statuses: Either a list of status names (backward compatible),
                            or a dict with 'include' and 'exclude' keys.
    
    Output:
        JQL clause string (e.g., 'status IN ("Open", "In Progress")' or
        'status NOT IN ("Closed")'), or empty string if no statuses specified.
    '''
    if not normalized_statuses:
        return ''
    
    clauses = []
    
    # Handle dict format (with exclusions)
    if isinstance(normalized_statuses, dict):
        includes = normalized_statuses.get('include', [])
        excludes = normalized_statuses.get('exclude', [])
        
        if includes:
            status_list = ', '.join([f'"{s}"' for s in includes])
            clauses.append(f'status IN ({status_list})')
        
        if excludes:
            status_list = ', '.join([f'"{s}"' for s in excludes])
            clauses.append(f'status NOT IN ({status_list})')
    else:
        # Handle list format (backward compatible, no exclusions)
        status_list = ', '.join([f'"{s}"' for s in normalized_statuses])
        clauses.append(f'status IN ({status_list})')
    
    return ' AND '.join(clauses)


def normalize_release(jira, project_key, release_name):
    '''
    Normalize a release/version name to match Jira's case (case-insensitive matching).

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project.
        release_name: Release/version name (case-insensitive).

    Output:
        Release name with correct case as defined in Jira.

    Raises:
        JiraProjectError: If the release is not found.
    '''
    log.debug(f'Entering normalize_release(project_key={project_key}, release_name={release_name})')
    if not release_name:
        log.debug('No release name provided, returning None')
        return None
    
    # Get all versions for the project
    versions = jira.project_versions(project_key)
    log.debug(f'Retrieved {len(versions)} versions from project {project_key}')
    # Build case-insensitive lookup
    version_lookup = {v.name.lower(): v.name for v in versions}
    
    name_lower = release_name.lower()
    if name_lower in version_lookup:
        normalized = version_lookup[name_lower]
        log.debug(f'Normalized release "{release_name}" -> "{normalized}"')
        return normalized
    else:
        available = ', '.join(sorted(version_lookup.values()))
        log.error(f'Release "{release_name}" not found')
        raise JiraProjectError(f'Release "{release_name}" not found. Available: {available}')


def parse_date_filter(date_arg):
    '''
    Parse the date filter argument and return a JQL date clause.

    Input:
        date_arg: String date filter. Options:
            - 'today': Issues created today
            - 'week': Issues created in the last 7 days
            - 'month': Issues created in the last 30 days
            - 'year': Issues created in the last 365 days
            - 'all': No date filter
            - 'MM-DD-YYYY:MM-DD-YYYY': Date range (start:end)

    Output:
        JQL date clause string, or empty string for 'all'.

    Raises:
        ValueError: If date format is invalid.
    '''
    log.debug(f'Entering parse_date_filter(date_arg={date_arg})')
    if not date_arg or date_arg.lower() == 'all':
        log.debug('No date filter or "all" specified, returning empty string')
        return ''
    
    date_arg_lower = date_arg.lower()
    
    if date_arg_lower == 'today':
        clause = 'AND created >= startOfDay()'
        log.debug(f'Date filter "today" -> {clause}')
        return clause
    elif date_arg_lower == 'week':
        clause = 'AND created >= -7d'
        log.debug(f'Date filter "week" -> {clause}')
        return clause
    elif date_arg_lower == 'month':
        clause = 'AND created >= -30d'
        log.debug(f'Date filter "month" -> {clause}')
        return clause
    elif date_arg_lower == 'year':
        clause = 'AND created >= -365d'
        log.debug(f'Date filter "year" -> {clause}')
        return clause
    elif ':' in date_arg:
        # Date range format: MM-DD-YYYY:MM-DD-YYYY
        log.debug(f'Parsing date range: {date_arg}')
        try:
            parts = date_arg.split(':')
            if len(parts) != 2:
                raise ValueError(f'Invalid date range format: {date_arg}')
            
            start_str, end_str = parts
            
            # Parse start date
            start_date = datetime.strptime(start_str.strip(), '%m-%d-%Y')
            # Parse end date
            end_date = datetime.strptime(end_str.strip(), '%m-%d-%Y')
            
            # Format for JQL (YYYY-MM-DD)
            start_jql = start_date.strftime('%Y-%m-%d')
            end_jql = end_date.strftime('%Y-%m-%d')
            
            clause = f'AND created >= "{start_jql}" AND created <= "{end_jql}"'
            log.debug(f'Date range parsed: {start_jql} to {end_jql} -> {clause}')
            return clause
        except ValueError as e:
            log.error(f'Invalid date range format: {date_arg}')
            raise ValueError(f'Invalid date range format: {date_arg}. Expected MM-DD-YYYY:MM-DD-YYYY. Error: {e}')
    else:
        log.error(f'Invalid date filter: {date_arg}')
        raise ValueError(f'Invalid date filter: {date_arg}. Use: today, week, month, year, all, or MM-DD-YYYY:MM-DD-YYYY')


def _find_sample_issue(jira, project_key, issue_type_name):
    '''
    Find a sample issue of the given type in the project.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project.
        issue_type_name: Name of the issue type to find.

    Output:
        Issue object if found, None otherwise.
    '''
    log.debug(f'Entering _find_sample_issue(project_key={project_key}, issue_type_name={issue_type_name})')
    try:
        jql = f'project = "{project_key}" AND issuetype = "{issue_type_name}" ORDER BY created DESC'
        log.debug(f'Sample issue JQL: {jql}')
        issues = jira.search_issues(jql, maxResults=1)
        return issues[0] if issues else None
    except Exception as e:
        log.debug(f'Could not find sample issue: {e}')
        return None


def _print_field_row(field_key, field_info, indent=4):
    '''
    Print a formatted field row.

    Input:
        field_key: The field key/ID.
        field_info: Dict containing field metadata.
        indent: Number of spaces to indent.
    '''
    log.debug(f'Entering _print_field_row(field_key={field_key}, indent={indent})')
    field_name = field_info.get('name', field_key)
    required = 'Yes' if field_info.get('required', False) else 'No'
    schema = field_info.get('schema', {})
    field_type = schema.get('type', 'unknown')
    
    # Truncate if too long
    if len(field_name) > 28:
        field_name = field_name[:28] + '..'
    if len(field_key) > 23:
        field_key_display = field_key[:23] + '..'
    else:
        field_key_display = field_key
    
    prefix = ' ' * indent
    output(f'{prefix}{field_name:<30} {field_key_display:<25} {required:<10} {field_type:<15}')


def get_project_versions(jira, project_key):
    '''
    Get and display the versions (releases) for a project.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').

    Output:
        None; prints versions/releases to stdout.

    Side Effects:
        Logs version information and prints formatted version list.
    '''
    log.debug(f'Entering get_project_versions(project_key={project_key})')
    
    project = validate_project(jira, project_key)
    
    try:
        versions = jira.project_versions(project_key)
        log.debug(f'Found {len(versions)} versions')
        
        output('')
        output('=' * 100)
        output(f'Versions (Releases) for Project: {project_key} ({project.name})')
        output('=' * 100)
        output(f'{"Name":<30} {"ID":<12} {"Released":<10} {"Archived":<10} {"Release Date":<15} {"Description":<20}')
        output('-' * 100)
        
        if not versions:
            output('  No versions defined for this project.')
        else:
            # Sort by release date (unreleased last), then by name
            sorted_versions = sorted(
                versions,
                key=lambda v: (
                    not getattr(v, 'released', False),
                    getattr(v, 'releaseDate', '9999-99-99') or '9999-99-99',
                    v.name
                )
            )
            
            for version in sorted_versions:
                name = version.name
                vid = version.id
                released = 'Yes' if getattr(version, 'released', False) else 'No'
                archived = 'Yes' if getattr(version, 'archived', False) else 'No'
                release_date = getattr(version, 'releaseDate', 'N/A') or 'N/A'
                description = getattr(version, 'description', '') or ''
                
                # Truncate if too long
                if len(name) > 28:
                    name = name[:28] + '..'
                if len(description) > 18:
                    description = description[:18] + '..'
                
                output(f'{name:<30} {vid:<12} {released:<10} {archived:<10} {release_date:<15} {description:<20}')
        
        output('=' * 100)
        output(f'Total versions: {len(versions)}')
        output('')
        
    except JiraProjectError:
        raise
    except Exception as e:
        log.error(f'Failed to get versions: {e}')
        raise


def get_project_components(jira, project_key, date_filter=None, dump_file=None, dump_format='csv'):
    '''
    Get and display the components for a project.
    If date_filter is specified, only shows components with tickets created in that date range.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        date_filter: Date filter string to show only active components, or None for all.
        dump_file: Output filename for dumping components, or None to skip.
        dump_format: Output format ('csv' or 'json').

    Output:
        None; prints components to stdout.

    Side Effects:
        Logs component information and prints formatted component list.
    '''
    log.debug(f'Entering get_project_components(project_key={project_key}, date_filter={date_filter}, dump_file={dump_file}, dump_format={dump_format})')
    
    project = validate_project(jira, project_key)
    
    try:
        components = jira.project_components(project_key)
        log.debug(f'Found {len(components)} components')
        
        # If date filter specified, find components with recent tickets
        if date_filter:
            active_components = _get_active_components(jira, project_key, date_filter)
            filtered_components = [c for c in components if c.name in active_components]
            log.debug(f'Filtered to {len(filtered_components)} components with activity in date range')
        else:
            filtered_components = components
        
        output('')
        output('=' * 110)
        output(f'Components for Project: {project_key} ({project.name})')
        if date_filter:
            output(f'Showing only components with tickets created: {date_filter}')
        output('=' * 110)
        output(f'{"Name":<30} {"ID":<12} {"Lead":<25} {"Ticket Count":<15} {"Description":<25}')
        output('-' * 110)
        
        if not filtered_components:
            if date_filter:
                output(f'  No components with tickets in the specified date range.')
            else:
                output('  No components defined for this project.')
        else:
            # Sort by name
            sorted_components = sorted(filtered_components, key=lambda c: c.name.lower())
            
            # Get ticket counts if date filter is active
            ticket_counts = {}
            if date_filter:
                ticket_counts = _get_active_components(jira, project_key, date_filter)
            
            for component in sorted_components:
                name = component.name
                cid = component.id
                lead = getattr(component, 'lead', None)
                lead_name = lead.displayName if lead else 'Unassigned'
                description = getattr(component, 'description', '') or ''
                count = ticket_counts.get(name, '-') if date_filter else '-'
                
                # Truncate for display
                display_name = name[:28] + '..' if len(name) > 28 else name
                display_lead = lead_name[:23] + '..' if len(lead_name) > 23 else lead_name
                display_desc = description[:23] + '..' if len(description) > 23 else description
                
                output(f'{display_name:<30} {cid:<12} {display_lead:<25} {str(count):<15} {display_desc:<25}')
        
        output('=' * 110)
        output(f'Total components: {len(filtered_components)}')
        output('')
        
        # Dump to file if requested
        if dump_file and filtered_components:
            sorted_components = sorted(filtered_components, key=lambda c: c.name.lower())
            ticket_counts = _get_active_components(jira, project_key, date_filter) if date_filter else {}
            _dump_components_to_file(sorted_components, dump_file, dump_format, ticket_counts)
        
    except JiraProjectError:
        raise
    except Exception as e:
        log.error(f'Failed to get components: {e}')
        raise


def _get_active_components(jira, project_key, date_filter):
    '''
    Get components that have tickets created in the specified date range.
    
    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project.
        date_filter: Date filter string.
    
    Output:
        Dict mapping component name to ticket count.
    '''
    log.debug(f'Getting active components for {project_key} with date filter {date_filter}')
    
    # Parse date filter
    date_clause = parse_date_filter(date_filter) if date_filter else ''
    
    # Build JQL to get all tickets with components in date range
    jql = f'project = "{project_key}" AND component is not EMPTY'
    if date_clause:
        jql = f'{jql} {date_clause}'
    
    log.debug(f'Active components JQL: {jql}')
    
    # Fetch tickets
    email, api_token = get_jira_credentials()
    
    component_counts = {}
    next_page_token = None
    batch_size = 100
    max_retries = 5
    
    while True:
        payload = {
            'jql': jql,
            'maxResults': batch_size,
            'fields': ['components']
        }
        if next_page_token:
            payload['nextPageToken'] = next_page_token
        
        for retry in range(max_retries):
            response = requests.post(
                f'{JIRA_URL}/rest/api/3/search/jql',
                auth=(email, api_token),
                headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
                json=payload
            )
            
            if response.status_code == 429:
                retry_after = int(response.headers.get('Retry-After', 5))
                log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                time.sleep(retry_after)
                continue
            break
        
        if response.status_code != 200:
            log.error(f'API request failed: {response.status_code} - {response.text}')
            raise Exception(f'Jira API error: {response.status_code} - {response.text}')
        
        data = response.json()
        issues = data.get('issues', [])
        
        for issue in issues:
            components = issue.get('fields', {}).get('components', [])
            for comp in components:
                comp_name = comp.get('name', '')
                if comp_name:
                    component_counts[comp_name] = component_counts.get(comp_name, 0) + 1
        
        next_page_token = data.get('nextPageToken')
        if not next_page_token:
            break
    
    log.debug(f'Found {len(component_counts)} active components')
    return component_counts


def _dump_components_to_file(components, dump_file, dump_format='csv', ticket_counts=None):
    '''
    Dump components to a file in CSV or JSON format.
    
    Input:
        components: List of component objects.
        dump_file: Output filename (extension added if not present).
        dump_format: Output format ('csv' or 'json').
        ticket_counts: Optional dict mapping component name to ticket count.
    '''
    # Add extension if not present
    if not dump_file.endswith(f'.{dump_format}'):
        output_path = f'{dump_file}.{dump_format}'
    else:
        output_path = dump_file
    
    log.debug(f'Writing {len(components)} components to {output_path}')
    
    rows = []
    for component in components:
        lead = getattr(component, 'lead', None)
        row = {
            'name': component.name,
            'id': component.id,
            'lead': lead.displayName if lead else '',
            'lead_email': lead.emailAddress if lead and hasattr(lead, 'emailAddress') else '',
            'description': getattr(component, 'description', '') or ''
        }
        if ticket_counts:
            row['ticket_count'] = ticket_counts.get(component.name, 0)
        rows.append(row)
    
    fieldnames = ['name', 'id', 'lead', 'lead_email', 'description']
    if ticket_counts:
        fieldnames.append('ticket_count')
    
    # Determine fieldnames (include any extra keys present in rows)
    base_fields = ['key', 'project', 'issue_type', 'status', 'priority', 'summary', 'assignee', 'reporter', 'created', 'updated', 'resolved', 'fix_version', 'affects_version']
    # Collect all keys that appear in any row
    all_keys = set(base_fields)
    for r in rows:
        all_keys.update(r.keys())
    # Preserve base field order, then append any extras deterministically
    extra_columns = [k for k in all_keys if k not in base_fields]
    extra_columns.sort()
    # ensure extra columns are present in all rows to satisfy DictWriter
    for r in rows:
        for col in extra_columns:
            r.setdefault(col, '')
    # Guarantee ordering and that DictWriter sees all columns: include extras in fieldnames and rows
    fieldnames = base_fields + extra_columns

    if dump_format == 'json':
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, indent=2, default=str)
    else:  # csv
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    
    log.info(f'Wrote {len(rows)} components to: {output_path}')


def match_pattern_with_exclusions(name, pattern):
    '''
    Match a name against a pattern that may include exclusions.
    
    Pattern syntax:
        - Simple glob: "12.*" matches names starting with "12."
        - With exclusions: "12.*,^*Samples*" matches 12.* but excludes names containing "Samples"
        - Multiple exclusions: "12.*,^*Samples*,^*Test*"
        - Multiple includes: "12.*,13.*" matches either pattern
        - Combined: "12.*,13.*,^*Samples*" matches 12.* or 13.* but excludes Samples
    
    Input:
        name: String name to match.
        pattern: Pattern string with optional comma-separated parts and ^ exclusions.
    
    Output:
        True if name matches the pattern (and doesn't match any exclusions), False otherwise.
    '''
    import fnmatch
    
    if not pattern or pattern == '*':
        return True
    
    # Split pattern by comma
    parts = [p.strip() for p in pattern.split(',')]
    
    # Separate includes and excludes (^ prefix for exclusions)
    includes = [p for p in parts if not p.startswith('^')]
    excludes = [p[1:] for p in parts if p.startswith('^')]  # Remove the ^ prefix
    
    # If no includes specified, default to match all
    if not includes:
        includes = ['*']
    
    # Check if name matches any include pattern
    matched = any(fnmatch.fnmatch(name, inc) for inc in includes)
    
    if not matched:
        return False
    
    # Check if name matches any exclude pattern
    excluded = any(fnmatch.fnmatch(name, exc) for exc in excludes)
    
    return not excluded


def get_children_hierarchy(jira, project_key=None, root_key=None, limit=None, dump_file=None, dump_format='csv', table_format='flat'):
    '''
    Recursively retrieve and display the full child hierarchy for a given ticket.

    Traversal strategy:
      - Uses JQL `parent = "<key>"` to find children (Advanced Roadmaps/parent links and subtasks).
      - Recurses depth-first until no more children or the optional --limit is reached.
      - Includes the root ticket in the output.

    Input:
        jira: JIRA object with active connection.
        project_key: Optional project key (validated if provided).
        root_key: Issue key to start from (required).
        limit: Optional maximum number of tickets to return (including root).
        dump_file: Optional output filename (extension added automatically).
        dump_format: Output format ('csv' or 'json').
        table_format: Table layout for CSV ('flat' or 'indented').

    Output:
        None; prints a hierarchy view and a table. Optionally writes to file.
    '''
    log.debug(f'Entering get_children_hierarchy(project_key={project_key}, root_key={root_key}, limit={limit}, dump_file={dump_file}, dump_format={dump_format}, table_format={table_format})')

    # Validate project if provided (root may be outside the project)
    if project_key:
        validate_project(jira, project_key)
    if not root_key:
        raise ValueError('root_key is required for get_children_hierarchy')

    try:
        # Helper: fetch a single issue and return its raw dict (matches search API structure)
        def _fetch_issue_raw(issue_key):
            log.debug(f'Fetching root/child issue via jira.issue: {issue_key}')
            issue_obj = jira.issue(issue_key)
            return issue_obj.raw

        # Helper: search for direct children of a parent key using the /rest/api/3/search/jql endpoint
        def _fetch_children(parent_key, remaining_limit=None):
            email, api_token = get_jira_credentials()
            all_children = []
            next_page_token = None
            batch_size = 100
            max_retries = 5

            # Fields needed for display/dump; keep aligned with print_ticket_row/dump_tickets_to_file
            fields_to_fetch = ['summary', 'status', 'issuetype', 'created', 'updated', 'assignee', 'priority', 'project', 'fixVersions', 'versions', 'parent']

            while True:
                if remaining_limit is not None and remaining_limit <= 0:
                    break

                current_batch = batch_size
                if remaining_limit is not None:
                    current_batch = min(batch_size, remaining_limit)

                payload = {
                    'jql': f'parent = "{parent_key}"',
                    'maxResults': current_batch,
                    'fields': fields_to_fetch
                }
                if next_page_token:
                    payload['nextPageToken'] = next_page_token

                # Record the top-level JQL (first page) for --show-jql visibility
                if not next_page_token:
                    show_jql(payload['jql'])

                for retry in range(max_retries):
                    response = requests.post(
                        f'{JIRA_URL}/rest/api/3/search/jql',
                        auth=(email, api_token),
                        headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
                        json=payload
                    )

                    if response.status_code == 429:
                        retry_after = int(response.headers.get('Retry-After', 5))
                        log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                        time.sleep(retry_after)
                        continue
                    break

                if response.status_code != 200:
                    log.error(f'Child search failed for {parent_key}: {response.status_code} - {response.text}')
                    raise Exception(f'Jira API error: {response.status_code} - {response.text}')

                data = response.json()
                issues = data.get('issues', [])
                all_children.extend(issues)

                next_page_token = data.get('nextPageToken')

                if remaining_limit is not None:
                    remaining_limit -= len(issues)
                    if remaining_limit <= 0:
                        break

                if not next_page_token:
                    break

            return all_children

        visited = set()
        ordered = []  # list of {'issue': issue_dict, 'depth': depth}

        def _recurse(current_key, depth, remaining_limit):
            if remaining_limit is not None and len(ordered) >= remaining_limit:
                return

            children = _fetch_children(current_key, None if remaining_limit is None else (remaining_limit - len(ordered)))

            for child in children:
                child_key = child.get('key')
                if not child_key or child_key in visited:
                    continue

                visited.add(child_key)
                ordered.append({'issue': child, 'depth': depth + 1})

                # Recurse further if limit not reached
                _recurse(child_key, depth + 1, remaining_limit)

        # Seed with root issue
        root_issue = _fetch_issue_raw(root_key)
        visited.add(root_key)
        ordered.append({'issue': root_issue, 'depth': 0})

        _recurse(root_key, 0, limit)

        # Apply limit after traversal (in case root counts toward limit)
        if limit is not None and len(ordered) > limit:
            ordered = ordered[:limit]

        # Output hierarchy view
        output('')
        output('=' * 120)
        output(f'Child Hierarchy for: {root_key}')
        if project_key:
            output(f'Project: {project_key}')
        if limit:
            output(f'Limit: {limit} (includes root)')
        output('=' * 120)

        output('Hierarchy (depth-indented):')
        for item in ordered:
            issue = item['issue']
            fields = issue.get('fields', {})
            indent = '  ' * item['depth']
            summary = fields.get('summary', 'N/A') or 'N/A'
            issue_type = fields.get('issuetype', {}).get('name', 'N/A') if fields.get('issuetype') else 'N/A'
            status = fields.get('status', {}).get('name', 'N/A') if fields.get('status') else 'N/A'
            output(f'{indent}- {issue.get("key", "N/A")} [{issue_type} | {status}] {summary}')

        output('')
        output('Table:')
        print_ticket_table_header()

        for item in ordered:
            issue = item['issue']
            fields = issue.get('fields', {})
            # Indent summary for table display to visualize hierarchy depth
            summary = fields.get('summary', 'N/A') or 'N/A'
            fields['summary'] = f'{"  " * item["depth"]}{summary}'
            print_ticket_row(issue)

        print_ticket_table_footer(len(ordered))

        if dump_file:
            # Build extra_fields with depth metadata for each ticket
            extras = {
                item['issue'].get('key', ''): {
                    'depth': item.get('depth'),
                }
                for item in ordered
            }
            dump_tickets_to_file([item['issue'] for item in ordered], dump_file, dump_format, extras, table_format=table_format)

    except Exception as e:
        log.error(f'Failed to get children hierarchy: {e}')
        raise


def get_related_issues(jira, project_key=None, root_key=None, hierarchy=None, limit=None, dump_file=None, dump_format='csv', table_format='flat'):
   '''
   Retrieve related issues for a given ticket.

   "Related" includes:
     - Linked issues (issuelinks)
     - Children discovered via JQL `parent = "<key>"`

   Traversal strategy:
     - Direct links + direct children when --hierarchy is not provided.
     - Depth-first traversal across both links and children when --hierarchy is provided.
     - Guards against cycles by tracking visited issue keys.
     - Includes the root ticket in the output.

   Input:
       jira: JIRA object with active connection.
       project_key: Optional project key (validated if provided).
       root_key: Issue key to start from (required).
       hierarchy: None for direct-only, -1 for unlimited depth, or positive int for depth limit.
       limit: Optional maximum number of tickets to return (including root).
       dump_file: Optional output filename (extension added automatically).
       dump_format: Output format ('csv' or 'json').

   Output:
       None; prints a linked view, a table, and optionally writes to file.
       Use drawio_utilities.py --create-map on the CSV output to generate a draw.io diagram.
   '''
   log.debug(f'Entering get_related_issues(project_key={project_key}, root_key={root_key}, hierarchy={hierarchy}, limit={limit}, dump_file={dump_file}, dump_format={dump_format})')

   if project_key:
       validate_project(jira, project_key)
   if not root_key:
       raise ValueError('root_key is required for get_related_issues')

   log.info(f'Getting related tickets to {root_key}. May take some time...')

   try:
       fields_to_fetch = ['summary', 'status', 'issuetype', 'created', 'updated', 'assignee', 'priority', 'project', 'fixVersions', 'versions', 'issuelinks']

       def _fetch_issue_raw(issue_key):
           log.debug(f'Fetching issue for related traversal: {issue_key}')
           issue_obj = jira.issue(issue_key, fields=','.join(fields_to_fetch))
           return issue_obj.raw

       def _collect_links(issue_raw):
           links = issue_raw.get('fields', {}).get('issuelinks', []) or []
           edges = []  # list of (target_key, via_label, relation)
           for link in links:
               link_type = link.get('type', {}) or {}
               inward_label = link_type.get('inward') or link_type.get('name') or 'linked'
               outward_label = link_type.get('outward') or link_type.get('name') or 'linked'

               if 'outwardIssue' in link:
                   tgt = link.get('outwardIssue') or {}
                   tgt_key = tgt.get('key')
                   if tgt_key:
                       edges.append((tgt_key, outward_label, 'link'))

               if 'inwardIssue' in link:
                   tgt = link.get('inwardIssue') or {}
                   tgt_key = tgt.get('key')
                   if tgt_key:
                       edges.append((tgt_key, inward_label, 'link'))
           return edges

       def _fetch_children_keys(parent_key, remaining_limit=None):
           '''
           Fetch direct children for a parent issue using JQL `parent = "<key>"`.

           Note:
               This mirrors the approach used in get_children_hierarchy(), but returns only keys
               so we can unify traversal across links + children without duplicating issue dicts.
           '''
           email, api_token = get_jira_credentials()
           all_child_keys = []
           next_page_token = None
           batch_size = 100
           max_retries = 5

           while True:
               if remaining_limit is not None and remaining_limit <= 0:
                   break

               current_batch = batch_size
               if remaining_limit is not None:
                   current_batch = min(batch_size, remaining_limit)

               payload = {
                   'jql': f'parent = "{parent_key}"',
                   'maxResults': current_batch,
                   # We only need keys; keep fields minimal to reduce payload size.
                   'fields': ['summary']
               }
               if next_page_token:
                   payload['nextPageToken'] = next_page_token

               # Record the top-level JQL (first page) for --show-jql visibility
               if not next_page_token:
                   show_jql(payload['jql'])

               for retry in range(max_retries):
                   response = requests.post(
                       f'{JIRA_URL}/rest/api/3/search/jql',
                       auth=(email, api_token),
                       headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
                       json=payload
                   )

                   if response.status_code == 429:
                       retry_after = int(response.headers.get('Retry-After', 5))
                       log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                       time.sleep(retry_after)
                       continue
                   break

               if response.status_code != 200:
                   log.error(f'Child search failed for {parent_key}: {response.status_code} - {response.text}')
                   raise Exception(f'Jira API error: {response.status_code} - {response.text}')

               data = response.json()
               issues = data.get('issues', [])

               for issue in issues:
                   child_key = issue.get('key')
                   if child_key:
                       all_child_keys.append(child_key)

               next_page_token = data.get('nextPageToken')

               if remaining_limit is not None:
                   remaining_limit -= len(issues)
                   if remaining_limit <= 0:
                       break

               if not next_page_token:
                   break

           return all_child_keys

       def _collect_children(parent_key, remaining_limit=None):
           # Children are modeled as edges from parent -> child
           child_keys = _fetch_children_keys(parent_key, remaining_limit)
           return [(k, 'child', 'child') for k in child_keys if k]

       visited = set()
       ordered = []  # list of {'issue': issue_dict, 'depth': depth, 'via': label, 'relation': str, 'from_key': str}

       def _traverse(issue_key, issue_raw, depth, remaining_limit, remaining_depth):
           if remaining_limit is not None and len(ordered) >= remaining_limit:
               return
           if remaining_depth is not None and remaining_depth <= 0:
               return

           # 1) Traverse linked issues (existing behavior)
           edges = _collect_links(issue_raw)
           for tgt_key, via, relation in edges:
               if remaining_limit is not None and len(ordered) >= remaining_limit:
                   break
               if tgt_key in visited:
                   continue

               tgt_raw = _fetch_issue_raw(tgt_key)
               visited.add(tgt_key)
               ordered.append({'issue': tgt_raw, 'depth': depth + 1, 'via': via, 'relation': relation, 'from_key': issue_key})

               if hierarchy is not None:
                   _traverse(tgt_key, tgt_raw, depth + 1, remaining_limit, None if remaining_depth is None else remaining_depth - 1)

           # 2) Also traverse child tickets (parent -> child) according to the same hierarchy depth
           if remaining_limit is None or len(ordered) < remaining_limit:
               remaining_slots = None if remaining_limit is None else (remaining_limit - len(ordered))
               child_edges = _collect_children(issue_key, remaining_slots)
               for tgt_key, via, relation in child_edges:
                   if remaining_limit is not None and len(ordered) >= remaining_limit:
                       break
                   if tgt_key in visited:
                       continue

                   tgt_raw = _fetch_issue_raw(tgt_key)
                   visited.add(tgt_key)
                   ordered.append({'issue': tgt_raw, 'depth': depth + 1, 'via': via, 'relation': relation, 'from_key': issue_key})

                   if hierarchy is not None:
                       _traverse(tgt_key, tgt_raw, depth + 1, remaining_limit, None if remaining_depth is None else remaining_depth - 1)

       # Seed with root issue
       root_issue = _fetch_issue_raw(root_key)
       visited.add(root_key)
       ordered.append({'issue': root_issue, 'depth': 0, 'via': None, 'relation': None, 'from_key': None})

       if hierarchy is not None:
           # hierarchy: None -> direct, -1 -> unlimited, n>0 -> depth-limited
           depth_limit = None if hierarchy == -1 else hierarchy
           _traverse(root_key, root_issue, 0, limit, depth_limit)
       else:
           # Only direct links + direct children
           edges = _collect_links(root_issue)
           for tgt_key, via, relation in edges:
               if limit is not None and len(ordered) >= limit:
                   break
               if tgt_key in visited:
                   continue
               tgt_raw = _fetch_issue_raw(tgt_key)
               visited.add(tgt_key)
               ordered.append({'issue': tgt_raw, 'depth': 1, 'via': via, 'relation': relation, 'from_key': root_key})

           if limit is None or len(ordered) < limit:
               remaining_slots = None if limit is None else (limit - len(ordered))
               child_edges = _collect_children(root_key, remaining_slots)
               for tgt_key, via, relation in child_edges:
                   if limit is not None and len(ordered) >= limit:
                       break
                   if tgt_key in visited:
                       continue
                   tgt_raw = _fetch_issue_raw(tgt_key)
                   visited.add(tgt_key)
                   ordered.append({'issue': tgt_raw, 'depth': 1, 'via': via, 'relation': relation, 'from_key': root_key})

       # Apply limit after traversal (in case root counts toward limit)
       if limit is not None and len(ordered) > limit:
           ordered = ordered[:limit]

       # Output linked view
       output('')
       output('=' * 120)
       output(f'Related Issues for: {root_key}')
       if project_key:
           output(f'Project: {project_key}')
       if hierarchy is None:
           output('Traversal: direct links + direct children')
       elif hierarchy == -1:
           output('Traversal: recursive across links + children (unlimited depth)')
       else:
           output(f'Traversal: recursive across links + children (depth <= {hierarchy})')
       if limit:
           output(f'Limit: {limit} (includes root)')
       output('=' * 120)

       output('Related issues (depth-indented):')
       for item in ordered:
           issue = item['issue']
           fields = issue.get('fields', {})
           indent = '  ' * item['depth']
           summary = fields.get('summary', 'N/A') or 'N/A'
           issue_type = fields.get('issuetype', {}).get('name', 'N/A') if fields.get('issuetype') else 'N/A'
           status = fields.get('status', {}).get('name', 'N/A') if fields.get('status') else 'N/A'
           via = item.get('via')
           via_label = f' via {via}' if via else ''
           output(f'{indent}- {issue.get("key", "N/A")} [{issue_type} | {status}]{via_label} {summary}')

       output('')
       output('Table:')
       print_ticket_table_header()

       for item in ordered:
           issue = item['issue']
           fields = issue.get('fields', {})
           summary = fields.get('summary', 'N/A') or 'N/A'
           via = item.get('via')
           if via:
               summary = f'{summary} (via {via})'
           fields['summary'] = f'{"  " * item["depth"]}{summary}'
           print_ticket_row(issue)

       print_ticket_table_footer(len(ordered))

       if dump_file:
           extras = {
               item['issue'].get('key', ''): {
                   'depth': item.get('depth'),
                   'via': item.get('via'),
                   'relation': item.get('relation'),
                   'from_key': item.get('from_key'),
               }
               for item in ordered
           }
           dump_tickets_to_file([item['issue'] for item in ordered], dump_file, dump_format, extras, table_format=table_format)

   except Exception as e:
       log.error(f'Failed to get related issues: {e}')
       raise


def get_releases(jira, project_key, pattern=None, dump_file=None, dump_format='csv'):
    '''
    Get and display the releases (versions) for a project.
    This is an alias for get_project_versions with simpler output.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        pattern: Optional glob pattern to filter releases (e.g., '12.*', '12.*,^*Samples*').
        dump_file: Output filename for dumping releases, or None to skip.
        dump_format: Output format ('csv' or 'json').

    Output:
        None; prints releases to stdout.
    '''
    log.debug(f'Entering get_releases(project_key={project_key}, pattern={pattern}, dump_file={dump_file}, dump_format={dump_format})')
    
    project = validate_project(jira, project_key)
    
    try:
        versions = jira.project_versions(project_key)
        log.debug(f'Found {len(versions)} releases')
        
        output('')
        output('=' * 80)
        output(f'Releases for Project: {project_key} ({project.name})')
        output('=' * 80)
        output(f'{"Release Name":<40} {"Released":<12} {"Release Date":<15}')
        output('-' * 80)
        
        # Filter by pattern if specified (supports exclusions with ^ prefix)
        if pattern and pattern != '*':
            filtered_versions = [v for v in versions if match_pattern_with_exclusions(v.name, pattern)]
            log.debug(f'Filtered to {len(filtered_versions)} releases matching pattern "{pattern}"')
        else:
            filtered_versions = versions
        
        if not filtered_versions:
            if pattern and pattern != '*':
                output(f'  No releases matching pattern "{pattern}".')
            else:
                output('  No releases defined for this project.')
        else:
            # Sort by release date (unreleased last), then by name
            sorted_versions = sorted(
                filtered_versions,
                key=lambda v: (
                    not getattr(v, 'released', False),
                    getattr(v, 'releaseDate', '9999-99-99') or '9999-99-99',
                    v.name
                )
            )
            
            for version in sorted_versions:
                name = version.name
                released = 'Yes' if getattr(version, 'released', False) else 'No'
                release_date = getattr(version, 'releaseDate', 'N/A') or 'N/A'
                
                # Truncate for display
                display_name = name[:38] + '..' if len(name) > 38 else name
                
                output(f'{display_name:<40} {released:<12} {release_date:<15}')
        
        output('=' * 80)
        if pattern and pattern != '*':
            output(f'Total releases matching "{pattern}": {len(filtered_versions)}')
        else:
            output(f'Total releases: {len(filtered_versions)}')
        output('')
        
        # Dump to file if requested
        if dump_file and filtered_versions:
            _dump_releases_to_file(sorted_versions, dump_file, dump_format)
        
    except JiraProjectError:
        raise
    except Exception as e:
        log.error(f'Failed to get releases: {e}')


def _dump_releases_to_file(versions, dump_file, dump_format='csv'):
    '''
    Dump releases to a file in CSV or JSON format.
    
    Input:
        versions: List of version objects.
        dump_file: Output filename (extension added if not present).
        dump_format: Output format ('csv' or 'json').
    '''
    # Add extension if not present
    if not dump_file.endswith(f'.{dump_format}'):
        output_path = f'{dump_file}.{dump_format}'
    else:
        output_path = dump_file
    
    log.debug(f'Writing {len(versions)} releases to {output_path}')
    
    rows = []
    for version in versions:
        rows.append({
            'name': version.name,
            'id': version.id,
            'released': 'Yes' if getattr(version, 'released', False) else 'No',
            'archived': 'Yes' if getattr(version, 'archived', False) else 'No',
            'release_date': getattr(version, 'releaseDate', '') or '',
            'description': getattr(version, 'description', '') or ''
        })
    
    if dump_format == 'json':
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, indent=2, default=str)
    else:  # csv
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['name', 'id', 'released', 'archived', 'release_date', 'description'])
            writer.writeheader()
            writer.writerows(rows)
    
    log.info(f'Wrote {len(rows)} releases to: {output_path}')


def get_release_tickets(jira, project_key, release_name, issue_types=None, statuses=None, date_filter=None, limit=None, dump_file=None, dump_format='csv'):
    '''
    Get and display tickets associated with a specific release or releases matching a pattern.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        release_name: Name of the release/version (case-insensitive), or a glob pattern (e.g., '12.1*').
        issue_types: List of issue type names to filter by, or None/empty for all.
        statuses: List of status names to filter by, or None/empty for all.
        date_filter: Date filter string.
        limit: Maximum number of tickets to retrieve, or None for all.
        dump_file: Output filename for dumping tickets, or None to skip.
        dump_format: Output format ('csv' or 'json').

    Output:
        None; prints ticket list to stdout.
    '''
    log.debug(f'Entering get_release_tickets(project_key={project_key}, release_name={release_name}, issue_types={issue_types}, statuses={statuses}, date_filter={date_filter}, limit={limit}, dump_file={dump_file}, dump_format={dump_format})')
    
    project = validate_project(jira, project_key)
    
    try:
        # Check if release_name contains wildcard characters - if so, treat as pattern
        # and delegate to get_releases_tickets for multi-release handling
        if '*' in release_name or '?' in release_name:
            log.debug(f'Release name "{release_name}" contains wildcards, treating as pattern')
            # Delegate to get_releases_tickets which handles patterns
            return get_releases_tickets(jira, project_key, release_name, issue_types, statuses,
                                        date_filter, limit, dump_file, dump_format)
        
        # Normalize release name (case-insensitive exact match)
        normalized_release = normalize_release(jira, project_key, release_name)
        log.debug(f'Normalized release name: {normalized_release}')
        
        # Normalize issue types and statuses
        normalized_types = normalize_issue_types(jira, project_key, issue_types) if issue_types else None
        normalized_statuses = normalize_statuses(jira, statuses) if statuses else None
        
        # Parse date filter
        date_clause = parse_date_filter(date_filter) if date_filter else ''
        
        # Build JQL query
        jql_parts = [f'project = "{project_key}"', f'fixVersion = "{normalized_release}"']
        
        if normalized_types:
            type_list = ', '.join([f'"{t}"' for t in normalized_types])
            jql_parts.append(f'issuetype IN ({type_list})')
        
        # Build status clause using helper (handles both includes and excludes)
        status_clause = _build_status_jql(normalized_statuses)
        if status_clause:
            jql_parts.append(status_clause)
        
        jql = ' AND '.join(jql_parts)
        if date_clause:
            jql = f'{jql} {date_clause}'
        jql = f'{jql} ORDER BY created DESC'
        
        log.debug(f'JQL query: {jql}')
        show_jql(jql)
        
        # Fetch tickets using the search API
        email, api_token = get_jira_credentials()
        
        all_issues = []
        next_page_token = None
        batch_size = min(100, limit) if limit else 100
        max_retries = 5
        
        while True:
            if limit and len(all_issues) >= limit:
                break
            
            if limit:
                remaining = limit - len(all_issues)
                current_batch = min(batch_size, remaining)
            else:
                current_batch = batch_size
            
            fields_to_fetch = ['summary', 'status', 'issuetype', 'created', 'updated', 'assignee', 'priority', 'project', 'fixVersions', 'versions']
            if dump_file:
                fields_to_fetch.extend(['reporter', 'resolutiondate'])
            
            payload = {
                'jql': jql,
                'maxResults': current_batch,
                'fields': fields_to_fetch
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token
            
            for retry in range(max_retries):
                response = requests.post(
                    f'{JIRA_URL}/rest/api/3/search/jql',
                    auth=(email, api_token),
                    headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
                    json=payload
                )
                
                if response.status_code == 429:
                    retry_after = int(response.headers.get('Retry-After', 5))
                    log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                    time.sleep(retry_after)
                    continue
                break
            
            if response.status_code != 200:
                log.error(f'API request failed: {response.status_code} - {response.text}')
                raise Exception(f'Jira API error: {response.status_code} - {response.text}')
            
            data = response.json()
            issues = data.get('issues', [])
            all_issues.extend(issues)
            
            next_page_token = data.get('nextPageToken')
            log.debug(f'Retrieved {len(all_issues)} issues so far...')
            
            if not next_page_token:
                break
            if limit and len(all_issues) >= limit:
                break
        
        if limit and len(all_issues) > limit:
            all_issues = all_issues[:limit]
        
        log.debug(f'Retrieved {len(all_issues)} issues total')
        
        # Print results
        output('')
        output('=' * 120)
        output(f'Tickets for Release: {normalized_release}')
        output(f'Project: {project_key} ({project.name})')
        output('=' * 120)
        
        if normalized_types:
            output(f'Issue types: {", ".join(normalized_types)}')
        if normalized_statuses:
            # Format status display for both includes and excludes
            if isinstance(normalized_statuses, dict):
                parts = []
                if normalized_statuses.get('include'):
                    parts.append(', '.join(normalized_statuses['include']))
                if normalized_statuses.get('exclude'):
                    parts.append('NOT: ' + ', '.join(normalized_statuses['exclude']))
                output(f'Statuses: {"; ".join(parts)}')
            else:
                output(f'Statuses: {", ".join(normalized_statuses)}')
        if date_filter and date_filter.lower() != 'all':
            output(f'Date filter: {date_filter}')
        if limit:
            output(f'Limit: {limit}')
        
        print_ticket_table_header()
        
        for issue in all_issues:
            print_ticket_row(issue)
        
        print_ticket_table_footer(len(all_issues))
        
        if dump_file:
            dump_tickets_to_file(all_issues, dump_file, dump_format)
        
    except JiraProjectError:
        raise
    except ValueError as e:
        log.error(f'Invalid parameter: {e}')
        raise
    except Exception as e:
        log.error(f'Failed to get release tickets: {e}')
        raise


def get_releases_tickets(jira, project_key, release_pattern, issue_types=None, statuses=None, date_filter=None, limit=None, dump_file=None, dump_format='csv'):
    '''
    Get and display tickets associated with releases matching a glob pattern.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        release_pattern: Glob pattern to match releases (e.g., '12.*').
        issue_types: List of issue type names to filter by, or None/empty for all.
        statuses: List of status names to filter by, or None/empty for all.
        date_filter: Date filter string.
        limit: Maximum number of tickets to retrieve, or None for all.
        dump_file: Output filename for dumping tickets, or None to skip.
        dump_format: Output format ('csv' or 'json').

    Output:
        None; prints ticket list to stdout.
    '''
    log.debug(f'Entering get_releases_tickets(project_key={project_key}, release_pattern={release_pattern}, issue_types={issue_types}, statuses={statuses}, date_filter={date_filter}, limit={limit}, dump_file={dump_file}, dump_format={dump_format})')
    
    project = validate_project(jira, project_key)
    
    try:
        # Get all versions and filter by pattern (supports exclusions with ! prefix)
        versions = jira.project_versions(project_key)
        if release_pattern and release_pattern != '*':
            matching_releases = [v.name for v in versions if match_pattern_with_exclusions(v.name, release_pattern)]
        else:
            matching_releases = [v.name for v in versions]
        
        if not matching_releases:
            output('')
            output(f'No releases matching pattern "{release_pattern}" found in project {project_key}.')
            output('')
            return
        
        log.info(f'Found {len(matching_releases)} releases matching pattern "{release_pattern}": {matching_releases}')
        
        # Normalize issue types and statuses
        normalized_types = normalize_issue_types(jira, project_key, issue_types) if issue_types else None
        normalized_statuses = normalize_statuses(jira, statuses) if statuses else None
        
        # Parse date filter
        date_clause = parse_date_filter(date_filter) if date_filter else ''
        
        # Build JQL query with all matching releases
        release_list = ', '.join([f'"{r}"' for r in matching_releases])
        jql_parts = [f'project = "{project_key}"', f'fixVersion IN ({release_list})']
        
        if normalized_types:
            type_list = ', '.join([f'"{t}"' for t in normalized_types])
            jql_parts.append(f'issuetype IN ({type_list})')
        
        # Build status clause using helper (handles both includes and excludes)
        status_clause = _build_status_jql(normalized_statuses)
        if status_clause:
            jql_parts.append(status_clause)
        
        jql = ' AND '.join(jql_parts)
        if date_clause:
            jql = f'{jql} {date_clause}'
        jql = f'{jql} ORDER BY fixVersion DESC, created DESC'
        
        log.debug(f'JQL query: {jql}')
        show_jql(jql)
        
        # Fetch tickets using the search API
        email, api_token = get_jira_credentials()
        
        all_issues = []
        next_page_token = None
        batch_size = min(100, limit) if limit else 100
        max_retries = 5
        
        while True:
            if limit and len(all_issues) >= limit:
                break
            
            if limit:
                remaining = limit - len(all_issues)
                current_batch = min(batch_size, remaining)
            else:
                current_batch = batch_size
            
            fields_to_fetch = ['summary', 'status', 'issuetype', 'created', 'updated', 'assignee', 'priority', 'project', 'fixVersions', 'versions']
            if dump_file:
                fields_to_fetch.extend(['reporter', 'resolutiondate'])
            
            payload = {
                'jql': jql,
                'maxResults': current_batch,
                'fields': fields_to_fetch
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token
            
            for retry in range(max_retries):
                response = requests.post(
                    f'{JIRA_URL}/rest/api/3/search/jql',
                    auth=(email, api_token),
                    headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
                    json=payload
                )
                
                if response.status_code == 429:
                    retry_after = int(response.headers.get('Retry-After', 5))
                    log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                    time.sleep(retry_after)
                    continue
                break
            
            if response.status_code != 200:
                log.error(f'API request failed: {response.status_code} - {response.text}')
                raise Exception(f'Jira API error: {response.status_code} - {response.text}')
            
            data = response.json()
            issues = data.get('issues', [])
            all_issues.extend(issues)
            
            next_page_token = data.get('nextPageToken')
            log.debug(f'Retrieved {len(all_issues)} issues so far...')
            
            if not next_page_token:
                break
            if limit and len(all_issues) >= limit:
                break
        
        if limit and len(all_issues) > limit:
            all_issues = all_issues[:limit]
        
        # Display results
        output('')
        output('=' * 130)
        output(f'Tickets for Releases matching "{release_pattern}" in Project: {project_key} ({project.name})')
        output('=' * 130)
        output(f'Matching releases: {len(matching_releases)}')
        if normalized_types:
            output(f'Issue types: {", ".join(normalized_types)}')
        if normalized_statuses:
            # Format status display for both includes and excludes
            if isinstance(normalized_statuses, dict):
                parts = []
                if normalized_statuses.get('include'):
                    parts.append(', '.join(normalized_statuses['include']))
                if normalized_statuses.get('exclude'):
                    parts.append('NOT: ' + ', '.join(normalized_statuses['exclude']))
                output(f'Statuses: {"; ".join(parts)}')
            else:
                output(f'Statuses: {", ".join(normalized_statuses)}')
        if date_filter:
            output(f'Date filter: {date_filter}')
        if limit:
            output(f'Limit: {limit}')
        print_ticket_table_header()
        
        for issue in all_issues:
            print_ticket_row(issue)
        
        print_ticket_table_footer(len(all_issues))
        
        # Dump to file if requested
        if dump_file and all_issues:
            dump_tickets_to_file(all_issues, dump_file, dump_format)
        
    except JiraProjectError:
        raise
    except ValueError as e:
        log.error(f'Invalid parameter: {e}')
        raise
    except Exception as e:
        log.error(f'Failed to get releases tickets: {e}')
        raise


def get_no_release_tickets(jira, project_key, issue_types=None, statuses=None, date_filter=None, limit=None, dump_file=None, dump_format='csv'):
    '''
    Get and display tickets that have no release/fixVersion assigned.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        issue_types: List of issue type names to filter by, or None/empty for all.
        statuses: List of status names to filter by, or None/empty for all.
        date_filter: Date filter string.
        limit: Maximum number of tickets to retrieve, or None for all.
        dump_file: Output filename for dumping tickets, or None to skip.
        dump_format: Output format ('csv' or 'json').

    Output:
        None; prints ticket list to stdout.
    '''
    log.debug(f'Entering get_no_release_tickets(project_key={project_key}, issue_types={issue_types}, statuses={statuses}, date_filter={date_filter}, limit={limit}, dump_file={dump_file}, dump_format={dump_format})')
    
    project = validate_project(jira, project_key)
    
    try:
        # Normalize issue types and statuses
        normalized_types = normalize_issue_types(jira, project_key, issue_types) if issue_types else None
        normalized_statuses = normalize_statuses(jira, statuses) if statuses else None
        
        # Parse date filter
        date_clause = parse_date_filter(date_filter) if date_filter else ''
        
        # Build JQL query - fixVersion is EMPTY means no release assigned
        jql_parts = [f'project = "{project_key}"', 'fixVersion is EMPTY']
        
        if normalized_types:
            type_list = ', '.join([f'"{t}"' for t in normalized_types])
            jql_parts.append(f'issuetype IN ({type_list})')
        
        # Build status clause using helper (handles both includes and excludes)
        status_clause = _build_status_jql(normalized_statuses)
        if status_clause:
            jql_parts.append(status_clause)
        
        jql = ' AND '.join(jql_parts)
        if date_clause:
            jql = f'{jql} {date_clause}'
        jql = f'{jql} ORDER BY created DESC'
        
        log.debug(f'JQL query: {jql}')
        show_jql(jql)
        
        # Fetch tickets using the search API
        email, api_token = get_jira_credentials()
        
        all_issues = []
        next_page_token = None
        batch_size = min(100, limit) if limit else 100
        max_retries = 5
        
        while True:
            if limit and len(all_issues) >= limit:
                break
            
            if limit:
                remaining = limit - len(all_issues)
                current_batch = min(batch_size, remaining)
            else:
                current_batch = batch_size
            
            fields_to_fetch = ['summary', 'status', 'issuetype', 'created', 'updated', 'assignee', 'priority', 'project', 'fixVersions', 'versions']
            if dump_file:
                fields_to_fetch.extend(['reporter', 'resolutiondate'])
            
            payload = {
                'jql': jql,
                'maxResults': current_batch,
                'fields': fields_to_fetch
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token
            
            for retry in range(max_retries):
                response = requests.post(
                    f'{JIRA_URL}/rest/api/3/search/jql',
                    auth=(email, api_token),
                    headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
                    json=payload
                )
                
                if response.status_code == 429:
                    retry_after = int(response.headers.get('Retry-After', 5))
                    log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                    time.sleep(retry_after)
                    continue
                break
            
            if response.status_code != 200:
                log.error(f'API request failed: {response.status_code} - {response.text}')
                raise Exception(f'Jira API error: {response.status_code} - {response.text}')
            
            data = response.json()
            issues = data.get('issues', [])
            all_issues.extend(issues)
            
            next_page_token = data.get('nextPageToken')
            log.debug(f'Retrieved {len(all_issues)} issues so far...')
            
            if not next_page_token:
                break
            if limit and len(all_issues) >= limit:
                break
        
        if limit and len(all_issues) > limit:
            all_issues = all_issues[:limit]
        
        log.debug(f'Retrieved {len(all_issues)} issues total')
        
        # Print results
        output('')
        output('=' * 120)
        output(f'Tickets with No Release Assigned')
        output(f'Project: {project_key} ({project.name})')
        output('=' * 120)
        
        if normalized_types:
            output(f'Issue types: {", ".join(normalized_types)}')
        if normalized_statuses:
            # Format status display for both includes and excludes
            if isinstance(normalized_statuses, dict):
                parts = []
                if normalized_statuses.get('include'):
                    parts.append(', '.join(normalized_statuses['include']))
                if normalized_statuses.get('exclude'):
                    parts.append('NOT: ' + ', '.join(normalized_statuses['exclude']))
                output(f'Statuses: {"; ".join(parts)}')
            else:
                output(f'Statuses: {", ".join(normalized_statuses)}')
        if date_filter and date_filter.lower() != 'all':
            output(f'Date filter: {date_filter}')
        if limit:
            output(f'Limit: {limit}')
        
        print_ticket_table_header()
        
        for issue in all_issues:
            print_ticket_row(issue)
        
        print_ticket_table_footer(len(all_issues))
        
        if dump_file:
            dump_tickets_to_file(all_issues, dump_file, dump_format)
        
    except JiraProjectError:
        raise
    except ValueError as e:
        log.error(f'Invalid parameter: {e}')
        raise
    except Exception as e:
        log.error(f'Failed to get no-release tickets: {e}')
        raise


def get_ticket_totals(jira, project_key, issue_types=None, statuses=None, date_filter=None):
    '''
    Get and display ticket counts for a project.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        issue_types: List of issue type names to filter by, or None/empty for all (case-insensitive).
        statuses: List of status names to filter by, or None/empty for all (case-insensitive).
        date_filter: Date filter string (today, week, month, year, all, or range).

    Output:
        None; prints ticket counts to stdout.

    Side Effects:
        Logs ticket count information and prints formatted totals.
    '''
    log.debug(f'Entering get_ticket_totals(project_key={project_key}, issue_types={issue_types}, statuses={statuses}, date_filter={date_filter})')
    
    project = validate_project(jira, project_key)
    
    try:
        # Normalize issue types (case-insensitive)
        normalized_types = normalize_issue_types(jira, project_key, issue_types) if issue_types else None
        
        # Normalize statuses (case-insensitive)
        normalized_statuses = normalize_statuses(jira, statuses) if statuses else None
        
        # Parse date filter
        date_clause = parse_date_filter(date_filter) if date_filter else ''
        
        # Build base JQL query
        jql_parts = [f'project = "{project_key}"']
        
        if normalized_types:
            type_list = ', '.join([f'"{t}"' for t in normalized_types])
            jql_parts.append(f'issuetype IN ({type_list})')
        
        # Build status clause using helper (handles both includes and excludes)
        status_clause = _build_status_jql(normalized_statuses)
        if status_clause:
            jql_parts.append(status_clause)
        
        jql = ' AND '.join(jql_parts)
        if date_clause:
            jql = f'{jql} {date_clause}'
        
        jql = jql.strip()
        log.debug(f'JQL query: {jql}')
        show_jql(jql)
        
        # Use the count API endpoint for efficiency
        email, api_token = get_jira_credentials()
        
        response = requests.post(
            f'{JIRA_URL}/rest/api/3/search/approximate-count',
            auth=(email, api_token),
            headers={
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            },
            json={'jql': jql}
        )
        
        if response.status_code != 200:
            log.error(f'Count API request failed: {response.status_code} - {response.text}')
            raise Exception(f'Jira API error: {response.status_code} - {response.text}')
        
        data = response.json()
        total_count = data.get('count', 0)
        
        # Print results
        output('')
        output('=' * 60)
        output(f'Ticket Count for Project: {project_key} ({project.name})')
        output('=' * 60)
        
        if normalized_types:
            output(f'Issue types: {", ".join(normalized_types)}')
        if normalized_statuses:
            # Format status display for both includes and excludes
            if isinstance(normalized_statuses, dict):
                parts = []
                if normalized_statuses.get('include'):
                    parts.append(', '.join(normalized_statuses['include']))
                if normalized_statuses.get('exclude'):
                    parts.append('NOT: ' + ', '.join(normalized_statuses['exclude']))
                output(f'Statuses: {"; ".join(parts)}')
            else:
                output(f'Statuses: {", ".join(normalized_statuses)}')
        if date_filter and date_filter.lower() != 'all':
            output(f'Date filter: {date_filter}')
        
        output('-' * 60)
        output(f'Total tickets: {total_count}')
        output('=' * 60)
        output('')
        
    except JiraProjectError:
        raise
    except ValueError as e:
        log.error(f'Invalid date filter: {e}')
        raise
    except Exception as e:
        log.error(f'Failed to get ticket totals: {e}')
        raise


def get_tickets(jira, project_key, issue_types=None, statuses=None, date_filter=None, limit=None, dump_file=None, dump_format='csv'):
    '''
    Get and display tickets for a project.

    Input:
        jira: JIRA object with active connection.
        project_key: String key of the project (e.g., 'PROJ').
        issue_types: List of issue type names to filter by, or None/empty for all (case-insensitive).
        statuses: List of status names to filter by, or None/empty for all (case-insensitive).
        date_filter: Date filter string (today, week, month, year, all, or range).
        limit: Maximum number of tickets to retrieve, or None for all.
        dump_file: Output filename for dumping tickets, or None to skip.
        dump_format: Output format ('csv' or 'json').

    Output:
        None; prints ticket list to stdout and optionally writes to file.

    Side Effects:
        Logs ticket information and prints formatted ticket list.
        If dump_file is specified, writes tickets to file.
    '''
    log.debug(f'Entering get_tickets(project_key={project_key}, issue_types={issue_types}, statuses={statuses}, date_filter={date_filter}, limit={limit}, dump_file={dump_file}, dump_format={dump_format})')
    
    project = validate_project(jira, project_key)
    
    try:
        # Normalize issue types (case-insensitive)
        normalized_types = normalize_issue_types(jira, project_key, issue_types) if issue_types else None
        
        # Normalize statuses (case-insensitive)
        normalized_statuses = normalize_statuses(jira, statuses) if statuses else None
        
        # Parse date filter
        date_clause = parse_date_filter(date_filter) if date_filter else ''
        
        # Build JQL query
        jql_parts = [f'project = "{project_key}"']
        
        if normalized_types:
            type_list = ', '.join([f'"{t}"' for t in normalized_types])
            jql_parts.append(f'issuetype IN ({type_list})')
        
        # Build status clause using helper (handles both includes and excludes)
        status_clause = _build_status_jql(normalized_statuses)
        if status_clause:
            jql_parts.append(status_clause)
        
        jql = ' AND '.join(jql_parts)
        if date_clause:
            jql = f'{jql} {date_clause}'
        jql = f'{jql} ORDER BY created DESC'
        
        jql = jql.strip()
        log.debug(f'JQL query: {jql}')
        show_jql(jql)
        
        # Use the new /rest/api/3/search/jql endpoint directly
        email, api_token = get_jira_credentials()
        
        all_issues = []
        next_page_token = None
        batch_size = min(100, limit) if limit else 100
        max_retries = 5
        
        while True:
            # Check if we've reached the limit
            if limit and len(all_issues) >= limit:
                break
            
            # Adjust batch size if near limit
            if limit:
                remaining = limit - len(all_issues)
                current_batch = min(batch_size, remaining)
            else:
                current_batch = batch_size
            
            # Build request payload - include extra fields if dumping
            fields_to_fetch = ['summary', 'status', 'issuetype', 'created', 'updated', 'assignee', 'priority', 'project', 'fixVersions', 'versions']
            if dump_file:
                fields_to_fetch.extend(['reporter', 'resolutiondate'])
            
            payload = {
                'jql': jql,
                'maxResults': current_batch,
                'fields': fields_to_fetch
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token
            
            # Retry logic for rate limiting
            for retry in range(max_retries):
                response = requests.post(
                    f'{JIRA_URL}/rest/api/3/search/jql',
                    auth=(email, api_token),
                    headers={
                        'Content-Type': 'application/json',
                        'Accept': 'application/json'
                    },
                    json=payload
                )
                
                if response.status_code == 429:
                    retry_after = int(response.headers.get('Retry-After', 5))
                    log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                    time.sleep(retry_after)
                    continue
                break
            
            if response.status_code != 200:
                log.error(f'API request failed: {response.status_code} - {response.text}')
                raise Exception(f'Jira API error: {response.status_code} - {response.text}')
            
            data = response.json()
            issues = data.get('issues', [])
            all_issues.extend(issues)
            
            next_page_token = data.get('nextPageToken')
            log.debug(f'Retrieved {len(all_issues)} issues so far...')
            
            # Check for next page token or limit reached
            if not next_page_token:
                break
            if limit and len(all_issues) >= limit:
                break
        
        # Trim to limit if we went over
        if limit and len(all_issues) > limit:
            all_issues = all_issues[:limit]
        
        log.debug(f'Retrieved {len(all_issues)} issues total')
        
        # Print results
        output('')
        output('=' * 120)
        output(f'Tickets for Project: {project_key} ({project.name})')
        output('=' * 120)
        
        if normalized_types:
            output(f'Issue types: {", ".join(normalized_types)}')
        if normalized_statuses:
            # Format status display for both includes and excludes
            if isinstance(normalized_statuses, dict):
                parts = []
                if normalized_statuses.get('include'):
                    parts.append(', '.join(normalized_statuses['include']))
                if normalized_statuses.get('exclude'):
                    parts.append('NOT: ' + ', '.join(normalized_statuses['exclude']))
                output(f'Statuses: {"; ".join(parts)}')
            else:
                output(f'Statuses: {", ".join(normalized_statuses)}')
        if date_filter and date_filter.lower() != 'all':
            output(f'Date filter: {date_filter}')
        if limit:
            output(f'Limit: {limit}')
        
        print_ticket_table_header()
        
        for issue in all_issues:
            print_ticket_row(issue)
        
        print_ticket_table_footer(len(all_issues))
        
        # Dump to file if requested
        if dump_file:
            dump_tickets_to_file(all_issues, dump_file, dump_format)
        
    except JiraProjectError:
        raise
    except ValueError as e:
        log.error(f'Invalid parameter: {e}')
        raise
    except Exception as e:
        log.error(f'Failed to get tickets: {e}')
        raise


def _write_excel(rows, output_path, extra_fields=None, table_format='flat'):
    '''
    Write ticket rows to an Excel (.xlsx) file using openpyxl.

    Ticket key cells are rendered as clickable Jira hyperlinks (display text is
    the ticket ID, e.g. "STL-71438", with the full browse URL behind it).

    Supports both "flat" and "indented" table formats (same logic as the CSV
    writer but targeting an Excel workbook).

    Input:
        rows: List of row dicts (already flattened by dump_tickets_to_file).
        output_path: Destination .xlsx file path.
        extra_fields: Original extra_fields mapping (used only for metadata awareness).
        table_format: 'flat' or 'indented'.
    '''
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.error('openpyxl is required for Excel output. Install with: pip install openpyxl')
        raise ImportError('openpyxl is required for --dump-format excel. Install with: pip install openpyxl')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tickets'

    # Jira base URL for building hyperlinks
    jira_url = JIRA_URL.rstrip('/')

    # ---------------------------------------------------------------
    # Determine column layout based on table_format
    # ---------------------------------------------------------------
    base_fields = ['key', 'project', 'issue_type', 'status', 'priority', 'summary',
                   'assignee', 'reporter', 'created', 'updated', 'resolved',
                   'fix_version', 'affects_version']

    is_indented = (table_format == 'indented' and any('depth' in r for r in rows))

    if is_indented:
        # Calculate max depth
        max_depth = 0
        for r in rows:
            d = r.get('depth')
            if d is not None:
                try:
                    max_depth = max(max_depth, int(d))
                except (ValueError, TypeError):
                    pass

        depth_columns = [f'Depth {i}' for i in range(max_depth + 1)]
        content_fields = [f for f in base_fields if f != 'key']

        # Extra columns excluding depth (already represented by depth columns)
        all_keys = set(base_fields)
        for r in rows:
            all_keys.update(r.keys())
        extra_columns = sorted(k for k in all_keys if k not in base_fields and k != 'depth')

        fieldnames = depth_columns + content_fields + extra_columns
    else:
        # Flat format
        all_keys = set(base_fields)
        for r in rows:
            all_keys.update(r.keys())
        extra_columns = sorted(k for k in all_keys if k not in base_fields)
        fieldnames = base_fields + extra_columns

    # ---------------------------------------------------------------
    # Header row styling
    # ---------------------------------------------------------------
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    link_font = Font(color='0563C1', underline='single')

    # Write header
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ---------------------------------------------------------------
    # Data rows
    # ---------------------------------------------------------------
    for row_idx, row_data in enumerate(rows, 2):
        if is_indented:
            # Build indented row: ticket key goes into the correct Depth column
            d = 0
            try:
                d = int(row_data.get('depth', 0))
            except (ValueError, TypeError):
                d = 0

            for col_idx, field in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                if field in depth_columns:
                    # Only populate the depth column matching this row's depth
                    if field == f'Depth {d}':
                        ticket_key = row_data.get('key', '')
                        cell.value = ticket_key
                        if ticket_key:
                            cell.hyperlink = f'{jira_url}/browse/{ticket_key}'
                            cell.font = link_font
                    else:
                        cell.value = ''
                else:
                    cell.value = row_data.get(field, '')
        else:
            # Flat row
            for col_idx, field in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                value = row_data.get(field, '')

                if field == 'key' and value:
                    # Render ticket key as a clickable hyperlink
                    cell.value = value
                    cell.hyperlink = f'{jira_url}/browse/{value}'
                    cell.font = link_font
                else:
                    cell.value = value

    # ---------------------------------------------------------------
    # Auto-fit column widths (approximate)
    # ---------------------------------------------------------------
    for col_idx, field in enumerate(fieldnames, 1):
        # Start with header width
        max_len = len(str(field))
        # Sample up to 50 data rows for width estimation
        for row_idx in range(2, min(len(rows) + 2, 52)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        # Cap at 50 characters, minimum 10
        adjusted_width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Enable auto-filter on the header row
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    log.info(f'Wrote {len(rows)} tickets (excel, table_format={table_format}) to: {output_path}')


def dump_tickets_to_file(issues, dump_file, dump_format, extra_fields=None, table_format='flat'):
    '''
    Write tickets to a file in the specified format.

    Input:
        issues: List of issue dicts from Jira API.
        dump_file: Output filename (without extension).
        dump_format: Output format ('csv' or 'json').
        extra_fields: Optional mapping of issue key -> dict of additional fields to include (e.g., {'depth': 1, 'via': 'blocks'}).
        table_format: Table layout for CSV ('flat' or 'indented'). 'flat' uses a depth column.
                      'indented' replaces the depth column with per-level columns (Depth 0, Depth 1, ...)
                      where the ticket key appears in the column matching its depth.

    Output:
        None; writes to file.

    Side Effects:
        Creates or overwrites the output file.
    '''
    log.debug(f'Entering dump_tickets_to_file(issues_count={len(issues)}, dump_file={dump_file}, dump_format={dump_format}, extra_fields_provided={extra_fields is not None}, table_format={table_format})')
    # Add extension if not present
    # Excel uses .xlsx extension rather than .excel
    ext = 'xlsx' if dump_format == 'excel' else dump_format
    if not dump_file.endswith(f'.{ext}'):
        output_path = f'{dump_file}.{ext}'
    else:
        output_path = dump_file
    
    log.debug(f'Writing {len(issues)} tickets to {output_path}')
    
    # Extract and flatten issue data
    rows = []
    for issue in issues:
        fields = issue.get('fields', {})
        
        # Extract fix versions
        fix_versions = fields.get('fixVersions', [])
        fix_version_str = ', '.join([v.get('name', '') for v in fix_versions]) if fix_versions else ''
        
        # Extract affects versions (for bugs)
        affects_versions = fields.get('versions', [])
        affects_version_str = ', '.join([v.get('name', '') for v in affects_versions]) if affects_versions else ''
        
        # Extract common fields
        row = {
            'key': issue.get('key', ''),
            'project': fields.get('project', {}).get('key', '') if fields.get('project') else '',
            'issue_type': fields.get('issuetype', {}).get('name', '') if fields.get('issuetype') else '',
            'status': fields.get('status', {}).get('name', '') if fields.get('status') else '',
            'priority': fields.get('priority', {}).get('name', '') if fields.get('priority') else '',
            'summary': fields.get('summary', '') or '',
            'assignee': fields.get('assignee', {}).get('displayName', '') if fields.get('assignee') else '',
            'reporter': fields.get('reporter', {}).get('displayName', '') if fields.get('reporter') else '',
            'created': '',
            'updated': '',
            'resolved': '',
            'fix_version': fix_version_str,
            'affects_version': affects_version_str,
        }

        # Optional extra fields (e.g., hierarchy depth, link type, traversal metadata)
        if extra_fields:
            meta = extra_fields.get(issue.get('key', '')) if isinstance(extra_fields, dict) else None
            if meta:
                # Backward compatible fields used by drawio_utilities.py
                if 'depth' in meta:
                    row['depth'] = meta.get('depth')
                if 'via' in meta:
                    row['link_via'] = meta.get('via') or ''

                # Pass through any additional metadata as extra CSV/JSON columns
                # (e.g., relation='child'|'link', from_key=<source issue key>)
                for k, v in meta.items():
                    if k in ('depth', 'via'):
                        continue
                    row[k] = v
        
        # Format dates
        for date_field in ['created', 'updated', 'resolutiondate']:
            date_val = fields.get(date_field)
            if date_val:
                try:
                    dt = datetime.fromisoformat(date_val.replace('Z', '+00:00'))
                    if date_field == 'resolutiondate':
                        row['resolved'] = dt.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        row[date_field] = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    if date_field == 'resolutiondate':
                        row['resolved'] = date_val
                    else:
                        row[date_field] = date_val
        
        rows.append(row)
    
    if dump_format == 'json':
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, indent=2, ensure_ascii=False)
    elif dump_format == 'excel':
        _write_excel(rows, output_path, extra_fields, table_format)
    elif dump_format == 'csv':
        if rows:
            base_fields = ['key', 'project', 'issue_type', 'status', 'priority', 'summary',
                           'assignee', 'reporter', 'created', 'updated', 'resolved',
                           'fix_version', 'affects_version']

            # ------------------------------------------------------------------
            # "indented" table format: replace the depth column with per-level
            # columns (Depth 0, Depth 1, ...) where the ticket key is placed in
            # the column matching its depth.  Content columns start after the
            # deepest depth column.
            # ------------------------------------------------------------------
            if table_format == 'indented' and any('depth' in r for r in rows):
                # Determine the maximum depth across all rows
                max_depth = 0
                for r in rows:
                    d = r.get('depth')
                    if d is not None:
                        try:
                            max_depth = max(max_depth, int(d))
                        except (ValueError, TypeError):
                            pass

                # Build depth column names: Depth 0, Depth 1, ...
                depth_columns = [f'Depth {i}' for i in range(max_depth + 1)]

                # Content columns are everything except 'key' and 'depth'
                # (key is moved into the depth columns)
                content_fields = [f for f in base_fields if f != 'key']

                # Collect extra columns (link_via, from_key, relation, etc.) but
                # exclude 'depth' since it is represented by the depth columns
                all_keys = set(base_fields)
                for r in rows:
                    all_keys.update(r.keys())
                extra_columns = sorted(k for k in all_keys if k not in base_fields and k != 'depth')

                fieldnames = depth_columns + content_fields + extra_columns

                indented_rows = []
                for r in rows:
                    new_row = {}
                    # Place the ticket key in the correct depth column
                    d = 0
                    try:
                        d = int(r.get('depth', 0))
                    except (ValueError, TypeError):
                        d = 0
                    for col in depth_columns:
                        new_row[col] = ''
                    new_row[f'Depth {d}'] = r.get('key', '')

                    # Copy content fields
                    for f in content_fields:
                        new_row[f] = r.get(f, '')

                    # Copy extra columns (excluding depth)
                    for col in extra_columns:
                        new_row[col] = r.get(col, '')

                    indented_rows.append(new_row)

                with open(output_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(indented_rows)

                log.info(f'Wrote {len(indented_rows)} tickets (indented format, max depth {max_depth}) to: {output_path}')
                return

            # ------------------------------------------------------------------
            # "flat" table format (default): depth is a regular column
            # ------------------------------------------------------------------
            # Collect all keys present across rows (to capture depth/link_via, etc.)
            all_keys = set(base_fields)
            for r in rows:
                all_keys.update(r.keys())
            extra_columns = [k for k in all_keys if k not in base_fields]
            extra_columns.sort()
            # Ensure every row has all fields
            for r in rows:
                for col in extra_columns:
                    r.setdefault(col, '')
            fieldnames = base_fields + extra_columns
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
        else:
            # Write empty file with headers
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['key', 'project', 'issue_type', 'status', 'priority', 'summary',
                                 'assignee', 'reporter', 'created', 'updated', 'resolved',
                                 'fix_version', 'affects_version'])
    
    log.info(f'Wrote {len(rows)} tickets to: {output_path}')


def load_tickets_from_csv(input_file):
    '''
    Load tickets from a CSV file into a list of dictionaries.

    Input:
        input_file: Path to the CSV file.

    Output:
        List of dictionaries, each containing ticket data with 'key' and other fields.

    Raises:
        FileNotFoundError: If the input file doesn't exist.
        ValueError: If the CSV doesn't have a 'key' column.
    '''
    log.debug(f'Entering load_tickets_from_csv(input_file={input_file})')
    if not os.path.exists(input_file):
        # Try adding .csv extension if not present
        if not input_file.endswith('.csv') and os.path.exists(f'{input_file}.csv'):
            input_file = f'{input_file}.csv'
            log.debug(f'Added .csv extension: {input_file}')
        else:
            raise FileNotFoundError(f'Input file not found: {input_file}')
    
    tickets = []
    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        
        # Verify 'key' column exists
        if 'key' not in reader.fieldnames:
            raise ValueError(f'CSV file must have a "key" column. Found columns: {reader.fieldnames}')
        
        for row in reader:
            # Create dict with all columns as key:value pairs
            ticket = {k: v for k, v in row.items()}
            tickets.append(ticket)
    
    log.debug(f'Loaded {len(tickets)} tickets from CSV')
    return tickets


def _adf_from_text(text):
    '''
    Convert plain text into Jira Cloud ADF (Atlassian Document Format).

    Jira Cloud REST API v3 expects ADF JSON for rich text fields (e.g. description).

    References:
      - Atlassian ADF overview: https://developer.atlassian.com/cloud/jira/platform/apis/document/structure/
      - Jira issue create (v3): https://developer.atlassian.com/cloud/jira/platform/rest/v3/api-group-issues/#api-rest-api-3-issue-post
    '''
    if text is None:
        return None

    # Keep it intentionally simple: one paragraph with one text node.
    return {
        'type': 'doc',
        'version': 1,
        'content': [
            {
                'type': 'paragraph',
                'content': [
                    {'type': 'text', 'text': str(text)}
                ],
            }
        ],
    }


def create_ticket(
    jira,
    project_key,
    summary,
    issue_type,
    description=None,
    assignee=None,
    components=None,
    fix_versions=None,
    labels=None,
    parent_key=None,
    dry_run=True,
):
    '''
    Create a Jira ticket.

    Safety model:
        - Dry-run by default (no changes).
        - Actual creation requires --execute (which disables dry-run).

    Input:
        jira: JIRA client instance.
        project_key: Jira project key.
        summary: Issue summary.
        issue_type: Issue type name (e.g. Task, Bug, Story, Epic).
        description: Optional plain-text description (converted to ADF for Jira Cloud).
        assignee: Optional assignee accountId.
        components: Optional list of component names.
        fix_versions: Optional list of fixVersion names.
        labels: Optional list of labels.
        parent_key: Optional parent issue key (sub-task parent / epic parent, depending on project configuration).
        dry_run: If True, do not create; only print what would be created.

    Output:
        None; prints results to stdout.

    Side Effects:
        If not dry_run, creates an issue in Jira.

    Notes:
        - For Jira Cloud, assignee must typically be an accountId, not an email.
        - Epic linkage varies across Jira configurations; some instances use a custom field.
          We set the standard `parent` field when parent_key is provided.
    '''
    log.debug(
        'Entering create_ticket('
        f'project_key={project_key}, summary={summary}, issue_type={issue_type}, '
        f'assignee={assignee}, components={components}, fix_versions={fix_versions}, '
        f'labels={labels}, parent_key={parent_key}, dry_run={dry_run})'
    )

    # Build issue fields
    fields = {
        'project': {'key': project_key},
        'summary': summary,
        'issuetype': {'name': issue_type},
    }

    if description:
        fields['description'] = _adf_from_text(description)

    if assignee:
        # Jira Cloud: `id` is the accountId.
        fields['assignee'] = {'id': assignee}

    if components:
        fields['components'] = [{'name': c} for c in components]

    if fix_versions:
        fields['fixVersions'] = [{'name': v} for v in fix_versions]

    if labels:
        fields['labels'] = labels

    if parent_key:
        fields['parent'] = {'key': parent_key}

    output('')
    output('=' * 80)
    if dry_run:
        output('CREATE TICKET - DRY RUN (no changes will be made)')
    else:
        output('CREATE TICKET - EXECUTING')
    output('=' * 80)
    output(f'Project:     {project_key}')
    output(f'Type:        {issue_type}')
    output(f'Summary:     {summary}')
    if parent_key:
        output(f'Parent:      {parent_key}')
    if assignee:
        output(f'Assignee ID: {assignee}')
    if components:
        output(f'Components:  {", ".join(components)}')
    if fix_versions:
        output(f'FixVersions: {", ".join(fix_versions)}')
    if labels:
        output(f'Labels:      {", ".join(labels)}')
    output('-' * 80)

    if dry_run:
        output('Would create ticket with the above fields.')
        output('To execute creation, add --execute.')
        output('=' * 80)
        output('')
        return

    try:
        issue = jira.create_issue(fields=fields)
        output(f'Created: {issue.key}')
        output(f'URL:     {JIRA_URL}/browse/{issue.key}')
        output('=' * 80)
        output('')
    except Exception as e:
        log.error(f'Failed to create ticket: {e}')
        raise


def bulk_update_tickets(jira, input_file, set_release=None, remove_release=False,
                        transition=None, assign=None, dry_run=True, max_updates=None):
    '''
    Perform bulk updates on tickets loaded from a CSV file.

    Input:
        jira: JIRA object with active connection.
        input_file: Path to the CSV file containing ticket keys.
        set_release: Release/version name to set on tickets, or None.
        remove_release: If True, remove all releases from tickets.
        transition: Status name to transition tickets to, or None.
        assign: Username or email to assign tickets to, or None.
        dry_run: If True, only preview changes without applying them.
        max_updates: Maximum number of tickets to update, or None for all.

    Output:
        None; prints results to stdout.

    Side Effects:
        If not dry_run, modifies tickets in Jira.
    '''
    log.debug(f'Entering bulk_update_tickets(input_file={input_file}, set_release={set_release}, remove_release={remove_release}, transition={transition}, assign={assign}, dry_run={dry_run}, max_updates={max_updates})')
    
    # Load tickets from CSV
    tickets = load_tickets_from_csv(input_file)
    
    if not tickets:
        output('No tickets found in input file.')
        return
    
    # Apply max_updates limit
    if max_updates and len(tickets) > max_updates:
        log.debug(f'Limiting updates to {max_updates} tickets (out of {len(tickets)})')
        tickets = tickets[:max_updates]
    
    # Determine what operations to perform
    operations = []
    if set_release:
        operations.append(f'Set release to: {set_release}')
    if remove_release:
        operations.append('Remove release')
    if transition:
        operations.append(f'Transition to: {transition}')
    if assign:
        operations.append(f'Assign to: {assign}')
    
    if not operations:
        output('ERROR: No update operations specified.')
        output('Use --set-release, --remove-release, --transition, or --assign')
        return
    
    # Print summary
    output()
    output('=' * 80)
    if dry_run:
        output('BULK UPDATE - DRY RUN (no changes will be made)')
    else:
        output('BULK UPDATE - EXECUTING CHANGES')
    output('=' * 80)
    output(f'Input file: {input_file}')
    output(f'Tickets to update: {len(tickets)}')
    output(f'Operations:')
    for op in operations:
        output(f'  - {op}')
    output('-' * 80)
    
    # Process each ticket
    log.debug(f'Starting to process {len(tickets)} tickets')
    success_count = 0
    error_count = 0
    errors = []
    
    for i, ticket in enumerate(tickets, 1):
        ticket_key = ticket.get('key')
        log.debug(f'Processing ticket {i}/{len(tickets)}: {ticket_key}')
        log.debug(f'Ticket data: {ticket}')
        
        if not ticket_key:
            log.warning(f'Skipping row {i}: no key found')
            error_count += 1
            errors.append((f'Row {i}', 'No key found'))
            continue
        
        # Show progress
        status_str = f'[{i}/{len(tickets)}] {ticket_key}'
        
        if dry_run:
            # Dry run - just show what would happen
            log.debug(f'{ticket_key}: Dry run - would apply operations')
            output(f'{status_str}: Would apply: {", ".join(operations)}')
            success_count += 1
        else:
            # Execute the updates
            try:
                log.debug(f'{ticket_key}: Fetching issue from Jira')
                issue = jira.issue(ticket_key)
                
                # Set release
                if set_release:
                    issue.update(fields={'fixVersions': [{'name': set_release}]})
                    log.debug(f'{ticket_key}: Set release to {set_release}')
                
                # Remove release
                if remove_release:
                    issue.update(fields={'fixVersions': []})
                    log.debug(f'{ticket_key}: Removed release')
                
                # Transition
                if transition:
                    # Find the transition ID
                    transitions = jira.transitions(issue)
                    trans_id = None
                    for t in transitions:
                        if t['name'].lower() == transition.lower():
                            trans_id = t['id']
                            break
                    
                    if trans_id:
                        jira.transition_issue(issue, trans_id)
                        log.debug(f'{ticket_key}: Transitioned to {transition}')
                    else:
                        available = [t['name'] for t in transitions]
                        raise Exception(f'Transition "{transition}" not available. Available: {available}')
                
                # Assign
                if assign:
                    # Handle special case of unassign
                    if assign.lower() in ['none', 'unassigned', '']:
                        jira.assign_issue(issue, None)
                    else:
                        jira.assign_issue(issue, assign)
                    log.debug(f'{ticket_key}: Assigned to {assign}')
                
                output(f'{status_str}: SUCCESS')
                success_count += 1
                
            except Exception as e:
                log.error(f'{ticket_key}: Failed - {e}')
                output(f'{status_str}: FAILED - {e}')
                error_count += 1
                errors.append((ticket_key, str(e)))
    
    # Print summary
    output('-' * 80)
    output(f'Completed: {success_count} successful, {error_count} failed')
    
    if errors and not dry_run:
        output()
        output('Errors:')
        for key, error in errors[:10]:  # Show first 10 errors
            output(f'  {key}: {error}')
        if len(errors) > 10:
            output(f'  ... and {len(errors) - 10} more errors')
    
    if dry_run:
        output()
        output('This was a DRY RUN. To execute changes, add --execute flag.')
    
    output('=' * 80)
    output()


def bulk_delete_tickets(jira, input_file, delete_subtasks=False, dry_run=True, max_deletes=None, force=False):
    '''
    Perform bulk deletion of tickets listed in a CSV file.

    Safety model:
        - Dry-run by default (no changes).
        - Actual deletes require --execute (which disables dry-run).
        - When executing, an additional explicit confirmation prompt is shown unless --force is set.

    Input:
        jira: JIRA object with active connection.
        input_file: Path to the CSV file containing ticket keys.
        delete_subtasks: If True, delete subtasks as well (Jira REST parameter deleteSubtasks=true).
        dry_run: If True, only preview deletions.
        max_deletes: Optional cap on how many tickets to delete.
        force: If True, skip the interactive confirmation prompt.

    Output:
        None; prints results to stdout.

    Side Effects:
        If not dry_run, deletes tickets in Jira.

    Notes:
        Jira Cloud does not provide a single "bulk delete" REST endpoint; deletion is performed
        per-issue via HTTP DELETE.
    '''
    log.debug(
        'Entering bulk_delete_tickets('
        f'input_file={input_file}, delete_subtasks={delete_subtasks}, '
        f'dry_run={dry_run}, max_deletes={max_deletes}, force={force})'
    )

    tickets = load_tickets_from_csv(input_file)

    if not tickets:
        output('No tickets found in input file.')
        return

    # Apply max_deletes cap
    if max_deletes and len(tickets) > max_deletes:
        log.debug(f'Limiting deletes to {max_deletes} tickets (out of {len(tickets)})')
        tickets = tickets[:max_deletes]

    # Print summary
    output()
    output('=' * 80)
    if dry_run:
        output('BULK DELETE - DRY RUN (no changes will be made)')
    else:
        output('BULK DELETE - EXECUTING DELETES')
    output('=' * 80)
    output(f'Input file:        {input_file}')
    output(f'Tickets to delete: {len(tickets)}')
    output(f'Delete subtasks:   {delete_subtasks}')
    output('-' * 80)

    if not dry_run and not force:
        output('WARNING: This operation will permanently delete issues in Jira.')
        output('To confirm, type DELETE and press Enter.')
        confirmation = input('CONFIRM DELETE> ').strip()
        if confirmation != 'DELETE':
            output('Aborting: confirmation did not match "DELETE".')
            output('=' * 80)
            output()
            return

    email, api_token = get_jira_credentials()

    success_count = 0
    error_count = 0
    errors = []

    max_retries = 5

    for i, ticket in enumerate(tickets, 1):
        ticket_key = (ticket.get('key') or '').strip()

        if not ticket_key:
            log.warning(f'Skipping row {i}: no key found')
            error_count += 1
            errors.append((f'Row {i}', 'No key found'))
            continue

        status_str = f'[{i}/{len(tickets)}] {ticket_key}'

        if dry_run:
            output(f'{status_str}: Would delete (delete_subtasks={delete_subtasks})')
            success_count += 1
            continue

        # Execute deletion via REST API (per-issue)
        params = {}
        if delete_subtasks:
            params['deleteSubtasks'] = 'true'

        for retry in range(max_retries):
            response = requests.delete(
                f'{JIRA_URL}/rest/api/3/issue/{ticket_key}',
                auth=(email, api_token),
                headers={'Accept': 'application/json'},
                params=params,
            )

            if response.status_code == 429:
                retry_after = int(response.headers.get('Retry-After', 5))
                log.warning(
                    f'{ticket_key}: Rate limited. Waiting {retry_after} seconds '
                    f'(retry {retry + 1}/{max_retries})...'
                )
                time.sleep(retry_after)
                continue

            break

        # Jira returns 204 No Content on success
        if response.status_code in (200, 202, 204):
            output(f'{status_str}: DELETED')
            success_count += 1
        else:
            msg = f'{response.status_code} - {response.text}'
            output(f'{status_str}: FAILED - {msg}')
            log.error(f'{ticket_key}: Delete failed: {msg}')
            error_count += 1
            errors.append((ticket_key, msg))

    # Print summary
    output('-' * 80)
    output(f'Completed: {success_count} successful, {error_count} failed')

    if errors and not dry_run:
        output()
        output('Errors:')
        for key, error in errors[:10]:  # Show first 10 errors
            output(f'  {key}: {error}')
        if len(errors) > 10:
            output(f'  ... and {len(errors) - 10} more errors')

    if dry_run:
        output()
        output('This was a DRY RUN. To execute deletions, add --execute flag.')

    output('=' * 80)
    output()


def run_jql_query(jira, jql_query, limit=None, dump_file=None, dump_format='csv'):
    '''
    Run a generic JQL query and display results.

    Input:
        jira: JIRA object with active connection.
        jql_query: JQL query string.
        limit: Maximum number of tickets to retrieve, or None for all.
        dump_file: Output filename for dumping tickets, or None to skip.
        dump_format: Output format ('csv' or 'json').

    Output:
        None; prints ticket list to stdout and optionally writes to file.

    Side Effects:
        Logs query information and prints formatted ticket list.
        If dump_file is specified, writes tickets to file.
    '''
    log.debug(f'Entering run_jql_query(jql_query={jql_query}, limit={limit}, dump_file={dump_file}, dump_format={dump_format})')
    
    try:
        show_jql(jql_query)
        
        # Use the new /rest/api/3/search/jql endpoint directly
        email, api_token = get_jira_credentials()
        
        all_issues = []
        next_page_token = None
        batch_size = min(100, limit) if limit else 100
        max_retries = 5
        
        while True:
            # Check if we've reached the limit
            if limit and len(all_issues) >= limit:
                break
            
            # Adjust batch size if near limit
            if limit:
                remaining = limit - len(all_issues)
                current_batch = min(batch_size, remaining)
            else:
                current_batch = batch_size
            
            # Build request payload - include extra fields if dumping
            fields_to_fetch = ['summary', 'status', 'issuetype', 'created', 'updated', 'assignee', 'priority', 'project', 'fixVersions', 'versions']
            if dump_file:
                fields_to_fetch.extend(['reporter', 'resolutiondate'])
            
            payload = {
                'jql': jql_query,
                'maxResults': current_batch,
                'fields': fields_to_fetch
            }
            if next_page_token:
                payload['nextPageToken'] = next_page_token
            
            # Retry logic for rate limiting
            for retry in range(max_retries):
                response = requests.post(
                    f'{JIRA_URL}/rest/api/3/search/jql',
                    auth=(email, api_token),
                    headers={
                        'Content-Type': 'application/json',
                        'Accept': 'application/json'
                    },
                    json=payload
                )
                
                if response.status_code == 429:
                    retry_after = int(response.headers.get('Retry-After', 5))
                    log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                    time.sleep(retry_after)
                    continue
                break
            
            if response.status_code != 200:
                log.error(f'API request failed: {response.status_code} - {response.text}')
                raise Exception(f'Jira API error: {response.status_code} - {response.text}')
            
            data = response.json()
            issues = data.get('issues', [])
            all_issues.extend(issues)
            
            next_page_token = data.get('nextPageToken')
            log.debug(f'Retrieved {len(all_issues)} issues so far...')
            
            # Check for next page token or limit reached
            if not next_page_token:
                break
            if limit and len(all_issues) >= limit:
                break
        
        # Trim to limit if we went over
        if limit and len(all_issues) > limit:
            all_issues = all_issues[:limit]
        
        log.debug(f'Retrieved {len(all_issues)} issues total')
        
        # Print results
        output('')
        output('=' * 130)
        output(f'JQL Query Results')
        output('=' * 130)
        output(f'Query: {jql_query}')
        if limit:
            output(f'Limit: {limit}')
        print_ticket_table_header()
        
        for issue in all_issues:
            print_ticket_row(issue)
        
        print_ticket_table_footer(len(all_issues))
        
        # Dump to file if requested
        if dump_file:
            dump_tickets_to_file(all_issues, dump_file, dump_format)
        
    except Exception as e:
        log.error(f'Failed to run JQL query: {e}')
        raise


# ****************************************************************************************
# Dashboard Management Functions
# ****************************************************************************************

def list_dashboards(jira, owner=None, shared=False):
    '''
    List accessible dashboards.

    Input:
        jira: JIRA object with active connection.
        owner: Filter by owner username/email (use "me" for current user), or None for all.
        shared: If True, show only dashboards shared with current user.

    Output:
        None; prints dashboard list to stdout.

    Raises:
        JiraDashboardError: If the API request fails.
    '''
    log.debug(f'Entering list_dashboards(owner={owner}, shared={shared})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # Use dashboard/search endpoint for filtering capabilities
        # Build query parameters
        params = {'maxResults': 100}
        
        if owner:
            # Handle "me" as current user
            if owner.lower() == 'me':
                params['accountId'] = 'me'
            else:
                # Try to find user by email or display name
                params['owner'] = owner
        
        if shared:
            # Filter to show only shared dashboards
            params['filter'] = 'sharedWithMe'
        
        all_dashboards = []
        start_at = 0
        max_retries = 5
        
        while True:
            params['startAt'] = start_at
            
            for retry in range(max_retries):
                response = requests.get(
                    f'{JIRA_URL}/rest/api/3/dashboard/search',
                    auth=(email, api_token),
                    headers={'Accept': 'application/json'},
                    params=params
                )
                
                if response.status_code == 429:
                    retry_after = int(response.headers.get('Retry-After', 5))
                    log.warning(f'Rate limited. Waiting {retry_after} seconds (retry {retry + 1}/{max_retries})...')
                    time.sleep(retry_after)
                    continue
                break
            
            if response.status_code != 200:
                log.error(f'Dashboard API request failed: {response.status_code} - {response.text}')
                raise JiraDashboardError(f'API error: {response.status_code} - {response.text}')
            
            data = response.json()
            dashboards = data.get('values', [])
            all_dashboards.extend(dashboards)
            
            # Check if there are more pages
            total = data.get('total', 0)
            if start_at + len(dashboards) >= total:
                break
            start_at += len(dashboards)
        
        log.debug(f'Retrieved {len(all_dashboards)} dashboards')
        
        # Print results
        output('')
        output('=' * 130)
        output('Accessible Dashboards')
        if owner:
            output(f'Owner filter: {owner}')
        if shared:
            output('Showing: Dashboards shared with current user')
        output('=' * 130)
        
        print_dashboard_table_header()
        
        for dashboard in all_dashboards:
            print_dashboard_row(dashboard)
        
        print_dashboard_table_footer(len(all_dashboards))
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to list dashboards: {e}')
        raise JiraDashboardError(str(e))


def get_dashboard(jira, dashboard_id):
    '''
    Get details of a specific dashboard by ID.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The dashboard ID.

    Output:
        None; prints dashboard details to stdout.

    Raises:
        JiraDashboardError: If the dashboard is not found or API request fails.
    '''
    log.debug(f'Entering get_dashboard(dashboard_id={dashboard_id})')
    
    try:
        email, api_token = get_jira_credentials()
        
        response = requests.get(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}',
            auth=(email, api_token),
            headers={'Accept': 'application/json'}
        )
        
        if response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} not found')
        
        if response.status_code != 200:
            log.error(f'Dashboard API request failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'API error: {response.status_code} - {response.text}')
        
        dashboard = response.json()
        log.debug(f'Retrieved dashboard: {dashboard.get("name")}')
        
        # Print detailed dashboard info
        output('')
        output('=' * 80)
        output(f'Dashboard Details: {dashboard.get("name", "N/A")}')
        output('=' * 80)
        output(f'ID:          {dashboard.get("id", "N/A")}')
        output(f'Name:        {dashboard.get("name", "N/A")}')
        output(f'Description: {dashboard.get("description", "N/A") or "N/A"}')
        
        owner = dashboard.get('owner', {})
        output(f'Owner:       {owner.get("displayName", "N/A") if owner else "N/A"}')
        
        output(f'Favourite:   {"Yes" if dashboard.get("isFavourite", False) else "No"}')
        output(f'View URL:    {dashboard.get("view", "N/A")}')
        
        # Share permissions
        share_permissions = dashboard.get('sharePermissions', [])
        if share_permissions:
            output('')
            output('Share Permissions:')
            for perm in share_permissions:
                perm_type = perm.get('type', 'unknown')
                if perm_type == 'global':
                    output(f'  - Global (all users)')
                elif perm_type == 'project':
                    project = perm.get('project', {})
                    output(f'  - Project: {project.get("name", "N/A")} ({project.get("key", "N/A")})')
                elif perm_type == 'group':
                    group = perm.get('group', {})
                    output(f'  - Group: {group.get("name", "N/A")}')
                elif perm_type == 'user':
                    user = perm.get('user', {})
                    output(f'  - User: {user.get("displayName", "N/A")}')
                else:
                    output(f'  - {perm_type}')
        else:
            output('')
            output('Share Permissions: Private (not shared)')
        
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to get dashboard: {e}')
        raise JiraDashboardError(str(e))


def create_dashboard(jira, name, description=None, share_permissions=None):
    '''
    Create a new dashboard.

    Input:
        jira: JIRA object with active connection.
        name: Name for the new dashboard.
        description: Optional description for the dashboard.
        share_permissions: Optional list of share permission dicts, or JSON string.

    Output:
        None; prints created dashboard details to stdout.

    Raises:
        JiraDashboardError: If the dashboard creation fails.
    '''
    log.debug(f'Entering create_dashboard(name={name}, description={description}, share_permissions={share_permissions})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # Build request payload
        payload = {'name': name}
        
        if description:
            payload['description'] = description
        
        # Parse share_permissions if provided as JSON string
        if share_permissions:
            if isinstance(share_permissions, str):
                try:
                    payload['sharePermissions'] = json.loads(share_permissions)
                except json.JSONDecodeError as e:
                    raise JiraDashboardError(f'Invalid JSON for share-permissions: {e}')
            else:
                payload['sharePermissions'] = share_permissions
        else:
            # Default to private (empty share permissions)
            payload['sharePermissions'] = []
        
        log.debug(f'Create dashboard payload: {payload}')
        
        response = requests.post(
            f'{JIRA_URL}/rest/api/3/dashboard',
            auth=(email, api_token),
            headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
            json=payload
        )
        
        if response.status_code not in [200, 201]:
            log.error(f'Dashboard create failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'Failed to create dashboard: {response.status_code} - {response.text}')
        
        dashboard = response.json()
        log.info(f'Created dashboard: {dashboard.get("id")} - {dashboard.get("name")}')
        
        # Print success message
        output('')
        output('=' * 80)
        output('Dashboard Created Successfully')
        output('=' * 80)
        output(f'ID:          {dashboard.get("id", "N/A")}')
        output(f'Name:        {dashboard.get("name", "N/A")}')
        output(f'Description: {dashboard.get("description", "N/A") or "N/A"}')
        output(f'View URL:    {dashboard.get("view", "N/A")}')
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to create dashboard: {e}')
        raise JiraDashboardError(str(e))


def update_dashboard(jira, dashboard_id, name=None, description=None, share_permissions=None):
    '''
    Update an existing dashboard.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The dashboard ID to update.
        name: New name for the dashboard, or None to keep existing.
        description: New description, or None to keep existing.
        share_permissions: New share permissions as list or JSON string, or None to keep existing.

    Output:
        None; prints updated dashboard details to stdout.

    Raises:
        JiraDashboardError: If the dashboard update fails.
    '''
    log.debug(f'Entering update_dashboard(dashboard_id={dashboard_id}, name={name}, description={description})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # First, get the current dashboard to preserve unchanged fields
        response = requests.get(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}',
            auth=(email, api_token),
            headers={'Accept': 'application/json'}
        )
        
        if response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} not found')
        
        if response.status_code != 200:
            raise JiraDashboardError(f'Failed to get dashboard: {response.status_code} - {response.text}')
        
        current = response.json()
        
        # Build update payload - name is required for PUT
        payload = {
            'name': name if name else current.get('name'),
            'sharePermissions': current.get('sharePermissions', [])
        }
        
        if description is not None:
            payload['description'] = description
        elif current.get('description'):
            payload['description'] = current.get('description')
        
        # Parse share_permissions if provided
        if share_permissions:
            if isinstance(share_permissions, str):
                try:
                    payload['sharePermissions'] = json.loads(share_permissions)
                except json.JSONDecodeError as e:
                    raise JiraDashboardError(f'Invalid JSON for share-permissions: {e}')
            else:
                payload['sharePermissions'] = share_permissions
        
        log.debug(f'Update dashboard payload: {payload}')
        
        response = requests.put(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}',
            auth=(email, api_token),
            headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
            json=payload
        )
        
        if response.status_code not in [200, 204]:
            log.error(f'Dashboard update failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'Failed to update dashboard: {response.status_code} - {response.text}')
        
        dashboard = response.json() if response.status_code == 200 else payload
        log.info(f'Updated dashboard: {dashboard_id}')
        
        # Print success message
        output('')
        output('=' * 80)
        output('Dashboard Updated Successfully')
        output('=' * 80)
        output(f'ID:          {dashboard_id}')
        output(f'Name:        {dashboard.get("name", "N/A")}')
        output(f'Description: {dashboard.get("description", "N/A") or "N/A"}')
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to update dashboard: {e}')
        raise JiraDashboardError(str(e))


def delete_dashboard(jira, dashboard_id, force=False):
    '''
    Delete a dashboard.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The dashboard ID to delete.
        force: If True, skip confirmation prompt.

    Output:
        None; prints deletion confirmation to stdout.

    Raises:
        JiraDashboardError: If the dashboard deletion fails.
    '''
    log.debug(f'Entering delete_dashboard(dashboard_id={dashboard_id}, force={force})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # First, get the dashboard to show what will be deleted
        response = requests.get(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}',
            auth=(email, api_token),
            headers={'Accept': 'application/json'}
        )
        
        if response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} not found')
        
        if response.status_code != 200:
            raise JiraDashboardError(f'Failed to get dashboard: {response.status_code} - {response.text}')
        
        dashboard = response.json()
        dashboard_name = dashboard.get('name', 'Unknown')
        
        # Confirm deletion unless force is True
        if not force:
            output('')
            output(f'WARNING: About to delete dashboard "{dashboard_name}" (ID: {dashboard_id})')
            output('This action cannot be undone.')
            output('')
            output('Use --force to skip this confirmation.')
            output('')
            return
        
        # Perform deletion
        response = requests.delete(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}',
            auth=(email, api_token),
            headers={'Accept': 'application/json'}
        )
        
        if response.status_code not in [200, 204]:
            log.error(f'Dashboard delete failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'Failed to delete dashboard: {response.status_code} - {response.text}')
        
        log.info(f'Deleted dashboard: {dashboard_id} - {dashboard_name}')
        
        # Print success message
        output('')
        output('=' * 80)
        output('Dashboard Deleted Successfully')
        output('=' * 80)
        output(f'ID:   {dashboard_id}')
        output(f'Name: {dashboard_name}')
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to delete dashboard: {e}')
        raise JiraDashboardError(str(e))


def copy_dashboard(jira, dashboard_id, name, description=None, share_permissions=None):
    '''
    Copy/clone an existing dashboard.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The source dashboard ID to copy.
        name: Name for the new dashboard copy.
        description: Optional description for the copy, or None to copy from source.
        share_permissions: Optional share permissions for the copy, or None for private.

    Output:
        None; prints created dashboard details to stdout.

    Raises:
        JiraDashboardError: If the dashboard copy fails.
    '''
    log.debug(f'Entering copy_dashboard(dashboard_id={dashboard_id}, name={name}, description={description})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # Build request payload
        payload = {'name': name}
        
        if description:
            payload['description'] = description
        
        # Parse share_permissions if provided
        if share_permissions:
            if isinstance(share_permissions, str):
                try:
                    payload['sharePermissions'] = json.loads(share_permissions)
                except json.JSONDecodeError as e:
                    raise JiraDashboardError(f'Invalid JSON for share-permissions: {e}')
            else:
                payload['sharePermissions'] = share_permissions
        else:
            # Default to private
            payload['sharePermissions'] = []
        
        log.debug(f'Copy dashboard payload: {payload}')
        
        response = requests.post(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}/copy',
            auth=(email, api_token),
            headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
            json=payload
        )
        
        if response.status_code not in [200, 201]:
            log.error(f'Dashboard copy failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'Failed to copy dashboard: {response.status_code} - {response.text}')
        
        dashboard = response.json()
        log.info(f'Copied dashboard {dashboard_id} to new dashboard: {dashboard.get("id")} - {dashboard.get("name")}')
        
        # Print success message
        output('')
        output('=' * 80)
        output('Dashboard Copied Successfully')
        output('=' * 80)
        output(f'Source ID:   {dashboard_id}')
        output(f'New ID:      {dashboard.get("id", "N/A")}')
        output(f'Name:        {dashboard.get("name", "N/A")}')
        output(f'Description: {dashboard.get("description", "N/A") or "N/A"}')
        output(f'View URL:    {dashboard.get("view", "N/A")}')
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to copy dashboard: {e}')
        raise JiraDashboardError(str(e))


# ****************************************************************************************
# Gadget Management Functions
# ****************************************************************************************

def list_gadgets(jira, dashboard_id):
    '''
    List gadgets on a dashboard.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The dashboard ID.

    Output:
        None; prints gadget list to stdout.

    Raises:
        JiraDashboardError: If the API request fails.
    '''
    log.debug(f'Entering list_gadgets(dashboard_id={dashboard_id})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # First get dashboard info for display
        dash_response = requests.get(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}',
            auth=(email, api_token),
            headers={'Accept': 'application/json'}
        )
        
        if dash_response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} not found')
        
        dashboard_name = 'Unknown'
        if dash_response.status_code == 200:
            dashboard_name = dash_response.json().get('name', 'Unknown')
        
        # Get gadgets
        response = requests.get(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}/gadget',
            auth=(email, api_token),
            headers={'Accept': 'application/json'}
        )
        
        if response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} not found')
        
        if response.status_code != 200:
            log.error(f'Gadget API request failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'API error: {response.status_code} - {response.text}')
        
        data = response.json()
        gadgets = data.get('gadgets', [])
        log.debug(f'Retrieved {len(gadgets)} gadgets')
        
        # Print results
        output('')
        output('=' * 120)
        output(f'Gadgets on Dashboard: {dashboard_name} (ID: {dashboard_id})')
        output('=' * 120)
        
        print_gadget_table_header()
        
        for gadget in gadgets:
            print_gadget_row(gadget)
        
        print_gadget_table_footer(len(gadgets))
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to list gadgets: {e}')
        raise JiraDashboardError(str(e))


def add_gadget(jira, dashboard_id, module_key, position=None, color=None, properties=None):
    '''
    Add a gadget to a dashboard.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The dashboard ID.
        module_key: The gadget module key (e.g., 'com.atlassian.jira.gadgets:filter-results-gadget').
        position: Optional position as 'row,column' string.
        color: Optional gadget color (blue, red, yellow, green, cyan, purple, gray, white).
        properties: Optional gadget properties as dict or JSON string.

    Output:
        None; prints added gadget details to stdout.

    Raises:
        JiraDashboardError: If the gadget addition fails.
    '''
    log.debug(f'Entering add_gadget(dashboard_id={dashboard_id}, module_key={module_key}, position={position}, color={color})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # Build request payload
        payload = {'moduleKey': module_key}
        
        # Parse position if provided
        if position:
            try:
                parts = position.split(',')
                if len(parts) != 2:
                    raise ValueError('Position must be in format row,column')
                row = int(parts[0].strip())
                col = int(parts[1].strip())
                payload['position'] = {'row': row, 'column': col}
            except ValueError as e:
                raise JiraDashboardError(f'Invalid position format: {e}')
        
        # Validate and set color
        valid_colors = ['blue', 'red', 'yellow', 'green', 'cyan', 'purple', 'gray', 'white']
        if color:
            color_lower = color.lower()
            if color_lower not in valid_colors:
                raise JiraDashboardError(f'Invalid color "{color}". Valid colors: {", ".join(valid_colors)}')
            payload['color'] = color_lower
        
        # Parse properties if provided
        if properties:
            if isinstance(properties, str):
                try:
                    payload['properties'] = json.loads(properties)
                except json.JSONDecodeError as e:
                    raise JiraDashboardError(f'Invalid JSON for gadget-properties: {e}')
            else:
                payload['properties'] = properties
        
        log.debug(f'Add gadget payload: {payload}')
        
        response = requests.post(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}/gadget',
            auth=(email, api_token),
            headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
            json=payload
        )
        
        if response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} not found')
        
        if response.status_code not in [200, 201]:
            log.error(f'Add gadget failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'Failed to add gadget: {response.status_code} - {response.text}')
        
        gadget = response.json()
        log.info(f'Added gadget {gadget.get("id")} to dashboard {dashboard_id}')
        
        # Print success message
        output('')
        output('=' * 80)
        output('Gadget Added Successfully')
        output('=' * 80)
        output(f'Dashboard ID: {dashboard_id}')
        output(f'Gadget ID:    {gadget.get("id", "N/A")}')
        output(f'Module Key:   {gadget.get("moduleKey", "N/A")}')
        output(f'Title:        {gadget.get("title", "N/A")}')
        
        pos = gadget.get('position', {})
        output(f'Position:     {pos.get("row", 0)},{pos.get("column", 0)}')
        output(f'Color:        {gadget.get("color", "N/A")}')
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to add gadget: {e}')
        raise JiraDashboardError(str(e))


def remove_gadget(jira, dashboard_id, gadget_id):
    '''
    Remove a gadget from a dashboard.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The dashboard ID.
        gadget_id: The gadget ID to remove.

    Output:
        None; prints removal confirmation to stdout.

    Raises:
        JiraDashboardError: If the gadget removal fails.
    '''
    log.debug(f'Entering remove_gadget(dashboard_id={dashboard_id}, gadget_id={gadget_id})')
    
    try:
        email, api_token = get_jira_credentials()
        
        response = requests.delete(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}/gadget/{gadget_id}',
            auth=(email, api_token),
            headers={'Accept': 'application/json'}
        )
        
        if response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} or gadget {gadget_id} not found')
        
        if response.status_code not in [200, 204]:
            log.error(f'Remove gadget failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'Failed to remove gadget: {response.status_code} - {response.text}')
        
        log.info(f'Removed gadget {gadget_id} from dashboard {dashboard_id}')
        
        # Print success message
        output('')
        output('=' * 80)
        output('Gadget Removed Successfully')
        output('=' * 80)
        output(f'Dashboard ID: {dashboard_id}')
        output(f'Gadget ID:    {gadget_id}')
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to remove gadget: {e}')
        raise JiraDashboardError(str(e))


def update_gadget(jira, dashboard_id, gadget_id, position=None, color=None):
    '''
    Update a gadget on a dashboard.

    Input:
        jira: JIRA object with active connection.
        dashboard_id: The dashboard ID.
        gadget_id: The gadget ID to update.
        position: New position as 'row,column' string, or None to keep existing.
        color: New color, or None to keep existing.

    Output:
        None; prints updated gadget details to stdout.

    Raises:
        JiraDashboardError: If the gadget update fails.
    '''
    log.debug(f'Entering update_gadget(dashboard_id={dashboard_id}, gadget_id={gadget_id}, position={position}, color={color})')
    
    try:
        email, api_token = get_jira_credentials()
        
        # Build request payload
        payload = {}
        
        # Parse position if provided
        if position:
            try:
                parts = position.split(',')
                if len(parts) != 2:
                    raise ValueError('Position must be in format row,column')
                row = int(parts[0].strip())
                col = int(parts[1].strip())
                payload['position'] = {'row': row, 'column': col}
            except ValueError as e:
                raise JiraDashboardError(f'Invalid position format: {e}')
        
        # Validate and set color
        valid_colors = ['blue', 'red', 'yellow', 'green', 'cyan', 'purple', 'gray', 'white']
        if color:
            color_lower = color.lower()
            if color_lower not in valid_colors:
                raise JiraDashboardError(f'Invalid color "{color}". Valid colors: {", ".join(valid_colors)}')
            payload['color'] = color_lower
        
        if not payload:
            raise JiraDashboardError('No updates specified. Use --position or --color.')
        
        log.debug(f'Update gadget payload: {payload}')
        
        response = requests.put(
            f'{JIRA_URL}/rest/api/3/dashboard/{dashboard_id}/gadget/{gadget_id}',
            auth=(email, api_token),
            headers={'Content-Type': 'application/json', 'Accept': 'application/json'},
            json=payload
        )
        
        if response.status_code == 404:
            raise JiraDashboardError(f'Dashboard {dashboard_id} or gadget {gadget_id} not found')
        
        if response.status_code not in [200, 204]:
            log.error(f'Update gadget failed: {response.status_code} - {response.text}')
            raise JiraDashboardError(f'Failed to update gadget: {response.status_code} - {response.text}')
        
        log.info(f'Updated gadget {gadget_id} on dashboard {dashboard_id}')
        
        # Print success message
        output('')
        output('=' * 80)
        output('Gadget Updated Successfully')
        output('=' * 80)
        output(f'Dashboard ID: {dashboard_id}')
        output(f'Gadget ID:    {gadget_id}')
        if position:
            output(f'Position:     {position}')
        if color:
            output(f'Color:        {color}')
        output('=' * 80)
        output('')
        
    except JiraDashboardError:
        raise
    except Exception as e:
        log.error(f'Failed to update gadget: {e}')
        raise JiraDashboardError(str(e))


# ****************************************************************************************
# Handle the arguments
# ****************************************************************************************
def handle_args():
    '''
    Parse CLI arguments and configure console logging handlers.

    Input:
        None directly; reads flags from sys.argv.

    Output:
        argparse.Namespace containing parsed arguments.

    Side Effects:
        Attaches a stream handler to the module logger with formatting and
        level derived from the parsed arguments.
    '''
    log.debug('Entering handle_args()')
    
    parser = argparse.ArgumentParser(
        description='Jira utilities for Cornelis Networks',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Credentials Setup:
  Set the following environment variables before running:
    export JIRA_EMAIL="your.email@cornelisnetworks.com"
    export JIRA_API_TOKEN="your_api_token_here"
  
  Generate an API token at:
    https://id.atlassian.com/manage-profile/security/api-tokens

Examples:
  %(prog)s --list                              List all Jira projects
  %(prog)s --project PROJ --get-workflow       Show workflow statuses for PROJ
  %(prog)s --project PROJ --get-issue-types    Show issue types for PROJ
  %(prog)s --project PROJ --get-fields         Show create/edit/transition fields for all types
  %(prog)s --project PROJ --get-fields --issue-types Bug Task
                                               Show fields for Bug and Task only
  %(prog)s --project PROJ --get-versions       Show versions/releases for PROJ
  %(prog)s --project PROJ --total              Show total ticket count for PROJ
  %(prog)s --project PROJ --total --issue-types Bug Task
                                               Show count for specific issue types
  %(prog)s --project PROJ --total --status Open
                                               Show count for Open tickets
  %(prog)s --project PROJ --total --issue-types Bug --status Open "In Progress"
                                               Show count for Open/In Progress Bugs
  %(prog)s --project PROJ --total --date week  Show count for tickets created this week
  %(prog)s --project PROJ --get-tickets        Get all tickets for PROJ
  %(prog)s --project PROJ --get-tickets --issue-types Bug --status Open --limit 50
                                               Get up to 50 Open Bugs
  %(prog)s --project PROJ --get-tickets --status Closed --date month
                                               Get Closed tickets from this month
  %(prog)s --project PROJ --get-tickets --date 01-01-2024:12-31-2024
                                               Get tickets created in date range
  %(prog)s --project PROJ --get-workflow --get-issue-types --get-fields --get-versions
                                               Show all info for PROJ
  %(prog)s --jql "project = PROJ AND status = Open"
                                               Run a custom JQL query
  %(prog)s --jql "assignee = currentUser() AND status != Done" --limit 20
                                               Run JQL with limit
  %(prog)s --project PROJ --get-tickets --dump-file tickets
                                               Dump all tickets to tickets.csv
  %(prog)s --project PROJ --get-tickets --issue-types Bug --status Open --dump-file bugs --dump-format json
                                               Dump Open Bugs to bugs.json
  %(prog)s --jql "project = PROJ" --dump-file results --dump-format csv
                                               Dump JQL results to results.csv
  %(prog)s --project PROJ --releases           List all releases for PROJ
  %(prog)s --project PROJ --releases "12.*"    List releases matching pattern
  %(prog)s --project PROJ --releases "12.*" --get-tickets
                                               Get tickets for all releases matching "12.*"
  %(prog)s --project PROJ --releases "12.*" --get-tickets --status Open --dump-file 12x_tickets
                                               Get Open tickets for 12.x releases, dump to file
  %(prog)s --project PROJ --release-tickets "v1.0"
                                               Get all tickets for release v1.0
  %(prog)s --project PROJ --release-tickets "12.1*" --issue-types Bug Story Task
                                               Get Bugs, Stories, Tasks for releases matching 12.1*
  %(prog)s --project PROJ --release-tickets "v1.0" --status Open --limit 50
                                               Get Open tickets for release v1.0
  %(prog)s --project PROJ --no-release         Get tickets with no release assigned
  %(prog)s --project PROJ --no-release --issue-types Bug --status Open
                                               Get Open Bugs with no release

Bulk Update Examples:
  # Step 1: Find tickets and dump to CSV
  %(prog)s --jql "project = PROJ AND fixVersion is EMPTY" --dump-file orphans
  
  # Step 2: Preview bulk update (dry-run is default)
  %(prog)s --bulk-update --input-file orphans.csv --set-release "12.1.1.x"
  
  # Step 3: Execute bulk update
  %(prog)s --bulk-update --input-file orphans.csv --set-release "12.1.1.x" --execute
  
  # Other bulk operations:
  %(prog)s --bulk-update --input-file tickets.csv --transition "Closed" --execute
  %(prog)s --bulk-update --input-file tickets.csv --assign "user@email.com" --execute
  %(prog)s --bulk-update --input-file tickets.csv --remove-release --execute
  %(prog)s --bulk-update --input-file tickets.csv --set-release "v2.0" --max-updates 10 --execute

Bulk Delete Examples:
  # Step 1: Find tickets and dump to CSV
  %(prog)s --jql "project = PROJ AND labels = temp" --dump-file to_delete

  # Step 2: Preview deletes (dry-run is default)
  %(prog)s --bulk-delete --input-file to_delete.csv

  # Step 3: Execute deletes (requires explicit DELETE confirmation)
  %(prog)s --bulk-delete --input-file to_delete.csv --execute

  # Delete parent issues and their subtasks
  %(prog)s --bulk-delete --input-file parents.csv --delete-subtasks --execute

  # Skip interactive confirmation prompt (DANGEROUS)
  %(prog)s --bulk-delete --input-file to_delete.csv --execute --force

Date Filters:
  today                    Tickets created today
  week                     Tickets created in the last 7 days
  month                    Tickets created in the last 30 days
  year                     Tickets created in the last 365 days
  all                      All tickets (no date filter)
  MM-DD-YYYY:MM-DD-YYYY    Tickets created within date range
        ''')
    parser.add_argument(
        '-v',
        '--verbose',
        action='store_true',
        help='Enable verbose output to stdout.')
    parser.add_argument(
        '-q',
        '--quiet',
        action='store_true',
        help='Minimal stdout.')
    parser.add_argument(
        '--env',
        type=str,
        default='.env',
        metavar='FILE',
        help='Path to dotenv file to load (default: .env). Use this to load a different env file at runtime.')
    parser.add_argument(
        '--list',
        action='store_true',
        dest='list_projects',
        help='List all available Jira projects.')
    parser.add_argument(
        '--project',
        type=str,
        metavar='KEY',
        help='Project key to operate on (e.g., PROJ).')
    parser.add_argument(
        '--get-workflow',
        action='store_true',
        dest='get_workflow',
        help='Display workflow statuses for the specified project.')
    parser.add_argument(
        '--get-issue-types',
        action='store_true',
        dest='get_issue_types',
        help='Display issue types for the specified project.')
    parser.add_argument(
        '--get-fields',
        nargs='*',
        metavar='TYPE',
        dest='get_fields',
        help='Display create/edit/transition fields. Optionally filter by issue types.')
    parser.add_argument(
        '--get-versions',
        action='store_true',
        dest='get_versions',
        help='Display versions (releases) with detailed info for the specified project.')
    parser.add_argument(
        '--get-components',
        action='store_true',
        dest='get_components',
        help='List all components for the specified project.')
    parser.add_argument(
        '--get-children',
        type=str,
        metavar='KEY',
        dest='get_children',
        help='Display full child hierarchy for the given ticket (recurses via parent links).')
    parser.add_argument(
        '--get-related',
        type=str,
        metavar='KEY',
        dest='get_related',
        help='Display related issues for the given ticket (linked issues + children); use --hierarchy [DEPTH] to recurse (DEPTH optional).')
    parser.add_argument(
        '--hierarchy',
        nargs='?',
        const=-1,
        type=int,
        metavar='DEPTH',
        dest='hierarchy',
        help='When used with --get-related, recursively traverse linked issues and children to build a hierarchy. Optional DEPTH limits traversal depth (1 = direct only). Omit DEPTH for unlimited depth.')
    parser.add_argument(
        '--table-format',
        type=str,
        choices=['flat', 'indented'],
        default='flat',
        metavar='FORMAT',
        dest='table_format',
        help='Table layout for hierarchy CSV output (default: flat). '
             '"flat" uses a depth column. '
             '"indented" replaces the depth column with per-level columns (Depth 0, Depth 1, ...) '
             'where the ticket key appears in the column matching its depth. '
             'Applies to --get-children and --get-related.')
    parser.add_argument(
        '--releases',
        nargs='?',
        const='*',
        default=None,
        metavar='PATTERN',
        help='List releases for the specified project. Optional glob pattern with exclusions using ^ (e.g., "12.*" or "12.*,^*Samples*").')
    parser.add_argument(
        '--release-tickets',
        type=str,
        metavar='RELEASE',
        dest='release_tickets',
        help='Get tickets for a release. Supports glob patterns (e.g., "12.1*"). Use --issue-types to filter by issue types.')
    parser.add_argument(
        '--issue-types',
        nargs='+',
        metavar='TYPE',
        dest='issue_types',
        help='Filter by issue types (e.g., Bug Story Task Sub-task). Used with --release-tickets, --no-release, --total, --get-tickets.')
    parser.add_argument(
        '--no-release',
        action='store_true',
        dest='no_release',
        help='Get tickets with no release assigned. Use --issue-types to filter by issue types.')
    parser.add_argument(
        '--total',
        action='store_true',
        help='Show ticket count. Use --issue-types to filter by specific issue types.')
    parser.add_argument(
        '--get-tickets',
        action='store_true',
        dest='get_tickets',
        help='Get tickets. Use --issue-types to filter by specific issue types.')
    parser.add_argument(
        '--status',
        nargs='+',
        metavar='STATUS',
        help='Filter by status (case-insensitive). Can specify multiple statuses. Use ^ prefix to exclude (e.g., ^Closed to exclude Closed).')
    parser.add_argument(
        '--date',
        type=str,
        metavar='FILTER',
        help='Date filter: today, week, month, year, all, or MM-DD-YYYY:MM-DD-YYYY range.')
    parser.add_argument(
        '--limit',
        type=int,
        metavar='N',
        help='Limit the number of tickets to retrieve.')
    parser.add_argument(
        '--jql',
        type=str,
        metavar='QUERY',
        help='Run a custom JQL query and display results.')

    # Ticket creation arguments
    parser.add_argument(
        '--create-ticket',
        nargs='?',
        const='',
        default=None,
        dest='create_ticket',
        metavar='FILE',
        help='Create a new ticket. If FILE is provided, load ticket fields from JSON. If omitted, use CLI flags. Dry-run by default; use --execute to actually create.')
    parser.add_argument(
        '--summary',
        type=str,
        metavar='TEXT',
        help='Ticket summary (required with --create-ticket).')
    parser.add_argument(
        '--issue-type',
        type=str,
        dest='issue_type',
        metavar='TYPE',
        help='Issue type name (required with --create-ticket), e.g., Task, Bug, Story.')
    parser.add_argument(
        '--ticket-description',
        type=str,
        dest='ticket_description',
        metavar='TEXT',
        help='Plain-text description to set on the ticket (optional).')
    parser.add_argument(
        '--assignee-id',
        type=str,
        dest='assignee_id',
        metavar='ACCOUNT_ID',
        help='Assignee accountId (optional).')
    parser.add_argument(
        '--components',
        nargs='+',
        metavar='NAME',
        help='Component names to set (optional).')
    parser.add_argument(
        '--fix-versions',
        nargs='+',
        dest='fix_versions',
        metavar='VERSION',
        help='Fix version(s) to set (optional).')
    parser.add_argument(
        '--labels',
        nargs='+',
        metavar='LABEL',
        help='Labels to set (optional).')
    parser.add_argument(
        '--parent',
        type=str,
        metavar='KEY',
        help='Parent ticket key (optional; for sub-tasks or parent-linked issue types).')

    parser.add_argument(
        '--dump-file',
        type=str,
        metavar='FILE',
        dest='dump_file',
        help='Output filename for dumping tickets (default: "out" with appropriate extension).')
    parser.add_argument(
        '--dump-format',
        type=str,
        choices=['csv', 'json', 'excel'],
        default='csv',
        dest='dump_format',
        metavar='FORMAT',
        help='Output format for dump: csv, json, or excel (default: csv). '
             'Excel format produces .xlsx files with ticket IDs as clickable Jira hyperlinks.')
    
    # Bulk update arguments
    parser.add_argument(
        '--bulk-update',
        action='store_true',
        dest='bulk_update',
        help='Perform bulk update on tickets from input file.')

    # Bulk delete arguments
    parser.add_argument(
        '--bulk-delete',
        action='store_true',
        dest='bulk_delete',
        help='Perform bulk delete on tickets from input file (dry-run by default).')

    parser.add_argument(
        '--input-file',
        type=str,
        metavar='FILE',
        dest='input_file',
        help='Input CSV file containing ticket keys for bulk update/delete.')
    parser.add_argument(
        '--set-release',
        type=str,
        metavar='RELEASE',
        dest='set_release',
        help='Set the release/fixVersion on tickets.')
    parser.add_argument(
        '--remove-release',
        action='store_true',
        dest='remove_release',
        help='Remove all releases from tickets.')
    parser.add_argument(
        '--transition',
        type=str,
        metavar='STATUS',
        help='Transition tickets to the specified status.')
    parser.add_argument(
        '--assign',
        type=str,
        metavar='USER',
        help='Assign tickets to the specified user (email or "unassigned").')
    parser.add_argument(
        '--dry-run',
        action='store_true',
        dest='dry_run',
        default=True,
        help='Preview changes without applying them (default: True).')
    parser.add_argument(
        '--execute',
        action='store_true',
        help='Execute the bulk update (disables dry-run).')
    parser.add_argument(
        '--max-updates',
        type=int,
        metavar='N',
        dest='max_updates',
        help='Maximum number of tickets to update in bulk operation.')

    parser.add_argument(
        '--max-deletes',
        type=int,
        metavar='N',
        dest='max_deletes',
        help='Maximum number of tickets to delete in bulk operation.')

    parser.add_argument(
        '--delete-subtasks',
        action='store_true',
        dest='delete_subtasks',
        help='When used with --bulk-delete, delete subtasks as well.')

    parser.add_argument(
        '--show-jql',
        action='store_true',
        dest='show_jql',
        help='Print the equivalent JQL statement for the operation.')
    
    # Dashboard management arguments
    parser.add_argument(
        '--dashboards',
        action='store_true',
        help='List accessible dashboards.')
    parser.add_argument(
        '--dashboard',
        type=str,
        metavar='ID',
        help='Get dashboard details by ID.')
    parser.add_argument(
        '--owner',
        type=str,
        metavar='USER',
        help='Filter dashboards by owner (use "me" for current user).')
    parser.add_argument(
        '--shared',
        action='store_true',
        help='Show only dashboards shared with current user.')
    parser.add_argument(
        '--create-dashboard',
        type=str,
        metavar='NAME',
        dest='create_dashboard',
        help='Create a new dashboard with the specified name.')
    parser.add_argument(
        '--update-dashboard',
        type=str,
        metavar='ID',
        dest='update_dashboard',
        help='Update dashboard by ID.')
    parser.add_argument(
        '--delete-dashboard',
        type=str,
        metavar='ID',
        dest='delete_dashboard',
        help='Delete dashboard by ID.')
    parser.add_argument(
        '--copy-dashboard',
        type=str,
        metavar='ID',
        dest='copy_dashboard',
        help='Copy/clone dashboard by ID.')
    parser.add_argument(
        '--description',
        type=str,
        metavar='TEXT',
        help='Dashboard description (for create/update/copy).')
    parser.add_argument(
        '--name',
        type=str,
        metavar='NAME',
        help='New name for dashboard (for update/copy).')
    parser.add_argument(
        '--share-permissions',
        type=str,
        metavar='JSON',
        dest='share_permissions',
        help='Share permissions as JSON array (e.g., \'[{"type":"global"}]\').')
    parser.add_argument(
        '--force',
        action='store_true',
        help='Skip confirmation prompts (for delete operations).')
    
    # Gadget management arguments
    parser.add_argument(
        '--gadgets',
        type=str,
        metavar='DASHBOARD_ID',
        help='List gadgets on the specified dashboard.')
    parser.add_argument(
        '--add-gadget',
        type=str,
        metavar='MODULE_KEY',
        dest='add_gadget',
        help='Add gadget to dashboard (requires --dashboard).')
    parser.add_argument(
        '--remove-gadget',
        type=str,
        metavar='GADGET_ID',
        dest='remove_gadget',
        help='Remove gadget from dashboard (requires --dashboard).')
    parser.add_argument(
        '--update-gadget',
        type=str,
        metavar='GADGET_ID',
        dest='update_gadget',
        help='Update gadget on dashboard (requires --dashboard).')
    parser.add_argument(
        '--position',
        type=str,
        metavar='ROW,COL',
        help='Gadget position as row,column (e.g., "0,1").')
    parser.add_argument(
        '--color',
        type=str,
        metavar='COLOR',
        help='Gadget color (blue, red, yellow, green, cyan, purple, gray, white).')
    parser.add_argument(
        '--gadget-properties',
        type=str,
        metavar='JSON',
        dest='gadget_properties',
        help='Gadget properties as JSON object.')
    
    args = parser.parse_args()

    # If user selected a non-default dotenv file, load it now.
    #
    # Note: we already load the default `.env` at import time with override=False.
    # When `--env` is provided, we reload from that file with override=True so it
    # can override values from the default `.env`.
    if args.env and args.env != '.env':
        if not os.path.exists(args.env):
            parser.error(f'--env file not found: {args.env}')
        load_dotenv(dotenv_path=args.env, override=True)

        # Refresh globals that were computed at import time.
        global JIRA_URL
        JIRA_URL = os.getenv('JIRA_URL', DEFAULT_JIRA_URL)

    # Ticket creation JSON (optional):
    #   - `--create-ticket` is always required to create a ticket
    #   - `--create-ticket FILE` loads fields from JSON
    #   - `--create-ticket` (no FILE) uses CLI flags only
    #
    # Precedence model:
    #   - CLI flags override JSON values
    #   - JSON can provide required fields so you don't have to repeat them
    args.ticket_json = {}
    args.create_ticket_path = None

    # When `--create-ticket` is specified with no FILE, argparse sets it to '' (empty string).
    # When not specified at all, it is None.
    if args.create_ticket is not None and args.create_ticket != '':
        args.create_ticket_path = args.create_ticket

        if not os.path.exists(args.create_ticket_path):
            parser.error(f'--create-ticket file not found: {args.create_ticket_path}')

        try:
            with open(args.create_ticket_path, 'r', encoding='utf-8') as f:
                ticket_json = json.load(f)
        except json.JSONDecodeError as e:
            parser.error(f'--create-ticket file is not valid JSON: {e}')
        except Exception as e:
            parser.error(f'Failed to read --create-ticket file: {e}')

        if not isinstance(ticket_json, dict):
            parser.error('--create-ticket file must contain a JSON object at the top level')

        args.ticket_json = ticket_json

        # Allow project to be specified via JSON.
        json_project = ticket_json.get('project') or ticket_json.get('project_key')
        if not args.project and json_project:
            args.project = str(json_project)

        # Fill create-ticket args from JSON (CLI overrides JSON).
        args.summary = args.summary or ticket_json.get('summary')
        args.issue_type = args.issue_type or ticket_json.get('issue_type')

        args.ticket_description = (
            args.ticket_description
            or ticket_json.get('description')
            or ticket_json.get('ticket_description')
        )

        args.assignee_id = args.assignee_id or ticket_json.get('assignee_id')

        def _normalize_str_list(value, field_name):
            """Normalize JSON field into a list[str] or None.

            Accepts either:
              - null / missing => None
              - string => [string]
              - list[string] => list[string]

            We keep this normalization in [`handle_args()`](jira_utils.py:4300) so later
            code can assume list semantics.
            """
            if value is None:
                return None
            if isinstance(value, str):
                return [value]
            if isinstance(value, list):
                try:
                    return [str(v) for v in value]
                except Exception:
                    parser.error(f'--create-ticket JSON field "{field_name}" must be a string or list of strings')
            parser.error(f'--create-ticket JSON field "{field_name}" must be a string or list of strings')

        args.components = args.components or _normalize_str_list(ticket_json.get('components'), 'components')
        args.fix_versions = args.fix_versions or _normalize_str_list(ticket_json.get('fix_versions'), 'fix_versions')
        args.labels = args.labels or _normalize_str_list(ticket_json.get('labels'), 'labels')

        args.parent = args.parent or ticket_json.get('parent') or ticket_json.get('parent_key')
        if args.parent is not None:
            args.parent = str(args.parent)

    # Configure stdout logging based on arguments
    ch = logging.StreamHandler(sys.stdout)
    if args.verbose:
        ch.setLevel(logging.DEBUG)
    elif args.quiet:
        ch.setLevel(logging.ERROR)
    else:
        ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    log.addHandler(ch)
    
    # Set quiet mode for output function
    global _quiet_mode
    _quiet_mode = args.quiet
    
    # Set show_jql mode
    global _show_jql
    _show_jql = args.show_jql
    
    log.debug(f'Checking script requirements...')
    # Check requirements to execute the script here
    if not args.verbose and not args.quiet:
        log.debug('No output level specified. Defaulting to INFO.')
    
    # Validate argument combinations
    # args.get_fields is None if not specified, [] if specified with no args, or a list if specified with args
    get_fields_specified = args.get_fields is not None
    release_tickets_specified = args.release_tickets is not None
    jql_specified = args.jql is not None
    children_specified = args.get_children is not None
    related_specified = args.get_related is not None
    create_ticket_specified = args.create_ticket is not None
    
    project_actions = [args.get_workflow, args.get_issue_types, get_fields_specified, args.get_versions, args.get_components, args.releases,
                       args.total, args.get_tickets, release_tickets_specified, args.no_release, create_ticket_specified]
    # --get-children and --get-related are allowed without --project
    if any(project_actions) and not args.project:
        parser.error('--project is required when using --get-workflow, --get-issue-types, --get-fields, --get-versions, --get-components, --releases, --total, --get-tickets, --release-tickets, --no-release, or --create-ticket')
    
    # Validate --issue-types is only used with appropriate commands
    if args.issue_types and not (args.total or args.get_tickets or release_tickets_specified or args.no_release or get_fields_specified):
        parser.error('--issue-types requires --total, --get-tickets, --release-tickets, --no-release, or --get-fields')
    
    # Validate --status is only used with appropriate commands
    if args.status and not (args.total or args.get_tickets or release_tickets_specified or args.no_release):
        parser.error('--status requires --total, --get-tickets, --release-tickets, or --no-release')
    
    # Validate --date and --limit usage
    if args.date and not (args.total or args.get_tickets or release_tickets_specified or args.no_release or args.get_components):
        parser.error('--date requires --total, --get-tickets, --release-tickets, --no-release, or --get-components')
    if args.limit and not (args.get_tickets or jql_specified or release_tickets_specified or args.no_release or args.get_children or args.get_related):
        parser.error('--limit requires --get-tickets, --jql, --release-tickets, --no-release, --get-children, or --get-related')

    # Validate ticket creation arguments
    if args.create_ticket is not None:
        if not args.summary:
            parser.error('--create-ticket requires --summary (or provide it via --create-ticket FILE)')
        if not args.issue_type:
            parser.error('--create-ticket requires --issue-type (or provide it via --create-ticket FILE)')
    else:
        create_only_args = [
            args.summary,
            args.issue_type,
            args.ticket_description,
            args.assignee_id,
            args.components,
            args.fix_versions,
            args.labels,
            args.parent,
        ]
        if any(create_only_args):
            parser.error('--summary, --issue-type, --ticket-description, --assignee-id, --components, --fix-versions, --labels, and --parent require --create-ticket')
 
    if args.hierarchy is not None and not args.get_related:
        parser.error('--hierarchy requires --get-related')
    if args.hierarchy is not None and args.hierarchy < -1:
        parser.error('--hierarchy DEPTH must be -1 (omit for unlimited) or a non-negative integer')

    # Validate --table-format usage
    if args.table_format != 'flat' and not (args.get_children or args.get_related):
        parser.error('--table-format requires --get-children or --get-related')

    # Validate --dump-file and --dump-format usage
    # Allow dump-file with any command that produces tabular/list data
    dump_compatible = (args.get_tickets or jql_specified or release_tickets_specified or
                       args.no_release or args.releases or args.get_versions or args.get_components or args.get_children or args.get_related)
    if (args.dump_file or args.dump_format != 'csv') and not dump_compatible:
        parser.error('--dump-file and --dump-format require a command that produces data (e.g., --get-tickets, --jql, --releases, --get-versions, --get-components, --get-children, --get-related)')
    
    # Validate bulk update arguments
    if args.bulk_update:
        if not args.input_file:
            parser.error('--bulk-update requires --input-file')
        if not any([args.set_release, args.remove_release, args.transition, args.assign]):
            parser.error('--bulk-update requires at least one operation: --set-release, --remove-release, --transition, or --assign')

    # Validate bulk delete arguments
    if args.bulk_delete:
        if not args.input_file:
            parser.error('--bulk-delete requires --input-file')
        # No extra operation flags required; deletion is the operation.

    # Validate bulk update operation args are only used with --bulk-update
    bulk_ops = [args.set_release, args.remove_release, args.transition, args.assign]
    if any(bulk_ops) and not args.bulk_update:
        parser.error('--set-release, --remove-release, --transition, and --assign require --bulk-update')

    if args.delete_subtasks and not args.bulk_delete:
        parser.error('--delete-subtasks requires --bulk-delete')

    # input-file may be used with either bulk-update or bulk-delete
    if args.input_file and not (args.bulk_update or args.bulk_delete):
        parser.error('--input-file requires --bulk-update or --bulk-delete')

    if args.execute and not (args.bulk_update or args.bulk_delete or (args.create_ticket is not None)):
        parser.error('--execute requires --bulk-update, --bulk-delete, or --create-ticket')

    if args.max_updates and not args.bulk_update:
        parser.error('--max-updates requires --bulk-update')

    if args.max_deletes and not args.bulk_delete:
        parser.error('--max-deletes requires --bulk-delete')

    # If --execute is specified, disable dry_run
    if args.execute:
        args.dry_run = False
    
    # Validate dashboard arguments
    dashboard_actions = [args.dashboards, args.dashboard, args.create_dashboard,
                        args.update_dashboard, args.delete_dashboard, args.copy_dashboard,
                        args.gadgets, args.add_gadget, args.remove_gadget, args.update_gadget]
    
    # --owner and --shared only valid with --dashboards
    if (args.owner or args.shared) and not args.dashboards:
        parser.error('--owner and --shared require --dashboards')
    
    # --description only valid with dashboard create/update/copy
    if args.description and not (args.create_dashboard or args.update_dashboard or args.copy_dashboard):
        parser.error('--description requires --create-dashboard, --update-dashboard, or --copy-dashboard')
    
    # --name only valid with dashboard update/copy
    if args.name and not (args.update_dashboard or args.copy_dashboard):
        parser.error('--name requires --update-dashboard or --copy-dashboard')
    
    # --share-permissions only valid with dashboard create/update/copy
    if args.share_permissions and not (args.create_dashboard or args.update_dashboard or args.copy_dashboard):
        parser.error('--share-permissions requires --create-dashboard, --update-dashboard, or --copy-dashboard')
    
    # --force only valid with --delete-dashboard or --bulk-delete
    if args.force and not (args.delete_dashboard or args.bulk_delete):
        parser.error('--force requires --delete-dashboard or --bulk-delete')
    
    # --copy-dashboard requires --name
    if args.copy_dashboard and not args.name:
        parser.error('--copy-dashboard requires --name to specify the new dashboard name')
    
    # Gadget operations require --dashboard
    if (args.add_gadget or args.remove_gadget or args.update_gadget) and not args.dashboard:
        parser.error('--add-gadget, --remove-gadget, and --update-gadget require --dashboard')
    
    # --position and --color only valid with gadget add/update
    if (args.position or args.color) and not (args.add_gadget or args.update_gadget):
        parser.error('--position and --color require --add-gadget or --update-gadget')
    
    # --gadget-properties only valid with --add-gadget
    if args.gadget_properties and not args.add_gadget:
        parser.error('--gadget-properties requires --add-gadget')
    
    # Validate that at least one action is specified
    if not args.list_projects and not any(project_actions) and not jql_specified and not (args.bulk_update or args.bulk_delete) and not any(dashboard_actions) and not children_specified and not related_specified:
        parser.print_help()
        sys.exit(1)
    
    # Store whether get_fields/jql/release_tickets was specified for easier checking
    args.get_fields_specified = get_fields_specified
    args.jql_specified = jql_specified
    args.release_tickets_specified = release_tickets_specified

    log.info('++++++++++++++++++++++++++++++++++++++++++++++')
    log.info(f'+  {os.path.basename(sys.argv[0])}')
    log.info(f'+  Python Version: {sys.version.split()[0]}')
    log.info(f'+  Today is: {date.today()}')
    log.info(f'+  Jira URL: {JIRA_URL}')
    if args.project:
        log.info(f'+  Project: {args.project}')
    log.info('++++++++++++++++++++++++++++++++++++++++++++++')        

    return args


# ****************************************************************************************
# Main
# ****************************************************************************************
def main():
    '''
    Entrypoint that wires together dependencies and launches the CLI.

    Sequence:
        1. Parse command line arguments
        2. Connect to Jira
        3. Execute requested action(s)

    Output:
        Exit code 0 on success, 1 on failure.
    '''
    args = handle_args()
    log.debug('Entering main()')
    
    try:
        jira = connect_to_jira()
        
        if args.list_projects:
            list_projects(jira)
        
        if args.get_workflow:
            get_project_workflows(jira, args.project)

        if args.get_issue_types:
            get_project_issue_types(jira, args.project)

        if args.get_fields_specified:
            # Use --issue-types for filtering if provided, otherwise use args from --get-fields
            issue_type_filter = args.issue_types if args.issue_types else args.get_fields
            get_project_fields(jira, args.project, issue_type_filter)

        if args.get_versions:
            get_project_versions(jira, args.project)

        if args.get_components:
            get_project_components(jira, args.project, args.date, args.dump_file, args.dump_format)

        if args.get_children:
            get_children_hierarchy(jira, args.project, args.get_children, args.limit, args.dump_file, args.dump_format, args.table_format)

        if args.get_related:
            get_related_issues(jira, args.project, args.get_related, args.hierarchy, args.limit, args.dump_file, args.dump_format, args.table_format)
        
        if args.releases:
            if args.get_tickets:
                # Get tickets for all releases matching the pattern
                get_releases_tickets(jira, args.project, args.releases, args.issue_types,
                                    args.status, args.date, args.limit, args.dump_file, args.dump_format)
            else:
                # Just list the releases
                get_releases(jira, args.project, args.releases, args.dump_file, args.dump_format)
        
        if args.release_tickets_specified:
            get_release_tickets(jira, args.project, args.release_tickets, args.issue_types,
                               args.status, args.date, args.limit, args.dump_file, args.dump_format)
        
        if args.no_release:
            get_no_release_tickets(jira, args.project, args.issue_types, args.status, args.date, args.limit,
                                  args.dump_file, args.dump_format)
        
        if args.total:
            get_ticket_totals(jira, args.project, args.issue_types, args.status, args.date)
        
        if args.get_tickets and not args.releases:
            # Only run standalone get_tickets if --releases is not specified
            # (when --releases is specified with --get-tickets, it's handled above)
            get_tickets(jira, args.project, args.issue_types, args.status, args.date, args.limit,
                       args.dump_file, args.dump_format)
        
        if args.jql_specified:
            run_jql_query(jira, args.jql, args.limit, args.dump_file, args.dump_format)

        if args.create_ticket is not None:
            # Note: args.* values may have been filled from `--create-ticket FILE`, with CLI overriding JSON.
            create_ticket(
                jira,
                args.project,
                args.summary,
                args.issue_type,
                description=args.ticket_description,
                assignee=args.assignee_id,
                components=args.components,
                fix_versions=args.fix_versions,
                labels=args.labels,
                parent_key=args.parent,
                dry_run=args.dry_run,
            )
        
        if args.bulk_update:
            bulk_update_tickets(jira, args.input_file, args.set_release, args.remove_release,
                               args.transition, args.assign, args.dry_run, args.max_updates)

        if args.bulk_delete:
            bulk_delete_tickets(
                jira,
                args.input_file,
                delete_subtasks=args.delete_subtasks,
                dry_run=args.dry_run,
                max_deletes=args.max_deletes,
                force=args.force,
            )
        
        # Dashboard management operations
        if args.dashboards:
            list_dashboards(jira, args.owner, args.shared)
        
        if args.dashboard:
            # Check if gadget operations are requested with --dashboard
            if args.add_gadget:
                add_gadget(jira, args.dashboard, args.add_gadget, args.position, args.color, args.gadget_properties)
            elif args.remove_gadget:
                remove_gadget(jira, args.dashboard, args.remove_gadget)
            elif args.update_gadget:
                update_gadget(jira, args.dashboard, args.update_gadget, args.position, args.color)
            else:
                # No gadget operation, just get dashboard details
                get_dashboard(jira, args.dashboard)
        
        if args.create_dashboard:
            create_dashboard(jira, args.create_dashboard, args.description, args.share_permissions)
        
        if args.update_dashboard:
            update_dashboard(jira, args.update_dashboard, args.name, args.description, args.share_permissions)
        
        if args.delete_dashboard:
            delete_dashboard(jira, args.delete_dashboard, args.force)
        
        if args.copy_dashboard:
            copy_dashboard(jira, args.copy_dashboard, args.name, args.description, args.share_permissions)
        
        if args.gadgets:
            list_gadgets(jira, args.gadgets)
            
    except JiraCredentialsError as e:
        log.error(e.message)
        output('')
        output('ERROR: ' + e.message)
        output('')
        output('Please set the required environment variables:')
        output('  export JIRA_EMAIL="your.email@cornelisnetworks.com"')
        output('  export JIRA_API_TOKEN="your_api_token_here"')
        output('')
        sys.exit(1)
    except JiraConnectionError as e:
        log.error(e.message)
        output('')
        output('ERROR: ' + e.message)
        output('')
        sys.exit(1)
    except JiraProjectError as e:
        log.error(e.message)
        output('')
        output('ERROR: ' + e.message)
        output('')
        sys.exit(1)
    except JiraDashboardError as e:
        log.error(e.message)
        output('')
        output('ERROR: ' + e.message)
        output('')
        sys.exit(1)
    except Exception as e:
        log.error(f'Unexpected error: {e}')
        output(f'ERROR: {e}')
        sys.exit(1)
    
    # Display JQL if --show-jql was specified
    display_jql()
    
    log.info('Operation complete.')


if __name__ == '__main__':
    main()