#!/usr/bin/env python3
##########################################################################################
#
# Script name: pm_agent.py
#
# Description: CLI entry point for Cornelis Agent Pipeline.
#              Provides commands for release planning workflow.
#
# Author: Cornelis Networks
#
# Usage:
#   python pm_agent.py --help
#   python pm_agent.py --plan --project PROJ --roadmap slides.pptx
#   python pm_agent.py --analyze --project PROJ
#   python pm_agent.py --resume --session abc123
#
##########################################################################################

import argparse
import logging
import re
import sys
import os
import threading
import time as _time
from datetime import date

from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# ****************************************************************************************
# Global data and configuration
# ****************************************************************************************

# Logging config - follows jira_utils.py pattern
log = logging.getLogger(os.path.basename(sys.argv[0]))
log.setLevel(logging.DEBUG)

# File handler for logging
fh = logging.FileHandler('cornelis_agent.log', mode='w')
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter(
    '%(asctime)-15s [%(funcName)25s:%(lineno)-5s] %(levelname)-8s %(message)s')
fh.setFormatter(formatter)
log.addHandler(fh)

# Output control
_quiet_mode = False


def output(message=''):
    '''
    Print user-facing output, respecting quiet mode.
    '''
    if message:
        record = logging.LogRecord(
            name=log.name,
            level=logging.INFO,
            pathname=__file__,
            lineno=0,
            msg=f'OUTPUT: {message}',
            args=(),
            exc_info=None,
            func='output'
        )
        fh.emit(record)
    
    if not _quiet_mode:
        print(message)


# ****************************************************************************************
# Command handlers
# ****************************************************************************************

def cmd_plan(args):
    '''
    Run the release planning workflow.
    '''
    log.debug(f'cmd_plan(project={args.project}, plan_mode={args.plan_mode})')
    
    from agents.orchestrator import ReleasePlanningOrchestrator
    from state.session import SessionManager
    from state.persistence import get_persistence
    
    output('')
    output('=' * 60)
    output('CORNELIS RELEASE PLANNING AGENT')
    output('=' * 60)
    output('')
    
    # Set up persistence
    persistence = None
    if args.save_session:
        persistence = get_persistence(args.persistence_format)
    
    session_manager = SessionManager(persistence=persistence)
    
    # Collect input files
    roadmap_files = []
    if args.roadmap:
        roadmap_files.extend(args.roadmap)
    
    # Create orchestrator and run
    orchestrator = ReleasePlanningOrchestrator()
    
    output(f'Project: {args.project}')
    output(f'Roadmap files: {len(roadmap_files)}')
    if args.org_chart:
        output(f'Org chart: {args.org_chart}')
    output('')
    
    # Run the workflow
    result = orchestrator.run({
        'project_key': args.project,
        'roadmap_files': roadmap_files,
        'org_chart_file': args.org_chart,
        'mode': args.plan_mode
    })
    
    if result.success:
        output(result.content)
        
        # Save session if requested
        if args.save_session and orchestrator.state:
            from state.session import SessionState
            session = SessionState(
                project_key=args.project,
                roadmap_files=roadmap_files,
                org_chart_file=args.org_chart,
                roadmap_data=orchestrator.state.roadmap_data,
                jira_state=orchestrator.state.jira_state,
                release_plan=orchestrator.state.release_plan,
                current_step=orchestrator.state.current_step
            )
            session_manager.current_session = session
            session_manager.save_session()
            output(f'\nSession saved: {session.session_id}')
    else:
        output(f'ERROR: {result.error}')
        return 1
    
    return 0


def cmd_analyze(args):
    '''
    Analyze Jira project state.
    '''
    log.debug(f'cmd_analyze(project={args.project}, quick={args.quick})')
    
    from agents.jira_analyst import JiraAnalystAgent
    
    output('')
    output('=' * 60)
    output('JIRA PROJECT ANALYSIS')
    output('=' * 60)
    output('')
    
    analyst = JiraAnalystAgent(project_key=args.project)
    
    if args.quick:
        # Quick analysis without LLM
        analysis = analyst.analyze_project(args.project)
        
        output(f"Project: {analysis.get('project_key')}")
        output('')
        
        summary = analysis.get('summary', {})
        output(f"Releases: {summary.get('total_releases', 0)} ({summary.get('unreleased_count', 0)} unreleased)")
        output(f"Components: {summary.get('component_count', 0)}")
        output(f"Issue Types: {summary.get('issue_type_count', 0)}")
        
        if analysis.get('errors'):
            output('\nErrors:')
            for error in analysis['errors']:
                output(f'  ! {error}')
    else:
        # Full LLM-powered analysis
        result = analyst.run(args.project)
        
        if result.success:
            output(result.content)
        else:
            output(f'ERROR: {result.error}')
            return 1
    
    return 0


def cmd_vision(args):
    '''
    Analyze roadmap files.
    '''
    log.debug(f'cmd_vision(files={args.vision})')
    
    from agents.vision_analyzer import VisionAnalyzerAgent
    
    output('')
    output('=' * 60)
    output('ROADMAP ANALYSIS')
    output('=' * 60)
    output('')
    
    analyzer = VisionAnalyzerAgent()
    
    if len(args.vision) == 1:
        result = analyzer.analyze_file(args.vision[0])
    else:
        result = analyzer.analyze_multiple(args.vision)
    
    if 'error' in result:
        output(f'ERROR: {result["error"]}')
        return 1
    
    output(f"Files analyzed: {len(result.get('files_analyzed', [args.vision[0]]))}")
    output(f"Releases found: {len(result.get('releases', []))}")
    output(f"Features found: {len(result.get('features', []))}")
    output(f"Timeline items: {len(result.get('timeline', []))}")
    
    if result.get('releases'):
        output('\nReleases:')
        for r in result['releases'][:10]:
            output(f"  - {r.get('version', 'Unknown')}")
    
    if result.get('features'):
        output('\nFeatures:')
        for f in result['features'][:10]:
            output(f"  - {f.get('text', '')[:60]}")
    
    return 0


def cmd_sessions(args):
    '''
    List or manage sessions.
    '''
    log.debug(f'cmd_sessions(list={args.list_sessions}, delete={args.delete_session})')
    
    from state.session import SessionManager
    from state.persistence import get_persistence
    
    persistence = get_persistence(args.persistence_format)
    session_manager = SessionManager(persistence=persistence)
    
    if args.delete_session:
        if session_manager.delete_session(args.delete_session):
            output(f'Deleted session: {args.delete_session}')
        else:
            output(f'Failed to delete session: {args.delete_session}')
        return 0
    
    # List sessions
    sessions = session_manager.list_sessions()
    
    output('')
    output('=' * 60)
    output('SAVED SESSIONS')
    output('=' * 60)
    output('')
    
    if not sessions:
        output('No saved sessions found.')
        return 0
    
    output(f'{"ID":<10} {"Project":<12} {"Step":<15} {"Updated":<20}')
    output('-' * 60)
    
    for session in sessions:
        output(f"{session['session_id']:<10} {session.get('project_key', 'N/A'):<12} {session.get('current_step', 'N/A'):<15} {session.get('updated_at', 'N/A')[:19]:<20}")
    
    output('')
    output(f'Total: {len(sessions)} sessions')
    
    return 0


def cmd_build_excel_map(args):
    '''
    Build a multi-sheet Excel workbook mapping one or more tickets and all
    their related issues' child hierarchies.

    Accepts a list of ticket keys. For each ticket, calls _get_related_data()
    and merges the results into a single Map sheet (deduplicating by key).
    Then gets children for every depth=1 ticket and adds each as a separate
    sheet.

    Steps:
        1. Connect to Jira
        2. For each root ticket, _get_related_data() with given hierarchy depth
        3. Merge all results into one Map sheet (indented format)
        4. For each depth=1 ticket, _get_children_data() (unlimited depth)
        5. Write each as a temp .xlsx via dump_tickets_to_file
        6. Assemble all into one workbook: Tickets (Map) + per-ticket sheets
        7. Cleanup temp files

    Uses jira_utils functions:
        connect_to_jira, _get_related_data, _get_children_data,
        dump_tickets_to_file
    '''
    import tempfile
    import shutil
    from copy import copy

    ticket_keys = [k.upper() for k in args.ticket_keys]

    log.debug(f'cmd_build_excel_map(ticket_keys={ticket_keys}, hierarchy={args.hierarchy}, limit={getattr(args, "limit", None)}, output={getattr(args, "output", None)})')

    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        output('ERROR: openpyxl is required for build-excel-map. Install with: pip install openpyxl')
        return 1

    import jira_utils

    hierarchy_depth = args.hierarchy
    limit = getattr(args, 'limit', None)
    # Default output filename: first ticket key, or combined if multiple
    if getattr(args, 'output', None):
        output_file = args.output
    elif len(ticket_keys) == 1:
        output_file = f'{ticket_keys[0]}.xlsx'
    else:
        output_file = f'{"_".join(ticket_keys)}.xlsx'
    keep_intermediates = getattr(args, 'keep_intermediates', False)

    # Ensure output has .xlsx extension
    if not output_file.endswith('.xlsx'):
        output_file = f'{output_file}.xlsx'

    output('')
    output('=' * 60)
    output('BUILD EXCEL MAP')
    output('=' * 60)
    output(f'Root ticket(s):  {", ".join(ticket_keys)}')
    output(f'Hierarchy depth: {hierarchy_depth}')
    output(f'Output file:     {output_file}')
    if limit:
        output(f'Limit:           {limit}')
    output('')

    # Create a temp directory for intermediate files
    temp_dir = tempfile.mkdtemp(prefix='excel_map_')
    temp_files = []

    try:
        # ---------------------------------------------------------------
        # Step 1: Connect to Jira
        # ---------------------------------------------------------------
        output('Step 1/4: Connecting to Jira...')
        jira = jira_utils.connect_to_jira()

        # Validate project if provided
        if getattr(args, 'project', None):
            jira_utils.validate_project(jira, args.project)

        # ---------------------------------------------------------------
        # Step 2: Get related issues for each root ticket (in-memory)
        #         Merge results, deduplicating by issue key while
        #         preserving insertion order.
        # ---------------------------------------------------------------
        output(f'Step 2/4: Getting related issues for {len(ticket_keys)} root ticket(s) (hierarchy={hierarchy_depth})...')

        merged_data = []       # combined ordered list (deduplicated)
        seen_keys = set()      # track keys already in merged_data

        for root_key in ticket_keys:
            output(f'  Fetching related for {root_key}...')
            related_data = jira_utils._get_related_data(jira, root_key, hierarchy=hierarchy_depth, limit=limit)

            added = 0
            for item in related_data:
                issue_key = item['issue'].get('key', '')
                if issue_key and issue_key not in seen_keys:
                    seen_keys.add(issue_key)
                    merged_data.append(item)
                    added += 1

            depth0_count = sum(1 for item in related_data if item['depth'] == 0)
            depth1_count = sum(1 for item in related_data if item['depth'] == 1)
            output(f'    {len(related_data)} issues ({depth0_count} root + {depth1_count} depth=1), {added} new after dedup')

        output(f'  Merged total: {len(merged_data)} unique issues')

        # Write Map sheet to temp file
        map_temp = os.path.join(temp_dir, '_map_temp.xlsx')
        temp_files.append(map_temp)

        # Build extras dict for dump_tickets_to_file (depth + via metadata)
        map_extras = {
            item['issue'].get('key', ''): {
                'depth': item.get('depth'),
                'via': item.get('via'),
                'relation': item.get('relation'),
                'from_key': item.get('from_key'),
            }
            for item in merged_data
        }
        jira_utils.dump_tickets_to_file(
            [item['issue'] for item in merged_data],
            map_temp, 'excel', map_extras, table_format='indented'
        )
        output(f'  Map sheet: {len(merged_data)} rows, indented format')

        # ---------------------------------------------------------------
        # Step 3: Get children for each depth=1 ticket
        # ---------------------------------------------------------------
        depth1_keys = [item['issue'].get('key', '') for item in merged_data if item['depth'] == 1]
        output(f'Step 3/4: Getting children for {len(depth1_keys)} depth=1 tickets...')

        children_temps = []  # list of (ticket_key, temp_file_path, row_count)
        for idx, ticket_key in enumerate(depth1_keys, 1):
            try:
                children_data = jira_utils._get_children_data(jira, ticket_key, limit=None)
                child_count = len(children_data) - 1  # subtract root itself

                # Write to temp file
                child_temp = os.path.join(temp_dir, f'temp_{ticket_key}.xlsx')
                temp_files.append(child_temp)

                child_extras = {
                    item['issue'].get('key', ''): {
                        'depth': item.get('depth'),
                    }
                    for item in children_data
                }
                jira_utils.dump_tickets_to_file(
                    [item['issue'] for item in children_data],
                    child_temp, 'excel', child_extras, table_format='indented'
                )

                children_temps.append((ticket_key, child_temp, len(children_data)))
                output(f'  [{idx}/{len(depth1_keys)}] {ticket_key}: {child_count} children')

            except Exception as e:
                log.warning(f'Failed to get children for {ticket_key}: {e}')
                output(f'  [{idx}/{len(depth1_keys)}] {ticket_key}: ERROR - {e}')

        # ---------------------------------------------------------------
        # Step 4: Assemble final workbook
        # ---------------------------------------------------------------
        output(f'Step 4/4: Assembling final workbook...')

        final_wb = Workbook()
        # Remove the default sheet created by Workbook()
        final_wb.remove(final_wb.active)

        total_rows = 0
        sheet_count = 0

        def _copy_sheet(src_wb_path, dest_wb, sheet_name):
            '''
            Copy the first sheet from a source workbook into the destination
            workbook as a new sheet with the given name. Preserves cell values,
            hyperlinks, font, fill, alignment, border, number_format, column
            widths, and conditional formatting.
            '''
            nonlocal total_rows, sheet_count

            src_wb = load_workbook(src_wb_path)
            src_ws = src_wb.active

            # Truncate sheet name to 31 chars (Excel limit)
            safe_name = sheet_name[:31]
            dest_ws = dest_wb.create_sheet(title=safe_name)

            # Copy cell values and formatting
            row_count = 0
            for row in src_ws.iter_rows():
                row_count += 1
                for cell in row:
                    dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                    # Copy formatting
                    if cell.font:
                        dest_cell.font = copy(cell.font)
                    if cell.fill:
                        dest_cell.fill = copy(cell.fill)
                    if cell.alignment:
                        dest_cell.alignment = copy(cell.alignment)
                    if cell.border:
                        dest_cell.border = copy(cell.border)
                    if cell.number_format:
                        dest_cell.number_format = cell.number_format

                    # Copy hyperlink
                    if cell.hyperlink:
                        dest_cell.hyperlink = cell.hyperlink

            # Copy column widths
            for col_letter, dim in src_ws.column_dimensions.items():
                dest_ws.column_dimensions[col_letter].width = dim.width

            # Copy merged cells
            for merged_range in src_ws.merged_cells.ranges:
                dest_ws.merge_cells(str(merged_range))

            # Copy conditional formatting rules.
            # openpyxl's add() expects a cell-range string *or* a
            # MultiCellRange, depending on the version.  Using the
            # sqref attribute (which is already a MultiCellRange) is
            # the safest approach; fall back to str() for older builds.
            for cf_rule in src_ws.conditional_formatting:
                try:
                    cell_range = cf_rule.sqref          # MultiCellRange
                except AttributeError:
                    cell_range = str(cf_rule)            # fallback
                for rule in cf_rule.rules:
                    try:
                        dest_ws.conditional_formatting.add(cell_range, rule)
                    except Exception as cf_err:
                        log.debug(f'Skipping conditional formatting rule: {cf_err}')

            # Copy freeze panes
            if src_ws.freeze_panes:
                dest_ws.freeze_panes = src_ws.freeze_panes

            src_wb.close()

            data_rows = max(0, row_count - 1)  # subtract header
            total_rows += data_rows
            sheet_count += 1
            return data_rows

        # Sheet 1: Tickets (merged get-related overview from all root tickets)
        map_rows = _copy_sheet(map_temp, final_wb, 'Tickets')
        output(f'  Sheet 1: Tickets ({map_rows} rows)')

        # Sheets 2..N: Per-ticket children
        for idx, (ticket_key, child_temp, child_row_count) in enumerate(children_temps, 2):
            child_rows = _copy_sheet(child_temp, final_wb, ticket_key)
            output(f'  Sheet {idx}: {ticket_key} ({child_rows} rows)')

        # Save the final workbook
        final_wb.save(output_file)
        final_wb.close()

        output('')
        output(f'Output: {output_file} ({sheet_count} sheets, {total_rows} total rows)')

    except Exception as e:
        log.error(f'Failed to build Excel map: {e}', exc_info=True)
        output(f'ERROR: {e}')
        return 1

    finally:
        # ---------------------------------------------------------------
        # Step 5: Cleanup
        # ---------------------------------------------------------------
        if keep_intermediates:
            output(f'Intermediate files kept in: {temp_dir}')
            for tf in temp_files:
                if os.path.exists(tf):
                    output(f'  {tf}')
        else:
            # Clean up temp directory
            try:
                shutil.rmtree(temp_dir)
                log.debug(f'Cleaned up {len(temp_files)} intermediate files in {temp_dir}')
            except Exception as cleanup_err:
                log.warning(f'Failed to clean up temp dir {temp_dir}: {cleanup_err}')

    return 0


def _validate_and_repair_csv(filepath):
    '''
    Validate a CSV file and repair rows that have more columns than the header.

    LLMs sometimes emit CSV with unquoted commas inside fields — for example
    a summary like ``12.1.0.0.72,78, 12.1.0.1.4 - hfi1_0: CPORT …`` ends up
    split across multiple columns.  This function detects such rows and merges
    the excess columns back into the widest text field (heuristically the
    "summary" or the longest-valued column in that row).

    The file is rewritten in-place only if repairs were needed.

    Returns the filepath (unchanged).
    '''
    import csv as _csv

    with open(filepath, 'r', encoding='utf-8') as f:
        reader = _csv.reader(f)
        header = next(reader, None)
        if not header:
            return filepath
        expected = len(header)
        rows = []
        needs_repair = False
        for row in reader:
            if len(row) > expected:
                needs_repair = True
                # Find the best column to merge excess into.
                # Prefer 'summary' if it exists; otherwise pick the column
                # with the longest value in this row.
                summary_idx = None
                for i, h in enumerate(header):
                    if h.strip().lower() == 'summary':
                        summary_idx = i
                        break
                if summary_idx is None:
                    # Fallback: pick the column with the longest value
                    summary_idx = max(range(min(expected, len(row))),
                                      key=lambda i: len(row[i]))

                # Merge: take the first `expected` cells, but fold the extra
                # cells into the summary column by joining with commas.
                extra_count = len(row) - expected
                # The extra cells are the ones that were split out of the
                # summary column.  They sit right after summary_idx.
                merged_value = ','.join(
                    row[summary_idx : summary_idx + 1 + extra_count])
                repaired = (row[:summary_idx]
                            + [merged_value]
                            + row[summary_idx + 1 + extra_count:])
                log.debug(
                    f'CSV repair: row had {len(row)} cols, merged '
                    f'{extra_count} extra into column "{header[summary_idx]}"')
                rows.append(repaired)
            elif len(row) < expected:
                # Pad short rows with empty strings
                needs_repair = True
                rows.append(row + [''] * (expected - len(row)))
            else:
                rows.append(row)

    if needs_repair:
        log.info(f'Repairing CSV: {filepath} ({len(rows)} data rows)')
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = _csv.writer(f)
            writer.writerow(header)
            writer.writerows(rows)

    return filepath


def _extract_and_save_files(response_text):
    '''
    Parse fenced code blocks from an LLM response and save any that specify
    a filename.

    Recognised patterns (the filename line immediately precedes the opening
    fence or is embedded in the fence info-string):

        **`path/to/file.ext`**          (bold-backtick markdown)
        `path/to/file.ext`              (backtick markdown)
        path/to/file.ext                (bare path on its own line)
        ```lang:path/to/file.ext        (colon-separated in info-string)
        ```path/to/file.ext             (info-string IS the path)

    Returns a list of (filepath, content) tuples that were written.
    '''
    saved = []

    # ---- Strategy 1: filename on the line before the opening fence ----------
    # Matches patterns like:
    #   **`somefile.py`**
    #   `somefile.py`
    #   somefile.py
    # followed by ```<optional lang>\n ... \n```
    pattern_pre = re.compile(
        r'(?:^|\n)'                          # start of string or newline
        r'[ \t]*'                            # optional leading whitespace
        r'(?:\*{0,2}`?)'                     # optional bold / backtick prefix
        r'([\w][\w./\\-]*\.\w+)'             # captured filename (must have extension)
        r'(?:`?\*{0,2})'                     # optional backtick / bold suffix
        r'[ \t]*\n'                          # end of filename line
        r'```[^\n]*\n'                       # opening fence (``` with optional info)
        r'(.*?)'                             # captured content (non-greedy)
        r'\n```',                            # closing fence
        re.DOTALL,
    )

    for m in pattern_pre.finditer(response_text):
        filepath = m.group(1).strip()
        content = m.group(2)
        if filepath and content is not None:
            saved.append((filepath, content))

    # ---- Strategy 2: filename embedded in the info-string -------------------
    # Matches ```lang:path/to/file.ext  or  ```path/to/file.ext
    # Only used for blocks NOT already captured by Strategy 1.
    pattern_info = re.compile(
        r'```'
        r'(?:[\w]+:)?'                       # optional lang: prefix
        r'([\w][\w./\\-]*\.\w+)'             # captured filename
        r'[ \t]*\n'
        r'(.*?)'                             # captured content
        r'\n```',
        re.DOTALL,
    )

    # Build a set of already-captured content spans to avoid duplicates
    captured_spans = {(m.start(), m.end()) for m in pattern_pre.finditer(response_text)}

    for m in pattern_info.finditer(response_text):
        # Skip if this block overlaps with one already captured
        if any(m.start() >= cs[0] and m.end() <= cs[1] for cs in captured_spans):
            continue
        filepath = m.group(1).strip()
        content = m.group(2)
        if filepath and content is not None:
            saved.append((filepath, content))

    # ---- Strategy 2b: filename as first line inside the code block ----------
    # Handles the pattern where the LLM places the filename as the first
    # non-empty line *inside* the fenced block (not on the fence line and
    # not on a preceding line).  Example:
    #   ```
    #   my_report.csv
    #   col1,col2,col3
    #   ...
    #   ```
    # We recognise a first line as a filename if it matches word.ext with a
    # known file extension and contains no spaces (to avoid false positives
    # on prose or data lines).
    KNOWN_EXTENSIONS = {
        'csv', 'json', 'xml', 'yaml', 'yml', 'md', 'txt', 'sql', 'html',
        'toml', 'ini', 'py', 'js', 'ts', 'sh', 'cfg', 'log', 'xlsx',
    }
    pattern_inner_name = re.compile(
        r'```[^\n]*\n'                       # opening fence (``` with optional info)
        r'(.*?)'                             # captured full content
        r'\n```',                            # closing fence
        re.DOTALL,
    )
    # Build set of already-captured spans (from Strategies 1 & 2)
    captured_spans_2b = set()
    for m in pattern_pre.finditer(response_text):
        captured_spans_2b.add((m.start(), m.end()))
    for m in pattern_info.finditer(response_text):
        if not any(m.start() >= cs[0] and m.end() <= cs[1] for cs in captured_spans_2b):
            captured_spans_2b.add((m.start(), m.end()))

    for m in pattern_inner_name.finditer(response_text):
        # Skip blocks already captured by earlier strategies
        if any(m.start() >= cs[0] and m.end() <= cs[1] for cs in captured_spans_2b):
            continue
        content = m.group(1)
        if not content or not content.strip():
            continue
        # Check if the first non-empty line looks like a filename
        lines = content.split('\n')
        first_line = ''
        first_line_idx = 0
        for i, line in enumerate(lines):
            if line.strip():
                first_line = line.strip()
                first_line_idx = i
                break
        if not first_line:
            continue
        # A filename candidate: no spaces, has a dot, extension is known
        if ' ' not in first_line and '.' in first_line:
            ext = first_line.rsplit('.', 1)[-1].lower()
            if ext in KNOWN_EXTENSIONS:
                filepath = first_line
                # Remaining lines (after the filename line) become the content
                remaining = '\n'.join(lines[first_line_idx + 1:])
                if remaining.strip():
                    saved.append((filepath, remaining))
                    log.debug(f'Strategy 2b: first-line filename "{filepath}" '
                              f'inside code block')

    # ---- Strategy 3: auto-save fallback for unnamed data blocks -------------
    # When no named files were found by Strategies 1-2, look for fenced blocks
    # whose info-string is a known data format and auto-save as llm_output.<ext>.
    # If multiple blocks share the same extension, number them (llm_output_2.csv, etc).
    AUTO_SAVE_EXTS = {
        'csv': 'csv', 'json': 'json', 'xml': 'xml', 'yaml': 'yaml',
        'yml': 'yaml', 'md': 'md', 'markdown': 'md', 'txt': 'txt',
        'sql': 'sql', 'html': 'html', 'toml': 'toml', 'ini': 'ini',
    }

    if not saved:
        # Find all fenced blocks with an info-string
        pattern_auto = re.compile(
            r'```(\w+)[ \t]*\n'              # opening fence with info-string
            r'(.*?)'                          # captured content
            r'\n```',                         # closing fence
            re.DOTALL,
        )
        ext_counts = {}  # track how many blocks per extension for numbering
        for m in pattern_auto.finditer(response_text):
            info = m.group(1).lower()
            content = m.group(2)
            if info in AUTO_SAVE_EXTS and content and content.strip():
                ext = AUTO_SAVE_EXTS[info]
                ext_counts[ext] = ext_counts.get(ext, 0) + 1
                count = ext_counts[ext]
                # Naming: llm_output_file1.csv, llm_output_file2.csv, etc.
                filepath = f'llm_output_file{count}.{ext}'
                saved.append((filepath, content))
                log.debug(f'Auto-save fallback: {info} block -> {filepath}')

    # ---- Write files --------------------------------------------------------
    written = []
    for filepath, content in saved:
        try:
            # Create parent directories if needed
            parent = os.path.dirname(filepath)
            if parent:
                os.makedirs(parent, exist_ok=True)

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content)
                # Ensure file ends with a newline
                if content and not content.endswith('\n'):
                    f.write('\n')

            # ---- CSV validation: repair rows with wrong column count --------
            # LLMs sometimes emit CSV with unquoted commas inside fields
            # (e.g. "12.1.0.0.72,78" in a summary).  Detect and repair by
            # re-reading with the csv module and merging excess columns back
            # into the widest text field (typically 'summary').
            if filepath.lower().endswith('.csv'):
                filepath = _validate_and_repair_csv(filepath)

            written.append(filepath)
            log.info(f'Saved extracted file: {filepath} ({len(content)} chars)')
        except Exception as e:
            log.warning(f'Failed to save extracted file {filepath}: {e}')
            output(f'  WARNING: could not save {filepath}: {e}')

    return written


def _invoke_llm(prompt_text, attachments=None, timeout=None, model=None):
    '''
    Send a prompt to the configured LLM with optional file attachments.

    This is the shared core logic extracted from cmd_invoke_llm() so it can
    be reused by workflow commands.

    Input:
        prompt_text: The prompt string to send.
        attachments: Optional list of file paths to attach.
        timeout: Optional timeout in seconds.
        model: Optional model name override.  When provided, this overrides
               the CORNELIS_LLM_MODEL / OPENAI_MODEL / etc. env-var default.

    Output:
        Tuple of (response_content, saved_files, token_info) where:
          - response_content: str, the full LLM response text
          - saved_files: list of file paths extracted/saved from the response
          - token_info: dict with keys prompt_tokens, completion_tokens, total_tokens,
                        model, finish_reason, elapsed, estimated_cost

    Raises:
        Exception if the LLM call fails.
    '''
    log.debug(f'Entering _invoke_llm(prompt_len={len(prompt_text)}, '
              f'attachments={attachments}, timeout={timeout}, model={model})')

    import base64
    import mimetypes
    from llm.config import get_llm_client
    from llm.base import Message

    # ---- classify attachments -----------------------------------------------
    IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.svg'}
    TEXT_EXTS = {
        '.md', '.txt', '.csv', '.json', '.yaml', '.yml', '.xml',
        '.py', '.js', '.ts', '.html', '.css', '.sh', '.sql',
        '.log', '.cfg', '.ini', '.toml',
    }

    image_data_uris = []   # base64 data URIs for vision API
    text_blocks = []       # fenced code blocks to append to prompt

    for filepath in (attachments or []):
        if not os.path.isfile(filepath):
            output(f'WARNING: attachment not found, skipping: {filepath}')
            log.warning(f'Attachment not found: {filepath}')
            continue

        ext = os.path.splitext(filepath)[1].lower()

        if ext in IMAGE_EXTS:
            # Encode image as base64 data URI for the vision API
            mime_type = mimetypes.guess_type(filepath)[0] or 'image/png'
            with open(filepath, 'rb') as img_f:
                b64 = base64.b64encode(img_f.read()).decode('utf-8')
            data_uri = f'data:{mime_type};base64,{b64}'
            image_data_uris.append(data_uri)
            log.debug(f'Attached image: {filepath} ({mime_type})')
            output(f'  Attached image: {filepath}')
        else:
            # Attempt to read as text
            try:
                with open(filepath, 'r', encoding='utf-8') as txt_f:
                    content = txt_f.read()
                # Determine language hint for fenced block
                lang = ext.lstrip('.') if ext in TEXT_EXTS else ''
                text_blocks.append(f'\n\n--- Attachment: {os.path.basename(filepath)} ---\n```{lang}\n{content}\n```')
                log.debug(f'Attached text file: {filepath} ({len(content)} chars)')
                output(f'  Attached text: {filepath} ({len(content)} chars)')
            except (UnicodeDecodeError, IOError) as e:
                output(f'WARNING: cannot read attachment as text, skipping: {filepath} ({e})')
                log.warning(f'Cannot read attachment {filepath}: {e}')

    # ---- build final prompt with inlined text attachments -------------------
    if text_blocks:
        prompt_text = prompt_text + '\n' + ''.join(text_blocks)

    # ---- send to LLM -------------------------------------------------------
    output('')
    total_chars = len(prompt_text)
    timeout_display = f'{timeout}s' if timeout else 'default'
    output(f'Sending to LLM... ({total_chars} chars, timeout={timeout_display})')

    # Use vision client if images are present, otherwise standard client.
    # When a model override is supplied via --model, pass it through to the
    # factory so it takes precedence over the env-var default.
    if image_data_uris:
        client = get_llm_client(for_vision=True, timeout=timeout, model=model)
        log.info(f'+  Using LLM model: {client.model}')
        output(f'Using vision model: {client.model} ({len(image_data_uris)} image(s))')
        messages = [Message.user(prompt_text)]
    else:
        client = get_llm_client(timeout=timeout, model=model)
        log.info(f'+  Using LLM model: {client.model}')
        output(f'Using model: {client.model}')
        messages = [Message.user(prompt_text)]

    # ---- call LLM with waiting-status heartbeat -----------------------------
    # Run the actual API call in a background thread so we can print
    # periodic "Still waiting..." messages every 10 seconds.
    llm_result = {}  # shared dict: {'response': ..., 'error': ...}

    def _llm_worker():
        try:
            if image_data_uris:
                llm_result['response'] = client.chat_with_vision(messages, image_data_uris)
            else:
                llm_result['response'] = client.chat(messages)
        except Exception as exc:
            llm_result['error'] = exc

    worker = threading.Thread(target=_llm_worker, daemon=True)
    start_time = _time.monotonic()
    worker.start()

    # Heartbeat loop — print status every 10 seconds while waiting
    heartbeat_interval = 10
    while worker.is_alive():
        worker.join(timeout=heartbeat_interval)
        if worker.is_alive():
            elapsed = int(_time.monotonic() - start_time)
            log.info(f'Still waiting on LLM return... {elapsed} seconds total')

    elapsed_total = _time.monotonic() - start_time

    # Re-raise any exception from the worker thread
    if 'error' in llm_result:
        raise llm_result['error']

    response = llm_result['response']
    log.info(f'LLM returned in {elapsed_total:.1f}s')

    # ---- display response ---------------------------------------------------
    output('')
    output('=' * 60)
    output('LLM RESPONSE')
    output('=' * 60)
    output('')
    output(response.content)
    output('')
    output('-' * 60)

    log.info(f'LLM response received: {len(response.content)} chars, '
             f'tokens={response.usage}')

    # ---- token usage & cost table -------------------------------------------
    prompt_tok = 0
    completion_tok = 0
    total_tok = 0
    if response.usage:
        prompt_tok = response.usage.get('prompt_tokens', 0)
        completion_tok = response.usage.get('completion_tokens', 0)
        total_tok = response.usage.get('total_tokens', 0)

    # Estimate cost using typical per-token pricing (rough approximation).
    # These rates are ballpark for GPT-4o-class models; adjust as needed.
    COST_PER_1K_PROMPT = 0.0025       # $2.50 / 1M prompt tokens
    COST_PER_1K_COMPLETION = 0.0100   # $10.00 / 1M completion tokens
    estimated_cost = (
        (prompt_tok / 1000.0) * COST_PER_1K_PROMPT
        + (completion_tok / 1000.0) * COST_PER_1K_COMPLETION
    )

    elapsed_str = f'{elapsed_total:.1f}s'

    # Build the summary box using ==== banner style
    banner = '=' * 80
    log.info(banner)
    log.info(f'  LLM totals: prompt_tokens={prompt_tok}, completion_tokens={completion_tok}')
    log.info(f'  Model: {response.model or "unknown"}')
    log.info(f'  Finish reason: {response.finish_reason or "n/a"}')
    log.info(f'  Elapsed: {elapsed_str}')
    log.info(f'  Estimated LLM cost: ${estimated_cost:.4f}')
    log.info(banner)

    # Token usage & metadata table — ==== banner style, key:  value alignment
    token_table = (
        f'\n'
        f'{banner}\n'
        f'LLM Token Usage\n'
        f'{banner}\n'
        f'{"Prompt tokens (in):":<30}  {prompt_tok:>12,}\n'
        f'{"Completion tokens (out):":<30}  {completion_tok:>12,}\n'
        f'{"Total tokens:":<30}  {total_tok:>12,}\n'
        f'{banner}\n'
        f'{"Model:":<30}  {(response.model or "unknown"):>12}\n'
        f'{"Finish reason:":<30}  {(response.finish_reason or "n/a"):>12}\n'
        f'{"Elapsed:":<30}  {elapsed_str:>12}\n'
        f'{"Estimated cost:":<30}  {f"${estimated_cost:.4f}":>12}\n'
        f'{banner}'
    )
    output(token_table)

    # ---- save full response as llm_output.md --------------------------------
    all_created_files = []   # (filepath, size_chars, source) tuples

    try:
        with open('llm_output.md', 'w', encoding='utf-8') as f:
            f.write(response.content)
            if response.content and not response.content.endswith('\n'):
                f.write('\n')
        all_created_files.append(('llm_output.md', len(response.content), 'full response'))
        log.info(f'Saved full LLM response: llm_output.md ({len(response.content)} chars)')
    except Exception as e:
        log.warning(f'Failed to save llm_output.md: {e}')
        output(f'WARNING: could not save llm_output.md: {e}')

    # ---- extract and save any file content from the response ----------------
    saved_files = _extract_and_save_files(response.content)
    for sf in saved_files:
        try:
            sz = os.path.getsize(sf)
        except OSError:
            sz = 0
        all_created_files.append((sf, sz, 'extracted'))

    # ---- file-creation summary table — ==== banner style --------------------
    if all_created_files:
        banner = '=' * 80
        max_name = max(len(f[0]) for f in all_created_files)
        max_name = max(max_name, 4)  # minimum width for header "File"
        row_fmt = f'{{:<4}}{{:<{max_name + 3}}}{{:>10}}   {{:<12}}'

        file_table_lines = [
            '',
            banner,
            'LLM Created Files',
            banner,
            row_fmt.format('#', 'File', 'Size', 'Source'),
        ]
        for idx, (fpath, fsize, fsource) in enumerate(all_created_files, 1):
            size_str = f'{fsize:,} ch' if fsize else '0 ch'
            file_table_lines.append(row_fmt.format(str(idx), fpath, size_str, fsource))
        file_table_lines.append(banner)

        file_table = '\n'.join(file_table_lines)
        output(file_table)

    # ---- build token_info dict for callers ----------------------------------
    token_info = {
        'prompt_tokens': prompt_tok,
        'completion_tokens': completion_tok,
        'total_tokens': total_tok,
        'model': response.model or 'unknown',
        'finish_reason': response.finish_reason or 'n/a',
        'elapsed': elapsed_total,
        'estimated_cost': estimated_cost,
    }

    return response.content, saved_files, token_info


def cmd_invoke_llm(args):
    '''
    Send a prompt to the configured LLM, optionally with file attachments.

    The prompt argument can be:
      - A path to a .md or .txt file whose contents become the prompt
      - An inline string used directly as the prompt

    Attachments are classified by extension:
      - Image files (.png, .jpg, .jpeg, .gif, .bmp, .webp, .svg) are sent
        via the vision API as base64-encoded images.
      - Text files (.md, .txt, .csv, .json, .yaml, .yml, .xml, .py, .js,
        .ts, .html, .css, .sh, .sql, .log, .cfg, .ini, .toml) are read
        and inlined into the prompt as fenced code blocks.
      - Other extensions are attempted as text; binary files are skipped
        with a warning.

    This is a thin wrapper around _invoke_llm() that resolves the prompt
    from args and delegates to the shared helper.
    '''
    log.debug(f'cmd_invoke_llm(prompt={args.prompt}, attachments={args.attachments})')

    # ---- resolve prompt text ------------------------------------------------
    prompt_arg = args.prompt
    if os.path.isfile(prompt_arg):
        log.debug(f'Reading prompt from file: {prompt_arg}')
        with open(prompt_arg, 'r', encoding='utf-8') as f:
            prompt_text = f.read().strip()
        output(f'Prompt loaded from {prompt_arg} ({len(prompt_text)} chars)')
    else:
        prompt_text = prompt_arg
        output(f'Using inline prompt ({len(prompt_text)} chars)')

    # ---- delegate to shared helper ------------------------------------------
    # Pass --model override if the user provided one on the CLI
    model_override = getattr(args, 'model', None)
    try:
        response_content, saved_files, token_info = _invoke_llm(
            prompt_text, attachments=args.attachments, timeout=args.timeout,
            model=model_override)
        return 0
    except Exception as e:
        log.error(f'LLM invocation failed: {e}', exc_info=True)
        output(f'ERROR: {e}')
        return 1


def cmd_workflow(args):
    '''
    Run a named workflow. Dispatches to the appropriate workflow handler.

    Input:
        args: Parsed argparse namespace with workflow_name and workflow-specific options.

    Output:
        int: Exit code (0 = success, 1 = error).

    Side Effects:
        Delegates to the workflow handler which may connect to Jira, invoke the LLM,
        and create output files.
    '''
    log.debug(f'Entering cmd_workflow(workflow_name={args.workflow_name})')

    # Registry of available workflows
    WORKFLOWS = {
        'bug-report': _workflow_bug_report,
    }

    handler = WORKFLOWS.get(args.workflow_name)
    if not handler:
        available = ', '.join(sorted(WORKFLOWS.keys()))
        output(f'ERROR: Unknown workflow "{args.workflow_name}". Available: {available}')
        return 1

    return handler(args)


def _workflow_bug_report(args):
    '''
    Bug report workflow: filter lookup → run filter → LLM analysis → Excel conversion.

    Steps:
      1. Connect to Jira
      2. Look up filter by name from favourite filters
      3. Run the filter to get tickets with latest comments
      4. Dump tickets to JSON file
      5. Send JSON + prompt to LLM
      6. Convert any extracted CSV to styled Excel

    Input:
        args: Parsed argparse namespace with:
            - workflow_filter: str, the Jira filter name to look up
            - workflow_prompt: str or None, path to the prompt file
            - timeout: float or None, LLM timeout in seconds
            - limit: int or None, max tickets to retrieve
            - output: str or None, output filename override

    Output:
        int: Exit code (0 = success, 1 = error).

    Side Effects:
        Creates JSON dump file, llm_output.md, extracted CSV/Excel files.
    '''
    log.debug(f'Entering _workflow_bug_report(filter={args.workflow_filter}, '
              f'prompt={args.workflow_prompt}, limit={args.limit}, timeout={args.timeout})')

    import jira_utils
    import excel_utils

    filter_name = args.workflow_filter
    prompt_path = args.workflow_prompt or 'agents/prompts/cn5000_bugs_clean.md'
    all_created_files = []  # (filepath, description) tuples for final summary

    # ---- Step 1/6: Connect to Jira ------------------------------------------
    output('')
    output('=' * 60)
    output('WORKFLOW: bug-report')
    output('=' * 60)
    output('')
    output('Step 1/6: Connecting to Jira...')
    try:
        jira = jira_utils.connect_to_jira()
        log.info('Jira connection established')
    except Exception as e:
        log.error(f'Step 1/6 failed: Jira connection error: {e}', exc_info=True)
        output(f'ERROR: Failed to connect to Jira: {e}')
        return 1

    # ---- Step 2/6: Look up filter by name -----------------------------------
    output('')
    output(f'Step 2/6: Looking up filter "{filter_name}" from favourite filters...')
    try:
        filters = jira_utils.list_filters(jira, favourite_only=True)
        # Search for exact match on filter name
        matched_filter = None
        for f in filters:
            if f.get('name') == filter_name:
                matched_filter = f
                break

        if not matched_filter:
            # Show available filter names to help the user
            available_names = [f.get('name', '(unnamed)') for f in filters]
            output(f'ERROR: Filter "{filter_name}" not found in favourite filters.')
            if available_names:
                output(f'  Available favourite filters:')
                for fn in available_names:
                    output(f'    - {fn}')
            return 1

        filter_id = matched_filter.get('id')
        filter_jql = matched_filter.get('jql', '')
        log.info(f'Found filter: id={filter_id}, name={filter_name}, jql={filter_jql}')
        output(f'  Found filter ID: {filter_id}')
        output(f'  JQL: {filter_jql}')
    except Exception as e:
        log.error(f'Step 2/6 failed: Filter lookup error: {e}', exc_info=True)
        output(f'ERROR: Failed to look up filter: {e}')
        return 1

    # ---- Step 3/6: Run the filter to get tickets ----------------------------
    output('')
    output(f'Step 3/6: Running filter to retrieve tickets...')
    try:
        # Set the global _include_comments flag so run_jql_query fetches comment
        # fields and dump_tickets_to_file applies the 'latest' filter.
        jira_utils._include_comments = 'latest'
        log.info('Set _include_comments = "latest" for comment extraction')

        issues = jira_utils.run_filter(jira, filter_id, limit=args.limit)
        if not issues:
            output('WARNING: Filter returned 0 tickets. Nothing to process.')
            return 0
        log.info(f'Filter returned {len(issues)} tickets')
        output(f'  Retrieved {len(issues)} tickets')
    except Exception as e:
        log.error(f'Step 3/6 failed: Filter execution error: {e}', exc_info=True)
        output(f'ERROR: Failed to run filter: {e}')
        return 1

    # ---- Step 4/6: Dump tickets to JSON file --------------------------------
    output('')
    output(f'Step 4/6: Dumping tickets to JSON...')
    try:
        # Derive dump filename from filter name (sanitize for filesystem)
        safe_name = re.sub(r'[^\w\s-]', '', filter_name).strip().lower()
        safe_name = re.sub(r'[\s]+', '_', safe_name)
        dump_basename = args.output or safe_name or 'bug_report'
        # Strip any extension the user may have provided — we force .json here
        dump_basename = os.path.splitext(dump_basename)[0]

        dump_path = jira_utils.dump_tickets_to_file(
            issues, dump_basename, 'json', include_comments='latest')
        log.info(f'Tickets dumped to: {dump_path}')
        output(f'  Saved: {dump_path}')
        all_created_files.append((dump_path, 'ticket JSON dump'))
    except Exception as e:
        log.error(f'Step 4/6 failed: Dump error: {e}', exc_info=True)
        output(f'ERROR: Failed to dump tickets: {e}')
        return 1

    # ---- Step 5/6: Send JSON + prompt to LLM --------------------------------
    output('')
    output(f'Step 5/6: Sending tickets + prompt to LLM...')
    try:
        # Read the prompt file
        if not os.path.isfile(prompt_path):
            output(f'ERROR: Prompt file not found: {prompt_path}')
            return 1
        with open(prompt_path, 'r', encoding='utf-8') as pf:
            prompt_text = pf.read().strip()
        log.info(f'Loaded prompt from {prompt_path} ({len(prompt_text)} chars)')
        output(f'  Prompt: {prompt_path} ({len(prompt_text)} chars)')

        # Pass --model override if the user provided one on the CLI
        model_override = getattr(args, 'model', None)
        response_content, saved_files, token_info = _invoke_llm(
            prompt_text, attachments=[dump_path], timeout=args.timeout,
            model=model_override)

        # Track llm_output.md if it was created
        if os.path.isfile('llm_output.md'):
            all_created_files.append(('llm_output.md', 'LLM full response'))
        for sf in saved_files:
            all_created_files.append((sf, 'LLM extracted'))

    except Exception as e:
        log.error(f'Step 5/6 failed: LLM invocation error: {e}', exc_info=True)
        output(f'ERROR: LLM invocation failed: {e}')
        return 1

    # ---- Step 6/6: Convert CSV files to styled Excel ------------------------
    output('')
    output(f'Step 6/6: Converting CSV output to Excel...')
    csv_files = [sf for sf in saved_files if sf.lower().endswith('.csv')]

    # Canonical CSV name derived from the filter / output basename.
    # We always rename the final CSV to this so the deliverable has a
    # predictable, clean filename regardless of what the LLM chose.
    canonical_csv = f'{dump_basename}.csv'

    # Deduplicate: when the LLM emits multiple CSV blocks (e.g. an original
    # and a "corrected" version), keep only the LAST one — it is typically the
    # most complete / corrected.
    if len(csv_files) > 1:
        keep = csv_files[-1]           # last CSV is the corrected one
        discard = csv_files[:-1]
        log.info(f'LLM emitted {len(csv_files)} CSV files; keeping last: {keep}')
        for d in discard:
            try:
                os.remove(d)
                log.info(f'Removed duplicate CSV: {d}')
            except OSError as rm_err:
                log.warning(f'Could not remove {d}: {rm_err}')
            # Also remove from all_created_files tracking
            all_created_files[:] = [
                (f, desc) for f, desc in all_created_files if f != d]
            # Remove from saved_files so it won't appear later
            saved_files = [sf for sf in saved_files if sf != d]

        csv_files = [keep]
        output(f'  Deduplicated: keeping {keep}')

    # Rename the surviving CSV to the canonical name so the deliverable
    # filename is predictable (e.g. sw_1211_p0p1_bugs.csv) regardless of
    # whatever name the LLM invented for the file.
    if len(csv_files) == 1 and csv_files[0] != canonical_csv:
        keep = csv_files[0]
        try:
            os.rename(keep, canonical_csv)
            log.info(f'Renamed {keep} -> {canonical_csv}')
            # Update tracking lists
            all_created_files[:] = [
                (canonical_csv if f == keep else f, desc)
                for f, desc in all_created_files]
            saved_files = [
                canonical_csv if sf == keep else sf for sf in saved_files]
            csv_files = [canonical_csv]
            output(f'  Renamed: {keep} -> {canonical_csv}')
        except OSError as ren_err:
            log.warning(f'Could not rename {keep} -> {canonical_csv}: {ren_err}')

    if not csv_files:
        output('  No CSV files found in LLM output — skipping Excel conversion.')
        log.info('No CSV files to convert to Excel')
    else:
        # Pass the Jira base URL so ticket-key cells become clickable links
        # in the Excel output (e.g. https://cornelisnetworks.atlassian.net/browse/STL-76582).
        jira_base_url = getattr(jira_utils, 'JIRA_URL', None)
        for csv_path in csv_files:
            try:
                xlsx_path = excel_utils.convert_from_csv(
                    csv_path, jira_base_url=jira_base_url)
                log.info(f'Converted {csv_path} -> {xlsx_path}')
                output(f'  Converted: {csv_path} -> {xlsx_path}')
                all_created_files.append((xlsx_path, 'Excel workbook'))
            except Exception as e:
                log.error(f'Excel conversion failed for {csv_path}: {e}', exc_info=True)
                output(f'  WARNING: Failed to convert {csv_path} to Excel: {e}')

    # ---- Final summary — ==== banner style ----------------------------------
    banner = '=' * 80
    output('')
    output(banner)
    output('WORKFLOW COMPLETE: bug-report')
    output(banner)

    if all_created_files:
        max_name = max(len(f[0]) for f in all_created_files)
        max_name = max(max_name, 4)
        row_fmt = f'{{:<4}}{{:<{max_name + 3}}}{{:<}}'

        file_table_lines = [
            '',
            banner,
            'Workflow Output Files',
            banner,
            row_fmt.format('#', 'File', 'Description'),
        ]
        for idx, (fpath, fdesc) in enumerate(all_created_files, 1):
            file_table_lines.append(row_fmt.format(str(idx), fpath, fdesc))
        file_table_lines.append(banner)

        file_table = '\n'.join(file_table_lines)
        output(file_table)

    output('')
    return 0


def cmd_resume(args):
    '''
    Resume a saved session.
    '''
    log.debug(f'cmd_resume(session={args.resume})')
    
    from state.session import SessionManager
    from state.persistence import get_persistence
    from agents.orchestrator import ReleasePlanningOrchestrator
    
    persistence = get_persistence(args.persistence_format)
    session_manager = SessionManager(persistence=persistence)
    
    session = session_manager.resume_session(args.resume)
    
    if not session:
        output(f'Session not found: {args.session}')
        return 1
    
    output('')
    output('=' * 60)
    output(f'RESUMING SESSION: {session.session_id}')
    output('=' * 60)
    output('')
    output(f'Project: {session.project_key}')
    output(f'Current step: {session.current_step}')
    output(f'Completed steps: {", ".join(session.completed_steps) or "None"}')
    output('')
    
    # Create orchestrator with session state
    orchestrator = ReleasePlanningOrchestrator()
    orchestrator.state.project_key = session.project_key
    orchestrator.state.roadmap_data = session.roadmap_data
    orchestrator.state.jira_state = session.jira_state
    orchestrator.state.release_plan = session.release_plan
    orchestrator.state.current_step = session.current_step
    
    # Determine what to do based on current step
    if session.current_step == 'analysis':
        output('Resuming from analysis step...')
        result = orchestrator._run_planning()
    elif session.current_step == 'planning':
        output('Plan is ready for review.')
        output(orchestrator._format_plan())
        result = None
    elif session.current_step == 'review':
        output('Resuming review...')
        result = orchestrator._run_execution()
    else:
        output(f'Unknown step: {session.current_step}')
        return 1
    
    if result:
        if result.success:
            output(result.content)
        else:
            output(f'ERROR: {result.error}')
            return 1
    
    return 0


# ****************************************************************************************
# Argument handling
# ****************************************************************************************

def handle_args():
    '''
    Parse and validate command line arguments.
    All commands use -- style flags (e.g. --invoke-llm, --build-excel-map).
    '''
    global _quiet_mode
    
    parser = argparse.ArgumentParser(
        description='Cornelis Agent Pipeline - Release Planning Automation',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  %(prog)s --plan --project PROJ --roadmap slides.pptx
  %(prog)s --analyze --project PROJ --quick
  %(prog)s --vision roadmap.png roadmap2.xlsx
  %(prog)s --build-excel-map STL-74071
  %(prog)s --build-excel-map STL-74071 STL-76297 --hierarchy 2 --output my_map.xlsx
  %(prog)s --invoke-llm prompt.md
  %(prog)s --invoke-llm "Summarize this data" --attachments data.csv
  %(prog)s --invoke-llm prompt.md --attachments screenshot.png report.csv
  %(prog)s --sessions --list-sessions
  %(prog)s --resume abc123
        '''
    )
    
    # Global options
    parser.add_argument('-q', '--quiet', action='store_true',
                       help='Suppress output to stdout')
    parser.add_argument('--persistence-format', choices=['json', 'sqlite', 'both'],
                       default='json', help='Session persistence format')
    
    # ---- Command flags (mutually exclusive) ------------------------------------
    # --plan: Run release planning workflow
    parser.add_argument('--plan', action='store_true',
                       help='Run release planning workflow')
    # --analyze: Analyze Jira project
    parser.add_argument('--analyze', action='store_true',
                       help='Analyze Jira project')
    # --vision: Analyze roadmap files (takes 1+ file paths)
    parser.add_argument('--vision', nargs='+', default=None, metavar='FILE',
                       help='Analyze roadmap file(s)')
    # --sessions: Manage saved sessions
    parser.add_argument('--sessions', action='store_true',
                       help='Manage saved sessions')
    # --build-excel-map: Build multi-sheet Excel map (takes 1+ ticket keys)
    parser.add_argument('--build-excel-map', nargs='+', default=None, metavar='TICKET',
                       dest='build_excel_map',
                       help='Build multi-sheet Excel map from root ticket key(s)')
    # --invoke-llm: Send a prompt to the configured LLM (takes prompt text or file path)
    parser.add_argument('--invoke-llm', default=None, metavar='PROMPT',
                       dest='invoke_llm',
                       help='Send a prompt (text or .md/.txt file path) to the configured LLM')
    # --resume: Resume a saved session (takes session ID)
    parser.add_argument('--resume', default=None, metavar='SESSION_ID',
                       help='Resume a saved session by ID')
    # --workflow: Run a named multi-step workflow
    parser.add_argument('--workflow', default=None, metavar='NAME',
                       dest='workflow_name',
                       help='Run a named workflow (e.g. "bug-report")')
    
    # ---- Options for --workflow ------------------------------------------------
    parser.add_argument('--filter', default=None, metavar='NAME',
                       dest='workflow_filter',
                       help='Jira filter name to look up (used by --workflow bug-report)')
    parser.add_argument('--prompt', default=None, metavar='FILE',
                       dest='workflow_prompt',
                       help='Prompt file for LLM step (used by --workflow bug-report, '
                            'default: agents/prompts/cn5000_bugs_clean.md)')
    
    # ---- Options for --plan ----------------------------------------------------
    parser.add_argument('--project', '-p', default=None,
                       help='Jira project key (used by --plan, --analyze, --build-excel-map)')
    parser.add_argument('--roadmap', '-r', action='append', default=None,
                       help='Roadmap file(s) to analyze (used by --plan)')
    parser.add_argument('--org-chart', default=None,
                       help='Organization chart file, draw.io (used by --plan)')
    parser.add_argument('--plan-mode', choices=['full', 'analyze', 'plan'],
                       default='full',
                       help='Workflow mode for --plan (default: full)')
    parser.add_argument('--save-session', action='store_true',
                       help='Save session for later resumption (used by --plan)')
    
    # ---- Options for --analyze -------------------------------------------------
    parser.add_argument('--quick', action='store_true',
                       help='Quick analysis without LLM (used by --analyze)')
    
    # ---- Options for --sessions ------------------------------------------------
    parser.add_argument('--list-sessions', action='store_true',
                       help='List all sessions (used by --sessions)')
    parser.add_argument('--delete-session', default=None, metavar='ID',
                       help='Delete a session by ID (used by --sessions)')
    
    # ---- Options for --build-excel-map -----------------------------------------
    parser.add_argument('--hierarchy', type=int, default=1,
                       help='Depth for related issue traversal (default: 1, used by --build-excel-map)')
    parser.add_argument('--limit', type=int, default=None,
                       help='Max tickets per step (used by --build-excel-map)')
    parser.add_argument('--output', '-o', default=None,
                       help='Output filename (used by --build-excel-map)')
    parser.add_argument('--keep-intermediates', action='store_true',
                       help='Keep temp files instead of cleaning up (used by --build-excel-map)')
    
    # ---- Options for --invoke-llm ----------------------------------------------
    parser.add_argument('--attachments', '-a', nargs='*', default=None,
                       help='File(s) to attach (images via vision API, text inlined; used by --invoke-llm)')
    parser.add_argument('--timeout', type=float, default=None,
                       help='LLM request timeout in seconds (default: 120; used by --invoke-llm)')
    
    # ---- Global LLM model override ---------------------------------------------
    # Allows the user (or a Jenkins Choice Parameter) to select a model at
    # run-time without editing the .env file.  Overrides CORNELIS_LLM_MODEL /
    # OPENAI_MODEL / etc.
    parser.add_argument('--model', '-m', default=None, metavar='MODEL',
                       help='LLM model name override (e.g. developer-opus, gpt-4o). '
                            'Overrides the env-var default for this run.')
    
    # ---- Global verbose flag ---------------------------------------------------
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Enable verbose (debug-level) logging to stdout')
    
    args = parser.parse_args()
    
    if args.quiet:
        _quiet_mode = True
    
    # ---- Verbose mode: add a stdout handler so debug messages appear on console
    if args.verbose:
        sh = logging.StreamHandler(sys.stdout)
        sh.setLevel(logging.DEBUG)
        sh.setFormatter(logging.Formatter(
            '%(asctime)-15s [%(funcName)25s:%(lineno)-5s] %(levelname)-8s %(message)s'))
        log.addHandler(sh)
        log.debug('Verbose mode enabled (--verbose)')
    
    # ---- Determine which command was requested ---------------------------------
    # Count how many command flags were set to enforce mutual exclusivity
    commands = []
    if args.plan:
        commands.append('plan')
    if args.analyze:
        commands.append('analyze')
    if args.vision is not None:
        commands.append('vision')
    if args.sessions:
        commands.append('sessions')
    if args.build_excel_map is not None:
        commands.append('build-excel-map')
    if args.invoke_llm is not None:
        commands.append('invoke-llm')
    if args.resume is not None:
        commands.append('resume')
    if args.workflow_name is not None:
        commands.append('workflow')
    
    if len(commands) == 0:
        parser.print_help()
        sys.exit(1)
    
    if len(commands) > 1:
        parser.error(f'Only one command may be specified at a time. Got: {", ".join("--" + c for c in commands)}')
    
    command = commands[0]
    
    # ---- Command-specific validation -------------------------------------------
    if command == 'plan':
        if not args.project:
            parser.error('--plan requires --project')
    
    if command == 'analyze':
        if not args.project:
            parser.error('--analyze requires --project')
    
    if command == 'workflow':
        if args.workflow_name == 'bug-report':
            if not args.workflow_filter:
                parser.error('--workflow bug-report requires --filter "FILTER_NAME"')
    
    # ---- Map ticket_keys for build-excel-map compatibility ---------------------
    # cmd_build_excel_map expects args.ticket_keys
    if command == 'build-excel-map':
        args.ticket_keys = args.build_excel_map
    
    # ---- Map prompt for invoke-llm compatibility -------------------------------
    # cmd_invoke_llm expects args.prompt
    if command == 'invoke-llm':
        args.prompt = args.invoke_llm
    
    # ---- Map session for resume compatibility ----------------------------------
    # cmd_resume expects args.resume (already set)
    
    # ---- Store resolved command name for dispatch ------------------------------
    args.command = command
    
    # ---- Startup logging box ---------------------------------------------------
    log.info('++++++++++++++++++++++++++++++++++++++++++++++')
    log.info(f'+  {os.path.basename(sys.argv[0])}')
    log.info(f'+  Python Version: {sys.version.split()[0]}')
    log.info(f'+  Today is: {date.today()}')
    log.info(f'+  Command: --{command}')
    # Command-specific details in the startup box
    if command == 'plan':
        log.info(f'+  Project: {args.project}')
        log.info(f'+  Plan mode: {args.plan_mode}')
        if args.roadmap:
            log.info(f'+  Roadmap files: {len(args.roadmap)}')
        if args.org_chart:
            log.info(f'+  Org chart: {args.org_chart}')
    elif command == 'analyze':
        log.info(f'+  Project: {args.project}')
        if args.quick:
            log.info(f'+  Quick mode: yes')
    elif command == 'vision':
        log.info(f'+  Files: {len(args.vision)}')
        for vf in args.vision:
            log.info(f'+    - {vf}')
    elif command == 'sessions':
        if args.list_sessions:
            log.info(f'+  Action: list')
        elif args.delete_session:
            log.info(f'+  Action: delete {args.delete_session}')
    elif command == 'build-excel-map':
        keys_str = ', '.join(k.upper() for k in args.ticket_keys)
        log.info(f'+  Root ticket(s): {keys_str}')
        log.info(f'+  Hierarchy depth: {args.hierarchy}')
        default_out = f'{args.ticket_keys[0].upper()}.xlsx' if len(args.ticket_keys) == 1 else f'{"_".join(k.upper() for k in args.ticket_keys)}.xlsx'
        log.info(f'+  Output file: {args.output or default_out}')
    elif command == 'invoke-llm':
        prompt_display = args.prompt if len(args.prompt) <= 60 else args.prompt[:57] + '...'
        log.info(f'+  Prompt: {prompt_display}')
        if args.attachments:
            log.info(f'+  Attachments: {len(args.attachments)} file(s)')
            for att in args.attachments:
                log.info(f'+    - {att}')
        if args.timeout:
            log.info(f'+  Timeout: {args.timeout}s')
    elif command == 'resume':
        log.info(f'+  Session: {args.resume}')
    elif command == 'workflow':
        log.info(f'+  Workflow: {args.workflow_name}')
        if args.workflow_filter:
            log.info(f'+  Filter: {args.workflow_filter}')
        if args.workflow_prompt:
            log.info(f'+  Prompt: {args.workflow_prompt}')
        if args.timeout:
            log.info(f'+  Timeout: {args.timeout}s')
        if args.limit:
            log.info(f'+  Limit: {args.limit}')
    log.info('++++++++++++++++++++++++++++++++++++++++++++++')
    
    return args


# ****************************************************************************************
# Main
# ****************************************************************************************

def main():
    '''
    Entrypoint for the CLI.
    Dispatches to the appropriate cmd_* handler based on args.command.
    '''
    args = handle_args()
    log.debug('Entering main()')
    
    # Dispatch table: command name -> handler function
    dispatch = {
        'plan':            cmd_plan,
        'analyze':         cmd_analyze,
        'vision':          cmd_vision,
        'sessions':        cmd_sessions,
        'build-excel-map': cmd_build_excel_map,
        'invoke-llm':      cmd_invoke_llm,
        'resume':          cmd_resume,
        'workflow':         cmd_workflow,
    }
    
    handler = dispatch.get(args.command)
    if not handler:
        output(f'ERROR: Unknown command: {args.command}')
        sys.exit(1)
    
    try:
        exit_code = handler(args)
        
    except KeyboardInterrupt:
        output('\nOperation cancelled.')
        exit_code = 130
        
    except Exception as e:
        log.error(f'Unexpected error: {e}', exc_info=True)
        output(f'ERROR: {e}')
        exit_code = 1
    
    log.info('Operation complete.')
    sys.exit(exit_code)


if __name__ == '__main__':
    main()
